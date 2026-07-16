import csv
import io
import os
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional, Tuple

# pylint: disable=too-many-locals,too-many-branches,too-many-arguments

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook

from comedores.models import (
    Comedor,
    EstadoActividad,
    EstadoDetalle,
    EstadoHistorial,
    EstadoProceso,
)
from comedores.services.estado_manager import registrar_cambio_estado
from expedientespagos.models import ExpedientePago
from importarexpediente.models import ArchivosImportados, RegistroImportado

# Formatos de fecha aceptados
DATE_FORMATS = ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]
ALIMENTAR_COMUNIDAD_PROGRAMA_ID = 2
ACTIVO = "Activo"
INACTIVO = "Inactivo"
EN_EJECUCION = "En ejecuci\u00f3n"
EN_PROCESO_RENOVACION = "En proceso - Renovaci\u00f3n"
BAJA = "Baja"
EN_PLAZO_RENOVACION = "En plazo de renovaci\u00f3n"
NO_RENOVACION_COMEDOR = "No renovaci\u00f3n (Comedor)"

# Mapeo de cabeceras posibles -> campo de modelo
HEADER_MAP = {
    # Expedientes
    "expediente de pago": "expediente_pago",
    # El modelo usa 'expediente_convenio' en lugar de 'resolucion_pago'
    "expediente del convenio": "expediente_convenio",
    # Comedor identificacion: SIEMPRE por ID del CSV/XLSX
    "anexo": "anexo",
    "id": "comedor_id",
    # Fechas
    "fecha de pago al banco": "fecha_pago_al_banco",
    "fecha de acreditacion": "fecha_acreditacion",
    "fecha de acreditaci\u00f3n": "fecha_acreditacion",
    # Organizacion creadora (variantes con y sin acento)
    "organizaci\u00f3n": "organizacion_creacion",
    "organizacion": "organizacion_creacion",
    # Total
    "total prestaciones": "total_prestaciones",
    "gastos accesorios 6%": "gastos_accesorios",
    "total": "total",
    # Mes y anio
    "mes de pago": "mes_pago",
    "mes de convenio": "mes_convenio",
    "a\u00f1o": "ano",
    "a\u00f1o de pago": "ano",
    # Prestaciones mensuales
    "prestaciones mensuales desayuno": "prestaciones_mensuales_desayuno",
    "prestaciones mensuales almuerzo": "prestaciones_mensuales_almuerzo",
    "prestaciones mensuales merienda": "prestaciones_mensuales_merienda",
    "prestaciones mensuales cena": "prestaciones_mensuales_cena",
    # Montos mensuales (singular)
    "monto mensual desayuno": "monto_mensual_desayuno",
    "monto mensual almuerzo": "monto_mensual_almuerzo",
    "monto mensual merienda": "monto_mensual_merienda",
    "monto mensual cena": "monto_mensual_cena",
    # Montos mensuales (plural variantes del CSV)
    "monto mensuales desayuno": "monto_mensual_desayuno",
    "monto mensuales almuerzo": "monto_mensual_almuerzo",
    "monto mensuales merienda": "monto_mensual_merienda",
    "monto mensuales cena": "monto_mensual_cena",
}

# Etiquetas amigables para usuarios no tecnicos
FIELD_LABELS = {
    "expediente_pago": "Expediente de pago",
    "expediente_convenio": "Expediente del convenio",
    "anexo": "Comedor (anexo)",
    "organizacion_creacion": "Organizaci\u00f3n",
    "total_prestaciones": "Total Prestaciones",
    "gastos_accesorios": "Gastos Accesorios 6%",
    "total": "Total",
    "mes_pago": "Mes de pago",
    "mes_convenio": "Mes de convenio",
    "ano": "A\u00f1o",
    "comedor_id": "ID",
    "fecha_pago_al_banco": "Fecha de pago al banco",
    "fecha_acreditacion": "Fecha de acreditaci\u00f3n",
    "prestaciones_mensuales_desayuno": "Prestaciones mensuales desayuno",
    "prestaciones_mensuales_almuerzo": "Prestaciones mensuales almuerzo",
    "prestaciones_mensuales_merienda": "Prestaciones mensuales merienda",
    "prestaciones_mensuales_cena": "Prestaciones mensuales cena",
    "monto_mensual_desayuno": "Monto mensual desayuno",
    "monto_mensual_almuerzo": "Monto mensual almuerzo",
    "monto_mensual_merienda": "Monto mensual merienda",
    "monto_mensual_cena": "Monto mensual cena",
}


@dataclass
class ParsedImportFile:
    headers: List[str]
    mapped_headers: List[Optional[str]]
    rows: List[Tuple[int, List]]
    detected_delimiter: Optional[str]
    file_format: str


class EmptyImportFileError(ValueError):
    pass


class HeaderlessImportFileError(ValueError):
    pass


@dataclass
class AcreditacionImportResult:
    filas_procesadas: int
    comedores_actualizados: int
    expedientes_actualizados: int


def _cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value):
    return _cell_text(value).replace('"', "").replace("'", "").lower()


def _file_extension(filename):
    return os.path.splitext(filename or "")[1].lower()


def _read_csv(data, delimiter):
    try:
        decoded = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        decoded = data.decode("latin-1")

    first_line = decoded.splitlines()[0] if decoded else ""
    detected_delimiter = ";" if ";" in first_line else ","
    reader = csv.reader(io.StringIO(decoded), delimiter=detected_delimiter or delimiter)
    return reader, detected_delimiter


def _read_xlsx(data):
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    worksheet = workbook.active
    return worksheet.iter_rows(values_only=True)


def parse_import_file(  # pylint: disable=too-many-locals
    data, filename, delimiter=",", has_header=True
):
    extension = _file_extension(filename)
    file_format = "xlsx" if extension == ".xlsx" else "csv"

    if file_format == "xlsx":
        raw_rows = _read_xlsx(data)
        detected_delimiter = None
    else:
        raw_rows, detected_delimiter = _read_csv(data, delimiter)

    try:
        first_row = next(raw_rows)
    except StopIteration as exc:
        raise EmptyImportFileError("El archivo est\u00e1 vac\u00edo.") from exc

    if not has_header:
        raise HeaderlessImportFileError(
            "Archivo sin cabecera no soportado por defecto."
        )

    headers = [normalize_header(cell) for cell in first_row]
    mapped_headers = [HEADER_MAP.get(header) for header in headers]
    expected_cols = len(headers)
    rows = []
    for row_number, row in enumerate(raw_rows, start=2):
        values = list(row)
        if file_format == "xlsx" and not any(_cell_text(value) for value in values):
            continue
        if len(values) < expected_cols:
            values += [""] * (expected_cols - len(values))
        elif len(values) > expected_cols:
            values = values[:expected_cols]
        rows.append((row_number, values))

    return ParsedImportFile(
        headers=headers,
        mapped_headers=mapped_headers,
        rows=rows,
        detected_delimiter=detected_delimiter,
        file_format=file_format,
    )


def parse_date(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def parse_decimal(value):
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    s = str(value).replace("$", "").replace(" ", "").replace("\u00a0", "").strip()
    if "." in s and "," in s:
        if s.rfind(".") > s.rfind(","):
            s = s.replace(",", "")
        else:
            s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        integer_part, decimal_part = s.rsplit(".", 1)
        if decimal_part.isdigit() and len(decimal_part) in (1, 2):
            s = integer_part.replace(".", "") + "." + decimal_part
        else:
            s = s.replace(".", "")
    try:
        return Decimal(s)
    except Exception:
        return None


def parse_int(value):
    if value is None or value == "":
        return None
    s = str(value).replace(".", "").replace(",", "").replace(" ", "")
    s = s.strip()
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None


def _column_warning(label, warning):
    return f'Error en validaci\u00f3n columna "{label}": ' f'Advertencia "{warning}"'


def _field_warning(field, warning):
    return _column_warning(FIELD_LABELS.get(field, field), warning)


def _required_column_indexes(parsed_file, required_fields):
    indexes = {}
    for index, field in enumerate(parsed_file.mapped_headers):
        if field in required_fields and field not in indexes:
            indexes[field] = index

    missing = [
        FIELD_LABELS.get(field, field)
        for field in required_fields
        if field not in indexes
    ]
    if missing:
        raise ValidationError(
            "El archivo debe incluir las columnas: " + ", ".join(missing)
        )
    return indexes


def _parse_acreditacion_updates(parsed_file):
    indexes = _required_column_indexes(
        parsed_file,
        ("comedor_id", "fecha_acreditacion"),
    )
    updates_by_comedor = {}
    errors = []

    for row_number, row in parsed_file.rows:
        raw_id = row[indexes["comedor_id"]]
        raw_fecha = row[indexes["fecha_acreditacion"]]
        id_text = _cell_text(raw_id)
        fecha_text = _cell_text(raw_fecha)

        if not id_text and not fecha_text:
            continue

        comedor_id = parse_int(raw_id)
        if comedor_id is None:
            errors.append(f"Fila {row_number}: ID debe ser num\u00e9rico.")
            continue

        fecha_acreditacion = parse_date(raw_fecha)
        if not fecha_text or fecha_acreditacion is None:
            errors.append(
                f"Fila {row_number}: Fecha de acreditaci\u00f3n debe tener formato DD/MM/AAAA."
            )
            continue

        previous_date = updates_by_comedor.get(comedor_id)
        if previous_date and previous_date != fecha_acreditacion:
            errors.append(
                f"Fila {row_number}: ID {comedor_id} aparece repetido con fechas distintas."
            )
            continue

        updates_by_comedor[comedor_id] = fecha_acreditacion

    if errors:
        raise ValidationError(errors)
    if not updates_by_comedor:
        raise ValidationError(
            "No se encontraron fechas de acreditaci\u00f3n para procesar."
        )
    if len(set(updates_by_comedor.values())) != 1:
        raise ValidationError(
            "El archivo debe informar una \u00fanica fecha de acreditaci\u00f3n "
            "para el lote."
        )
    return updates_by_comedor


def actualizar_fechas_acreditacion_por_lote(
    batch,
    data,
    filename,
    delimiter=",",
):
    if not getattr(batch, "importacion_completada", False):
        raise ValidationError("Primero se debe completar la importaci\u00f3n del lote.")

    parsed_file = parse_import_file(
        data,
        filename,
        delimiter,
        has_header=True,
    )
    updates_by_comedor = _parse_acreditacion_updates(parsed_file)

    expediente_rows = list(
        ExpedientePago.objects.filter(
            registros_importados__exito_importacion__archivo_importado=batch,
        ).values_list("id", "comedor_id")
    )
    expediente_ids_by_comedor = {}
    for expediente_id, comedor_id in expediente_rows:
        expediente_ids_by_comedor.setdefault(comedor_id, set()).add(expediente_id)

    matched_comedor_ids = set(expediente_ids_by_comedor)
    missing_ids = sorted(set(updates_by_comedor) - matched_comedor_ids)
    if missing_ids:
        missing = ", ".join(str(comedor_id) for comedor_id in missing_ids)
        raise ValidationError(
            "Los siguientes ID no corresponden a expedientes importados en este lote: "
            + missing
        )

    fecha_acreditacion = next(iter(set(updates_by_comedor.values())))
    with transaction.atomic():
        expedientes_actualizados = (
            ExpedientePago.objects.select_for_update()
            .filter(id__in={expediente_id for expediente_id, _ in expediente_rows})
            .update(fecha_acreditacion=fecha_acreditacion)
        )

    return AcreditacionImportResult(
        filas_procesadas=len(updates_by_comedor),
        comedores_actualizados=len(expediente_ids_by_comedor),
        expedientes_actualizados=expedientes_actualizados,
    )


def expediente_pago_from_row(parsed_file, row):  # pylint: disable=too-many-branches
    kwargs = {}
    specific_errors = []
    for col_idx, cell in enumerate(row):
        field = parsed_file.mapped_headers[col_idx]
        if not field:
            continue
        val = _cell_text(cell)
        if field in ("total", "total_prestaciones", "gastos_accesorios"):
            parsed = parse_decimal(cell)
            if val and parsed is None:
                specific_errors.append(
                    _field_warning(field, "El campo debe ser num\u00e9rico")
                )
            kwargs[field] = parsed
        elif field.startswith("monto_mensual_"):
            parsed = parse_decimal(cell)
            if val and parsed is None:
                specific_errors.append(
                    _field_warning(field, "El campo debe ser num\u00e9rico")
                )
            kwargs[field] = parsed
        elif field in ("fecha_pago_al_banco", "fecha_acreditacion"):
            kwargs[field] = parse_date(val)
        elif field == "comedor_id":
            parsed_id = parse_int(val)
            if val and parsed_id is None:
                specific_errors.append(
                    _column_warning("ID", "El campo debe ser num\u00e9rico")
                )
            kwargs["comedor_id"] = parsed_id
        elif field.startswith("prestaciones_mensuales_"):
            parsed_int = parse_int(val)
            if val and parsed_int is None:
                specific_errors.append(
                    _field_warning(field, "El campo debe ser num\u00e9rico")
                )
            kwargs[field] = parsed_int
        elif field == "mes_convenio":
            parsed_int = parse_int(val)
            if parsed_int not in range(1, 7):
                specific_errors.append(
                    _field_warning(field, "Debe ser un n\u00famero entre 1 y 6")
                )
            kwargs[field] = parsed_int
        elif field == "ano":
            val_digits = val.replace(" ", "")
            if val and not val_digits.isdigit():
                specific_errors.append(
                    _field_warning(field, "El campo debe ser num\u00e9rico")
                )
            kwargs[field] = val or None
        else:
            kwargs[field] = val or None
    return kwargs, specific_errors


def extract_numero_expediente_pago(parsed_file, row):
    for col_idx, field in enumerate(parsed_file.mapped_headers):
        if field == "expediente_pago":
            value = _cell_text(row[col_idx])
            if value:
                return value
    return None


def extract_periodo_pago(parsed_file, row):
    mes_pago = None
    ano_pago = None
    for col_idx, field in enumerate(parsed_file.mapped_headers):
        value = _cell_text(row[col_idx])
        if not value:
            continue
        if field == "mes_pago" and mes_pago is None:
            mes_pago = value
        elif field == "ano" and ano_pago is None:
            ano_pago = value
    return mes_pago, ano_pago


def ensure_import_required_defaults(kwargs):
    for field in (
        "prestaciones_mensuales_desayuno",
        "prestaciones_mensuales_almuerzo",
        "prestaciones_mensuales_merienda",
        "prestaciones_mensuales_cena",
    ):
        kwargs.setdefault(field, 0)
    for field in (
        "monto_mensual_desayuno",
        "monto_mensual_almuerzo",
        "monto_mensual_merienda",
        "monto_mensual_cena",
    ):
        kwargs.setdefault(field, Decimal("0"))


def _estado_catalogo():
    activo = EstadoActividad.objects.get(estado=ACTIVO)
    inactivo = EstadoActividad.objects.get(estado=INACTIVO)
    en_ejecucion = EstadoProceso.objects.get(
        estado=EN_EJECUCION,
        estado_actividad=activo,
    )
    en_proceso_renovacion = EstadoProceso.objects.get(
        estado=EN_PROCESO_RENOVACION,
        estado_actividad=activo,
    )
    baja = EstadoProceso.objects.get(estado=BAJA, estado_actividad=inactivo)
    en_plazo_renovacion = EstadoDetalle.objects.get(
        estado=EN_PLAZO_RENOVACION,
        estado_proceso=en_ejecucion,
    )
    no_renovacion_comedor = EstadoDetalle.objects.get(
        estado=NO_RENOVACION_COMEDOR,
        estado_proceso=baja,
    )
    return {
        "activo": activo,
        "inactivo": inactivo,
        "en_ejecucion": en_ejecucion,
        "en_proceso_renovacion": en_proceso_renovacion,
        "baja": baja,
        "en_plazo_renovacion": en_plazo_renovacion,
        "no_renovacion_comedor": no_renovacion_comedor,
    }


def _new_imported_batches_before(batch):
    return list(
        ArchivosImportados.objects.filter(
            Q(fecha_subida__lt=batch.fecha_subida)
            | Q(fecha_subida=batch.fecha_subida, id__lt=batch.id),
            importacion_completada=True,
            exitos__registros__expediente_pago__mes_convenio__isnull=False,
        )
        .exclude(pk=batch.pk)
        .distinct()
        .order_by("-fecha_subida", "-id")
        .values_list("id", "fecha_subida")
    )


def _previous_presence_by_batch(batch_ids):
    presence_by_batch = {batch_id: set() for batch_id in batch_ids}
    rows = (
        RegistroImportado.objects.filter(
            exito_importacion__archivo_importado_id__in=batch_ids,
            expediente_pago__mes_convenio__isnull=False,
            expediente_pago__comedor__programa_id=ALIMENTAR_COMUNIDAD_PROGRAMA_ID,
        )
        .values_list(
            "exito_importacion__archivo_importado_id",
            "expediente_pago__comedor_id",
        )
        .distinct()
    )
    for batch_id, comedor_id in rows:
        presence_by_batch.setdefault(batch_id, set()).add(comedor_id)
    return presence_by_batch


def _batch_ids_since(previous_batches, since):
    if since is None:
        return [batch_id for batch_id, _fecha_subida in previous_batches]
    return [
        batch_id for batch_id, fecha_subida in previous_batches if fecha_subida >= since
    ]


def _latest_execution_state_by_comedor(comedor_ids):
    latest_by_comedor = {}
    rows = (
        EstadoHistorial.objects.filter(
            comedor_id__in=comedor_ids,
            estado_general__estado_actividad__estado=ACTIVO,
            estado_general__estado_proceso__estado=EN_EJECUCION,
        )
        .order_by("comedor_id", "-fecha_cambio", "-id")
        .values_list("comedor_id", "fecha_cambio")
    )
    for comedor_id, fecha_cambio in rows:
        latest_by_comedor.setdefault(comedor_id, fecha_cambio)
    return latest_by_comedor


def _inicio_conteo_ausencias(comedor, latest_execution_at):
    fechas = []
    if getattr(comedor, "fecha_creacion", None):
        fechas.append(comedor.fecha_creacion)
    if latest_execution_at:
        fechas.append(latest_execution_at)
    if not fechas:
        return None
    return max(fechas)


def _ausencias_consecutivas(comedor_id, previous_batch_ids, presence_by_batch):
    count = 1
    for batch_id in previous_batch_ids:
        if comedor_id in presence_by_batch.get(batch_id, set()):
            break
        count += 1
    return count


def _registrar_estado(  # pylint: disable=too-many-arguments
    comedor, catalogo, actividad_key, proceso_key, detalle_key, usuario
):
    detalle = catalogo[detalle_key] if detalle_key else None
    registrar_cambio_estado(
        comedor=comedor,
        actividad=catalogo[actividad_key],
        proceso=catalogo[proceso_key],
        detalle=detalle,
        usuario=usuario,
    )


def aplicar_estados_por_lote(batch, usuario=None):  # pylint: disable=too-many-locals
    registros_actuales = list(
        RegistroImportado.objects.filter(
            exito_importacion__archivo_importado=batch,
            expediente_pago__mes_convenio__isnull=False,
            expediente_pago__comedor__programa_id=ALIMENTAR_COMUNIDAD_PROGRAMA_ID,
        )
        .select_related("expediente_pago__comedor")
        .order_by("id")
    )
    if not registros_actuales:
        return 0

    catalogo = _estado_catalogo()
    present_mes_by_comedor = {}
    for registro in registros_actuales:
        expediente = registro.expediente_pago
        present_mes_by_comedor.setdefault(
            expediente.comedor_id,
            expediente.mes_convenio,
        )

    present_ids = set(present_mes_by_comedor)
    updated_count = 0
    comedores_presentes = Comedor.objects.filter(id__in=present_ids).select_related(
        "ultimo_estado__estado_general__estado_actividad"
    )
    for comedor in comedores_presentes:
        mes_convenio = present_mes_by_comedor[comedor.id]
        if mes_convenio in (1, 2, 3):
            _registrar_estado(
                comedor,
                catalogo,
                "activo",
                "en_ejecucion",
                None,
                usuario,
            )
        else:
            _registrar_estado(
                comedor,
                catalogo,
                "activo",
                "en_ejecucion",
                "en_plazo_renovacion",
                usuario,
            )
        updated_count += 1

    previous_batches = _new_imported_batches_before(batch)
    previous_batch_ids = [batch_id for batch_id, _fecha_subida in previous_batches]
    presence_by_batch = _previous_presence_by_batch(previous_batch_ids)
    comedores_ausentes = list(
        Comedor.objects.filter(programa_id=ALIMENTAR_COMUNIDAD_PROGRAMA_ID)
        .exclude(id__in=present_ids)
        .select_related("ultimo_estado__estado_general__estado_actividad")
    )
    latest_execution_by_comedor = _latest_execution_state_by_comedor(
        [comedor.id for comedor in comedores_ausentes]
    )
    for comedor in comedores_ausentes:
        ultimo_estado = getattr(comedor, "ultimo_estado", None)
        actividad_actual = None
        if ultimo_estado and ultimo_estado.estado_general_id:
            actividad_actual = ultimo_estado.estado_general.estado_actividad.estado

        inicio_conteo = _inicio_conteo_ausencias(
            comedor,
            latest_execution_by_comedor.get(comedor.id),
        )
        previous_batch_ids_for_comedor = _batch_ids_since(
            previous_batches,
            inicio_conteo,
        )
        ausencias = _ausencias_consecutivas(
            comedor.id,
            previous_batch_ids_for_comedor,
            presence_by_batch,
        )
        if actividad_actual == INACTIVO or ausencias >= 3:
            _registrar_estado(
                comedor,
                catalogo,
                "inactivo",
                "baja",
                "no_renovacion_comedor",
                usuario,
            )
        else:
            _registrar_estado(
                comedor,
                catalogo,
                "activo",
                "en_proceso_renovacion",
                None,
                usuario,
            )
        updated_count += 1

    return updated_count


def friendly_error_message(exc: Exception) -> str:
    """Convierte errores tecnicos en mensajes claros para usuarios finales."""
    if isinstance(exc, ValidationError):
        if hasattr(exc, "message_dict") and exc.message_dict:
            partes = []
            for campo, mensajes in exc.message_dict.items():
                etiqueta = FIELD_LABELS.get(campo, campo)
                detalle = "; ".join(str(m) for m in mensajes)
                partes.append(f"{etiqueta}: {detalle}")
            return ". ".join(partes)
        if hasattr(exc, "messages"):
            return "; ".join(str(m) for m in exc.messages)
    return (
        "No se pudo procesar la fila. Verifica que las fechas tengan formato DD/MM/AAAA, "
        "los montos sean numericos y el comedor exista. Detalle: " + str(exc)
    )
