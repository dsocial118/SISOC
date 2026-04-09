from io import BytesIO
import smtplib
from datetime import timedelta

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.contrib.contenttypes.models import ContentType
from django.contrib.messages.storage.fallback import FallbackStorage
from django.core.management import call_command
from django.core import mail
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import RequestFactory, override_settings
from django.contrib.sessions.middleware import SessionMiddleware
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework.authtoken.models import Token

from users.forms import BulkCredentialsUploadForm
from users.models import BulkCredentialsJob, BulkCredentialsJobRow
from users.services_bulk_credentials import (
    BulkCredentialsEmailTimeoutError,
    process_bulk_credentials_file,
)
from users.services_bulk_credentials_jobs import (
    STALE_JOB_ERROR_MESSAGE,
    can_resume_bulk_credentials_job,
    create_bulk_credentials_job,
    mark_stale_bulk_credentials_jobs_as_failed,
    process_bulk_credentials_job,
    request_resume_bulk_credentials_job,
    run_bulk_credentials_jobs_worker,
)
from users.services import UsuariosService
from users.views import (
    BulkCredentialsJobDetailView,
    BulkCredentialsJobResumeView,
    BulkCredentialsTemplateView,
    BulkCredentialsUploadView,
)

User = get_user_model()


def _build_excel_file(rows, headers=("usuario", "mail", "password")):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "credenciales"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))

    output = BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        "credenciales.xlsx",
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


def _bulk_credentials_permission():
    content_type = ContentType.objects.get_for_model(Group)
    permission, _ = Permission.objects.get_or_create(
        content_type=content_type,
        codename="role_enviar_credenciales_masivas",
        defaults={"name": "Enviar credenciales masivas"},
    )
    return permission


def _build_request(method: str, path: str, user, *, data=None, files=None):
    factory = RequestFactory()
    request_method = getattr(factory, method.lower())
    request = request_method(path, data=data or {})
    SessionMiddleware(lambda req: None).process_request(request)
    request.session.save()
    setattr(request, "_messages", FallbackStorage(request))
    request.user = user
    if files:
        request.FILES.update(files)
    return request


def _grant_bulk_credentials_permissions(user):
    change_user_permission = Permission.objects.get(
        content_type__app_label="auth",
        codename="change_user",
    )
    user.user_permissions.add(change_user_permission, _bulk_credentials_permission())
    return user


@pytest.mark.django_db
def test_bulk_credentials_template_download_requires_both_permissions():
    user = User.objects.create_user(
        username="users_bulk_reader",
        email="reader@example.com",
        password="Secreta123!",
    )
    change_user_permission = Permission.objects.get(
        content_type__app_label="auth",
        codename="change_user",
    )
    user.user_permissions.add(change_user_permission)

    with pytest.raises(PermissionDenied):
        BulkCredentialsTemplateView.as_view()(
            _build_request("get", "/usuarios/credenciales-masivas/plantilla/", user),
        )

    user.user_permissions.add(_bulk_credentials_permission())
    user = User.objects.get(pk=user.pk)
    response = BulkCredentialsTemplateView.as_view()(
        _build_request("get", "/usuarios/credenciales-masivas/plantilla/", user),
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    assert header == ["usuario", "mail", "password"]


@pytest.mark.django_db
def test_bulk_credentials_template_download_supports_inet_template():
    user = User.objects.create_superuser(
        username="users_bulk_inet_template",
        email="inet-template@example.com",
        password="Secreta123!",
    )

    response = BulkCredentialsTemplateView.as_view()(
        _build_request(
            "get",
            "/usuarios/credenciales-masivas/plantilla/",
            user,
            data={"tipo_envio": "inet"},
        ),
    )

    assert response.status_code == 200
    assert (
        response["Content-Disposition"]
        == 'attachment; filename="plantilla_credenciales_usuarios_inet.xlsx"'
    )
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    assert header == ["usuario", "mail", "password", "Nombre del Centro"]


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_users_bulk_credentials")
def test_user_list_context_shows_bulk_credentials_button_only_with_role_permission(
    mocker,
):
    user = User.objects.create_user(
        username="users_bulk_button",
        email="button@example.com",
        password="Secreta123!",
    )
    mocker.patch(
        "users.services.build_columns_context",
        return_value={"table_headers": [], "table_fields": []},
    )
    view_user_permission = Permission.objects.get(
        content_type__app_label="auth",
        codename="view_user",
    )
    change_user_permission = Permission.objects.get(
        content_type__app_label="auth",
        codename="change_user",
    )
    user.user_permissions.add(view_user_permission, change_user_permission)

    request = _build_request("get", "/usuarios/", user)
    context = UsuariosService.get_usuarios_list_context(request)
    assert context["additional_buttons"] == []

    user.user_permissions.add(_bulk_credentials_permission())
    user = User.objects.get(pk=user.pk)
    request = _build_request("get", "/usuarios/", user)
    context = UsuariosService.get_usuarios_list_context(request)
    assert context["additional_buttons"] == [
        {
            "label": "ENVIO DE CREDENCIALES",
            "url": "/usuarios/credenciales-masivas/",
            "class": "btn btn-lg btn-export-csv",
            "title": "Actualizar password y enviar credenciales desde Excel",
        }
    ]


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_updates_password_and_sends_to_mail_from_excel():
    user = User.objects.create_user(
        username="bulk_target",
        email="viejo@example.com",
        password="Vieja123!",
    )
    Token.objects.create(user=user)

    upload = _build_excel_file(
        [("bulk_target", "nuevo@example.com", "Nueva123!")],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    user.refresh_from_db()
    profile = user.profile
    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 1,
        "actualizadas": 1,
        "sin_cambios": 0,
        "rechazadas": 0,
    }
    assert result["rows"][0]["estado"] == "enviada"
    assert result["rows"][0]["mail_destino"] == "nuevo@example.com"
    assert result["rows"][0]["password_actualizada"] is True
    assert user.email == "viejo@example.com"
    assert user.check_password("Nueva123!") is True
    assert profile.must_change_password is True
    assert profile.password_changed_at is None
    assert profile.password_reset_requested_at is None
    assert profile.temporary_password_plaintext == "Nueva123!"
    assert Token.objects.filter(user=user).exists() is False
    assert len(mail.outbox) == 1
    assert mail.outbox[0].to == ["nuevo@example.com"]


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_same_data_sends_email_without_updating():
    user = User.objects.create_user(
        username="bulk_same",
        email="same@example.com",
        password="Misma123!",
    )

    upload = _build_excel_file(
        [("bulk_same", "same@example.com", "Misma123!")],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    user.refresh_from_db()
    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 1,
        "actualizadas": 0,
        "sin_cambios": 1,
        "rechazadas": 0,
    }
    assert result["rows"][0]["mail_destino"] == "same@example.com"
    assert result["rows"][0]["password_actualizada"] is False
    assert user.email == "same@example.com"
    assert user.check_password("Misma123!") is True
    assert len(mail.outbox) == 1


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_rejects_unknown_user_and_continues():
    known_user = User.objects.create_user(
        username="bulk_known",
        email="known@example.com",
        password="Inicial123!",
    )
    upload = _build_excel_file(
        [
            ("missing_user", "missing@example.com", "Temporal123!"),
            ("bulk_known", "known@example.com", "Inicial123!"),
        ],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    known_user.refresh_from_db()
    assert result["summary"] == {
        "procesadas": 2,
        "enviadas": 1,
        "actualizadas": 0,
        "sin_cambios": 1,
        "rechazadas": 1,
    }
    assert result["rows"][0]["estado"] == "rechazada"
    assert "No existe un usuario" in result["rows"][0]["mensaje"]
    assert result["rows"][1]["estado"] == "enviada"
    assert known_user.email == "known@example.com"
    assert len(mail.outbox) == 1


@pytest.mark.django_db
def test_process_bulk_credentials_rejects_missing_required_headers():
    upload = _build_excel_file(
        [("bulk_user", "user@example.com", "Temporal123!")],
        headers=("usuario", "mail"),
    )

    with pytest.raises(ValidationError) as exc:
        process_bulk_credentials_file(uploaded_file=upload, send_type="standard")

    assert "columnas obligatorias" in " ".join(exc.value.messages)


@pytest.mark.django_db
def test_process_bulk_credentials_rejects_missing_inet_center_column():
    upload = _build_excel_file(
        [("bulk_user", "user@example.com", "Temporal123!")],
        headers=("usuario", "mail", "password"),
    )

    with pytest.raises(ValidationError) as exc:
        process_bulk_credentials_file(uploaded_file=upload, send_type="inet")

    assert "nombre_del_centro" in " ".join(exc.value.messages)


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_reports_missing_mail_and_continues():
    User.objects.create_user(
        username="bulk_invalid",
        email="bulk_invalid@example.com",
        password="Temporal123!",
    )
    User.objects.create_user(
        username="bulk_valid",
        email="bulk_valid@example.com",
        password="Temporal123!",
    )
    upload = _build_excel_file(
        [
            ("bulk_invalid", "", "Temporal123!"),
            ("bulk_valid", "destino@example.com", "Temporal123!"),
        ],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    assert result["summary"] == {
        "procesadas": 2,
        "enviadas": 1,
        "actualizadas": 0,
        "sin_cambios": 1,
        "rechazadas": 1,
    }
    assert "mail es obligatoria" in result["rows"][0]["mensaje"]
    assert result["rows"][0]["mail_destino"] == ""
    assert result["rows"][1]["estado"] == "enviada"
    assert result["rows"][1]["mail_destino"] == "destino@example.com"
    assert len(mail.outbox) == 1


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_allows_shared_recipient_email():
    first_user = User.objects.create_user(
        username="bulk_shared_one",
        email="one@example.com",
        password="Inicial123!",
    )
    second_user = User.objects.create_user(
        username="bulk_shared_two",
        email="two@example.com",
        password="Inicial123!",
    )
    upload = _build_excel_file(
        [
            ("bulk_shared_one", "shared@example.com", "NuevaOne123!"),
            ("bulk_shared_two", "shared@example.com", "NuevaTwo123!"),
        ],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    first_user.refresh_from_db()
    second_user.refresh_from_db()
    assert result["summary"] == {
        "procesadas": 2,
        "enviadas": 2,
        "actualizadas": 2,
        "sin_cambios": 0,
        "rechazadas": 0,
    }
    assert first_user.email == "one@example.com"
    assert second_user.email == "two@example.com"
    assert first_user.check_password("NuevaOne123!") is True
    assert second_user.check_password("NuevaTwo123!") is True
    assert [email.to for email in mail.outbox] == [
        ["shared@example.com"],
        ["shared@example.com"],
    ]


@pytest.mark.django_db
def test_process_bulk_credentials_rolls_back_row_when_email_send_fails(mocker):
    user = User.objects.create_user(
        username="bulk_rollback",
        email="rollback@example.com",
        password="Inicial123!",
    )
    Token.objects.create(user=user)
    original_password = user.password
    mocker.patch(
        "users.services_bulk_credentials.send_bulk_credentials_email",
        side_effect=RuntimeError("smtp down"),
    )
    upload = _build_excel_file(
        [("bulk_rollback", "nuevo_rollback@example.com", "TemporalNueva123!")],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    user.refresh_from_db()
    profile = user.profile
    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 0,
        "actualizadas": 0,
        "sin_cambios": 0,
        "rechazadas": 1,
    }
    assert user.email == "rollback@example.com"
    assert user.password == original_password
    assert user.check_password("Inicial123!") is True
    assert profile.temporary_password_plaintext is None
    assert Token.objects.filter(user=user).exists() is True


@pytest.mark.django_db
def test_process_bulk_credentials_retries_email_timeout_and_succeeds(mocker):
    user = User.objects.create_user(
        username="bulk_retry",
        email="retry@example.com",
        password="Inicial123!",
    )
    send_once = mocker.patch(
        "users.services_bulk_credentials._send_bulk_credentials_email_once",
        side_effect=[BulkCredentialsEmailTimeoutError("smtp timeout"), None],
    )
    sleep = mocker.patch("users.services_bulk_credentials.time.sleep")
    upload = _build_excel_file(
        [("bulk_retry", "retry@example.com", "Inicial123!")],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 1,
        "actualizadas": 0,
        "sin_cambios": 1,
        "rechazadas": 0,
    }
    assert result["rows"][0]["estado"] == "enviada"
    assert send_once.call_count == 2
    sleep.assert_called_once()


@pytest.mark.django_db
def test_process_bulk_credentials_rejects_row_when_email_timeout_persists(mocker):
    user = User.objects.create_user(
        username="bulk_timeout",
        email="timeout@example.com",
        password="Inicial123!",
    )
    Token.objects.create(user=user)
    original_password = user.password
    send_once = mocker.patch(
        "users.services_bulk_credentials._send_bulk_credentials_email_once",
        side_effect=BulkCredentialsEmailTimeoutError("smtp timeout"),
    )
    sleep = mocker.patch("users.services_bulk_credentials.time.sleep")
    upload = _build_excel_file(
        [("bulk_timeout", "timeout@example.com", "TemporalNueva123!")],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    user.refresh_from_db()
    profile = user.profile
    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 0,
        "actualizadas": 0,
        "sin_cambios": 0,
        "rechazadas": 1,
    }
    assert (
        result["rows"][0]["mensaje"]
        == "No se pudo enviar el correo para esta fila luego de reintentar."
    )
    assert send_once.call_count == 2
    sleep.assert_called_once()
    assert user.password == original_password
    assert user.check_password("Inicial123!") is True
    assert profile.temporary_password_plaintext is None
    assert Token.objects.filter(user=user).exists() is True


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_avoids_hash_check_when_temporary_password_matches(
    mocker,
):
    user = User.objects.create_user(
        username="bulk_temp_password",
        email="bulk-temp@example.com",
        password="Temporal123!",
    )
    profile = user.profile
    profile.must_change_password = True
    profile.temporary_password_plaintext = "Temporal123!"
    profile.save(
        update_fields=[
            "must_change_password",
            "temporary_password_plaintext",
        ]
    )
    mocker.patch.object(
        User,
        "check_password",
        side_effect=AssertionError("No deberia calcular hash en este caso"),
    )
    upload = _build_excel_file(
        [("bulk_temp_password", "bulk-temp@example.com", "Temporal123!")],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 1,
        "actualizadas": 0,
        "sin_cambios": 1,
        "rechazadas": 0,
    }


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_rejects_pending_rows_when_batch_budget_is_exhausted(
    mocker,
):
    User.objects.create_user(
        username="bulk_budget_first",
        email="first@example.com",
        password="Temporal123!",
    )
    User.objects.create_user(
        username="bulk_budget_second",
        email="second@example.com",
        password="Temporal123!",
    )
    budget_check = mocker.patch(
        "users.services_bulk_credentials._has_enough_batch_time",
        side_effect=[True, False],
    )
    upload = _build_excel_file(
        [
            ("bulk_budget_first", "first@example.com", "Temporal123!"),
            ("bulk_budget_second", "second@example.com", "Temporal123!"),
        ],
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="standard",
    )

    assert result["summary"] == {
        "procesadas": 2,
        "enviadas": 1,
        "actualizadas": 0,
        "sin_cambios": 1,
        "rechazadas": 1,
    }
    assert result["rows"][0]["estado"] == "enviada"
    assert result["rows"][1]["estado"] == "rechazada"
    assert (
        result["rows"][1]["mensaje"]
        == "Se alcanzo el tiempo maximo del lote. Reintente con una planilla mas chica."
    )
    assert budget_check.call_count == 2


@pytest.mark.django_db
@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    ROOT_URLCONF="tests.urls_users_bulk_credentials",
)
def test_bulk_credentials_upload_view_creates_job_and_redirects_to_detail(
    settings,
    tmp_path,
):
    settings.MEDIA_ROOT = tmp_path
    user = User.objects.create_user(
        username="bulk_operator",
        email="operator@example.com",
        password="Secreta123!",
    )
    target = User.objects.create_user(
        username="bulk_view_target",
        email="view_target@example.com",
        password="Inicial123!",
    )
    _grant_bulk_credentials_permissions(user)

    upload = _build_excel_file(
        [("bulk_view_target", "nuevo_view_target@example.com", "NuevaView123!")]
    )
    form = BulkCredentialsUploadForm(
        data={"tipo_envio": "standard"},
        files={"archivo": upload},
    )
    assert form.is_valid(), form.errors

    request = _build_request(
        "post",
        "/usuarios/credenciales-masivas/",
        user,
        data={"tipo_envio": "standard"},
        files={"archivo": upload},
    )
    view = BulkCredentialsUploadView()
    view.setup(request)

    response = view.form_valid(form)
    job = BulkCredentialsJob.objects.get()

    target.refresh_from_db()
    assert response.status_code == 302
    assert response.url == f"/usuarios/credenciales-masivas/lotes/{job.pk}/"
    assert job.requested_by == user
    assert job.status == BulkCredentialsJob.Status.PENDING
    assert job.send_type == "standard"
    assert target.email == "view_target@example.com"
    assert target.check_password("Inicial123!") is True
    assert len(mail.outbox) == 0


@pytest.mark.django_db
def test_create_bulk_credentials_job_persists_upload_and_sets_pending(
    settings,
    tmp_path,
):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_job_operator",
        email="job-operator@example.com",
        password="Secreta123!",
    )
    upload = _build_excel_file(
        [("bulk_job_target", "target@example.com", "Temporal123!")],
    )

    job = create_bulk_credentials_job(
        uploaded_file=upload,
        send_type="standard",
        requested_by=operator,
    )

    assert job.status == BulkCredentialsJob.Status.PENDING
    assert job.requested_by == operator
    assert job.original_filename == "credenciales.xlsx"
    assert job.send_type == "standard"
    assert job.total_rows == 0
    assert job.last_error_message == ""
    assert job.archivo.name.endswith("credenciales.xlsx")
    assert job.archivo.storage.exists(job.archivo.name) is True


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_job_stops_on_first_failure_and_tracks_checkpoint(
    settings,
    tmp_path,
    mocker,
):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_job_runner",
        email="job-runner@example.com",
        password="Secreta123!",
    )
    first_user = User.objects.create_user(
        username="bulk_job_first",
        email="first@example.com",
        password="Inicial123!",
    )
    second_user = User.objects.create_user(
        username="bulk_job_second",
        email="second@example.com",
        password="Inicial123!",
    )
    third_user = User.objects.create_user(
        username="bulk_job_third",
        email="third@example.com",
        password="Inicial123!",
    )
    upload = _build_excel_file(
        [
            ("bulk_job_first", "destino-first@example.com", "Inicial123!"),
            ("bulk_job_second", "destino-second@example.com", "Inicial123!"),
            ("bulk_job_third", "destino-third@example.com", "Inicial123!"),
        ],
    )
    job = create_bulk_credentials_job(
        uploaded_file=upload,
        send_type="standard",
        requested_by=operator,
    )

    def _send_side_effect(*args, **kwargs):
        user = kwargs["user"]
        if user.username == "bulk_job_second":
            raise smtplib.SMTPAuthenticationError(535, b"auth failed")
        return None

    mocker.patch(
        "users.services_bulk_credentials._send_bulk_credentials_email_once",
        side_effect=_send_side_effect,
    )
    mocker.patch("users.services_bulk_credentials.time.sleep")

    process_bulk_credentials_job(job)

    job.refresh_from_db()
    first_user.refresh_from_db()
    second_user.refresh_from_db()
    third_user.refresh_from_db()
    rows = list(job.rows.order_by("fila"))

    assert job.status == BulkCredentialsJob.Status.FAILED
    assert job.total_rows == 3
    assert job.processed_rows == 2
    assert job.sent_rows == 1
    assert job.updated_password_rows == 0
    assert job.unchanged_password_rows == 1
    assert job.rejected_rows == 1
    assert job.next_row_index == 1
    assert job.last_successful_username == "bulk_job_first"
    assert job.last_successful_row == 2
    assert job.last_attempted_username == "bulk_job_second"
    assert job.last_attempted_row == 3
    assert (
        job.last_error_message
        == (
            "El servidor de correo rechazo la autenticacion. "
            "El envio se reintento sin exito."
        )
    )
    assert can_resume_bulk_credentials_job(job) is True
    assert len(rows) == 2
    assert rows[0].status == BulkCredentialsJobRow.Status.SENT
    assert rows[0].usuario == "bulk_job_first"
    assert rows[1].status == BulkCredentialsJobRow.Status.FAILED
    assert rows[1].usuario == "bulk_job_second"
    assert rows[1].attempts == 1
    assert len(mail.outbox) == 1
    assert mail.outbox[0].to == ["destino-first@example.com"]
    assert first_user.check_password("Inicial123!") is True
    assert second_user.check_password("Inicial123!") is True
    assert third_user.check_password("Inicial123!") is True


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_job_can_resume_from_failed_row(
    settings,
    tmp_path,
    mocker,
):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_resume_runner",
        email="resume-runner@example.com",
        password="Secreta123!",
    )
    User.objects.create_user(
        username="bulk_resume_first",
        email="first@example.com",
        password="Inicial123!",
    )
    User.objects.create_user(
        username="bulk_resume_second",
        email="second@example.com",
        password="Inicial123!",
    )
    User.objects.create_user(
        username="bulk_resume_third",
        email="third@example.com",
        password="Inicial123!",
    )
    upload = _build_excel_file(
        [
            ("bulk_resume_first", "destino-first@example.com", "Inicial123!"),
            ("bulk_resume_second", "destino-second@example.com", "Inicial123!"),
            ("bulk_resume_third", "destino-third@example.com", "Inicial123!"),
        ],
    )
    job = create_bulk_credentials_job(
        uploaded_file=upload,
        send_type="standard",
        requested_by=operator,
    )
    failure_state = {"enabled": True}

    def _send_side_effect(*args, **kwargs):
        user = kwargs["user"]
        if failure_state["enabled"] and user.username == "bulk_resume_second":
            raise smtplib.SMTPAuthenticationError(535, b"auth failed")
        return None

    mocker.patch(
        "users.services_bulk_credentials._send_bulk_credentials_email_once",
        side_effect=_send_side_effect,
    )
    mocker.patch("users.services_bulk_credentials.time.sleep")

    process_bulk_credentials_job(job)
    request_resume_bulk_credentials_job(job=job)
    failure_state["enabled"] = False

    job.refresh_from_db()
    assert job.status == BulkCredentialsJob.Status.PENDING
    assert job.resume_count == 1
    assert job.last_error_message == ""

    process_bulk_credentials_job(job)

    job.refresh_from_db()
    second_row = job.rows.get(fila=3)

    assert job.status == BulkCredentialsJob.Status.COMPLETED
    assert job.processed_rows == 3
    assert job.sent_rows == 3
    assert job.updated_password_rows == 0
    assert job.unchanged_password_rows == 3
    assert job.rejected_rows == 0
    assert job.next_row_index == 3
    assert job.last_successful_username == "bulk_resume_third"
    assert job.last_successful_row == 4
    assert second_row.status == BulkCredentialsJobRow.Status.SENT
    assert second_row.attempts == 2
    assert len(mail.outbox) == 3


@pytest.mark.django_db
def test_mark_stale_bulk_credentials_jobs_as_failed(settings, tmp_path, monkeypatch):
    settings.MEDIA_ROOT = tmp_path
    monkeypatch.setenv("BULK_CREDENTIALS_JOB_STALE_SECONDS", "1")
    operator = User.objects.create_user(
        username="bulk_stale_runner",
        email="stale-runner@example.com",
        password="Secreta123!",
    )
    upload = _build_excel_file(
        [("bulk_stale_target", "target@example.com", "Temporal123!")],
    )
    job = create_bulk_credentials_job(
        uploaded_file=upload,
        send_type="standard",
        requested_by=operator,
    )
    job.status = BulkCredentialsJob.Status.PROCESSING
    job.last_activity_at = timezone.now() - timedelta(minutes=5)
    job.save(update_fields=["status", "last_activity_at"])

    updated = mark_stale_bulk_credentials_jobs_as_failed()

    job.refresh_from_db()
    assert updated == 1
    assert job.status == BulkCredentialsJob.Status.FAILED
    assert job.last_error_message == STALE_JOB_ERROR_MESSAGE
    assert can_resume_bulk_credentials_job(job) is True


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_users_bulk_credentials")
def test_bulk_credentials_upload_view_lists_only_request_user_jobs(
    settings,
    tmp_path,
):
    settings.MEDIA_ROOT = tmp_path
    current_user = User.objects.create_user(
        username="bulk_jobs_current",
        email="current@example.com",
        password="Secreta123!",
    )
    other_user = User.objects.create_user(
        username="bulk_jobs_other",
        email="other@example.com",
        password="Secreta123!",
    )
    _grant_bulk_credentials_permissions(current_user)
    create_bulk_credentials_job(
        uploaded_file=_build_excel_file(
            [("bulk_jobs_current", "current@example.com", "Temporal123!")]
        ),
        send_type="standard",
        requested_by=current_user,
    )
    create_bulk_credentials_job(
        uploaded_file=_build_excel_file(
            [("bulk_jobs_other", "other@example.com", "Temporal123!")]
        ),
        send_type="standard",
        requested_by=other_user,
    )
    request = _build_request("get", "/usuarios/credenciales-masivas/", current_user)
    view = BulkCredentialsUploadView()
    view.setup(request)

    context = view.get_context_data(form=BulkCredentialsUploadForm())

    assert len(context["recent_jobs"]) == 1
    assert context["recent_jobs"][0].requested_by == current_user


@pytest.mark.django_db
@override_settings(
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    ROOT_URLCONF="tests.urls_users_bulk_credentials",
)
def test_bulk_credentials_job_detail_view_shows_failure_context(
    settings,
    tmp_path,
    mocker,
):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_detail_operator",
        email="detail-operator@example.com",
        password="Secreta123!",
    )
    _grant_bulk_credentials_permissions(operator)
    User.objects.create_user(
        username="bulk_detail_first",
        email="first@example.com",
        password="Inicial123!",
    )
    User.objects.create_user(
        username="bulk_detail_second",
        email="second@example.com",
        password="Inicial123!",
    )
    job = create_bulk_credentials_job(
        uploaded_file=_build_excel_file(
            [
                ("bulk_detail_first", "destino-first@example.com", "Inicial123!"),
                ("bulk_detail_second", "destino-second@example.com", "Inicial123!"),
            ]
        ),
        send_type="standard",
        requested_by=operator,
    )

    def _send_side_effect(*args, **kwargs):
        if kwargs["user"].username == "bulk_detail_second":
            raise smtplib.SMTPAuthenticationError(535, b"auth failed")
        return None

    mocker.patch(
        "users.services_bulk_credentials._send_bulk_credentials_email_once",
        side_effect=_send_side_effect,
    )
    mocker.patch("users.services_bulk_credentials.time.sleep")
    process_bulk_credentials_job(job)

    response = BulkCredentialsJobDetailView.as_view()(
        _build_request(
            "get",
            f"/usuarios/credenciales-masivas/lotes/{job.pk}/",
            operator,
        ),
        pk=job.pk,
    )
    content = response.content.decode("utf-8")

    assert response.status_code == 200
    assert "bulk_detail_first" in content
    assert "bulk_detail_second" in content
    assert "El servidor de correo rechazo la autenticacion." in content
    assert "Reanudar" in content


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_users_bulk_credentials")
def test_bulk_credentials_job_resume_view_sets_job_pending(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_resume_operator",
        email="resume-operator@example.com",
        password="Secreta123!",
    )
    _grant_bulk_credentials_permissions(operator)
    job = create_bulk_credentials_job(
        uploaded_file=_build_excel_file(
            [("bulk_resume_user", "resume@example.com", "Temporal123!")]
        ),
        send_type="standard",
        requested_by=operator,
    )
    job.status = BulkCredentialsJob.Status.FAILED
    job.last_error_message = "Fallo previo"
    job.last_error_at = timezone.now()
    job.finished_at = timezone.now()
    job.save(
        update_fields=[
            "status",
            "last_error_message",
            "last_error_at",
            "finished_at",
        ]
    )

    response = BulkCredentialsJobResumeView.as_view()(
        _build_request(
            "post",
            f"/usuarios/credenciales-masivas/lotes/{job.pk}/reanudar/",
            operator,
        ),
        pk=job.pk,
    )

    job.refresh_from_db()
    assert response.status_code == 302
    assert response.url == f"/usuarios/credenciales-masivas/lotes/{job.pk}/"
    assert job.status == BulkCredentialsJob.Status.PENDING
    assert job.resume_count == 1
    assert job.last_error_message == ""


def test_process_bulk_credentials_jobs_command_invokes_worker_once(mocker):
    run_worker = mocker.patch(
        "users.management.commands.process_bulk_credentials_jobs."
        "run_bulk_credentials_jobs_worker"
    )

    call_command("process_bulk_credentials_jobs", "--once")

    run_worker.assert_called_once_with(once=True)


def test_run_bulk_credentials_jobs_worker_once_processes_single_cycle(mocker):
    process_next = mocker.patch(
        "users.services_bulk_credentials_jobs.process_next_bulk_credentials_job",
        return_value=True,
    )

    run_bulk_credentials_jobs_worker(once=True)

    process_next.assert_called_once_with()


def test_run_bulk_credentials_jobs_worker_once_reraises_unexpected_error(mocker):
    mocker.patch(
        "users.services_bulk_credentials_jobs.process_next_bulk_credentials_job",
        side_effect=RuntimeError("boom"),
    )

    with pytest.raises(RuntimeError, match="boom"):
        run_bulk_credentials_jobs_worker(once=True)


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_process_bulk_credentials_inet_uses_inet_email_template():
    user = User.objects.create_user(
        username="bulk_inet",
        email="inet-old@example.com",
        password="ViejaInet123!",
    )
    upload = _build_excel_file(
        [("bulk_inet", "inet-new@example.com", "NuevaInet123!", "CFP INET 401")],
        headers=("usuario", "mail", "password", "Nombre del Centro"),
    )

    result = process_bulk_credentials_file(
        uploaded_file=upload,
        send_type="inet",
    )

    user.refresh_from_db()
    assert result["summary"] == {
        "procesadas": 1,
        "enviadas": 1,
        "actualizadas": 1,
        "sin_cambios": 0,
        "rechazadas": 0,
    }
    assert result["send_type"] == "inet"
    assert result["send_type_label"] == "INET"
    assert user.email == "inet-old@example.com"
    assert len(mail.outbox) == 1
    assert mail.outbox[0].subject == (
        "Acceso a la plataforma y capacitación virtual – INET"
    )
    assert mail.outbox[0].to == ["inet-new@example.com"]
    assert "CFP INET 401" in mail.outbox[0].body
    assert (
        "Capacitacion a instituciones de FP para beneficiarios de VAT (1)"
        in mail.outbox[0].body
    )
    assert "En todas se abordaran los mismos temas." in mail.outbox[0].body
    assert "https://youtu.be/vR_lODbOdJg" in mail.outbox[0].body
    assert "Nos vemos pronto." in mail.outbox[0].body
