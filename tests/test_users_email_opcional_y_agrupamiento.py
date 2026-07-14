"""Tests para issues #1979 y #1936.

Cubren:
- Usuarios sin email y con email repetido en forms.
- User import username-first con fallback a email.
- Actualización masiva de grupos con acciones agregar / quitar / reemplazar.
- Enforcement de alcance via Profile.grupos_asignables.
- Bulk credentials agrupado por mail destinatario.
- Nombre y apellido en el cuerpo del email.
"""

from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from openpyxl import Workbook

from users.forms import CustomUserChangeForm, UserCreationForm
from users.models import (
    BulkCredentialsJob,
    BulkCredentialsJobRow,
    UserImportJob,
    UserImportJobRow,
)
from users.services_bulk_credentials import (
    process_bulk_credentials_file,
)
from users.services_bulk_credentials_jobs import (
    create_bulk_credentials_job,
    process_bulk_credentials_job,
)
from users.services_user_import import (
    USER_IMPORT_TEMPLATE_HEADERS,
    create_user_import_job,
    process_single_user_import_row,
)
from users.services_user_import_jobs import process_user_import_job

User = get_user_model()


def _build_excel_file(rows, headers=("usuario", "mail")):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "credenciales"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    output = BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        "credenciales.xlsx",
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


def _build_user_import_excel_file(rows):
    return _build_excel_file(rows, headers=USER_IMPORT_TEMPLATE_HEADERS)


def _set_visible_temporary_password(user, plain_password):
    profile = user.profile
    profile.must_change_password = True
    profile.temporary_password_plaintext = plain_password
    profile.save(update_fields=["must_change_password", "temporary_password_plaintext"])
    return user


# ---------- A1: forms permiten email vacío y repetido ----------


@pytest.mark.django_db
def test_user_creation_form_acepta_email_vacio():
    form = UserCreationForm(
        data={
            "username": "sin_email_user",
            "email": "",
            "password": "Inicial123!",
            "first_name": "Juan",
            "last_name": "Pérez",
        },
    )
    assert form.is_valid(), form.errors
    user = form.save()
    assert user.email == ""
    assert user.username == "sin_email_user"


@pytest.mark.django_db
def test_user_creation_form_acepta_email_repetido():
    User.objects.create_user(
        username="primer_usuario",
        email="repetido@example.com",
        password="Pass123!",
    )
    form = UserCreationForm(
        data={
            "username": "segundo_usuario",
            "email": "repetido@example.com",
            "password": "Pass123!",
            "first_name": "Ana",
            "last_name": "García",
        },
    )
    assert form.is_valid(), form.errors
    nuevo = form.save()
    assert nuevo.email == "repetido@example.com"
    assert User.objects.filter(email__iexact="repetido@example.com").count() == 2


@pytest.mark.django_db
def test_user_change_form_acepta_email_vacio():
    existing = User.objects.create_user(
        username="cambia_email",
        email="viejo@example.com",
        password="Pass123!",
    )
    form = CustomUserChangeForm(
        data={
            "username": "cambia_email",
            "email": "",
            "first_name": "Pedro",
            "last_name": "Lopez",
        },
        instance=existing,
    )
    assert form.is_valid(), form.errors


# ---------- A2: user import – stub de job ----------


class _StubImportJob:
    """Job mínimo para process_single_user_import_row. Actor=None => sin restricción."""

    is_pwa_import = False
    send_credentials = False
    requested_by = None


def _row(extra=None, **kwargs):
    base = {
        "nombre": "Ana",
        "apellido": "García",
        "correo": "",
        "username": "",
        "permisos": "",
        "provincias": "",
        "rol": "operadora",
        "accion_grupos": "",
        "fila": 2,
    }
    base.update(kwargs)
    if extra:
        base.update(extra)
    return base


# ---------- Caso 1: username existente + grupos nuevos => CREATED ----------


@pytest.mark.django_db
def test_import_username_existente_grupos_nuevos():
    grupo = Group.objects.create(name="Grupo Test 1")
    user = User.objects.create_user(username="juan.perez", email="j@e.com")

    resultado = process_single_user_import_row(
        row_data=_row(username="juan.perez", permisos="Grupo Test 1"),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    user.refresh_from_db()
    assert grupo in user.groups.all()


# ---------- Caso 2: username existente + sin cambios => SKIPPED ----------


@pytest.mark.django_db
def test_import_username_existente_sin_cambios():
    grupo = Group.objects.create(name="Grupo Test 2")
    user = User.objects.create_user(username="sin.cambios", email="sc@e.com")
    user.groups.add(grupo)

    resultado = process_single_user_import_row(
        row_data=_row(username="sin.cambios", permisos="Grupo Test 2"),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "skipped", resultado


# ---------- Caso 3: sin username, con email existente y acción vacía => crea ----------


@pytest.mark.django_db
def test_import_sin_username_con_email_existente():
    grupo_existente = Group.objects.create(name="Grupo Existente")
    grupo_nuevo = Group.objects.create(name="Grupo Test 3")
    existente = User.objects.create_user(username="email.user", email="email@e.com")
    existente.groups.add(grupo_existente)

    resultado = process_single_user_import_row(
        row_data=_row(correo="email@e.com", permisos="Grupo Test 3"),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    existente.refresh_from_db()
    assert existente.username == "email.user"
    assert existente.email == "email@e.com"
    assert list(existente.groups.all()) == [grupo_existente]
    creado = User.objects.get(pk=resultado["created_user_id"])
    assert creado.email == "email@e.com"
    assert grupo_nuevo in creado.groups.all()


@pytest.mark.django_db
def test_import_sin_username_con_email_existente_y_accion_explicita_actualiza():
    grupo = Group.objects.create(name="Grupo Explícito")
    existente = User.objects.create_user(
        username="email.explicito",
        email="explicito@example.com",
    )

    resultado = process_single_user_import_row(
        row_data=_row(
            correo="explicito@example.com",
            permisos="Grupo Explícito",
            accion_grupos="agregar",
        ),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    assert "created_user_id" not in resultado
    assert User.objects.filter(email__iexact="explicito@example.com").count() == 1
    existente.refresh_from_db()
    assert grupo in existente.groups.all()


# ---------- Caso 4: sin username ni email => error ----------


@pytest.mark.django_db
def test_import_sin_username_ni_email():
    from django.core.exceptions import ValidationError

    with pytest.raises(ValidationError, match="Username o Correo"):
        process_single_user_import_row(
            row_data=_row(),
            job=_StubImportJob(),
        )


# ---------- Caso 5: username y email apuntan a usuarios distintos => gana username, se actualiza email ----------


@pytest.mark.django_db
def test_import_username_y_email_usuarios_distintos():
    user_a = User.objects.create_user(username="user.a", email="a@e.com")
    User.objects.create_user(username="user.b", email="b@e.com")

    resultado = process_single_user_import_row(
        row_data=_row(username="user.a", correo="b@e.com"),
        job=_StubImportJob(),
    )

    user_a.refresh_from_db()
    assert user_a.email == "b@e.com"
    assert resultado["status"].value == "created", resultado
    assert User.objects.filter(username="user.b").exists()


# ---------- Caso 6: accion agregar ----------


@pytest.mark.django_db
def test_import_accion_agregar():
    g1 = Group.objects.create(name="Grupo Agregar 1")
    g2 = Group.objects.create(name="Grupo Agregar 2")
    user = User.objects.create_user(username="agregar.user")
    user.groups.add(g1)

    resultado = process_single_user_import_row(
        row_data=_row(
            username="agregar.user", permisos="Grupo Agregar 2", accion_grupos="agregar"
        ),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    user.refresh_from_db()
    group_ids = set(user.groups.values_list("id", flat=True))
    assert g1.pk in group_ids
    assert g2.pk in group_ids


# ---------- Caso 7: accion quitar ----------


@pytest.mark.django_db
def test_import_accion_quitar():
    g1 = Group.objects.create(name="Grupo Quitar 1")
    g2 = Group.objects.create(name="Grupo Quitar 2")
    user = User.objects.create_user(username="quitar.user")
    user.groups.add(g1, g2)

    resultado = process_single_user_import_row(
        row_data=_row(
            username="quitar.user", permisos="Grupo Quitar 1", accion_grupos="quitar"
        ),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    user.refresh_from_db()
    group_ids = set(user.groups.values_list("id", flat=True))
    assert g1.pk not in group_ids
    assert g2.pk in group_ids


# ---------- Caso 8: accion reemplazar ----------


@pytest.mark.django_db
def test_import_accion_reemplazar():
    g1 = Group.objects.create(name="Grupo Reemplazar 1")
    g2 = Group.objects.create(name="Grupo Reemplazar 2")
    user = User.objects.create_user(username="reemplazar.user")
    user.groups.add(g1)

    resultado = process_single_user_import_row(
        row_data=_row(
            username="reemplazar.user",
            permisos="Grupo Reemplazar 2",
            accion_grupos="reemplazar",
        ),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    user.refresh_from_db()
    group_ids = set(user.groups.values_list("id", flat=True))
    assert g1.pk not in group_ids
    assert g2.pk in group_ids


# ---------- Caso 9: reemplazar preserva grupos fuera de alcance del actor ----------


@pytest.mark.django_db
def test_import_reemplazar_preserva_grupos_fuera_de_alcance():
    g_dentro = Group.objects.create(name="Grupo Dentro Alcance")
    g_fuera = Group.objects.create(name="Grupo Fuera Alcance")

    actor = User.objects.create_user(username="actor.restringido")
    actor.profile.grupos_asignables.set([g_dentro])

    target = User.objects.create_user(username="target.reemplazar")
    target.groups.add(g_fuera)

    class _RestrictedJob:
        is_pwa_import = False
        send_credentials = False
        requested_by = actor

    resultado = process_single_user_import_row(
        row_data=_row(
            username="target.reemplazar",
            permisos="Grupo Dentro Alcance",
            accion_grupos="reemplazar",
        ),
        job=_RestrictedJob(),
    )

    assert resultado["status"].value == "created", resultado
    target.refresh_from_db()
    group_ids = set(target.groups.values_list("id", flat=True))
    assert g_fuera.pk in group_ids
    assert g_dentro.pk in group_ids


# ---------- Caso 10: grupo fuera de grupos_asignables => error ----------


@pytest.mark.django_db
def test_import_grupo_fuera_de_asignables_error():
    from django.core.exceptions import ValidationError

    g_permitido = Group.objects.create(name="Grupo Permitido")
    g_prohibido = Group.objects.create(name="Grupo Prohibido")

    actor = User.objects.create_user(username="actor.limitado")
    actor.profile.grupos_asignables.set([g_permitido])

    User.objects.create_user(username="target.limitado")

    class _LimitedJob:
        is_pwa_import = False
        send_credentials = False
        requested_by = actor

    with pytest.raises(ValidationError, match="permiso"):
        process_single_user_import_row(
            row_data=_row(
                username="target.limitado",
                permisos="Grupo Prohibido",
                accion_grupos="agregar",
            ),
            job=_LimitedJob(),
        )


# ---------- A2 actualizados: creacion requiere username ----------


@pytest.mark.django_db
def test_user_import_row_con_username_crea_usuario():
    resultado = process_single_user_import_row(
        row_data=_row(username="nuevo.usuario", nombre="Ana", apellido="García"),
        job=_StubImportJob(),
    )
    assert resultado["status"].value == "created", resultado
    user = User.objects.filter(username="nuevo.usuario").first()
    assert user is not None
    assert user.first_name == "Ana"


@pytest.mark.django_db
def test_user_import_row_email_existente_crea_usuario_nuevo():
    grupo_existente = Group.objects.create(name="Grupo Existente Email")
    grupo_nuevo = Group.objects.create(name="Grupo Nuevo")
    existente = User.objects.create_user(
        username="existente.email",
        email="shared@example.com",
    )
    existente.groups.add(grupo_existente)

    resultado = process_single_user_import_row(
        row_data=_row(correo="shared@example.com", permisos="Grupo Nuevo"),
        job=_StubImportJob(),
    )

    assert resultado["status"].value == "created", resultado
    assert User.objects.filter(email__iexact="shared@example.com").count() == 2
    existente.refresh_from_db()
    assert existente.username == "existente.email"
    assert existente.email == "shared@example.com"
    assert list(existente.groups.all()) == [grupo_existente]
    creado = User.objects.get(pk=resultado["created_user_id"])
    assert grupo_nuevo in creado.groups.all()
    assert creado.profile.temporary_password_plaintext


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_user_import_job_agrupa_credenciales_para_email_duplicado(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(username="import_operator")
    upload = _build_user_import_excel_file(
        [
            (
                "",
                "Ana",
                "García",
                "duplicado@example.com",
                "",
                "",
                "",
                "operadora",
                "",
                "",
            ),
            (
                "",
                "Beto",
                "García",
                "duplicado@example.com",
                "",
                "",
                "",
                "operadora",
                "",
                "",
            ),
        ]
    )
    job = create_user_import_job(
        uploaded_file=upload,
        requested_by=operator,
        send_credentials=True,
    )

    process_user_import_job(job)
    job.refresh_from_db()

    assert job.status == UserImportJob.Status.COMPLETED
    assert User.objects.filter(email="duplicado@example.com").count() == 2
    assert len(mail.outbox) == 1
    assert mail.outbox[0].to == ["duplicado@example.com"]
    assert "garcia.ana" in mail.outbox[0].body
    assert "garcia.beto" in mail.outbox[0].body
    rows = list(job.rows.order_by("fila"))
    assert all(row.created_user_id for row in rows)
    assert all(row.credentials_sent_at for row in rows)
    assert all(row.created_user.profile.temporary_password_plaintext for row in rows)


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_user_import_job_reprocesado_no_reenvia_credenciales(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(username="import_resume_operator")
    upload = _build_user_import_excel_file(
        [
            ("", "Ana", "López", "resume@example.com", "", "", "", "operadora", "", ""),
        ]
    )
    job = create_user_import_job(
        uploaded_file=upload,
        requested_by=operator,
        send_credentials=True,
    )

    process_user_import_job(job)
    sent_at = UserImportJobRow.objects.get(job=job).credentials_sent_at
    job.status = UserImportJob.Status.PROCESSING
    job.save(update_fields=["status"])
    process_user_import_job(job)

    assert len(mail.outbox) == 1
    row = UserImportJobRow.objects.get(job=job)
    assert row.credentials_sent_at == sent_at


@pytest.mark.django_db
def test_user_import_template_headers_incluye_username_y_correo():
    assert "Username" in USER_IMPORT_TEMPLATE_HEADERS
    assert "Correo" in USER_IMPORT_TEMPLATE_HEADERS


# ---------- C: nombre y apellido en el email ----------


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_bulk_credentials_email_incluye_nombre_y_apellido():
    user = User.objects.create_user(
        username="con_datos",
        email="con_datos@example.com",
        password="Inicial123!",
        first_name="Lucía",
        last_name="Fernández",
    )
    _set_visible_temporary_password(user, "Inicial123!")

    upload = _build_excel_file([("con_datos", "con_datos@example.com")])
    process_bulk_credentials_file(uploaded_file=upload, send_type="standard")

    assert len(mail.outbox) == 1
    body = mail.outbox[0].body
    assert "Lucía Fernández" in body


# ---------- D: agrupamiento por destinatario ----------


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_bulk_credentials_agrupa_credenciales_por_mismo_destinatario():
    user1 = User.objects.create_user(
        username="user_uno",
        email="x@example.com",
        password="Pass1!",
        first_name="Uno",
        last_name="Apellido",
    )
    user2 = User.objects.create_user(
        username="user_dos",
        email="x@example.com",
        password="Pass2!",
        first_name="Dos",
        last_name="Apellido",
    )
    user3 = User.objects.create_user(
        username="user_solo",
        email="z@example.com",
        password="Pass3!",
        first_name="Solo",
        last_name="Otra",
    )
    _set_visible_temporary_password(user1, "Pass1!")
    _set_visible_temporary_password(user2, "Pass2!")
    _set_visible_temporary_password(user3, "Pass3!")

    upload = _build_excel_file(
        [
            ("user_uno", "compartido@example.com"),
            ("user_dos", "compartido@example.com"),
            ("user_solo", "solo@example.com"),
        ]
    )

    result = process_bulk_credentials_file(uploaded_file=upload, send_type="standard")

    assert result["summary"]["procesadas"] == 3
    assert result["summary"]["enviadas"] == 3
    assert len(mail.outbox) == 2

    agrupado = next(msg for msg in mail.outbox if msg.to == ["compartido@example.com"])
    assert "user_uno" in agrupado.body
    assert "user_dos" in agrupado.body
    assert "Pass1!" in agrupado.body
    assert "Pass2!" in agrupado.body

    individual = next(msg for msg in mail.outbox if msg.to == ["solo@example.com"])
    assert "user_solo" in individual.body
    assert "Pass3!" in individual.body
    assert "user_uno" not in individual.body


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_bulk_credentials_agrupa_cuando_destinatario_se_resuelve_desde_user_email():
    """Si la columna mail está vacía, se usa user.email y debe agruparse igual."""
    user1 = User.objects.create_user(
        username="vacio_uno",
        email="comp@example.com",
        password="Aaa1!",
        first_name="A",
        last_name="A",
    )
    user2 = User.objects.create_user(
        username="vacio_dos",
        email="comp@example.com",
        password="Bbb2!",
        first_name="B",
        last_name="B",
    )
    _set_visible_temporary_password(user1, "Aaa1!")
    _set_visible_temporary_password(user2, "Bbb2!")

    upload = _build_excel_file([("vacio_uno", ""), ("vacio_dos", "")])
    result = process_bulk_credentials_file(uploaded_file=upload, send_type="standard")

    assert result["summary"]["enviadas"] == 2
    assert len(mail.outbox) == 1
    assert mail.outbox[0].to == ["comp@example.com"]
    assert "vacio_uno" in mail.outbox[0].body
    assert "vacio_dos" in mail.outbox[0].body


# ---------- Defensivos R3: INET no agrupa con centros distintos ----------


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_inet_no_agrupa_cuando_centros_son_distintos():
    """Aunque dos filas INET compartan destinatario, no se agrupan si los
    nombres de centro difieren: evita mezclar datos en el correo."""
    user_a = User.objects.create_user(
        username="inet_a",
        email="inet_a@example.com",
        password="Pass!",
        first_name="A",
        last_name="A",
    )
    user_b = User.objects.create_user(
        username="inet_b",
        email="inet_b@example.com",
        password="Pass!",
        first_name="B",
        last_name="B",
    )
    _set_visible_temporary_password(user_a, "Pass!")
    _set_visible_temporary_password(user_b, "Pass!")

    upload = _build_excel_file(
        [
            ("inet_a", "compartido@example.com", "CFP 401"),
            ("inet_b", "compartido@example.com", "CFP 402"),
        ],
        headers=("usuario", "mail", "Nombre del Centro"),
    )
    result = process_bulk_credentials_file(uploaded_file=upload, send_type="inet")

    assert result["summary"]["enviadas"] == 2
    assert len(mail.outbox) == 2
    assert all(msg.to == ["compartido@example.com"] for msg in mail.outbox)
    bodies = [msg.body for msg in mail.outbox]
    assert any("CFP 401" in body and "inet_a" in body for body in bodies)
    assert any("CFP 402" in body and "inet_b" in body for body in bodies)


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_inet_agrupa_cuando_centro_y_destinatario_coinciden():
    user_a = User.objects.create_user(
        username="inet_mismo_a",
        email="a@example.com",
        password="Pass!",
        first_name="A",
        last_name="A",
    )
    user_b = User.objects.create_user(
        username="inet_mismo_b",
        email="b@example.com",
        password="Pass!",
        first_name="B",
        last_name="B",
    )
    _set_visible_temporary_password(user_a, "Pass!")
    _set_visible_temporary_password(user_b, "Pass!")

    upload = _build_excel_file(
        [
            ("inet_mismo_a", "centro@example.com", "CFP 401"),
            ("inet_mismo_b", "centro@example.com", "CFP 401"),
        ],
        headers=("usuario", "mail", "Nombre del Centro"),
    )
    result = process_bulk_credentials_file(uploaded_file=upload, send_type="inet")

    assert result["summary"]["enviadas"] == 2
    assert len(mail.outbox) == 1
    body = mail.outbox[0].body
    assert "inet_mismo_a" in body
    assert "inet_mismo_b" in body
    assert "CFP 401" in body


# ---------- Defensivos R1: resume no duplica correos al destinatario ----------


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_worker_no_re_agrupa_si_destinatario_ya_recibio_envio_previo(
    settings, tmp_path
):
    """Si en un job ya hay una fila SENT con cierto destinatario y se reanuda
    con filas pendientes para el mismo destinatario, no se vuelve a agrupar:
    el destinatario ya recibió un correo, no debe recibir uno repetido con
    todas las credenciales."""
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_operator_r1",
        email="op@example.com",
        password="Pass!",
    )
    user_a = User.objects.create_user(
        username="reagr_a", email="reagr_a@x.com", password="Pass!"
    )
    user_b = User.objects.create_user(
        username="reagr_b", email="reagr_b@x.com", password="Pass!"
    )
    user_c = User.objects.create_user(
        username="reagr_c", email="reagr_c@x.com", password="Pass!"
    )
    for user in (user_a, user_b, user_c):
        _set_visible_temporary_password(user, "Pass!")

    upload = _build_excel_file(
        [
            ("reagr_a", "compartido@example.com"),
            ("reagr_b", "compartido@example.com"),
            ("reagr_c", "compartido@example.com"),
        ]
    )
    job = create_bulk_credentials_job(
        uploaded_file=upload, send_type="standard", requested_by=operator
    )
    BulkCredentialsJobRow.objects.create(
        job=job,
        fila=2,
        usuario="reagr_a",
        mail_destino="compartido@example.com",
        status=BulkCredentialsJobRow.Status.SENT,
        password_actualizada=False,
        mensaje="OK previo",
        attempts=1,
    )
    job.next_row_index = 1
    job.sent_rows = 1
    job.unchanged_password_rows = 1
    job.processed_rows = 1
    job.save(
        update_fields=[
            "next_row_index",
            "sent_rows",
            "unchanged_password_rows",
            "processed_rows",
        ]
    )

    process_bulk_credentials_job(job)
    job.refresh_from_db()

    assert job.status == BulkCredentialsJob.Status.COMPLETED
    assert len(mail.outbox) == 2
    assert all(msg.to == ["compartido@example.com"] for msg in mail.outbox)
    body_b = next(m.body for m in mail.outbox if "reagr_b" in m.body)
    assert "reagr_c" not in body_b
    body_c = next(m.body for m in mail.outbox if "reagr_c" in m.body)
    assert "reagr_b" not in body_c


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_worker_no_agrupa_cuando_primary_tiene_attempts_previos(settings, tmp_path):
    """Si la fila primaria ya tenía attempts > 0 (resume tras fallo posible
    post-envío), debe procesarse sola para acotar duplicación al destinatario."""
    settings.MEDIA_ROOT = tmp_path
    operator = User.objects.create_user(
        username="bulk_operator_attempts",
        email="op2@example.com",
        password="Pass!",
    )
    user_a = User.objects.create_user(
        username="att_a", email="att_a@x.com", password="Pass!"
    )
    user_b = User.objects.create_user(
        username="att_b", email="att_b@x.com", password="Pass!"
    )
    _set_visible_temporary_password(user_a, "Pass!")
    _set_visible_temporary_password(user_b, "Pass!")

    upload = _build_excel_file(
        [
            ("att_a", "compartido2@example.com"),
            ("att_b", "compartido2@example.com"),
        ]
    )
    job = create_bulk_credentials_job(
        uploaded_file=upload, send_type="standard", requested_by=operator
    )
    BulkCredentialsJobRow.objects.create(
        job=job,
        fila=2,
        usuario="att_a",
        mail_destino="compartido2@example.com",
        status=BulkCredentialsJobRow.Status.FAILED,
        password_actualizada=False,
        mensaje="fallo previo",
        attempts=1,
    )
    job.rejected_rows = 1
    job.processed_rows = 1
    job.save(update_fields=["rejected_rows", "processed_rows"])

    process_bulk_credentials_job(job)
    job.refresh_from_db()

    assert job.status == BulkCredentialsJob.Status.COMPLETED
    assert len(mail.outbox) == 2
    bodies = [m.body for m in mail.outbox]
    assert any("att_a" in b and "att_b" not in b for b in bodies)
    assert any("att_b" in b and "att_a" not in b for b in bodies)


# ---------- Defensivos R2: el cache evita N queries por destinatario ----------


@pytest.mark.django_db
@override_settings(EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend")
def test_recipient_cache_resuelve_destinatarios_con_una_query(
    django_assert_num_queries,
):
    """Cuando las filas no informan mail explícito, todos los emails de los
    usuarios se resuelven en una sola query, no una por fila."""
    from users.services_bulk_credentials import _build_recipient_cache
    from users.services_bulk_credentials import (
        _load_workbook_rows,
        get_bulk_credentials_send_type_config,
    )

    for index in range(5):
        User.objects.create_user(
            username=f"cache_user_{index}",
            email=f"cache_user_{index}@example.com",
            password="Pass!",
        )
    upload = _build_excel_file([(f"cache_user_{i}", "") for i in range(5)])
    config = get_bulk_credentials_send_type_config("standard")
    rows = _load_workbook_rows(upload, send_type_config=config)

    with django_assert_num_queries(1):
        cache = _build_recipient_cache(rows)

    assert len(cache) == 5
    assert cache["cache_user_0"] == "cache_user_0@example.com"
