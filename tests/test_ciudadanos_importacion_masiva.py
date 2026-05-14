from datetime import timedelta
from io import BytesIO

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.contrib.messages.storage.fallback import FallbackStorage
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import RequestFactory, override_settings
from django.contrib.sessions.middleware import SessionMiddleware
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from ciudadanos.models import Ciudadano, CiudadanosImportJob, CiudadanosImportJobRow
from ciudadanos.services_importacion_masiva import (
    generate_ciudadanos_import_template,
    generate_ciudadanos_import_job_results_workbook,
    load_ciudadanos_import_rows,
    parse_cuil_o_dni,
)
from ciudadanos.services_importacion_masiva_jobs import (
    STALE_JOB_ERROR_MESSAGE,
    can_resume_ciudadanos_import_job,
    create_ciudadanos_import_job,
    mark_stale_ciudadanos_import_jobs_as_failed,
    process_ciudadanos_import_job,
    request_resume_ciudadanos_import_job,
    run_ciudadanos_import_jobs_worker,
)
from ciudadanos.views import CiudadanosListView
from ciudadanos.views_importacion_masiva import (
    CiudadanosImportJobExportView,
    CiudadanosImportJobDetailView,
    CiudadanosImportJobResumeView,
    CiudadanosImportTemplateView,
    CiudadanosImportUploadView,
)
from core.models import Sexo

User = get_user_model()


@pytest.fixture(autouse=True)
def _ciudadanos_import_media_root(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path


def _build_excel_file(rows, headers=("cuil_o_dni", "sexo")):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ciudadanos"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))

    output = BytesIO()
    workbook.save(output)
    return SimpleUploadedFile(
        "ciudadanos.xlsx",
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


def _build_request(method: str, path: str, user, *, data=None, files=None):
    factory = RequestFactory()
    request_method = getattr(factory, method.lower())
    request = request_method(path, data=data or {})
    SessionMiddleware(lambda req: None).process_request(request)
    request.session.save()
    setattr(request, "_messages", FallbackStorage(request))
    request.user = user
    if files:
        request.FILES.update(files)
    return request


def _grant_add_ciudadano(user):
    permission = Permission.objects.get(
        content_type__app_label="ciudadanos",
        codename="add_ciudadano",
    )
    user.user_permissions.add(permission)
    return User.objects.get(pk=user.pk)


def _renaper_success(*, dni, sexo="M", cuil=None, nombre="JUAN", apellido="PEREZ"):
    return {
        "success": True,
        "data": {
            "cuil": int(cuil) if cuil else None,
            "dni": int(dni),
            "apellido": apellido,
            "nombre": nombre,
            "genero": sexo,
            "sexo": None,
            "tipo_documento": Ciudadano.DOCUMENTO_DNI,
            "fecha_nacimiento": "1990-01-02",
        },
        "datos_api": {
            "cuil": str(cuil or ""),
            "apellido": apellido,
            "nombres": nombre,
            "fechaNacimiento": "1990-01-02",
        },
    }


def _renaper_error(message, error_type):
    return {"success": False, "error": message, "error_type": error_type}


@pytest.mark.parametrize(
    "raw",
    [
        "20-44535030-4",
        "20.44535030.4",
        "20 44535030 4",
        "20445350304",
    ],
)
def test_parse_cuil_o_dni_accepts_cuil_formats(raw):
    parsed = parse_cuil_o_dni(raw)

    assert parsed.dni == "44535030"
    assert parsed.cuil == "20445350304"
    assert parsed.input_type == "cuil"


def test_parse_cuil_o_dni_accepts_direct_dni():
    parsed = parse_cuil_o_dni("44535030")

    assert parsed.dni == "44535030"
    assert parsed.cuil == ""
    assert parsed.input_type == "dni"


def test_parse_cuil_o_dni_rejects_short_direct_dni():
    with pytest.raises(ValidationError, match="DNI"):
        parse_cuil_o_dni("4435031")


def test_parse_cuil_o_dni_rejects_invalid_verifier():
    with pytest.raises(ValidationError, match="CUIL"):
        parse_cuil_o_dni("20-44535030-5")


def test_load_ciudadanos_import_rows_accepts_header_aliases_and_optional_sexo():
    upload = _build_excel_file(
        [
            ("20445350304", "Femenino"),
            ("30111222", ""),
        ],
        headers=("documento", "sexo"),
    )

    rows = load_ciudadanos_import_rows(upload)

    assert [row.fila for row in rows] == [2, 3]
    assert rows[0].dni == "44535030"
    assert rows[0].cuil == "20445350304"
    assert rows[0].sexo == "F"
    assert rows[1].dni == "30111222"
    assert rows[1].cuil == ""
    assert rows[1].sexo == ""


def test_load_ciudadanos_import_rows_keeps_invalid_row_for_history():
    upload = _build_excel_file([("20-44535030-5", "M")])

    rows = load_ciudadanos_import_rows(upload)

    assert len(rows) == 1
    assert rows[0].parse_error
    assert rows[0].error_type == "invalid_cuil"


def test_load_ciudadanos_import_rows_marks_short_dni_as_invalid_row():
    upload = _build_excel_file([("4435031", "M")])

    rows = load_ciudadanos_import_rows(upload)

    assert len(rows) == 1
    assert rows[0].parse_error
    assert rows[0].error_type == "invalid_dni"


def test_generate_ciudadanos_import_template_headers():
    content = generate_ciudadanos_import_template()
    workbook = load_workbook(BytesIO(content))
    worksheet = workbook.active

    header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    assert header == ["cuil_o_dni", "sexo"]


@pytest.mark.django_db
def test_create_ciudadanos_import_job_persists_pending_job():
    user = User.objects.create_user(
        username="ciudadanos_import_owner",
        email="owner@example.com",
        password="Secreta123!",
    )
    upload = _build_excel_file([("44535030", "M")])

    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)

    assert job.status == CiudadanosImportJob.Status.PENDING
    assert job.total_rows == 0
    assert job.original_filename == "ciudadanos.xlsx"
    assert job.requested_by == user
    assert job.archivo.name.endswith("ciudadanos.xlsx")


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_creates_existing_and_failed_rows(mocker):
    user = User.objects.create_user(username="ciudadanos_import_processor")
    existing = Ciudadano.objects.create(
        apellido="Existente",
        nombre="Persona",
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=44535030,
        cuil_cuit="20445350304",
        tipo_registro_identidad=Ciudadano.TIPO_REGISTRO_ESTANDAR,
    )
    upload = _build_excel_file(
        [
            ("44535030", "M"),
            ("30111222", "F"),
            ("20-44535030-5", "M"),
        ]
    )
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mock_consultar = mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        return_value=_renaper_success(
            dni="30111222",
            sexo="F",
            cuil="20301112220",
            nombre="Nueva",
            apellido="Ciudadana",
        ),
    )

    result = process_ciudadanos_import_job(job)
    result.refresh_from_db()

    assert result.status == CiudadanosImportJob.Status.COMPLETED_WITH_ERRORS
    assert result.total_rows == 3
    assert result.processed_rows == 3
    assert result.created_rows == 1
    assert result.existing_rows == 1
    assert result.failed_rows == 1
    assert result.pending_rows == 0
    mock_consultar.assert_called_once_with("30111222", "F")

    existing_row = result.rows.get(fila=2)
    created_row = result.rows.get(fila=3)
    failed_row = result.rows.get(fila=4)
    assert existing_row.status == CiudadanosImportJobRow.Status.EXISTING
    assert existing_row.ciudadano == existing
    assert created_row.status == CiudadanosImportJobRow.Status.CREATED
    assert created_row.ciudadano.documento == 30111222
    assert created_row.ciudadano.origen_dato == "renaper"
    assert created_row.ciudadano.estado_validacion_renaper == Ciudadano.RENAPER_VALIDADO
    assert created_row.ciudadano.datos_renaper["cuil"] == "20301112220"
    assert failed_row.status == CiudadanosImportJobRow.Status.FAILED
    assert failed_row.error_type == "invalid_cuil"


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_tries_all_sexes_when_missing(mocker):
    user = User.objects.create_user(username="ciudadanos_import_fallback")
    upload = _build_excel_file([("30111222", "")])
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mock_consultar = mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        side_effect=[
            _renaper_error("No se encontraron datos.", "not_found"),
            _renaper_error("No se encontraron datos.", "not_found"),
            _renaper_success(dni="30111222", sexo="X", cuil="20301112220"),
        ],
    )

    process_ciudadanos_import_job(job)
    row = CiudadanosImportJobRow.objects.get(job=job)

    assert [call.args for call in mock_consultar.call_args_list] == [
        ("30111222", "M"),
        ("30111222", "F"),
        ("30111222", "X"),
    ]
    assert row.status == CiudadanosImportJobRow.Status.CREATED
    assert row.sexos_intentados == "M,F,X"


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_fails_row_on_cuil_mismatch_and_continues(
    mocker,
):
    user = User.objects.create_user(username="ciudadanos_import_cuil_mismatch")
    upload = _build_excel_file(
        [
            ("20-44535030-4", "M"),
            ("30111222", "M"),
        ]
    )
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        side_effect=[
            _renaper_success(dni="44535030", cuil="20301112220"),
            _renaper_success(dni="30111222", cuil="20301112220"),
        ],
    )

    process_ciudadanos_import_job(job)
    job.refresh_from_db()

    assert job.status == CiudadanosImportJob.Status.COMPLETED_WITH_ERRORS
    assert job.created_rows == 1
    assert job.failed_rows == 1
    mismatch_row = job.rows.get(fila=2)
    created_row = job.rows.get(fila=3)
    assert mismatch_row.status == CiudadanosImportJobRow.Status.FAILED
    assert mismatch_row.error_type == "cuil_mismatch"
    assert created_row.status == CiudadanosImportJobRow.Status.CREATED


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_continues_after_invalid_short_dni(mocker):
    user = User.objects.create_user(username="ciudadanos_import_invalid_dni")
    upload = _build_excel_file(
        [
            ("4435031", "M"),
            ("30111222", "M"),
        ]
    )
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mock_consultar = mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        return_value=_renaper_success(dni="30111222", cuil="20301112220"),
    )

    process_ciudadanos_import_job(job)
    job.refresh_from_db()

    assert job.status == CiudadanosImportJob.Status.COMPLETED_WITH_ERRORS
    assert job.created_rows == 1
    assert job.failed_rows == 1
    mock_consultar.assert_called_once_with("30111222", "M")
    failed_row = job.rows.get(fila=2)
    assert failed_row.status == CiudadanosImportJobRow.Status.FAILED
    assert failed_row.error_type == "invalid_dni"


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_continues_after_renaper_unexpected_error(
    mocker,
):
    user = User.objects.create_user(username="ciudadanos_import_unexpected")
    upload = _build_excel_file(
        [
            ("44535032", "M"),
            ("30111222", "M"),
        ]
    )
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        side_effect=[
            _renaper_error(
                "Ocurrio un error inesperado al consultar RENAPER.",
                "unexpected_error",
            ),
            _renaper_success(dni="30111222", cuil="20301112220"),
        ],
    )

    process_ciudadanos_import_job(job)
    job.refresh_from_db()

    assert job.status == CiudadanosImportJob.Status.COMPLETED_WITH_ERRORS
    assert job.created_rows == 1
    assert job.failed_rows == 1
    assert job.pending_rows == 0
    failed_row = job.rows.get(fila=2)
    assert failed_row.status == CiudadanosImportJobRow.Status.FAILED
    assert failed_row.error_type == "unexpected_error"


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_records_unexpected_row_error_detail(mocker):
    user = User.objects.create_user(username="ciudadanos_import_row_exception")
    upload = _build_excel_file(
        [
            ("44535032", "M"),
            ("44535030", "M"),
        ]
    )
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mocker.patch(
        "ciudadanos.services_importacion_masiva_jobs.process_ciudadanos_import_row",
        side_effect=[
            ValueError("RENAPER devolvio DNI invalido para la fila."),
            {
                "status": "failed",
                "mensaje": "No se encontro coincidencia.",
                "error_type": "no_match",
                "sexos_intentados": "M",
                "ciudadano": None,
                "systemic": False,
                "contacted_renaper": True,
            },
        ],
    )

    process_ciudadanos_import_job(job)
    job.refresh_from_db()

    assert job.status == CiudadanosImportJob.Status.COMPLETED_WITH_ERRORS
    assert job.processed_rows == 2
    failed_row = job.rows.get(fila=2)
    assert failed_row.status == CiudadanosImportJobRow.Status.FAILED
    assert failed_row.error_type == "unexpected_row_error"
    assert "RENAPER devolvio DNI invalido" in failed_row.mensaje


@pytest.mark.django_db
@override_settings(CIUDADANOS_IMPORT_RENAPER_SLEEP_SECONDS=0)
def test_process_ciudadanos_import_job_pauses_on_systemic_error_and_resumes(mocker):
    user = User.objects.create_user(username="ciudadanos_import_pause")
    upload = _build_excel_file([("30111222", "M")])
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        return_value=_renaper_error("Timeout RENAPER.", "timeout"),
    )

    process_ciudadanos_import_job(job)
    job.refresh_from_db()

    assert job.status == CiudadanosImportJob.Status.FAILED
    assert job.processed_rows == 0
    assert job.pending_rows == 1
    assert job.next_row_index == 0
    assert can_resume_ciudadanos_import_job(job) is True
    row = job.rows.get(fila=2)
    assert row.status == CiudadanosImportJobRow.Status.PENDING
    assert row.error_type == "timeout"

    mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        return_value=_renaper_success(dni="30111222", cuil="20301112220"),
    )
    request_resume_ciudadanos_import_job(job=job)
    job.refresh_from_db()
    process_ciudadanos_import_job(job)
    job.refresh_from_db()

    assert job.status == CiudadanosImportJob.Status.COMPLETED
    row.refresh_from_db()
    assert row.status == CiudadanosImportJobRow.Status.CREATED
    assert row.attempts == 2


@pytest.mark.django_db
def test_mark_stale_ciudadanos_import_jobs_as_failed():
    user = User.objects.create_user(username="ciudadanos_import_stale")
    job = CiudadanosImportJob.objects.create(
        requested_by=user,
        original_filename="ciudadanos.xlsx",
        archivo="ciudadanos/import_jobs/test/ciudadanos.xlsx",
        status=CiudadanosImportJob.Status.PROCESSING,
        last_activity_at=timezone.now() - timedelta(seconds=901),
    )

    updated_count = mark_stale_ciudadanos_import_jobs_as_failed()
    job.refresh_from_db()

    assert updated_count == 1
    assert job.status == CiudadanosImportJob.Status.FAILED
    assert job.last_error_message == STALE_JOB_ERROR_MESSAGE


@pytest.mark.django_db
def test_ciudadanos_import_template_view_requires_add_permission():
    user = User.objects.create_user(username="ciudadanos_import_no_perm")

    with pytest.raises(PermissionDenied):
        CiudadanosImportTemplateView.as_view()(
            _build_request("get", "/ciudadanos/importacion-masiva/plantilla/", user),
        )

    user = _grant_add_ciudadano(user)
    response = CiudadanosImportTemplateView.as_view()(
        _build_request("get", "/ciudadanos/importacion-masiva/plantilla/", user),
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        response["Content-Disposition"]
        == 'attachment; filename="plantilla_ciudadanos_importacion_masiva.xlsx"'
    )


@pytest.mark.django_db
def test_ciudadanos_import_upload_view_creates_job_without_processing():
    user = _grant_add_ciudadano(
        User.objects.create_user(username="ciudadanos_uploader")
    )
    upload = _build_excel_file([("44535030", "M")])
    request = _build_request(
        "post",
        "/ciudadanos/importacion-masiva/",
        user,
        files={"archivo": upload},
    )

    response = CiudadanosImportUploadView.as_view()(request)

    assert response.status_code == 302
    job = CiudadanosImportJob.objects.get()
    assert response.url.endswith(f"/ciudadanos/importacion-masiva/lotes/{job.pk}/")
    assert job.status == CiudadanosImportJob.Status.PENDING
    assert CiudadanosImportJobRow.objects.count() == 0


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="config.urls")
def test_ciudadanos_list_context_shows_import_button_only_with_add_permission():
    user = User.objects.create_user(username="ciudadanos_button")
    request = _build_request("get", "/ciudadanos/listar", user)

    view = CiudadanosListView()
    view.request = request
    context = view.get_context_data(object_list=Ciudadano.objects.none())
    assert context.get("additional_buttons") in (None, [])

    user = _grant_add_ciudadano(user)
    request = _build_request("get", "/ciudadanos/listar", user)
    view.request = request
    context = view.get_context_data(object_list=Ciudadano.objects.none())

    assert context["additional_buttons"] == [
        {
            "label": "IMPORTACION MASIVA",
            "url": "/ciudadanos/importacion-masiva/",
            "class": "btn btn-lg btn-export-csv",
            "title": "Importar ciudadanos desde Excel",
        }
    ]


@pytest.mark.django_db
def test_ciudadanos_import_job_detail_view_lists_global_history_row():
    user = _grant_add_ciudadano(User.objects.create_user(username="ciudadanos_detail"))
    other = User.objects.create_user(username="ciudadanos_detail_other")
    ciudadano = Ciudadano.objects.create(
        apellido="Detalle",
        nombre="Ciudadano",
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111222,
        tipo_registro_identidad=Ciudadano.TIPO_REGISTRO_ESTANDAR,
    )
    job = CiudadanosImportJob.objects.create(
        requested_by=other,
        original_filename="ciudadanos.xlsx",
        archivo="ciudadanos/import_jobs/test/ciudadanos.xlsx",
        status=CiudadanosImportJob.Status.COMPLETED,
        total_rows=1,
        processed_rows=1,
        created_rows=1,
    )
    row = CiudadanosImportJobRow.objects.create(
        job=job,
        fila=2,
        documento_raw="30111222",
        dni="30111222",
        status=CiudadanosImportJobRow.Status.CREATED,
        ciudadano=ciudadano,
    )

    response = CiudadanosImportJobDetailView.as_view()(
        _build_request(
            "get",
            f"/ciudadanos/importacion-masiva/lotes/{job.pk}/",
            user,
        ),
        pk=job.pk,
    )

    assert response.status_code == 200
    assert response.context_data["job"] == job
    assert list(response.context_data["rows_page"].object_list) == [row]


@pytest.mark.django_db
def test_generate_ciudadanos_import_job_results_workbook_matches_visible_rows():
    user = User.objects.create_user(username="ciudadanos_export_results")
    ciudadano = Ciudadano.objects.create(
        apellido="Exportado",
        nombre="Ciudadano",
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111222,
        tipo_registro_identidad=Ciudadano.TIPO_REGISTRO_ESTANDAR,
    )
    job = CiudadanosImportJob.objects.create(
        requested_by=user,
        original_filename="ciudadanos.xlsx",
        archivo="ciudadanos/import_jobs/test/ciudadanos.xlsx",
        status=CiudadanosImportJob.Status.COMPLETED_WITH_ERRORS,
        total_rows=4,
        processed_rows=3,
        created_rows=1,
        existing_rows=1,
        failed_rows=1,
        pending_rows=1,
    )
    CiudadanosImportJobRow.objects.create(
        job=job,
        fila=2,
        documento_raw="30111222",
        dni="30111222",
        status=CiudadanosImportJobRow.Status.CREATED,
        ciudadano=ciudadano,
        mensaje="Ciudadano creado desde RENAPER.",
    )
    CiudadanosImportJobRow.objects.create(
        job=job,
        fila=3,
        documento_raw="44535030",
        dni="44535030",
        status=CiudadanosImportJobRow.Status.EXISTING,
        ciudadano=ciudadano,
        mensaje="Ya existe.",
    )
    CiudadanosImportJobRow.objects.create(
        job=job,
        fila=4,
        documento_raw="44535032",
        dni="44535032",
        status=CiudadanosImportJobRow.Status.FAILED,
        error_type="unexpected_error",
        mensaje="Ocurrio un error inesperado al consultar RENAPER.",
    )
    CiudadanosImportJobRow.objects.create(
        job=job,
        fila=5,
        documento_raw="40111222",
        dni="40111222",
        status=CiudadanosImportJobRow.Status.PENDING,
    )

    content = generate_ciudadanos_import_job_results_workbook(job)
    workbook = load_workbook(BytesIO(content))

    assert workbook.sheetnames == ["filas"]
    worksheet = workbook["filas"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Fila",
        "CUIL/DNI",
        "DNI",
        "Sexo",
        "Resultado",
        "Estado",
        "Intentos sexo",
        "Intentos",
        "Detalle",
        "Ciudadano",
    )
    assert "error_type" not in rows[0]
    assert "ciudadano_id" not in rows[0]
    assert [row[4] for row in rows[1:]] == ["OK", "OK", "Fallo", "Pendiente"]
    assert [row[5] for row in rows[1:]] == [
        "Creado",
        "Existente",
        "Fallido",
        "Pendiente",
    ]
    assert (
        rows[3][8]
        == "Ocurrio un error inesperado al consultar RENAPER. (unexpected_error)"
    )
    assert worksheet["J2"].value == "Ver"
    assert worksheet["J2"].hyperlink.target.endswith(f"/ciudadanos/ver/{ciudadano.pk}")


@pytest.mark.django_db
def test_ciudadanos_import_job_export_view_downloads_xlsx():
    user = _grant_add_ciudadano(User.objects.create_user(username="ciudadanos_export"))
    job = CiudadanosImportJob.objects.create(
        requested_by=user,
        original_filename="ciudadanos.xlsx",
        archivo="ciudadanos/import_jobs/test/ciudadanos.xlsx",
        status=CiudadanosImportJob.Status.COMPLETED,
    )
    CiudadanosImportJobRow.objects.create(
        job=job,
        fila=2,
        documento_raw="30111222",
        dni="30111222",
        status=CiudadanosImportJobRow.Status.CREATED,
    )

    response = CiudadanosImportJobExportView.as_view()(
        _build_request(
            "get",
            f"/ciudadanos/importacion-masiva/lotes/{job.pk}/exportar/",
            user,
        ),
        pk=job.pk,
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response["Content-Disposition"].startswith(
        'attachment; filename="resultado_importacion_ciudadanos_'
    )


@pytest.mark.django_db
def test_ciudadanos_import_resume_view_marks_failed_job_pending():
    user = _grant_add_ciudadano(User.objects.create_user(username="ciudadanos_resume"))
    job = CiudadanosImportJob.objects.create(
        requested_by=user,
        original_filename="ciudadanos.xlsx",
        archivo="ciudadanos/import_jobs/test/ciudadanos.xlsx",
        status=CiudadanosImportJob.Status.FAILED,
        last_error_message="Timeout.",
        last_error_at=timezone.now(),
        finished_at=timezone.now(),
    )

    response = CiudadanosImportJobResumeView.as_view()(
        _build_request(
            "post",
            f"/ciudadanos/importacion-masiva/lotes/{job.pk}/reanudar/",
            user,
        ),
        pk=job.pk,
    )
    job.refresh_from_db()

    assert response.status_code == 302
    assert job.status == CiudadanosImportJob.Status.PENDING
    assert job.resume_count == 1


@pytest.mark.django_db
def test_process_ciudadanos_import_jobs_command_invokes_worker_once(mocker):
    mock_worker = mocker.patch(
        "ciudadanos.management.commands.process_ciudadanos_import_jobs."
        "run_ciudadanos_import_jobs_worker"
    )

    call_command("process_ciudadanos_import_jobs", "--once")

    mock_worker.assert_called_once_with(once=True)


@pytest.mark.django_db
def test_run_ciudadanos_import_jobs_worker_once_processes_one_cycle(mocker):
    mock_process_next = mocker.patch(
        "ciudadanos.services_importacion_masiva_jobs."
        "process_next_ciudadanos_import_job",
        return_value=False,
    )

    run_ciudadanos_import_jobs_worker(once=True)

    mock_process_next.assert_called_once_with()


@pytest.mark.django_db
def test_process_ciudadanos_import_job_uses_existing_sexo_catalog(mocker):
    sexo = Sexo.objects.create(sexo="Masculino")
    user = User.objects.create_user(username="ciudadanos_sexo_catalog")
    upload = _build_excel_file([("30111222", "M")])
    job = create_ciudadanos_import_job(uploaded_file=upload, requested_by=user)
    mocker.patch(
        "ciudadanos.services_importacion_masiva.consultar_datos_renaper",
        return_value=_renaper_success(
            dni="30111222",
            sexo="M",
            cuil="20301112220",
        ),
    )

    process_ciudadanos_import_job(job)

    ciudadano = Ciudadano.objects.get(documento=30111222)
    assert ciudadano.sexo == sexo
