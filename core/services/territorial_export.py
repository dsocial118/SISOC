"""Exporta la jerarquía territorial desde fixtures JSON a Excel."""

from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PROVINCIA_MODEL = "core.provincia"
MUNICIPIO_MODEL = "core.municipio"
LOCALIDAD_MODEL = "core.localidad"


def _sort_records(records):
    return sorted(records, key=lambda item: (item["nombre"].lower(), item["id"]))


def _load_fixture_records(fixture_path: Path):
    with fixture_path.open(encoding="utf-8") as fixture_file:
        payload = json.load(fixture_file)

    provincias = {}
    municipios = {}
    localidades = {}

    for row in payload:
        model = row.get("model")
        pk = row.get("pk")
        fields = row.get("fields", {})

        if model == PROVINCIA_MODEL:
            provincias[pk] = {"id": pk, "nombre": fields.get("nombre", "")}
        elif model == MUNICIPIO_MODEL:
            municipios[pk] = {
                "id": pk,
                "nombre": fields.get("nombre", ""),
                "provincia_id": fields.get("provincia"),
            }
        elif model == LOCALIDAD_MODEL:
            localidades[pk] = {
                "id": pk,
                "nombre": fields.get("nombre", ""),
                "municipio_id": fields.get("municipio"),
            }

    return provincias, municipios, localidades


def _index_municipios_by_provincia(municipios, provincias):
    municipios_by_provincia = defaultdict(list)
    for municipio in municipios.values():
        if municipio["provincia_id"] in provincias:
            municipios_by_provincia[municipio["provincia_id"]].append(municipio)
    return municipios_by_provincia


def _index_localidades_by_municipio(localidades, municipios):
    localidades_by_municipio = defaultdict(list)
    for localidad in localidades.values():
        if localidad["municipio_id"] in municipios:
            localidades_by_municipio[localidad["municipio_id"]].append(localidad)
    return localidades_by_municipio


def _build_hierarchy_rows(
    provincias, municipios_by_provincia, localidades_by_municipio
):
    jerarquia_rows = []
    municipios_sin_localidades_rows = []
    provincias_sin_municipios_rows = []
    resumen_rows = []

    for provincia in _sort_records(provincias.values()):
        municipios_validos = _sort_records(
            municipios_by_provincia.get(provincia["id"], [])
        )
        if not municipios_validos:
            provincias_sin_municipios_rows.append(
                (provincia["id"], provincia["nombre"])
            )
            jerarquia_rows.append(
                (
                    provincia["id"],
                    provincia["nombre"],
                    None,
                    None,
                    None,
                    None,
                    "provincia_sin_municipios",
                )
            )
            resumen_rows.append((provincia["id"], provincia["nombre"], 0, 0, 0))
            continue

        total_localidades = 0
        total_municipios_sin_localidades = 0
        for municipio in municipios_validos:
            localidades_validas = _sort_records(
                localidades_by_municipio.get(municipio["id"], [])
            )
            if not localidades_validas:
                total_municipios_sin_localidades += 1
                municipios_sin_localidades_rows.append(
                    (
                        provincia["id"],
                        provincia["nombre"],
                        municipio["id"],
                        municipio["nombre"],
                    )
                )
                jerarquia_rows.append(
                    (
                        provincia["id"],
                        provincia["nombre"],
                        municipio["id"],
                        municipio["nombre"],
                        None,
                        None,
                        "municipio_sin_localidades",
                    )
                )
                continue

            total_localidades += len(localidades_validas)
            for localidad in localidades_validas:
                jerarquia_rows.append(
                    (
                        provincia["id"],
                        provincia["nombre"],
                        municipio["id"],
                        municipio["nombre"],
                        localidad["id"],
                        localidad["nombre"],
                        "ok",
                    )
                )

        resumen_rows.append(
            (
                provincia["id"],
                provincia["nombre"],
                len(municipios_validos),
                total_localidades,
                total_municipios_sin_localidades,
            )
        )

    return (
        jerarquia_rows,
        municipios_sin_localidades_rows,
        provincias_sin_municipios_rows,
        resumen_rows,
    )


def _build_inconsistencias_rows(municipios, provincias, localidades):
    inconsistencias_rows = []
    for municipio in _sort_records(municipios.values()):
        if municipio["provincia_id"] not in provincias:
            inconsistencias_rows.append(
                (
                    "municipio_sin_provincia_valida",
                    municipio["id"],
                    municipio["nombre"],
                    municipio["provincia_id"],
                    "La provincia referenciada no existe en el fixture.",
                )
            )

    for localidad in _sort_records(localidades.values()):
        if localidad["municipio_id"] not in municipios:
            inconsistencias_rows.append(
                (
                    "localidad_sin_municipio_valido",
                    localidad["id"],
                    localidad["nombre"],
                    localidad["municipio_id"],
                    "El municipio referenciado no existe en el fixture.",
                )
            )

    return inconsistencias_rows


def build_fixture_relations_report(fixture_path: Path):
    """Construye las filas del reporte territorial a partir del fixture."""

    provincias, municipios, localidades = _load_fixture_records(fixture_path)
    municipios_by_provincia = _index_municipios_by_provincia(municipios, provincias)
    localidades_by_municipio = _index_localidades_by_municipio(localidades, municipios)
    (
        jerarquia_rows,
        municipios_sin_localidades_rows,
        provincias_sin_municipios_rows,
        resumen_rows,
    ) = _build_hierarchy_rows(
        provincias,
        municipios_by_provincia,
        localidades_by_municipio,
    )
    inconsistencias_rows = _build_inconsistencias_rows(
        municipios,
        provincias,
        localidades,
    )

    return {
        "jerarquia": {
            "headers": (
                "provincia_id",
                "provincia",
                "municipio_id",
                "municipio",
                "localidad_id",
                "localidad",
                "estado_relacion",
            ),
            "rows": jerarquia_rows,
        },
        "municipios_sin_localidades": {
            "headers": ("provincia_id", "provincia", "municipio_id", "municipio"),
            "rows": municipios_sin_localidades_rows,
        },
        "provincias_sin_municipios": {
            "headers": ("provincia_id", "provincia"),
            "rows": provincias_sin_municipios_rows,
        },
        "inconsistencias": {
            "headers": (
                "tipo",
                "registro_id",
                "nombre",
                "referencia_id",
                "detalle",
            ),
            "rows": inconsistencias_rows,
        },
        "resumen": {
            "headers": (
                "provincia_id",
                "provincia",
                "municipios_validos",
                "localidades_validas",
                "municipios_sin_localidades",
            ),
            "rows": resumen_rows,
        },
        "summary": {
            "provincias": len(provincias),
            "municipios": len(municipios),
            "localidades": len(localidades),
            "provincias_sin_municipios": len(provincias_sin_municipios_rows),
            "municipios_sin_localidades": len(municipios_sin_localidades_rows),
            "inconsistencias": len(inconsistencias_rows),
        },
    }


def _apply_header_style(worksheet, headers):
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"


def _set_column_widths(worksheet):
    for column_cells in worksheet.columns:
        values = [
            "" if cell.value is None else str(cell.value) for cell in column_cells
        ]
        max_length = max(len(value) for value in values)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(
            max(max_length + 2, 14),
            60,
        )


def export_fixture_relations_workbook(fixture_path: Path, output_path: Path):
    """Genera un workbook con la jerarquía territorial y devuelve su resumen."""

    fixture_path = Path(fixture_path)
    output_path = Path(output_path)

    report = build_fixture_relations_report(fixture_path)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name in (
        "jerarquia",
        "municipios_sin_localidades",
        "provincias_sin_municipios",
        "inconsistencias",
        "resumen",
    ):
        sheet_report = report[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name)
        _apply_header_style(worksheet, sheet_report["headers"])
        for row in sheet_report["rows"]:
            worksheet.append(row)
        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        _set_column_widths(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return report["summary"]
