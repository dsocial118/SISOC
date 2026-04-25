import json
import tempfile
import unittest
from io import StringIO
from pathlib import Path

from openpyxl import load_workbook

from core.management.commands.export_relaciones_territoriales_fixture import Command
from core.services.territorial_export import export_fixture_relations_workbook


def _write_fixture(base_dir: Path, payload):
    fixture_path = base_dir / "territorial_fixture.json"
    fixture_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return fixture_path


class ExportRelacionesTerritorialesTests(unittest.TestCase):
    def test_export_fixture_relations_workbook_creates_expected_sheets_and_gaps(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            fixture_path = _write_fixture(
                tmp_path,
                [
                    {
                        "model": "core.provincia",
                        "pk": 1,
                        "fields": {"nombre": "Buenos Aires"},
                    },
                    {
                        "model": "core.provincia",
                        "pk": 2,
                        "fields": {"nombre": "Catamarca"},
                    },
                    {
                        "model": "core.municipio",
                        "pk": 10,
                        "fields": {"nombre": "La Plata", "provincia": 1},
                    },
                    {
                        "model": "core.municipio",
                        "pk": 11,
                        "fields": {"nombre": "Rosario", "provincia": 99},
                    },
                    {
                        "model": "core.localidad",
                        "pk": 100,
                        "fields": {"nombre": "Tolosa", "municipio": 10},
                    },
                    {
                        "model": "core.localidad",
                        "pk": 101,
                        "fields": {"nombre": "Sin Municipio", "municipio": 999},
                    },
                ],
            )
            output_path = tmp_path / "relaciones.xlsx"

            summary = export_fixture_relations_workbook(fixture_path, output_path)

            workbook = load_workbook(output_path)

            self.assertEqual(
                workbook.sheetnames,
                [
                    "jerarquia",
                    "municipios_sin_localidades",
                    "provincias_sin_municipios",
                    "inconsistencias",
                    "resumen",
                ],
            )

            hierarchy_rows = list(workbook["jerarquia"].iter_rows(values_only=True))
            self.assertEqual(
                hierarchy_rows[0],
                (
                    "provincia_id",
                    "provincia",
                    "municipio_id",
                    "municipio",
                    "localidad_id",
                    "localidad",
                    "estado_relacion",
                ),
            )
            self.assertEqual(
                hierarchy_rows[1],
                (1, "Buenos Aires", 10, "La Plata", 100, "Tolosa", "ok"),
            )
            self.assertEqual(
                hierarchy_rows[2],
                (
                    2,
                    "Catamarca",
                    None,
                    None,
                    None,
                    None,
                    "provincia_sin_municipios",
                ),
            )

            provinces_without_municipios = list(
                workbook["provincias_sin_municipios"].iter_rows(values_only=True)
            )
            self.assertEqual(provinces_without_municipios[1], (2, "Catamarca"))

            inconsistencies = list(
                workbook["inconsistencias"].iter_rows(values_only=True)
            )
            self.assertIn(
                (
                    "municipio_sin_provincia_valida",
                    11,
                    "Rosario",
                    99,
                    "La provincia referenciada no existe en el fixture.",
                ),
                inconsistencies,
            )
            self.assertIn(
                (
                    "localidad_sin_municipio_valido",
                    101,
                    "Sin Municipio",
                    999,
                    "El municipio referenciado no existe en el fixture.",
                ),
                inconsistencies,
            )

            self.assertEqual(
                summary,
                {
                    "provincias": 2,
                    "municipios": 2,
                    "localidades": 2,
                    "provincias_sin_municipios": 1,
                    "municipios_sin_localidades": 0,
                    "inconsistencias": 2,
                },
            )

    def test_command_generates_workbook(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            fixture_path = _write_fixture(
                tmp_path,
                [
                    {
                        "model": "core.provincia",
                        "pk": 1,
                        "fields": {"nombre": "Buenos Aires"},
                    },
                    {
                        "model": "core.municipio",
                        "pk": 10,
                        "fields": {"nombre": "La Plata", "provincia": 1},
                    },
                    {
                        "model": "core.localidad",
                        "pk": 100,
                        "fields": {"nombre": "Tolosa", "municipio": 10},
                    },
                ],
            )
            output_path = tmp_path / "salida" / "relaciones.xlsx"
            out = StringIO()

            Command(stdout=out).handle(
                fixture=str(fixture_path), output=str(output_path)
            )

            self.assertTrue(output_path.exists())
            self.assertIn("Archivo generado", out.getvalue())
