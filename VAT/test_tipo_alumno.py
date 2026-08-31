"""Clasificación "Tipo de alumno" (VAT / Sin Plan) en listados y exportaciones.

La regla vive en `VAT/services/tipo_alumno_service.py`: voucher con estado
"activo" y vencimiento vigente → VAT; cualquier otro caso → Sin Plan. Estos
tests cubren el servicio, la nómina Excel, el detalle nominal del reporte y el
render de asistencia/detalle de comisión (camino Curso).

Se usa `RequestFactory` en lugar del test client a propósito: el venv local
corre Python 3.14 con Django 4.2 y la instrumentación de templates del test
client falla ahí (ver test_comision_detail_template_compartido.py).
"""

import codecs
import csv
from datetime import time, timedelta
from io import BytesIO
from types import SimpleNamespace

import pytest
from django.contrib.auth.models import Group, Permission, User
from django.contrib.contenttypes.models import ContentType
from django.test import RequestFactory
from django.utils import timezone
from openpyxl import load_workbook

from ciudadanos.models import Ciudadano
from core.models import Dia, Localidad, Municipio, Programa, Provincia, Sexo
from VAT.models import (
    Centro,
    ComisionCurso,
    ComisionHorario,
    Curso,
    Inscripcion,
    InstitucionUbicacion,
    ModalidadCursada,
    SesionComision,
    Voucher,
)
from VAT.services.nomina_export import build_comision_curso_nomina_excel
from VAT.services.reportes_inscripciones_asistencia import (
    ReporteFiltros,
    build_detalle_queryset,
    export_detalle_to_csv,
    export_rows_to_csv,
)
from VAT.services.tipo_alumno_service import (
    TIPO_ALUMNO_SIN_PLAN,
    TIPO_ALUMNO_VAT,
    anotar_tipo_alumno,
)
from VAT.views.curso import AsistenciaSesionCursoView, ComisionCursoDetailView


def test_reporte_agrupado_csv_declara_utf8_e_incluye_bom_para_excel():
    response = export_rows_to_csv(
        [{"grupo": "Comisión Ñ · Córdoba", "inscripciones_total": 1}],
        group_by="comision",
    )

    assert response["Content-Type"] == "text/csv; charset=utf-8"
    assert response.content.startswith(codecs.BOM_UTF8)
    assert "Comisión Ñ · Córdoba" in response.content.decode("utf-8-sig")


def _crear_ciudadano(documento, apellido="Alumno"):
    return Ciudadano.objects.create(
        apellido=apellido,
        nombre="Tipo",
        fecha_nacimiento=timezone.localdate() - timedelta(days=9000),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=documento,
    )


def _crear_voucher(ciudadano, programa, *, estado="activo", dias_vencimiento=30):
    return Voucher.objects.create(
        ciudadano=ciudadano,
        programa=programa,
        cantidad_inicial=5,
        cantidad_usada=0,
        cantidad_disponible=5,
        fecha_vencimiento=timezone.localdate() + timedelta(days=dias_vencimiento),
        estado=estado,
    )


@pytest.mark.django_db
def test_anotar_tipo_alumno_clasifica_segun_voucher_activo():
    programa = Programa.objects.create(nombre="Programa tipo alumno")
    con_voucher = _crear_ciudadano(45100001)
    voucher_vencido_por_fecha = _crear_ciudadano(45100002)
    voucher_agotado = _crear_ciudadano(45100003)
    sin_voucher = _crear_ciudadano(45100004)

    _crear_voucher(con_voucher, programa)
    # Estado "activo" pero vencido por fecha: el estado se estampa perezosamente.
    _crear_voucher(voucher_vencido_por_fecha, programa, dias_vencimiento=-1)
    _crear_voucher(voucher_agotado, programa, estado="agotado")

    filas = anotar_tipo_alumno(
        [
            SimpleNamespace(ciudadano_id=con_voucher.pk),
            SimpleNamespace(ciudadano_id=voucher_vencido_por_fecha.pk),
            SimpleNamespace(ciudadano_id=voucher_agotado.pk),
            SimpleNamespace(ciudadano_id=sin_voucher.pk),
        ]
    )

    assert [fila.tipo_alumno for fila in filas] == [
        TIPO_ALUMNO_VAT,
        TIPO_ALUMNO_SIN_PLAN,
        TIPO_ALUMNO_SIN_PLAN,
        TIPO_ALUMNO_SIN_PLAN,
    ]
    assert filas[0].es_alumno_vat is True
    assert filas[1].es_alumno_vat is False


@pytest.fixture
def escenario_tipo_alumno(db):
    provincia = Provincia.objects.create(nombre="BA tipo")
    municipio = Municipio.objects.create(nombre="LP tipo", provincia=provincia)
    localidad = Localidad.objects.create(nombre="Tolosa tipo", municipio=municipio)
    modalidad = ModalidadCursada.objects.create(nombre="Pres tipo", activo=True)
    programa = Programa.objects.create(nombre="Prog tipo")
    Sexo.objects.get_or_create(sexo="F tipo")
    Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="tipo-admin", email="tipo@vat.test", password="x"
    )
    permission, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Group),
        codename="role_centroreferentevat",
        defaults={"name": "ReferenteCentroVAT legacy"},
    )
    user.user_permissions.add(permission)
    centro = Centro.objects.create(
        nombre="CFP tipo",
        codigo="CFP-TIPO",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="1",
        numero=1,
        domicilio_actividad="C 1",
        telefono="1",
        celular="1",
        correo="tipo@a.test",
        nombre_referente="A",
        apellido_referente="B",
        telefono_referente="1",
        correo_referente="tipo@b.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="C 1",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro, nombre="Curso tipo", modalidad=modalidad, estado="activo"
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="TIPO-CC",
        nombre="CC tipo",
        cupo_total=10,
        fecha_inicio=timezone.localdate() - timedelta(days=10),
        fecha_fin=timezone.localdate() + timedelta(days=10),
        estado="activa",
    )
    alumno_vat = _crear_ciudadano(45200001, apellido="Convoucher")
    alumno_sin_plan = _crear_ciudadano(45200002, apellido="Sinvoucher")
    _crear_voucher(alumno_vat, programa)
    for ciudadano in (alumno_vat, alumno_sin_plan):
        Inscripcion.objects.create(
            ciudadano=ciudadano,
            comision_curso=comision,
            programa=programa,
            estado="inscripta",
            origen_canal="backoffice",
        )
    return user, comision


def _crear_sesion(comision):
    dia, _ = Dia.objects.get_or_create(nombre="Lunes")
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(10, 0),
        hora_hasta=time(12, 0),
        aula_espacio="Aula 1",
        vigente=True,
    )
    return SesionComision.objects.create(
        comision_curso=comision,
        horario=horario,
        numero_sesion=1,
        fecha=timezone.localdate(),
        estado="programada",
    )


@pytest.mark.django_db
def test_nomina_excel_incluye_tipo_alumno(escenario_tipo_alumno):
    _, comision = escenario_tipo_alumno

    content = build_comision_curso_nomina_excel(
        comision,
        Inscripcion.objects.filter(comision_curso=comision).select_related(
            "ciudadano__sexo"
        ),
    )

    worksheet = load_workbook(BytesIO(content)).active
    header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    rows = [
        dict(zip(header, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    tipos = {row["Apellido"]: row["Tipo de alumno"] for row in rows}

    assert "Tipo de alumno" in header
    assert tipos == {
        "Convoucher": TIPO_ALUMNO_VAT,
        "Sinvoucher": TIPO_ALUMNO_SIN_PLAN,
    }


@pytest.mark.django_db
def test_detalle_reporte_y_export_clasifican_tipo_alumno(escenario_tipo_alumno):
    user, _ = escenario_tipo_alumno

    filas = {
        row["ciudadano__apellido"]: row["tiene_voucher_activo"]
        for row in build_detalle_queryset(user, ReporteFiltros())
    }
    assert filas == {"Convoucher": True, "Sinvoucher": False}

    response = export_detalle_to_csv(user, ReporteFiltros())
    lineas = list(csv.reader(response.content.decode("utf-8-sig").splitlines()))
    header = lineas[0]
    indice_apellido = header.index("Apellido")
    indice_tipo = header.index("Tipo de alumno")
    tipos = {linea[indice_apellido]: linea[indice_tipo] for linea in lineas[1:]}

    assert tipos == {
        "Convoucher": TIPO_ALUMNO_VAT,
        "Sinvoucher": TIPO_ALUMNO_SIN_PLAN,
    }


def _render(view, user, **kwargs):
    request = RequestFactory().get("/x/")
    request.user = user
    request.csp_nonce = "n"
    return view.as_view()(request, **kwargs).render().content.decode()


@pytest.mark.django_db
def test_asistencia_sesion_curso_muestra_tipo_alumno(escenario_tipo_alumno):
    user, comision = escenario_tipo_alumno
    sesion = _crear_sesion(comision)

    html = _render(AsistenciaSesionCursoView, user, sesion_pk=sesion.pk)

    assert "<th>Tipo de alumno</th>" in html
    assert html.count('class="asis-chip-vat"') == 1
    assert html.count('class="asis-tipo-sin-plan"') == 1


@pytest.mark.django_db
def test_comision_curso_detail_muestra_tipo_alumno(escenario_tipo_alumno):
    user, comision = escenario_tipo_alumno

    html = _render(ComisionCursoDetailView, user, pk=comision.pk)
    panel = html.split('data-sisoc-panel="inscriptos"')[1].split(
        'data-sisoc-panel="sesiones"'
    )[0]

    assert "<th>Tipo de alumno</th>" in panel
    assert panel.count('class="ci-pill ci-pill-teal"') == 1
    assert f'class="ci-pill ci-pill-teal">{TIPO_ALUMNO_VAT}</span>' in panel
    assert f'class="ci-pill ci-pill-gray">{TIPO_ALUMNO_SIN_PLAN}</span>' in panel
