from __future__ import annotations

import csv
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from openpyxl import load_workbook

from VAT.models import Centro, InstitucionContacto, InstitucionIdentificadorHist
from VAT.models import InstitucionUbicacion
from core.models import Localidad, Municipio, Provincia

HEADER_ALIASES = {
    "nombre": {"nombre", "centro", "nombre_centro"},
    "codigo": {"codigo", "cue", "codigo_cue"},
    "provincia_id": {"provincia_id", "jurisdiccion_id"},
    "municipio_id": {"municipio_id", "partido_id"},
    "localidad_id": {"localidad_id"},
    "domicilio_actividad": {"domicilio_actividad", "direccion", "domicilio"},
    "telefono": {"telefono", "telefono_institucional"},
    "correo": {"correo", "email", "correo_institucional"},
    "nombre_referente": {"nombre_referente"},
    "apellido_referente": {"apellido_referente"},
    "telefono_referente": {"telefono_referente"},
    "correo_referente": {"correo_referente"},
    "referente_id": {"referente_id"},
    "tipo_gestion": {"tipo_gestion"},
    "clase_institucion": {"clase_institucion"},
    "situacion": {"situacion", "estado_etp"},
    "autoridad_dni": {"autoridad_dni", "documento_autoridad", "director_dni"},
    "contacto_nombre": {"contacto_nombre", "nombre_contacto"},
    "contacto_rol_area": {"contacto_rol_area", "contacto_area", "rol_area"},
    "contacto_telefono": {"contacto_telefono", "telefono_contacto"},
    "contacto_email": {"contacto_email", "email_contacto"},
    "contacto_es_principal": {"contacto_es_principal", "es_principal"},
}
REQUIRED_FIELDS = ("nombre", "codigo")
FANTASY_DOMAIN = "vat.local"
TIPO_GESTION_MAP = {
    "estatal": "Estatal",
    "privada": "Privada",
    "privado": "Privada",
}
CLASE_INSTITUCION_MAP = {
    "formacion profesional": "Formación Profesional",
    "secundario tecnico": "Secundario Técnico",
    "superior formacion docente": "Superior Formación Docente",
    "superior tecnico": "Superior Técnico",
    "secundario orientado": "Secundario Orientado",
}
SITUACION_MAP = {
    "institucion de etp": "Institución de ETP",
    "institucion de otro nivel y/o modalidad": (
        "Institución de Otro Nivel y/o Modalidad"
    ),
}


@dataclass(frozen=True)
class ParsedCentroRow:
    line_number: int
    nombre: str
    codigo: str
    provincia_id: int | None
    municipio_id: int | None
    localidad_id: int | None
    domicilio_actividad: str
    telefono: str
    correo: str
    nombre_referente: str
    apellido_referente: str
    telefono_referente: str
    correo_referente: str
    referente_id: int | None
    tipo_gestion: str
    clase_institucion: str
    situacion: str
    autoridad_dni: str
    contacto_nombre: str
    contacto_rol_area: str
    contacto_telefono: str
    contacto_email: str
    contacto_es_principal: bool


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    normalized = normalized.lower().replace(" ", "_").replace("-", "_")
    normalized = re.sub(r"[^a-z0-9_]+", "", normalized)
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _build_normalized_choice_map(raw_mapping: dict[str, str]) -> dict[str, str]:
    normalized_mapping: dict[str, str] = {}
    for raw_key, value in raw_mapping.items():
        normalized_mapping[normalize_header(raw_key).replace("_", " ")] = value
    return normalized_mapping


NORMALIZED_TIPO_GESTION_MAP = _build_normalized_choice_map(TIPO_GESTION_MAP)
NORMALIZED_CLASE_INSTITUCION_MAP = _build_normalized_choice_map(CLASE_INSTITUCION_MAP)
NORMALIZED_SITUACION_MAP = _build_normalized_choice_map(SITUACION_MAP)


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def resolve_header_mapping(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field_name not in mapping:
                mapping[field_name] = index
                break

    missing_fields = [
        field_name for field_name in REQUIRED_FIELDS if field_name not in mapping
    ]
    if missing_fields:
        raise CommandError(
            "El archivo no contiene las columnas requeridas: "
            f"{', '.join(missing_fields)}. Encabezados detectados: {', '.join(headers)}"
        )

    return mapping


def iter_csv_rows(file_path: Path) -> tuple[list[str], list[tuple[int, list[str]]]]:
    with file_path.open(newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    if not rows:
        raise CommandError("El archivo CSV está vacío.")

    headers = [clean_cell(value) for value in rows[0]]
    data_rows = [
        (index, [clean_cell(value) for value in row])
        for index, row in enumerate(rows[1:], start=2)
    ]
    return headers, data_rows


def iter_excel_rows(
    file_path: Path, sheet_name: str | None
) -> tuple[list[str], list[tuple[int, list[str]]]]:
    workbook = load_workbook(filename=file_path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"La hoja '{sheet_name}' no existe. Hojas disponibles: "
                f"{', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise CommandError("El archivo Excel está vacío.")

    headers = [clean_cell(value) for value in rows[0]]
    data_rows = [
        (index, [clean_cell(value) for value in row])
        for index, row in enumerate(rows[1:], start=2)
    ]
    return headers, data_rows


def load_rows(
    file_path: Path, sheet_name: str | None
) -> tuple[list[str], list[tuple[int, list[str]]]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return iter_csv_rows(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return iter_excel_rows(file_path, sheet_name)
    raise CommandError("Formato no soportado. Use un archivo .csv, .xlsx o .xlsm.")


def parse_int(value: str, field_label: str) -> int | None:
    cleaned_value = (value or "").strip()
    if not cleaned_value:
        return None
    if not cleaned_value.isdigit():
        raise ValueError(f"{field_label} debe ser numérico")
    return int(cleaned_value)


def normalize_choice(value: str, mapping: dict[str, str], field_label: str) -> str:
    cleaned_value = (value or "").strip()
    if not cleaned_value:
        return ""
    normalized_key = normalize_header(cleaned_value).replace("_", " ")
    resolved = mapping.get(normalized_key)
    if not resolved:
        raise ValueError(f"{field_label} no reconocido: '{cleaned_value}'")
    return resolved


def parse_bool(value: str) -> bool:
    return (value or "").strip().lower() in {"1", "si", "sí", "true", "on", "x"}


def build_fantasy_email(*, prefix: str, codigo: str) -> str:
    base = slugify(prefix).replace("-", "")[:20] or "centro"
    return f"{base}{codigo[-4:]}@{FANTASY_DOMAIN}"


def build_contact_fantasy_email(*, codigo: str) -> str:
    return f"contacto{codigo[-6:]}@{FANTASY_DOMAIN}"


def normalize_codigo(codigo: str) -> str:
    cleaned_codigo = (codigo or "").strip()
    if not cleaned_codigo or not cleaned_codigo.isdigit():
        raise ValueError("el código/CUE debe ser numérico de hasta 9 dígitos")
    if len(cleaned_codigo) > 9:
        raise ValueError("el código/CUE debe ser numérico de hasta 9 dígitos")
    return cleaned_codigo.zfill(9)


def build_parsed_row(
    *,
    line_number: int,
    raw_values: list[str],
    header_mapping: dict[str, int],
) -> ParsedCentroRow:
    def get_value(field_name: str) -> str:
        index = header_mapping.get(field_name)
        if index is None or index >= len(raw_values):
            return ""
        return raw_values[index]

    nombre = get_value("nombre")
    codigo = normalize_codigo(get_value("codigo"))
    if not nombre:
        raise ValueError("el nombre es obligatorio")

    tipo_gestion = normalize_choice(
        get_value("tipo_gestion"),
        NORMALIZED_TIPO_GESTION_MAP,
        "tipo_gestion",
    )
    clase_institucion = normalize_choice(
        get_value("clase_institucion"),
        NORMALIZED_CLASE_INSTITUCION_MAP,
        "clase_institucion",
    )
    situacion = normalize_choice(
        get_value("situacion"),
        NORMALIZED_SITUACION_MAP,
        "situacion",
    )

    correo = get_value("correo") or build_fantasy_email(prefix="centro", codigo=codigo)

    return ParsedCentroRow(
        line_number=line_number,
        nombre=nombre,
        codigo=codigo,
        provincia_id=parse_int(get_value("provincia_id"), "provincia_id"),
        municipio_id=parse_int(get_value("municipio_id"), "municipio_id"),
        localidad_id=parse_int(get_value("localidad_id"), "localidad_id"),
        domicilio_actividad=get_value("domicilio_actividad"),
        telefono=get_value("telefono"),
        correo=correo,
        nombre_referente=get_value("nombre_referente"),
        apellido_referente=get_value("apellido_referente"),
        telefono_referente=get_value("telefono_referente"),
        correo_referente=get_value("correo_referente"),
        referente_id=parse_int(get_value("referente_id"), "referente_id"),
        tipo_gestion=tipo_gestion,
        clase_institucion=clase_institucion,
        situacion=situacion,
        autoridad_dni=get_value("autoridad_dni"),
        contacto_nombre=get_value("contacto_nombre"),
        contacto_rol_area=get_value("contacto_rol_area"),
        contacto_telefono=get_value("contacto_telefono"),
        contacto_email=get_value("contacto_email"),
        contacto_es_principal=parse_bool(get_value("contacto_es_principal")),
    )


def resolve_foreign_keys(parsed_row: ParsedCentroRow):
    provincia = None
    municipio = None
    localidad = None
    referente = None

    if parsed_row.provincia_id is not None:
        provincia = Provincia.objects.filter(pk=parsed_row.provincia_id).first()
        if provincia is None:
            raise ValueError(f"provincia_id inexistente: {parsed_row.provincia_id}")

    if parsed_row.municipio_id is not None:
        municipio = Municipio.objects.filter(pk=parsed_row.municipio_id).first()
        if municipio is None:
            raise ValueError(f"municipio_id inexistente: {parsed_row.municipio_id}")

    if parsed_row.localidad_id is not None:
        localidad = Localidad.objects.filter(pk=parsed_row.localidad_id).first()
        if localidad is None:
            raise ValueError(f"localidad_id inexistente: {parsed_row.localidad_id}")

    if provincia and municipio and municipio.provincia_id != provincia.id:
        raise ValueError("municipio_id no pertenece a provincia_id")
    if municipio and localidad and localidad.municipio_id != municipio.id:
        raise ValueError("localidad_id no pertenece a municipio_id")
    if provincia and localidad and localidad.municipio.provincia_id != provincia.id:
        raise ValueError("localidad_id no pertenece a provincia_id")

    if parsed_row.referente_id is not None:
        referente = User.objects.filter(pk=parsed_row.referente_id).first()
        if referente is None:
            raise ValueError(f"referente_id inexistente: {parsed_row.referente_id}")
        if not referente.groups.filter(name="CFP").exists():
            raise ValueError("referente_id no pertenece al grupo CFP")

    return provincia, municipio, localidad, referente


def sync_principal_contact(centro: Centro, contacto: InstitucionContacto) -> None:
    centro.nombre_referente = contacto.nombre_contacto or ""
    centro.apellido_referente = ""
    centro.telefono_referente = contacto.telefono_contacto or ""
    centro.correo_referente = contacto.email_contacto or ""
    centro.save(
        update_fields=[
            "nombre_referente",
            "apellido_referente",
            "telefono_referente",
            "correo_referente",
        ]
    )


def should_create_contact(parsed_row: ParsedCentroRow) -> bool:
    return any(
        [
            parsed_row.contacto_nombre,
            parsed_row.contacto_rol_area,
            parsed_row.autoridad_dni,
            parsed_row.contacto_telefono,
            parsed_row.contacto_email,
        ]
    )


class Command(BaseCommand):
    help = "Importa centros VAT desde un archivo Excel/CSV con layout institucional."

    progress_step = 100

    def _write_status(self, message: str) -> None:
        self.stdout.write(message)
        if hasattr(self.stdout, "flush"):
            self.stdout.flush()

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)
        parser.add_argument("--sheet-name", type=str, default=None)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        sheet_name = options["sheet_name"]
        dry_run = options["dry_run"]

        if not file_path.exists():
            raise CommandError(f"El archivo '{file_path}' no existe.")

        headers, raw_rows = load_rows(file_path, sheet_name)
        header_mapping = resolve_header_mapping(headers)
        total_rows = sum(1 for _, row in raw_rows if any(row))

        created_count = 0
        updated_count = 0
        skipped_count = 0
        processed_count = 0

        self._write_status(
            f"Inicio importación VAT centros. Filas a procesar: {total_rows}."
        )

        for line_number, raw_values in raw_rows:
            if not any(raw_values):
                continue

            try:
                parsed_row = build_parsed_row(
                    line_number=line_number,
                    raw_values=raw_values,
                    header_mapping=header_mapping,
                )
                provincia, municipio, localidad, referente = resolve_foreign_keys(
                    parsed_row
                )
            except ValueError as exc:
                skipped_count += 1
                self.stdout.write(
                    self.style.WARNING(f"Fila {line_number}: {exc}. Se omite.")
                )
                continue

            with transaction.atomic():
                centro = Centro.objects.filter(codigo=parsed_row.codigo).first()
                created = centro is None
                if centro is None:
                    centro = Centro(codigo=parsed_row.codigo)

                centro.nombre = parsed_row.nombre
                centro.provincia = provincia
                centro.municipio = municipio
                centro.localidad = localidad
                centro.domicilio_actividad = parsed_row.domicilio_actividad
                centro.telefono = parsed_row.telefono
                centro.celular = ""
                centro.correo = parsed_row.correo
                centro.nombre_referente = parsed_row.nombre_referente
                centro.apellido_referente = parsed_row.apellido_referente
                centro.telefono_referente = parsed_row.telefono_referente
                centro.correo_referente = parsed_row.correo_referente
                centro.referente = referente
                centro.tipo_gestion = parsed_row.tipo_gestion
                centro.clase_institucion = parsed_row.clase_institucion
                centro.situacion = parsed_row.situacion
                centro.activo = True

                if not dry_run:
                    centro.save()

                    if localidad is not None:
                        ubicacion = (
                            centro.ubicaciones.filter(rol_ubicacion="sede_principal")
                            .order_by("-es_principal", "id")
                            .first()
                        )
                        if ubicacion is None:
                            ubicacion = InstitucionUbicacion(centro=centro)
                        ubicacion.localidad = localidad
                        ubicacion.rol_ubicacion = "sede_principal"
                        ubicacion.domicilio = parsed_row.domicilio_actividad
                        ubicacion.es_principal = True
                        ubicacion.save()
                    else:
                        ubicacion = None

                    identificador = (
                        centro.identificadores_hist.filter(tipo_identificador="cue")
                        .order_by("-es_actual", "id")
                        .first()
                    )
                    if identificador is None:
                        identificador = InstitucionIdentificadorHist(
                            centro=centro,
                            tipo_identificador="cue",
                        )
                    identificador.valor_identificador = parsed_row.codigo
                    identificador.rol_institucional = "sede"
                    identificador.ubicacion = ubicacion
                    identificador.es_actual = True
                    identificador.save()

                    if should_create_contact(parsed_row):
                        contact_email = (
                            parsed_row.contacto_email
                            or build_contact_fantasy_email(codigo=parsed_row.codigo)
                        )
                        if not parsed_row.contacto_telefono and not contact_email:
                            raise CommandError(
                                f"Fila {line_number}: el contacto requiere teléfono o email."
                            )

                        contacto = centro.contactos_adicionales.filter(
                            email_contacto=contact_email
                        ).first()
                        if contacto is None and parsed_row.contacto_telefono:
                            contacto = centro.contactos_adicionales.filter(
                                telefono_contacto=parsed_row.contacto_telefono
                            ).first()
                        if contacto is None:
                            contacto = InstitucionContacto(centro=centro)

                        contacto.nombre_contacto = parsed_row.contacto_nombre
                        contacto.rol_area = parsed_row.contacto_rol_area
                        contacto.documento = parsed_row.autoridad_dni
                        contacto.telefono_contacto = parsed_row.contacto_telefono
                        contacto.email_contacto = contact_email
                        contacto.tipo = "email" if contact_email else "telefono"
                        contacto.valor = contact_email or parsed_row.contacto_telefono
                        contacto.es_principal = parsed_row.contacto_es_principal
                        contacto.save()

                        if contacto.es_principal:
                            centro.contactos_adicionales.exclude(pk=contacto.pk).filter(
                                es_principal=True
                            ).update(es_principal=False)
                            sync_principal_contact(centro, contacto)

                processed_count += 1
                if created:
                    created_count += 1
                else:
                    updated_count += 1

                if processed_count % self.progress_step == 0:
                    self._write_status(
                        "Avance centros VAT: "
                        f"{processed_count}/{total_rows} procesados, "
                        f"{created_count} creados, {updated_count} actualizados, "
                        f"{skipped_count} omitidos."
                    )

        mode = "Simulación finalizada" if dry_run else "Proceso finalizado"
        self._write_status(
            self.style.SUCCESS(
                f"{mode}. Procesados: {processed_count}, creados: {created_count}, "
                f"actualizados: {updated_count}, omitidos: {skipped_count}."
            )
        )
