import importlib
from io import BytesIO
from datetime import date, time
import json

import pytest
from bs4 import BeautifulSoup
from django.contrib.auth.models import Group, User
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from django.core.exceptions import ValidationError
from django.test import RequestFactory, override_settings
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework.test import APIClient
from rest_framework_api_key.models import APIKey

from VAT import serializers as vat_serializers
from VAT.api_views import CursoViewSet
from VAT.forms import (
    ComisionCursoForm,
    CentroAltaForm,
    CursoForm,
    InstitucionContactoAltaForm,
    InstitucionContactoForm,
    PlanVersionCurricularForm,
)
from VAT.views import centro as centro_views
from VAT.models import (
    Centro,
    Sector,
    Subsector,
    TituloReferencia,
    Curso,
    OfertaInstitucional,
    Comision,
    ComisionCurso,
    ComisionHorario,
    Inscripcion,
    SolicitudInscripcionPublica,
    InstitucionContacto,
    InstitucionIdentificadorHist,
    InstitucionUbicacion,
    ModalidadCursada,
    PlanVersionCurricular,
    SesionComision,
    Voucher,
    VoucherParametria,
)
from VAT.services.access_scope import filter_centros_queryset_for_user, is_vat_referente
from ciudadanos.models import Ciudadano
from core.models import Dia, Localidad, Municipio, Provincia, Programa, Sexo
from users.models import Profile


@pytest.fixture
def vat_geo_data(db):
    provincia = Provincia.objects.create(nombre="Buenos Aires")
    municipio = Municipio.objects.create(nombre="La Plata", provincia=provincia)
    localidad = Localidad.objects.create(nombre="Tolosa", municipio=municipio)
    return provincia, municipio, localidad


@pytest.fixture
def vat_referente_user(db):
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_user(username="referente-vat", password="test1234")
    user.groups.add(group)
    return user


@pytest.fixture
def vat_admin_client(client, db):
    user = User.objects.create_superuser(
        username="admin-vat",
        email="admin@vat.test",
        password="test1234",
    )
    client.force_login(user)
    return client


def _grant_vat_referente_access(user, *permissions):
    legacy_permission, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Group),
        codename="role_centroreferentevat",
        defaults={"name": "ReferenteCentroVAT legacy"},
    )
    user.user_permissions.add(legacy_permission, *permissions)


@pytest.fixture
def vat_api_key(db):
    _, key = APIKey.objects.create_key(name="vat-tests")
    return key


@pytest.fixture
def vat_api_client(vat_api_key):
    client = APIClient()
    client.credentials(HTTP_AUTHORIZATION=f"Api-Key {vat_api_key}")
    return client


def _build_centro_payload(referente_user, provincia, municipio, localidad, **overrides):
    payload = {
        "nombre": "Centro de Formación 401",
        "codigo": "500144900",
        "provincia": str(provincia.pk),
        "municipio": str(municipio.pk),
        "localidad": str(localidad.pk),
        "calle": "7",
        "numero": "1234",
        "domicilio_actividad": "Calle 7 N° 1234",
        "codigo_postal": "1900",
        "lote": "12",
        "manzana": "B",
        "entre_calles": "45 y 46",
        "telefono": "221-4000000",
        "celular": "221-5000000",
        "correo": "institucion@vat.test",
        "sitio_web": "https://vat.test",
        "nombre_referente": "Ana",
        "apellido_referente": "Pérez",
        "autoridad_dni": "30111222",
        "telefono_referente": "221-4111111",
        "correo_referente": "direccion@vat.test",
        "referente": str(referente_user.pk),
        "activo": "on",
        "tipo_gestion": "Estatal",
        "clase_institucion": "Formación Profesional",
        "situacion": "Institución de ETP",
        "contactos-TOTAL_FORMS": "1",
        "contactos-INITIAL_FORMS": "0",
        "contactos-MIN_NUM_FORMS": "0",
        "contactos-MAX_NUM_FORMS": "1000",
        "contactos-0-nombre_contacto": "María Gómez",
        "contactos-0-rol_area": "Administración",
        "contactos-0-documento": "30111222",
        "contactos-0-telefono_contacto": "221-4222222",
        "contactos-0-email_contacto": "maria@vat.test",
        "contactos-0-es_principal": "on",
    }
    payload.update(overrides)
    return payload


def _assign_user_profile_provincia(user, provincia, es_usuario_provincial=False):
    Profile.objects.update_or_create(
        user=user,
        defaults={
            "es_usuario_provincial": es_usuario_provincial,
            "provincia": provincia,
        },
    )


@pytest.mark.django_db
def test_centro_create_crea_entidades_relacionadas(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    payload = _build_centro_payload(
        vat_referente_user, provincia, municipio, localidad, save_continue="1"
    )

    response = vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")

    assert response.status_code == 302
    assert response.url == reverse("vat_centro_detail", kwargs={"pk": centro.pk})
    contacto = InstitucionContacto.objects.get(
        centro=centro,
        nombre_contacto="María Gómez",
    )
    assert contacto.es_principal is True
    assert contacto.documento == "30111222"
    assert centro.nombre_referente == "María Gómez"
    assert centro.correo_referente == "maria@vat.test"
    assert InstitucionIdentificadorHist.objects.filter(
        centro=centro,
        tipo_identificador="cue",
        valor_identificador="500144900",
    ).exists()
    assert InstitucionUbicacion.objects.filter(
        centro=centro,
        rol_ubicacion="sede_principal",
        es_principal=True,
    ).exists()


@pytest.mark.django_db
def test_centro_create_permite_guardar_sin_contactos_institucionales(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    payload = _build_centro_payload(
        vat_referente_user,
        provincia,
        municipio,
        localidad,
        **{
            "nombre_referente": "",
            "apellido_referente": "",
            "telefono_referente": "",
            "correo_referente": "",
            "autoridad_dni": "",
            "contactos-TOTAL_FORMS": "0",
            "contactos-INITIAL_FORMS": "0",
        },
    )

    response = vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")

    assert response.status_code == 302
    assert centro.contactos_adicionales.count() == 0
    assert centro.nombre_referente == ""
    assert centro.correo_referente == ""


@pytest.mark.django_db
def test_centro_create_rechaza_referente_sin_grupo_cfp(vat_admin_client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    user_sin_cfp = User.objects.create_user(
        username="sin-cfp",
        email="sin-cfp@vat.test",
        password="test1234",
    )
    payload = _build_centro_payload(
        user_sin_cfp,
        provincia,
        municipio,
        localidad,
    )
    payload.pop("provincia")

    response = vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    assert response.status_code == 200
    assert Centro.objects.filter(codigo="500144900").count() == 0
    assert (
        "El referente seleccionado debe tener un rol valido de referente VAT."
        in response.content.decode("utf-8")
    )


@pytest.mark.django_db
def test_centro_alta_form_configura_referente_como_buscador_simple(vat_referente_user):
    vat_referente_user.first_name = "Ana"
    vat_referente_user.last_name = "Pérez"
    vat_referente_user.save(update_fields=["first_name", "last_name"])

    form = CentroAltaForm()
    referente_field = form.fields["referente"]

    assert referente_field.widget.attrs["class"] == "referente-hidden-select"
    assert referente_field.widget.attrs["tabindex"] == "-1"
    assert referente_field.empty_label == "Seleccionar referente..."
    assert (
        referente_field.label_from_instance(vat_referente_user)
        == "referente-vat - Ana Pérez"
    )
    assert form.referente_search_options == [
        {
            "id": str(vat_referente_user.pk),
            "username": "referente-vat",
            "label": "referente-vat - Ana Pérez",
        }
    ]


@pytest.mark.django_db
def test_centro_alta_form_no_incluye_grupos_legacy_de_referente():
    legacy_group_vat, _ = Group.objects.get_or_create(name="ReferenteCentroVAT")
    legacy_group_centro, _ = Group.objects.get_or_create(name="ReferenteCentro")
    legacy_user_vat = User.objects.create_user(
        username="referente-legacy-form",
        password="test1234",
    )
    legacy_user_centro = User.objects.create_user(
        username="referente-centro-form",
        password="test1234",
    )
    legacy_user_vat.groups.add(legacy_group_vat)
    legacy_user_centro.groups.add(legacy_group_centro)

    form = CentroAltaForm()
    queryset_usernames = list(
        form.fields["referente"]
        .queryset.order_by("username")
        .values_list("username", flat=True)
    )

    assert "referente-legacy-form" not in queryset_usernames
    assert "referente-centro-form" not in queryset_usernames


@pytest.mark.django_db
def test_centro_update_renderiza_mismo_formulario_extendido_que_alta(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    payload = _build_centro_payload(
        vat_referente_user, provincia, municipio, localidad, save_continue="1"
    )
    vat_admin_client.post(reverse("vat_centro_create"), data=payload)
    centro = Centro.objects.get(codigo="500144900")

    response = vat_admin_client.get(
        reverse("vat_centro_update", kwargs={"pk": centro.pk})
    )

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert "contactos-TOTAL_FORMS" in content
    assert "3.2 Contactos de la institución" in content
    assert "4. Autoridades" not in content
    assert 'name="contactos-0-documento"' in content
    assert 'name="save_continue"' not in content
    assert 'for="id_provincia"' not in content
    assert 'name="provincia"' in content
    assert 'name="activo_present"' in content
    assert "4. Estado de la sede" in content


@pytest.mark.django_db
def test_centro_update_conserva_activo_si_el_form_no_envia_el_switch(
    client, vat_geo_data
):
    provincia_ba, municipio_ba, localidad_ba = vat_geo_data
    user = User.objects.create_superuser(
        username="admin-vat-update-activo",
        email="admin-vat-update-activo@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba)
    group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-update-activo",
        email="referente-update-activo@vat.test",
        password="test1234",
    )
    referente.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Estado Editable",
        codigo="500144999",
        provincia=provincia_ba,
        municipio=municipio_ba,
        localidad=localidad_ba,
        calle="12",
        numero=100,
        domicilio_actividad="Calle 12 N° 100",
        telefono="221-1111111",
        celular="221-2222222",
        correo="cfp-estado@vat.test",
        nombre_referente="Ana",
        apellido_referente="Perez",
        telefono_referente="221-3333333",
        correo_referente="ana@vat.test",
        referente=referente,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    contacto = InstitucionContacto.objects.create(
        centro=centro,
        nombre_contacto="Ana Perez",
        documento="30111223",
        rol_area="Dirección",
        telefono_contacto="221-3333333",
        email_contacto="ana@vat.test",
        es_principal=True,
    )
    client.force_login(user)

    update_payload = _build_centro_payload(
        referente,
        provincia_ba,
        municipio_ba,
        localidad_ba,
        nombre="CFP Estado Editable",
        codigo="500144999",
        **{
            "contactos-TOTAL_FORMS": "1",
            "contactos-INITIAL_FORMS": "1",
            "contactos-MIN_NUM_FORMS": "0",
            "contactos-MAX_NUM_FORMS": "1000",
            "contactos-0-id": str(contacto.id),
            "contactos-0-centro": str(centro.id),
            "contactos-0-nombre_contacto": "Ana Perez",
            "contactos-0-rol_area": "Dirección",
            "contactos-0-documento": "30111223",
            "contactos-0-telefono_contacto": "221-3333333",
            "contactos-0-email_contacto": "ana@vat.test",
            "contactos-0-es_principal": "on",
        },
    )
    update_payload.pop("provincia", None)
    update_payload.pop("activo", None)

    response = client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    centro.refresh_from_db()

    assert response.status_code == 302
    assert centro.activo is True


@pytest.mark.django_db
def test_centro_update_permite_desactivar_activo_desde_switch(client, vat_geo_data):
    provincia_ba, municipio_ba, localidad_ba = vat_geo_data
    user = User.objects.create_superuser(
        username="admin-vat-desactiva-centro",
        email="admin-vat-desactiva-centro@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba)
    group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-desactiva-centro",
        email="referente-desactiva-centro@vat.test",
        password="test1234",
    )
    referente.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Estado Editable",
        codigo="500144998",
        provincia=provincia_ba,
        municipio=municipio_ba,
        localidad=localidad_ba,
        calle="12",
        numero=100,
        domicilio_actividad="Calle 12 N° 100",
        telefono="221-1111111",
        celular="221-2222222",
        correo="cfp-estado-off@vat.test",
        nombre_referente="Ana",
        apellido_referente="Perez",
        telefono_referente="221-3333333",
        correo_referente="ana@vat.test",
        referente=referente,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    contacto = InstitucionContacto.objects.create(
        centro=centro,
        nombre_contacto="Ana Perez",
        documento="30111223",
        rol_area="Dirección",
        telefono_contacto="221-3333333",
        email_contacto="ana@vat.test",
        es_principal=True,
    )
    client.force_login(user)

    update_payload = _build_centro_payload(
        referente,
        provincia_ba,
        municipio_ba,
        localidad_ba,
        nombre="CFP Estado Editable",
        codigo="500144998",
        activo_present="1",
        **{
            "contactos-TOTAL_FORMS": "1",
            "contactos-INITIAL_FORMS": "1",
            "contactos-MIN_NUM_FORMS": "0",
            "contactos-MAX_NUM_FORMS": "1000",
            "contactos-0-id": str(contacto.id),
            "contactos-0-centro": str(centro.id),
            "contactos-0-nombre_contacto": "Ana Perez",
            "contactos-0-rol_area": "Dirección",
            "contactos-0-documento": "30111223",
            "contactos-0-telefono_contacto": "221-3333333",
            "contactos-0-email_contacto": "ana@vat.test",
            "contactos-0-es_principal": "on",
        },
    )
    update_payload.pop("provincia", None)
    update_payload.pop("activo", None)

    response = client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    centro.refresh_from_db()

    assert response.status_code == 302
    assert centro.activo is False


@pytest.mark.django_db
def test_centro_update_rechaza_formset_contactos_sin_ids_existentes(
    client, vat_geo_data
):
    provincia_ba, municipio_ba, localidad_ba = vat_geo_data
    user = User.objects.create_superuser(
        username="admin-vat-contactos-sin-ids",
        email="admin-vat-contactos-sin-ids@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba)
    group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-contactos-sin-ids",
        email="referente-contactos-sin-ids@vat.test",
        password="test1234",
    )
    referente.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Contactos Seguros",
        codigo="500145000",
        provincia=provincia_ba,
        municipio=municipio_ba,
        localidad=localidad_ba,
        calle="12",
        numero=100,
        domicilio_actividad="Calle 12 N° 100",
        telefono="221-1111111",
        celular="221-2222222",
        correo="cfp-contactos@vat.test",
        nombre_referente="Ana",
        apellido_referente="Perez",
        telefono_referente="221-3333333",
        correo_referente="ana@vat.test",
        referente=referente,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    contacto = InstitucionContacto.objects.create(
        centro=centro,
        nombre_contacto="Contacto Original",
        documento="30111224",
        rol_area="Dirección",
        telefono_contacto="221-3333333",
        email_contacto="original@vat.test",
        es_principal=True,
    )
    client.force_login(user)

    update_payload = _build_centro_payload(
        referente,
        provincia_ba,
        municipio_ba,
        localidad_ba,
        nombre="CFP Contactos Seguros",
        codigo="500145000",
        **{
            "contactos-TOTAL_FORMS": "1",
            "contactos-INITIAL_FORMS": "0",
            "contactos-MIN_NUM_FORMS": "0",
            "contactos-MAX_NUM_FORMS": "1000",
            "contactos-0-centro": str(centro.id),
            "contactos-0-nombre_contacto": "Contacto Editado Sin ID",
            "contactos-0-rol_area": "Dirección",
            "contactos-0-documento": "30111224",
            "contactos-0-telefono_contacto": "221-4444444",
            "contactos-0-email_contacto": "editado@vat.test",
            "contactos-0-es_principal": "on",
        },
    )
    update_payload.pop("provincia", None)

    response = client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    contacto.refresh_from_db()

    assert response.status_code == 200
    assert response.context["form"].non_field_errors() == [
        (
            "No se pudo guardar la edición de contactos porque faltan los "
            "identificadores de filas existentes. Recargá la página e intentá "
            "nuevamente."
        )
    ]
    assert contacto.nombre_contacto == "Contacto Original"


@pytest.mark.django_db
def test_centro_create_oculta_jurisdiccion_y_asigna_provincia_del_usuario(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)

    response = vat_admin_client.get(reverse("vat_centro_create"))

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert 'for="id_provincia"' not in content
    assert "Seleccionar jurisdicción" not in content

    payload = _build_centro_payload(vat_referente_user, provincia, municipio, localidad)
    payload.pop("provincia")

    post_response = vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")
    assert post_response.status_code == 302
    assert centro.provincia_id == provincia.id


@pytest.mark.django_db
def test_centro_create_filtra_municipios_por_provincia_del_usuario(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    provincia_sf = Provincia.objects.create(nombre="Santa Fe")
    municipio_ba = Municipio.objects.create(nombre="La Plata", provincia=provincia_ba)
    Municipio.objects.create(nombre="Rosario", provincia=provincia_sf)
    Localidad.objects.create(nombre="Tolosa", municipio=municipio_ba)

    user = User.objects.create_superuser(
        username="admin-vat-filtro-municipio",
        email="admin-vat-filtro-municipio@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba)

    client.force_login(user)
    response = client.get(reverse("vat_centro_create"))

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert "La Plata" in content
    assert "Rosario" not in content


@pytest.mark.django_db
def test_centro_create_muestra_provincia_y_permite_alta_si_usuario_no_tiene_provincia(
    client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    user = User.objects.create_superuser(
        username="admin-vat-sin-provincia",
        email="admin-vat-sin-provincia@vat.test",
        password="test1234",
    )

    client.force_login(user)
    response = client.get(reverse("vat_centro_create"))

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert 'for="id_provincia"' in content

    payload = _build_centro_payload(vat_referente_user, provincia, municipio, localidad)
    post_response = client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")
    assert post_response.status_code == 302
    assert centro.provincia_id == provincia.id


@pytest.mark.django_db
def test_centro_update_actualiza_entidades_relacionadas_del_formulario_extendido(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    payload = _build_centro_payload(
        vat_referente_user, provincia, municipio, localidad, save_continue="1"
    )
    vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")
    contacto = centro.contactos_adicionales.get(nombre_contacto="María Gómez")

    update_payload = _build_centro_payload(
        vat_referente_user,
        provincia,
        municipio,
        localidad,
        nombre="Centro de Formación 402",
        codigo="500144901",
        domicilio_actividad="Calle 8 N° 4321",
        nombre_referente="Laura",
        apellido_referente="Gómez",
        telefono_referente="221-4999999",
        correo_referente="laura@vat.test",
        **{
            "contactos-TOTAL_FORMS": "1",
            "contactos-INITIAL_FORMS": "1",
            "contactos-MIN_NUM_FORMS": "0",
            "contactos-MAX_NUM_FORMS": "1000",
            "contactos-0-id": str(contacto.id),
            "contactos-0-centro": str(centro.id),
            "contactos-0-nombre_contacto": "Laura Gómez",
            "contactos-0-rol_area": "Dirección",
            "contactos-0-documento": "30999888",
            "contactos-0-telefono_contacto": "221-4333333",
            "contactos-0-email_contacto": "direccion2@vat.test",
            "contactos-0-es_principal": "on",
        },
    )

    response = vat_admin_client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    centro.refresh_from_db()
    identificador = centro.identificadores_hist.get(tipo_identificador="cue")
    ubicacion = centro.ubicaciones.get(rol_ubicacion="sede_principal")
    contacto.refresh_from_db()

    assert response.status_code == 302
    assert centro.nombre == "Centro de Formación 402"
    assert centro.nombre_referente == "Laura Gómez"
    assert centro.correo_referente == "direccion2@vat.test"
    assert identificador.valor_identificador == "500144901"
    assert ubicacion.domicilio == "Calle 8 N° 4321"
    assert contacto.nombre_contacto == "Laura Gómez"
    assert contacto.documento == "30999888"
    assert contacto.email_contacto == "direccion2@vat.test"


@pytest.mark.django_db
def test_institucion_contacto_alta_form_rechaza_documento_no_numerico():
    form = InstitucionContactoAltaForm(
        data={
            "nombre_contacto": "María Gómez",
            "rol_area": "Administración",
            "documento": "30A11222",
            "telefono_contacto": "221-4222222",
            "email_contacto": "maria@vat.test",
            "es_principal": "on",
        }
    )

    assert not form.is_valid()
    assert form.errors["documento"] == ["El documento debe contener solo números."]


@pytest.mark.django_db
def test_institucion_contacto_forms_renderizan_documento_como_numerico():
    alta_form = InstitucionContactoAltaForm()
    admin_form = InstitucionContactoForm()

    for form in [alta_form, admin_form]:
        attrs = form.fields["documento"].widget.attrs
        assert attrs["inputmode"] == "numeric"
        assert attrs["pattern"] == "[0-9]*"
        assert attrs["maxlength"] == "20"


def test_institucion_contacto_forms_no_requieren_campos_visibles():
    alta_form = InstitucionContactoAltaForm()
    admin_form = InstitucionContactoForm()

    for form in [alta_form, admin_form]:
        assert form.fields["nombre_contacto"].required is False
        assert form.fields["rol_area"].required is False
        assert form.fields["documento"].required is False
        assert form.fields["telefono_contacto"].required is False
        assert form.fields["email_contacto"].required is False


@pytest.mark.django_db
def test_institucion_contacto_forms_requieren_un_canal_de_contacto():
    alta_form = InstitucionContactoAltaForm(
        data={
            "nombre_contacto": "María Gómez",
            "rol_area": "Administración",
            "documento": "30111222",
            "telefono_contacto": "",
            "email_contacto": "",
            "es_principal": "on",
        }
    )
    admin_form = InstitucionContactoForm(
        data={
            "centro": "",
            "nombre_contacto": "María Gómez",
            "rol_area": "Administración",
            "documento": "30111222",
            "telefono_contacto": "",
            "email_contacto": "",
            "es_principal": "on",
        }
    )

    assert not alta_form.is_valid()
    assert alta_form.non_field_errors() == [
        "Debe informar al menos un teléfono o correo electrónico para el contacto."
    ]
    assert not admin_form.is_valid()
    assert admin_form.non_field_errors() == [
        "Debe informar al menos un teléfono o correo electrónico para el contacto."
    ]


@pytest.mark.django_db
def test_institucion_contacto_create_renderiza_campos_unificados(client):
    user = User.objects.create_superuser(
        username="admin-contacto-form",
        email="admin-contacto-form@vat.test",
        password="test1234",
    )
    client.force_login(user)

    response = client.get(reverse("vat_institucion_contacto_create"))

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert 'name="nombre_contacto"' in content
    assert 'name="rol_area"' in content
    assert 'name="documento"' in content
    assert 'name="telefono_contacto"' in content
    assert 'name="email_contacto"' in content
    assert 'name="tipo"' not in content
    assert 'name="valor"' not in content


@pytest.mark.django_db
def test_sync_responsable_principal_prioriza_contacto_marcado(vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-principal-sync",
        email="referente-principal-sync@vat.test",
        password="test1234",
    )
    referente.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Contactos Principales",
        codigo="500145111",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="12",
        numero=100,
        domicilio_actividad="Calle 12 N° 100",
        telefono="221-1111111",
        celular="221-2222222",
        correo="cfp-principal@vat.test",
        nombre_referente="Ana",
        apellido_referente="Pérez",
        telefono_referente="221-3333333",
        correo_referente="ana@vat.test",
        referente=referente,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    primer_contacto = InstitucionContacto.objects.create(
        centro=centro,
        nombre_contacto="Primer Contacto",
        rol_area="Administración",
        documento="30111001",
        telefono_contacto="221-4000001",
        email_contacto="primer@vat.test",
        tipo="email",
        valor="primer@vat.test",
        es_principal=False,
    )
    contacto_principal = InstitucionContacto.objects.create(
        centro=centro,
        nombre_contacto="Contacto Principal",
        rol_area="Dirección",
        documento="30111002",
        telefono_contacto="221-4000002",
        email_contacto="principal@vat.test",
        tipo="email",
        valor="principal@vat.test",
        es_principal=True,
    )

    centro_views._sync_responsable_principal(centro)
    primer_contacto.refresh_from_db()
    contacto_principal.refresh_from_db()
    centro.refresh_from_db()

    assert contacto_principal.es_principal is True
    assert primer_contacto.es_principal is False
    assert centro.nombre_referente == "Contacto Principal"
    assert centro.correo_referente == "principal@vat.test"


@pytest.mark.django_db
def test_centro_update_no_reactiva_centros_inactivos(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    payload = _build_centro_payload(
        vat_referente_user, provincia, municipio, localidad, save_continue="1"
    )
    vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")
    centro.activo = False
    centro.save(update_fields=["activo"])

    update_payload = _build_centro_payload(
        vat_referente_user,
        provincia,
        municipio,
        localidad,
        nombre="Centro inactivo editado",
        codigo="500144902",
    )

    response = vat_admin_client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    centro.refresh_from_db()

    assert response.status_code == 302
    assert centro.nombre == "Centro inactivo editado"
    assert centro.activo is False


@pytest.mark.django_db
def test_centro_update_permite_reactivar_centros_inactivos_desde_switch(
    vat_admin_client, vat_referente_user, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    admin_user = User.objects.get(username="admin-vat")
    _assign_user_profile_provincia(admin_user, provincia)
    payload = _build_centro_payload(
        vat_referente_user, provincia, municipio, localidad, save_continue="1"
    )
    vat_admin_client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")
    centro.activo = False
    centro.save(update_fields=["activo"])

    update_payload = _build_centro_payload(
        vat_referente_user,
        provincia,
        municipio,
        localidad,
        nombre="Centro reactivado",
        codigo="500144902",
        activo_present="1",
    )

    response = vat_admin_client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    centro.refresh_from_db()

    assert response.status_code == 302
    assert centro.nombre == "Centro reactivado"
    assert centro.activo is True


@pytest.mark.django_db
def test_centro_list_usuario_provincial_solo_ve_su_provincia(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    provincia_sf = Provincia.objects.create(nombre="Santa Fe")

    municipio_ba = Municipio.objects.create(nombre="La Plata", provincia=provincia_ba)
    municipio_sf = Municipio.objects.create(nombre="Rosario", provincia=provincia_sf)

    localidad_ba = Localidad.objects.create(nombre="Tolosa", municipio=municipio_ba)
    localidad_sf = Localidad.objects.create(nombre="Centro", municipio=municipio_sf)

    user = User.objects.create_user(username="provincial-vat", password="test1234")
    Profile.objects.update_or_create(
        user=user,
        defaults={
            "es_usuario_provincial": True,
            "provincia": provincia_ba,
        },
    )
    permiso_add_plan = Permission.objects.get(
        content_type__app_label="VAT",
        codename="add_planversioncurricular",
    )
    user.user_permissions.add(permiso_add_plan)

    permiso_view_centro = Permission.objects.get(
        content_type__app_label="VAT",
        codename="view_centro",
    )
    user.user_permissions.add(permiso_view_centro)

    Centro.objects.create(
        nombre="Centro BA",
        codigo="BA-001",
        provincia=provincia_ba,
        municipio=municipio_ba,
        localidad=localidad_ba,
        calle="7",
        numero=123,
        domicilio_actividad="Calle 7",
        telefono="221-111111",
        celular="221-111112",
        correo="ba@vat.test",
        nombre_referente="Ana",
        apellido_referente="Perez",
        telefono_referente="221-111113",
        correo_referente="refba@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    Centro.objects.create(
        nombre="Centro SF",
        codigo="SF-001",
        provincia=provincia_sf,
        municipio=municipio_sf,
        localidad=localidad_sf,
        calle="Córdoba",
        numero=456,
        domicilio_actividad="Córdoba 456",
        telefono="341-222221",
        celular="341-222222",
        correo="sf@vat.test",
        nombre_referente="Juan",
        apellido_referente="Gomez",
        telefono_referente="341-222223",
        correo_referente="refsf@vat.test",
        tipo_gestion="Privada",
        clase_institucion="Capacitación Laboral",
        situacion="Institución de ETP",
        activo=True,
    )

    client.force_login(user)
    response = client.get(reverse("vat_centro_list"))

    assert response.status_code == 200
    centros = list(response.context["centros"])
    assert len(centros) == 1
    assert centros[0].nombre == "Centro BA"


@pytest.mark.django_db
def test_filter_centros_queryset_usuario_con_role_provincia_vat_aplica_scope():
    provincia_corrientes = Provincia.objects.create(nombre="Corrientes")
    provincia_chaco = Provincia.objects.create(nombre="Chaco")

    municipio_corrientes = Municipio.objects.create(
        nombre="Capital",
        provincia=provincia_corrientes,
    )
    municipio_chaco = Municipio.objects.create(
        nombre="Resistencia",
        provincia=provincia_chaco,
    )

    localidad_corrientes = Localidad.objects.create(
        nombre="Corrientes",
        municipio=municipio_corrientes,
    )
    localidad_chaco = Localidad.objects.create(
        nombre="Resistencia",
        municipio=municipio_chaco,
    )

    user = User.objects.create_user(username="provincia-vat-role", password="test1234")
    Profile.objects.update_or_create(
        user=user,
        defaults={
            "es_usuario_provincial": True,
            "provincia": provincia_corrientes,
        },
    )

    permiso_role_provincia, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Group),
        codename="role_provincia_vat",
        defaults={"name": "Puede role provincia vat"},
    )
    user.user_permissions.add(permiso_role_provincia)
    user = User.objects.get(pk=user.pk)

    centro_corrientes = Centro.objects.create(
        nombre="Centro Corrientes",
        codigo="CTES-001",
        provincia=provincia_corrientes,
        municipio=municipio_corrientes,
        localidad=localidad_corrientes,
        calle="Mendoza",
        numero=100,
        domicilio_actividad="Mendoza 100",
        telefono="379-111111",
        celular="379-111112",
        correo="corrientes@vat.test",
        nombre_referente="Ana",
        apellido_referente="Acosta",
        telefono_referente="379-111113",
        correo_referente="refcorrientes@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    Centro.objects.create(
        nombre="Centro Chaco",
        codigo="CHA-001",
        provincia=provincia_chaco,
        municipio=municipio_chaco,
        localidad=localidad_chaco,
        calle="Santa María",
        numero=200,
        domicilio_actividad="Santa María 200",
        telefono="362-222221",
        celular="362-222222",
        correo="chaco@vat.test",
        nombre_referente="Luis",
        apellido_referente="Benitez",
        telefono_referente="362-222223",
        correo_referente="refchaco@vat.test",
        tipo_gestion="Privada",
        clase_institucion="Capacitación Laboral",
        situacion="Institución de ETP",
        activo=True,
    )

    centros = list(filter_centros_queryset_for_user(Centro.objects.all(), user))

    assert centros == [centro_corrientes]


@pytest.mark.django_db
def test_vat_centro_list_filters_config_expone_solo_nombre_y_codigo(mocker):
    user = User.objects.create_superuser(
        username="admin-vat-filtros",
        email="admin-filtros@vat.test",
        password="test1234",
    )
    request = RequestFactory().get("/vat/centros/")
    request.user = user

    mocker.patch.object(centro_views, "reverse", side_effect=lambda name: f"/{name}/")

    view = centro_views.CentroListView()
    view.request = request
    view.kwargs = {}
    view.object_list = Centro.objects.none()
    context = view.get_context_data()

    assert [field["name"] for field in context["filters_config"]["fields"]] == [
        "nombre",
        "codigo",
    ]


@pytest.mark.django_db
def test_vat_centro_list_filtra_por_cue_actual_en_codigo(vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    user = User.objects.create_superuser(
        username="centro-filtro-cue",
        email="centro-filtro-cue@vat.test",
        password="test1234",
    )

    centro_match = Centro.objects.create(
        nombre="Centro CUE vigente",
        codigo="LEG-001",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="7",
        numero=123,
        domicilio_actividad="Calle 7",
        telefono="221-111111",
        celular="221-111112",
        correo="cue@vat.test",
        nombre_referente="Ana",
        apellido_referente="Perez",
        telefono_referente="221-111113",
        correo_referente="refcue@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    InstitucionIdentificadorHist.objects.create(
        centro=centro_match,
        tipo_identificador="cue",
        valor_identificador="500144900",
        rol_institucional="sede",
        es_actual=True,
    )
    centro_other = Centro.objects.create(
        nombre="Centro sin match",
        codigo="LEG-002",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="8",
        numero=456,
        domicilio_actividad="Calle 8",
        telefono="221-222221",
        celular="221-222222",
        correo="other@vat.test",
        nombre_referente="Juan",
        apellido_referente="Gomez",
        telefono_referente="221-222223",
        correo_referente="refother@vat.test",
        tipo_gestion="Privada",
        clase_institucion="Capacitación Laboral",
        situacion="Institución de ETP",
        activo=True,
    )
    InstitucionIdentificadorHist.objects.create(
        centro=centro_other,
        tipo_identificador="cue",
        valor_identificador="500144901",
        rol_institucional="sede",
        es_actual=True,
    )

    request = RequestFactory().get(
        "/vat/centros/",
        data={
            "filters": json.dumps(
                {
                    "logic": "AND",
                    "items": [
                        {
                            "field": "codigo",
                            "op": "contains",
                            "value": "500144900",
                        }
                    ],
                }
            )
        },
    )
    request.user = user

    view = centro_views.CentroListView()
    view.request = request

    assert list(view.get_queryset()) == [centro_match]


@pytest.mark.django_db
def test_is_vat_referente_reconoce_alias_legacy_centroreferentevat():
    user = User.objects.create_user(username="referente-legacy", password="test1234")
    permiso_role_referente_legacy, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Group),
        codename="role_centroreferentevat",
        defaults={"name": "ReferenteCentroVAT legacy"},
    )
    user.user_permissions.add(permiso_role_referente_legacy)

    assert is_vat_referente(user) is True


@pytest.mark.django_db
def test_centro_create_usuario_provincial_puede_crear_con_su_provincia(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    municipio_ba = Municipio.objects.create(nombre="La Plata", provincia=provincia_ba)
    localidad_ba = Localidad.objects.create(nombre="Tolosa", municipio=municipio_ba)
    group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-provincial",
        password="test1234",
    )
    referente.groups.add(group)
    user = User.objects.create_user(username="provincial-create", password="test1234")
    Profile.objects.update_or_create(
        user=user,
        defaults={
            "es_usuario_provincial": True,
            "provincia": provincia_ba,
        },
    )
    permiso_view_centro = Permission.objects.get(
        content_type__app_label="VAT",
        codename="view_centro",
    )
    permiso_add_centro = Permission.objects.get(
        content_type__app_label="VAT",
        codename="add_centro",
    )
    user.user_permissions.add(permiso_view_centro, permiso_add_centro)

    client.force_login(user)
    payload = _build_centro_payload(
        referente,
        provincia_ba,
        municipio_ba,
        localidad_ba,
    )
    payload.pop("provincia")
    response = client.post(reverse("vat_centro_create"), data=payload)

    centro = Centro.objects.get(codigo="500144900")

    assert response.status_code == 302
    assert centro.provincia_id == provincia_ba.id


@pytest.mark.django_db
def test_centro_update_usuario_provincial_puede_editar_dentro_de_su_provincia(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    municipio_ba = Municipio.objects.create(nombre="La Plata", provincia=provincia_ba)
    localidad_ba = Localidad.objects.create(nombre="Tolosa", municipio=municipio_ba)
    group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-update-provincial",
        password="test1234",
    )
    referente.groups.add(group)

    centro = Centro.objects.create(
        nombre="Centro Provincial",
        codigo="CTRO-001",
        provincia=provincia_ba,
        municipio=municipio_ba,
        localidad=localidad_ba,
        calle="7",
        numero=123,
        domicilio_actividad="Calle 7 N° 123",
        telefono="221-4000000",
        celular="221-5000000",
        correo="institucion@vat.test",
        sitio_web="https://vat.test",
        nombre_referente="Ana",
        apellido_referente="Pérez",
        telefono_referente="221-4111111",
        correo_referente="direccion@vat.test",
        referente=referente,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    contacto = InstitucionContacto.objects.create(
        centro=centro,
        nombre_contacto="María Gómez",
        rol_area="Administración",
        telefono_contacto="221-4222222",
        email_contacto="maria@vat.test",
        es_principal=True,
    )
    InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad_ba,
        rol_ubicacion="sede_principal",
        domicilio="Calle 7 N° 123",
        es_principal=True,
    )
    InstitucionIdentificadorHist.objects.create(
        centro=centro,
        tipo_identificador="cue",
        valor_identificador="CTRO-001",
        rol_institucional="sede",
        es_actual=True,
    )
    user = User.objects.create_user(
        username="provincial-update",
        password="test1234",
    )
    Profile.objects.update_or_create(
        user=user,
        defaults={
            "es_usuario_provincial": True,
            "provincia": provincia_ba,
        },
    )
    permiso_view_centro = Permission.objects.get(
        content_type__app_label="VAT",
        codename="view_centro",
    )
    permiso_change_centro = Permission.objects.get(
        content_type__app_label="VAT",
        codename="change_centro",
    )
    user.user_permissions.add(permiso_view_centro, permiso_change_centro)

    client.force_login(user)
    update_payload = _build_centro_payload(
        referente,
        provincia_ba,
        municipio_ba,
        localidad_ba,
        nombre="Centro Provincial Editado",
        codigo="500144902",
        domicilio_actividad="Calle 8 N° 4321",
        autoridad_dni="30999888",
        nombre_referente="Laura",
        apellido_referente="Gómez",
        telefono_referente="221-4999999",
        correo_referente="laura@vat.test",
        **{
            "contactos-TOTAL_FORMS": "1",
            "contactos-INITIAL_FORMS": "1",
            "contactos-MIN_NUM_FORMS": "0",
            "contactos-MAX_NUM_FORMS": "1000",
            "contactos-0-id": str(contacto.id),
            "contactos-0-centro": str(centro.id),
            "contactos-0-nombre_contacto": "Laura Gómez",
            "contactos-0-rol_area": "Dirección",
            "contactos-0-telefono_contacto": "221-4333333",
            "contactos-0-email_contacto": "direccion2@vat.test",
            "contactos-0-es_principal": "on",
        },
    )

    response = client.post(
        reverse("vat_centro_update", kwargs={"pk": centro.pk}),
        data=update_payload,
    )

    centro.refresh_from_db()
    contacto.refresh_from_db()

    assert response.status_code == 302
    assert centro.nombre == "Centro Provincial Editado"
    assert contacto.nombre_contacto == "Laura Gómez"


@pytest.mark.django_db
def test_centro_detail_muestra_boton_editar_para_referente_cfp(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    referente_group, _ = Group.objects.get_or_create(name="CFP")
    referente = User.objects.create_user(
        username="referente-centro-detail",
        password="test1234",
    )
    referente.groups.add(referente_group)

    centro = Centro.objects.create(
        nombre="Centro CFP",
        codigo="CFP-001",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="7",
        numero=123,
        domicilio_actividad="Calle 7 N° 123",
        telefono="221-4000000",
        celular="221-5000000",
        correo="institucion@vat.test",
        sitio_web="https://vat.test",
        nombre_referente="Ana",
        apellido_referente="Pérez",
        telefono_referente="221-4111111",
        correo_referente="direccion@vat.test",
        referente=referente,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )

    client.force_login(referente)
    response = client.get(reverse("vat_centro_detail", kwargs={"pk": centro.pk}))

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert reverse("vat_centro_update", kwargs={"pk": centro.pk}) in content
    assert "Editar" in content
    assert "Datos del Establecimiento" in content
    assert "CUE" in content
    assert "Estructura Institucional" not in content
    assert "Ubicacion Principal" in content
    assert "Identificadores" in content


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="config.urls")
def test_institucion_ubicacion_update_renderiza_con_volver_al_detalle_del_centro(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-ubicacion-update-get",
        email="admin-ubicacion-update-get@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="Centro Ubicaciones GET",
        codigo="CFP-UBI-GET",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="8",
        numero=456,
        domicilio_actividad="Calle 8 N° 456",
        telefono="221-4100000",
        celular="221-5100000",
        correo="centro-ubicaciones-get@vat.test",
        nombre_referente="Luisa",
        apellido_referente="Martinez",
        telefono_referente="221-6100000",
        correo_referente="luisa-get@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="anexo",
        nombre_ubicacion="Anexo Norte",
        domicilio="Calle 8 N° 460",
        es_principal=False,
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_institucion_ubicacion_update", kwargs={"pk": ubicacion.pk})
    )

    return_url = reverse("vat_centro_detail", kwargs={"pk": centro.pk})
    content = response.content.decode("utf-8")

    assert response.status_code == 200
    assert response.context["return_url"] == return_url
    assert f'href="{return_url}"' in content


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="config.urls")
def test_institucion_ubicacion_update_redirige_al_detalle_del_centro(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-ubicacion-update",
        email="admin-ubicacion-update@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="Centro Ubicaciones",
        codigo="CFP-UBI-001",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="8",
        numero=456,
        domicilio_actividad="Calle 8 N° 456",
        telefono="221-4100000",
        celular="221-5100000",
        correo="centro-ubicaciones@vat.test",
        nombre_referente="Luisa",
        apellido_referente="Martinez",
        telefono_referente="221-6100000",
        correo_referente="luisa@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="anexo",
        nombre_ubicacion="Anexo Norte",
        domicilio="Calle 8 N° 460",
        es_principal=False,
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_institucion_ubicacion_update", kwargs={"pk": ubicacion.pk}),
        data={
            "centro": str(centro.pk),
            "localidad": str(localidad.pk),
            "rol_ubicacion": "anexo",
            "nombre_ubicacion": "Anexo Norte Actualizado",
            "domicilio": "Calle 8 N° 999",
            "observaciones": "Actualización desde el legajo del centro",
        },
    )

    ubicacion.refresh_from_db()

    assert response.status_code == 302
    assert response.url == reverse("vat_centro_detail", kwargs={"pk": centro.pk})
    assert ubicacion.nombre_ubicacion == "Anexo Norte Actualizado"
    assert ubicacion.domicilio == "Calle 8 N° 999"
    assert ubicacion.observaciones == "Actualización desde el legajo del centro"


@pytest.mark.django_db
def test_plan_curricular_list_usuario_no_provincial_recibe_403(client):
    user = User.objects.create_user(username="no-provincial-plan", password="test1234")
    permiso_view_plan = Permission.objects.get(
        content_type__app_label="VAT",
        codename="view_planversioncurricular",
    )
    user.user_permissions.add(permiso_view_plan)

    client.force_login(user)
    response = client.get(reverse("vat_planversioncurricular_list"))

    assert response.status_code == 403


@pytest.mark.django_db
def test_plan_curricular_list_superuser_puede_acceder(client):
    user = User.objects.create_superuser(
        username="super-plan-list",
        email="super-plan-list@vat.test",
        password="test1234",
    )

    client.force_login(user)
    response = client.get(reverse("vat_planversioncurricular_list"))

    assert response.status_code == 200


@pytest.fixture
def vat_plan_estudio_base(db):
    sector = Sector.objects.create(nombre="Industria")
    subsector = Subsector.objects.create(sector=sector, nombre="Metalúrgica")
    otro_sector = Sector.objects.create(nombre="Servicios")
    otro_subsector = Subsector.objects.create(
        sector=otro_sector,
        nombre="Administración",
    )
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        nombre="Soldador Básico",
        sector=sector,
        subsector=subsector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    titulo = TituloReferencia.objects.create(
        plan_estudio=plan,
        nombre="Soldador Básico",
        activo=True,
    )
    return sector, subsector, otro_sector, otro_subsector, titulo, modalidad


@pytest.mark.django_db
def test_plan_estudio_rechaza_subsector_fuera_del_sector(vat_plan_estudio_base):
    sector, _, _, otro_subsector, titulo, modalidad = vat_plan_estudio_base
    plan = PlanVersionCurricular(
        sector=sector,
        subsector=otro_subsector,
        modalidad_cursada=modalidad,
        activo=True,
    )

    with pytest.raises(ValidationError):
        plan.full_clean()


@pytest.mark.django_db
def test_plan_estudio_rechaza_sector_distinto_al_titulo(vat_plan_estudio_base):
    _, _, otro_sector, otro_subsector, titulo, modalidad = vat_plan_estudio_base
    # Tras la inversión de la relación, el plan ya no valida coherencia con
    # el título. Título de Referencia ya no tiene sector/subsector propios.
    plan = PlanVersionCurricular(
        sector=otro_sector,
        subsector=otro_subsector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    assert titulo.plan_estudio.sector_id != plan.sector_id
    plan.full_clean()  # ya no debe lanzar error por sector del titulo


@pytest.mark.django_db
def test_plan_estudio_backward_compat_devuelve_primer_titulo(vat_plan_estudio_base):
    _, _, _, _, titulo, _ = vat_plan_estudio_base

    assert titulo.plan_estudio.titulo_referencia == titulo
    assert titulo.plan_estudio.titulo_referencia_id == titulo.id


@pytest.mark.django_db
def test_plan_estudio_create_usuario_provincial_asigna_provincia(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_superuser(
        username="provincial-plan",
        email="provincial-plan@vat.test",
        password="test1234",
    )
    Profile.objects.update_or_create(
        user=user,
        defaults={
            "es_usuario_provincial": True,
            "provincia": provincia_ba,
        },
    )

    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)

    client.force_login(user)
    response = client.post(
        reverse("vat_planversioncurricular_create"),
        data={
            "nombre": "Plan Industrial Inicial",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa_tipo": "Resolución",
            "normativa_numero": "123",
            "normativa_anio": "2026",
            "horas_reloj": "120",
            "nivel_requerido": "sin_requisito",
            "nivel_certifica": "nivel_1",
            "activo": "on",
        },
    )

    assert response.status_code == 302
    plan = PlanVersionCurricular.objects.get(normativa="Resolución 123/2026")
    assert plan.provincia_id == provincia_ba.id
    assert plan.nombre == "Plan Industrial Inicial"
    assert plan.titulo_referencia.nombre == "Plan Industrial Inicial"


@pytest.mark.django_db
def test_plan_version_curricular_form_inicializa_campos_compuestos_de_normativa():
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        nombre="Plan de Prueba",
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Disposición 55/2024",
    )

    form = PlanVersionCurricularForm(instance=plan)

    assert form.initial["nombre"] == "Plan de Prueba"
    assert form.initial["normativa_tipo"] == "Disposición"
    assert form.initial["normativa_numero"] == "55"
    assert form.initial["normativa_anio"] == "2024"


@pytest.mark.django_db
def test_plan_version_curricular_form_no_duplica_normativa_estructurada_al_editar():
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Disposicion 55/2024",
        horas_reloj=120,
        nivel_requerido="sin_requisito",
        nivel_certifica="nivel_1",
        activo=True,
    )

    form = PlanVersionCurricularForm(
        data={
            "nombre": "Plan Industrial",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa_tipo": "Disposición",
            "normativa_numero": "55",
            "normativa_anio": "2024",
            "horas_reloj": "120",
            "nivel_requerido": "sin_requisito",
            "nivel_certifica": "nivel_1",
            "activo": "on",
        },
        instance=plan,
    )

    assert form.is_valid(), form.errors
    assert form.normativa_texto_actual == ""

    plan = form.save()

    assert plan.normativa == "Disposición 55/2024"


@pytest.mark.django_db
def test_plan_version_curricular_create_solo_guarda_normativa_estructurada(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_superuser(
        username="super-plan-estructurado",
        email="super-plan-estructurado@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)

    client.force_login(user)
    response = client.post(
        reverse("vat_planversioncurricular_create"),
        data={
            "nombre": "Plan con normativa estructurada",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa_tipo": "Resolución",
            "normativa_numero": "321",
            "normativa_anio": "2025",
            "horas_reloj": "120",
            "nivel_requerido": "sin_requisito",
            "nivel_certifica": "nivel_1",
            "activo": "on",
        },
    )

    assert response.status_code == 302
    plan = PlanVersionCurricular.objects.get(normativa="Resolución 321/2025")
    assert plan.provincia_id == provincia_ba.id
    assert plan.nombre == "Plan con normativa estructurada"
    assert plan.titulo_referencia.nombre == "Plan con normativa estructurada"


@pytest.mark.django_db
def test_plan_version_curricular_form_preserva_normativa_libre_existente():
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        nombre="Plan con texto libre persistido",
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Texto libre cargado por base || Resolución 123/2024",
        activo=True,
    )

    form = PlanVersionCurricularForm(instance=plan)

    assert form.normativa_texto_actual == "Texto libre cargado por base"
    assert "normativa" not in form.fields


@pytest.mark.django_db
def test_plan_version_curricular_form_guarda_normativa_libre_existente_y_actualiza_estructurada():
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        nombre="Plan con texto libre persistido",
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Texto libre cargado por base || Resolución 123/2024",
        horas_reloj=120,
        nivel_requerido="sin_requisito",
        nivel_certifica="nivel_1",
        activo=True,
    )
    form = PlanVersionCurricularForm(
        data={
            "nombre": "Plan con texto libre persistido",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa_tipo": "Disposición",
            "normativa_numero": "55",
            "normativa_anio": "2024",
            "horas_reloj": "120",
            "nivel_requerido": "sin_requisito",
            "nivel_certifica": "nivel_1",
            "activo": "on",
        },
        instance=plan,
    )

    assert form.is_valid(), form.errors

    plan = form.save()

    assert plan.normativa == "Texto libre cargado por base || Disposición 55/2024"


@pytest.mark.django_db
def test_plan_version_curricular_create_no_mezcla_texto_libre_en_alta(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_superuser(
        username="super-plan-mixta",
        email="super-plan-mixta@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)

    client.force_login(user)
    response = client.post(
        reverse("vat_planversioncurricular_create"),
        data={
            "nombre": "Plan mixto",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa_tipo": "Disposición",
            "normativa_numero": "55",
            "normativa_anio": "2024",
            "horas_reloj": "120",
            "nivel_requerido": "sin_requisito",
            "nivel_certifica": "nivel_1",
            "activo": "on",
        },
    )

    assert response.status_code == 302
    plan = PlanVersionCurricular.objects.get(normativa="Disposición 55/2024")
    assert plan.provincia_id == provincia_ba.id
    assert plan.nombre == "Plan mixto"
    assert plan.titulo_referencia.nombre == "Plan mixto"


@pytest.mark.django_db
def test_plan_version_curricular_create_requiere_nombre(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_superuser(
        username="super-plan-sin-nombre",
        email="super-plan-sin-nombre@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)

    client.force_login(user)
    response = client.post(
        reverse("vat_planversioncurricular_create"),
        data={
            "nombre": "   ",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa_tipo": "Resolución",
            "normativa_numero": "123",
            "normativa_anio": "2026",
            "horas_reloj": "120",
            "nivel_requerido": "sin_requisito",
            "nivel_certifica": "nivel_1",
            "activo": "on",
        },
    )

    assert response.status_code == 200
    assert response.context["form"].errors["nombre"] == [
        "El nombre no puede contener solo espacios."
    ]


@pytest.mark.django_db
def test_plan_version_curricular_form_save_actualiza_titulo_asociado():
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        nombre="Nombre anterior",
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Disposición 55/2024",
        activo=True,
    )
    titulo = TituloReferencia.objects.create(
        plan_estudio=plan,
        nombre="Nombre anterior",
        activo=True,
    )

    form = PlanVersionCurricularForm(
        data={
            "nombre": "Nombre actualizado",
            "sector": str(sector.id),
            "subsector": "",
            "modalidad_cursada": str(modalidad.id),
            "normativa": "Disposición 55/2024",
            "normativa_tipo": "",
            "normativa_numero": "",
            "normativa_anio": "",
            "horas_reloj": "",
            "nivel_requerido": "",
            "nivel_certifica": "",
            "activo": "on",
        },
        instance=plan,
    )

    assert form.is_valid(), form.errors
    form.save()
    plan.refresh_from_db()
    titulo.refresh_from_db()

    assert plan.nombre == "Nombre actualizado"
    assert titulo.nombre == "Nombre actualizado"


@pytest.mark.django_db
def test_plan_version_curricular_detail_muestra_los_cuatro_campos_de_normativa(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_superuser(
        username="super-plan-detail",
        email="super-plan-detail@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Disposición 55/2024",
        horas_reloj=120,
        nivel_requerido="sin_requisito",
        nivel_certifica="nivel_1",
        activo=True,
        provincia=provincia_ba,
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_planversioncurricular_detail", kwargs={"pk": plan.pk})
    )

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert "Normativa - Tipo" in content
    assert "Disposición" in content
    assert "Normativa - Número" in content
    assert ">55<" in content
    assert "Normativa - Año" in content
    assert ">2024<" in content


@pytest.mark.django_db
def test_plan_version_curricular_detail_muestra_normativa_libre_y_estructurada(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_superuser(
        username="super-plan-detail-mixto",
        email="super-plan-detail-mixto@vat.test",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="asdedas || Disposición 55/2024",
        horas_reloj=120,
        nivel_requerido="sin_requisito",
        nivel_certifica="nivel_1",
        activo=True,
        provincia=provincia_ba,
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_planversioncurricular_detail", kwargs={"pk": plan.pk})
    )

    content = response.content.decode("utf-8")
    assert response.status_code == 200
    assert ">asdedas<" in content
    assert "Disposición" in content
    assert ">55<" in content
    assert ">2024<" in content


@pytest.mark.django_db
def test_plan_version_curricular_usuario_provincial_sin_change_no_puede_editar(client):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_user(
        username="provincial-plan-sin-change",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    permiso_role_provincia, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Group),
        codename="role_provincia_vat",
        defaults={"name": "Puede role provincia vat"},
    )
    permiso_add_plan = Permission.objects.get(
        content_type__app_label="VAT",
        codename="add_planversioncurricular",
    )
    user.user_permissions.add(permiso_role_provincia, permiso_add_plan)

    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        provincia=provincia_ba,
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Resolución 12/2026",
        horas_reloj=120,
        nivel_requerido="sin_requisito",
        nivel_certifica="nivel_1",
        activo=True,
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_planversioncurricular_update", kwargs={"pk": plan.pk})
    )

    assert response.status_code == 403


@pytest.mark.django_db
def test_plan_version_curricular_usuario_provincial_sin_delete_no_puede_eliminar(
    client,
):
    provincia_ba = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_user(
        username="provincial-plan-sin-delete",
        password="test1234",
    )
    _assign_user_profile_provincia(user, provincia_ba, es_usuario_provincial=True)
    permiso_role_provincia, _ = Permission.objects.get_or_create(
        content_type=ContentType.objects.get_for_model(Group),
        codename="role_provincia_vat",
        defaults={"name": "Puede role provincia vat"},
    )
    permiso_add_plan = Permission.objects.get(
        content_type__app_label="VAT",
        codename="add_planversioncurricular",
    )
    user.user_permissions.add(permiso_role_provincia, permiso_add_plan)

    sector = Sector.objects.create(nombre="Industria")
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    plan = PlanVersionCurricular.objects.create(
        provincia=provincia_ba,
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Resolución 13/2026",
        horas_reloj=120,
        nivel_requerido="sin_requisito",
        nivel_certifica="nivel_1",
        activo=True,
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_planversioncurricular_delete", kwargs={"pk": plan.pk}),
        data={},
    )

    assert response.status_code == 403
    assert PlanVersionCurricular.objects.filter(pk=plan.pk).exists()


@pytest.mark.django_db
def test_titulo_referencia_serializer_expone_clasificacion_via_plan(
    vat_plan_estudio_base,
):
    sector, subsector, _, _, titulo, _ = vat_plan_estudio_base

    data = vat_serializers.TituloReferenciaSerializer(instance=titulo).data

    assert data["plan_estudio"] == titulo.plan_estudio_id
    assert data["sector"] == sector.id
    assert data["sector_nombre"] == sector.nombre
    assert data["subsector"] == subsector.id
    assert data["subsector_nombre"] == subsector.nombre


@pytest.mark.django_db
def test_plan_version_curricular_serializer_omite_campos_eliminados(
    vat_plan_estudio_base,
):
    _, _, _, _, titulo, _ = vat_plan_estudio_base

    data = vat_serializers.PlanVersionCurricularSerializer(
        instance=titulo.plan_estudio
    ).data

    assert data["nombre"] == titulo.plan_estudio.nombre
    assert data["titulo_referencia"] == titulo.id
    assert data["titulo_referencia_nombre"] == titulo.plan_estudio.nombre
    assert "version" not in data
    assert "frecuencia" not in data


def test_migracion_0021_falla_si_un_titulo_tiene_multiples_planes():
    migration = importlib.import_module(
        "VAT.migrations.0021_invert_titulo_plan_relation"
    )

    with pytest.raises(RuntimeError, match="múltiples planes históricos"):
        migration._raise_if_ambiguous_title_plan_rows([(7, 2, "11,12")])


def test_migracion_0021_droppea_fk_antes_que_indices_de_titulo_referencia():
    migration = importlib.import_module(
        "VAT.migrations.0021_invert_titulo_plan_relation"
    )
    executed_sql = []

    class FakeCursor:
        def execute(self, sql, params=None):
            executed_sql.append((sql, params))
            if "CONSTRAINT_TYPE = 'FOREIGN KEY'" in sql:
                self._rows = [("vat_plan_titulo_fk",)]
                self._row = None
            elif "CONSTRAINT_TYPE = 'UNIQUE'" in sql:
                self._rows = [
                    ("VAT_planversioncurricula_titulo_referencia_id_mod_uniq",)
                ]
                self._row = None
            elif "FROM information_schema.STATISTICS" in sql:
                self._rows = []
                self._row = None
            elif "FROM information_schema.COLUMNS" in sql:
                self._rows = None
                self._row = (1,)
            else:
                self._rows = None
                self._row = None

        def fetchall(self):
            return self._rows

        def fetchone(self):
            return self._row

    class FakeConnection:
        def cursor(self):
            return FakeCursor()

    schema_editor = type(
        "FakeSchemaEditor",
        (),
        {"connection": FakeConnection()},
    )()

    migration._drop_titulo_referencia(None, schema_editor)

    drop_fk_index = next(
        i for i, (sql, _) in enumerate(executed_sql) if "DROP FOREIGN KEY" in sql
    )
    drop_unique_index = next(
        i for i, (sql, _) in enumerate(executed_sql) if "DROP INDEX" in sql
    )

    assert drop_fk_index < drop_unique_index


@pytest.fixture
def vat_curso_base(db, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    centro = Centro.objects.create(
        nombre="CFP 501",
        codigo="CFP-501",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="8",
        numero=111,
        domicilio_actividad="Calle 8 N° 111",
        telefono="221-1111111",
        celular="221-1111112",
        correo="cfp501@vat.test",
        nombre_referente="Marta",
        apellido_referente="Lopez",
        telefono_referente="221-1111113",
        correo_referente="marta@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 8 N° 111",
        es_principal=True,
    )
    return centro, ubicacion, modalidad


@pytest.mark.django_db
def test_api_vat_provincias_lista_sin_paginacion(vat_api_client):
    Provincia.objects.bulk_create(
        [Provincia(nombre=f"Provincia VAT {index:02d}") for index in range(12)]
    )

    response = vat_api_client.get("/api/vat/provincias/")

    assert response.status_code == 200
    payload = response.json()
    assert isinstance(payload, list)
    assert len(payload) == 12
    assert payload[0]["nombre"] == "Provincia VAT 00"
    assert payload[-1]["nombre"] == "Provincia VAT 11"


@pytest.mark.django_db
def test_api_vat_localidades_lista_sin_paginacion(vat_api_client):
    provincia = Provincia.objects.create(nombre="Provincia VAT")
    municipio = Municipio.objects.create(nombre="Municipio VAT", provincia=provincia)
    Localidad.objects.bulk_create(
        [
            Localidad(nombre=f"Localidad VAT {index:02d}", municipio=municipio)
            for index in range(12)
        ]
    )

    response = vat_api_client.get("/api/vat/localidades/")

    assert response.status_code == 200
    payload = response.json()
    assert isinstance(payload, list)
    assert len(payload) == 12
    assert payload[0]["nombre"] == "Localidad VAT 00"
    assert payload[-1]["nombre"] == "Localidad VAT 11"


@pytest.mark.django_db
def test_api_vat_centros_lista_con_api_key(vat_api_client, vat_curso_base):
    centro, _, _ = vat_curso_base

    response = vat_api_client.get("/api/vat/centros/?activo=true")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] >= 1
    assert payload["results"][0]["id"] == centro.id
    assert payload["results"][0]["provincia"] == centro.provincia_id


@pytest.mark.django_db
def test_api_vat_cursos_lista_por_centro(vat_api_client, vat_curso_base):
    centro, _, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API VAT",
        modalidad=modalidad,
        estado="activo",
    )

    response = vat_api_client.get(f"/api/vat/cursos/?centro_id={centro.id}")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] == 1
    assert payload["results"][0]["id"] == curso.id
    assert payload["results"][0]["centro"] == centro.id


@pytest.mark.django_db
def test_api_vat_cursos_lista_por_provincia_y_municipio(vat_api_client, vat_curso_base):
    centro, _, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API VAT Geografico",
        modalidad=modalidad,
        estado="activo",
    )
    otra_provincia = Provincia.objects.create(nombre="Otra Provincia VAT")
    otro_municipio = Municipio.objects.create(
        nombre="Otro Municipio VAT",
        provincia=otra_provincia,
    )
    otra_localidad = Localidad.objects.create(
        nombre="Otra Localidad VAT",
        municipio=otro_municipio,
    )
    otro_centro = Centro.objects.create(
        nombre="Centro API VAT Alternativo",
        codigo="CFP-ALT-VAT-1",
        provincia=otra_provincia,
        municipio=otro_municipio,
        localidad=otra_localidad,
        calle="9",
        numero=10,
        domicilio_actividad="Calle 9 N° 10",
        telefono="221-9999991",
        celular="221-9999992",
        correo="alternativo@vat.test",
        nombre_referente="Luis",
        apellido_referente="Perez",
        telefono_referente="221-9999993",
        correo_referente="luis@vat.test",
        tipo_gestion="Privada",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    Curso.objects.create(
        centro=otro_centro,
        nombre="Curso API VAT Excluido",
        modalidad=modalidad,
        estado="activo",
    )

    response = vat_api_client.get(
        f"/api/vat/cursos/?provincia_id={centro.provincia_id}&municipio_id={centro.municipio_id}"
    )

    assert response.status_code == 200
    payload = response.json()
    result_ids = {item["id"] for item in payload["results"]}
    assert curso.id in result_ids
    assert all(item["centro"] == centro.id for item in payload["results"])


@pytest.mark.django_db
def test_api_vat_cursos_rechaza_voucher_e_inscripcion_libre_al_mismo_tiempo(
    vat_api_client, vat_curso_base
):
    centro, _, modalidad = vat_curso_base

    response = vat_api_client.post(
        "/api/vat/cursos/",
        {
            "centro": centro.id,
            "nombre": "Curso API Invalido",
            "modalidad": modalidad.id,
            "estado": "planificado",
            "usa_voucher": True,
            "inscripcion_libre": True,
            "costo_creditos": 1,
        },
        format="json",
    )

    assert response.status_code == 400
    assert response.json() == {
        "inscripcion_libre": [
            "Un curso no puede usar voucher e inscripción libre al mismo tiempo."
        ]
    }
    assert not Curso.objects.filter(nombre="Curso API Invalido").exists()


@pytest.mark.django_db
def test_api_vat_cursos_buscar_por_texto_devuelve_info_enriquecida(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa API Búsqueda Curso")
    usuario = User.objects.create_user(
        username="api-busqueda-curso",
        password="test1234",
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Administración Contable Avanzada",
        modalidad=modalidad,
        estado="activo",
        inscripcion_libre=True,
        usa_voucher=True,
        costo_creditos=2,
    )
    voucher_parametria = VoucherParametria.objects.create(
        nombre="Voucher Búsqueda Curso",
        programa=programa,
        cantidad_inicial=8,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher_parametria)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="BUSQ-CUR-01",
        nombre="Comisión Búsqueda Curso",
        cupo_total=12,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 4, 15),
        fecha_fin=date(2026, 5, 15),
        estado="activa",
    )
    dia = Dia.objects.create(nombre="Martes")
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(9, 0),
        hora_hasta=time(11, 0),
        aula_espacio="Aula 2",
        vigente=True,
    )
    SesionComision.objects.create(
        comision_curso=comision,
        horario=horario,
        numero_sesion=1,
        fecha=date(2026, 4, 22),
        estado="programada",
    )

    response = vat_api_client.get("/api/vat/cursos/buscar/?q=con")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] == 1
    result = payload["results"][0]
    assert result["id"] == curso.id
    assert result["inscripcion_libre"] is True
    assert result["centro"]["id"] == centro.id
    assert result["centro"]["provincia"]["id"] == centro.provincia_id
    assert result["centro"]["provincia"]["nombre"] == centro.provincia.nombre
    assert result["centro"]["ciudad"]["provincia"]["id"] == centro.provincia_id
    assert result["centro"]["ciudad"]["municipio"]["id"] == centro.municipio_id
    assert result["centro"]["ciudad"]["localidad"]["id"] == centro.localidad_id
    assert result["centro"]["ciudad"]["direccion"] == centro.domicilio_actividad
    assert result["programa"] == {"id": programa.id, "nombre": programa.nombre}
    assert result["voucher_parametrias"][0]["id"] == voucher_parametria.id
    assert result["comisiones"][0]["id"] == comision.id
    assert result["comisiones"][0]["acepta_lista_espera"] is True
    assert result["comisiones"][0]["total_inscriptos"] == 0
    assert result["comisiones"][0]["cupos_disponibles"] == 12
    assert result["comisiones"][0]["horarios"][0]["id"] == horario.id
    assert result["comisiones"][0]["ubicacion"]["id"] == ubicacion.id


@pytest.mark.django_db
def test_api_vat_cursos_buscar_sin_texto_devuelve_listado_paginado(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Operador de Primera Carga",
        modalidad=modalidad,
        estado="activo",
    )
    ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="PC-001",
        nombre="Comisión Primera Carga",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 10),
        fecha_fin=date(2026, 5, 10),
        estado="activa",
    )

    response = vat_api_client.get("/api/vat/cursos/buscar/")

    assert response.status_code == 200
    payload = response.json()
    result_ids = {item["id"] for item in payload["results"]}
    assert payload["count"] == 1
    assert curso.id in result_ids


@pytest.mark.django_db
def test_api_vat_cursos_buscar_con_texto_vacio_devuelve_listado_paginado(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Operador de Campo Limpio",
        modalidad=modalidad,
        estado="activo",
    )
    ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="CL-001",
        nombre="Comisión Campo Limpio",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 10),
        fecha_fin=date(2026, 5, 10),
        estado="activa",
    )

    response = vat_api_client.get("/api/vat/cursos/buscar/?q=   ")

    assert response.status_code == 200
    payload = response.json()
    result_ids = {item["id"] for item in payload["results"]}
    assert payload["count"] == 1
    assert curso.id in result_ids


@pytest.mark.django_db
def test_api_vat_cursos_buscar_devuelve_solo_cursos_activos_con_comisiones_activas(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso_visible = Curso.objects.create(
        centro=centro,
        nombre="Curso Visible",
        modalidad=modalidad,
        estado="activo",
    )
    Curso.objects.create(
        centro=centro,
        nombre="Curso Inactivo",
        modalidad=modalidad,
        estado="finalizado",
    )
    curso_sin_comision_activa = Curso.objects.create(
        centro=centro,
        nombre="Curso Sin Comisión Activa",
        modalidad=modalidad,
        estado="activo",
    )
    ComisionCurso.objects.create(
        curso=curso_visible,
        ubicacion=ubicacion,
        codigo_comision="VIS-001",
        nombre="Comisión Visible",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 10),
        fecha_fin=date(2026, 5, 10),
        estado="activa",
    )
    ComisionCurso.objects.create(
        curso=curso_sin_comision_activa,
        ubicacion=ubicacion,
        codigo_comision="NOACT-001",
        nombre="Comisión Cerrada",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 10),
        fecha_fin=date(2026, 5, 10),
        estado="cerrada",
    )

    response = vat_api_client.get("/api/vat/cursos/buscar/")

    assert response.status_code == 200
    payload = response.json()
    result_ids = {item["id"] for item in payload["results"]}
    assert curso_visible.id in result_ids
    assert curso_sin_comision_activa.id not in result_ids


@pytest.mark.django_db
def test_api_vat_cursos_buscar_requiere_minimo_tres_caracteres(vat_api_client):
    response = vat_api_client.get("/api/vat/cursos/buscar/?q=he")

    assert response.status_code == 400
    assert response.json() == {"q": ["Debe enviar al menos 3 caracteres para buscar."]}


@pytest.mark.django_db
def test_api_vat_cursos_prioritarios_devuelve_info_enriquecida(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa API Curso Prioritario")
    usuario = User.objects.create_user(
        username="api-prioritario-curso",
        password="test1234",
    )
    curso_prioritario = Curso.objects.create(
        centro=centro,
        nombre="Herramientas de Gestión Prioritaria",
        modalidad=modalidad,
        estado="activo",
        prioritario=True,
        usa_voucher=True,
        costo_creditos=3,
    )
    Curso.objects.create(
        centro=centro,
        nombre="Curso No Prioritario",
        modalidad=modalidad,
        estado="activo",
        prioritario=False,
    )
    voucher_parametria = VoucherParametria.objects.create(
        nombre="Voucher Curso Prioritario",
        programa=programa,
        cantidad_inicial=10,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso_prioritario.voucher_parametrias.add(voucher_parametria)
    comision = ComisionCurso.objects.create(
        curso=curso_prioritario,
        ubicacion=ubicacion,
        codigo_comision="PRIO-CUR-01",
        nombre="Comisión Curso Prioritario",
        cupo_total=15,
        fecha_inicio=date(2026, 4, 20),
        fecha_fin=date(2026, 5, 20),
        estado="activa",
    )
    dia = Dia.objects.create(nombre="Jueves")
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(14, 0),
        hora_hasta=time(16, 0),
        aula_espacio="Aula Prioritaria",
        vigente=True,
    )
    SesionComision.objects.create(
        comision_curso=comision,
        horario=horario,
        numero_sesion=1,
        fecha=date(2026, 4, 24),
        estado="programada",
    )

    response = vat_api_client.get("/api/vat/cursos/prioritarios/")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] == 1
    result = payload["results"][0]
    assert result["id"] == curso_prioritario.id
    assert result["prioritario"] is True


@pytest.mark.django_db
def test_api_vat_cursos_prioritarios_devuelve_solo_cursos_activos_con_comisiones_activas(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso_visible = Curso.objects.create(
        centro=centro,
        nombre="Curso Prioritario Visible",
        modalidad=modalidad,
        estado="activo",
        prioritario=True,
    )
    curso_inactivo = Curso.objects.create(
        centro=centro,
        nombre="Curso Prioritario Inactivo",
        modalidad=modalidad,
        estado="finalizado",
        prioritario=True,
    )
    curso_sin_comision_activa = Curso.objects.create(
        centro=centro,
        nombre="Curso Prioritario Sin Comisión Activa",
        modalidad=modalidad,
        estado="activo",
        prioritario=True,
    )
    primera_comision_visible = ComisionCurso.objects.create(
        curso=curso_visible,
        ubicacion=ubicacion,
        codigo_comision="PRIOVIS-001",
        nombre="Comisión Prioritaria Visible",
        cupo_total=10,
        fecha_inicio=date(2026, 4, 12),
        fecha_fin=date(2026, 5, 12),
        estado="activa",
    )
    segunda_comision_visible = ComisionCurso.objects.create(
        curso=curso_visible,
        ubicacion=ubicacion,
        codigo_comision="PRIOVIS-002",
        nombre="Comisión Prioritaria Visible 2",
        cupo_total=10,
        fecha_inicio=date(2026, 4, 13),
        fecha_fin=date(2026, 5, 13),
        estado="activa",
    )
    ComisionCurso.objects.create(
        curso=curso_inactivo,
        ubicacion=ubicacion,
        codigo_comision="PRIOINAC-001",
        nombre="Comisión Prioritaria Inactiva",
        cupo_total=10,
        fecha_inicio=date(2026, 4, 12),
        fecha_fin=date(2026, 5, 12),
        estado="activa",
    )
    ComisionCurso.objects.create(
        curso=curso_sin_comision_activa,
        ubicacion=ubicacion,
        codigo_comision="PRIONOACT-001",
        nombre="Comisión Prioritaria Cerrada",
        cupo_total=10,
        fecha_inicio=date(2026, 4, 12),
        fecha_fin=date(2026, 5, 12),
        estado="cerrada",
    )

    response = vat_api_client.get("/api/vat/cursos/prioritarios/")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] == 1
    result_ids = {item["id"] for item in payload["results"]}
    result = payload["results"][0]
    assert curso_visible.id in result_ids
    assert curso_inactivo.id not in result_ids
    assert curso_sin_comision_activa.id not in result_ids
    assert result["id"] == curso_visible.id
    assert {comision["id"] for comision in result["comisiones"]} == {
        primera_comision_visible.id,
        segunda_comision_visible.id,
    }
    assert result["centro"]["id"] == centro.id
    assert result["centro"]["provincia"]["id"] == centro.provincia_id
    assert result["centro"]["ciudad"]["municipio"]["id"] == centro.municipio_id
    assert result["centro"]["ciudad"]["localidad"]["id"] == centro.localidad_id
    assert result["centro"]["ciudad"]["direccion"] == centro.domicilio_actividad
    assert result["programa"] is None
    assert result["voucher_parametrias"] == []
    assert all(
        comision["ubicacion"]["id"] == ubicacion.id for comision in result["comisiones"]
    )


@pytest.mark.django_db
def test_api_vat_comisiones_curso_lista_por_curso(vat_api_client, vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API Comision VAT",
        modalidad=modalidad,
        estado="activo",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="API-COM-01",
        nombre="Comision API VAT",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )

    response = vat_api_client.get(f"/api/vat/comisiones-curso/?curso_id={curso.id}")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] == 1
    assert payload["results"][0]["id"] == comision.id
    assert payload["results"][0]["curso"] == curso.id


@pytest.mark.django_db
def test_api_vat_comisiones_curso_lista_incluye_horarios_y_sesiones(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API Horarios VAT",
        modalidad=modalidad,
        estado="activo",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="API-COM-HOR-01",
        nombre="Comision API Horarios VAT",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    dia = Dia.objects.create(nombre="Lunes")
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(18, 0),
        hora_hasta=time(20, 0),
        aula_espacio="Taller 1",
        vigente=True,
    )
    sesion = SesionComision.objects.create(
        comision_curso=comision,
        horario=horario,
        numero_sesion=1,
        fecha=date(2026, 4, 14),
        estado="programada",
    )

    response = vat_api_client.get(f"/api/vat/comisiones-curso/?curso_id={curso.id}")

    assert response.status_code == 200
    payload = response.json()
    result = payload["results"][0]
    assert result["horarios"] == [
        {
            "id": horario.id,
            "dia_semana": dia.id,
            "dia_nombre": "Lunes",
            "hora_desde": "18:00:00",
            "hora_hasta": "20:00:00",
            "aula_espacio": "Taller 1",
            "vigente": True,
        }
    ]
    assert result["sesiones"] == [
        {
            "id": sesion.id,
            "horario": horario.id,
            "numero_sesion": 1,
            "fecha": "2026-04-14",
            "estado": "programada",
            "observaciones": None,
            "dia_semana": dia.id,
            "dia_nombre": "Lunes",
            "hora_desde": "18:00:00",
            "hora_hasta": "20:00:00",
            "aula_espacio": "Taller 1",
        }
    ]


@pytest.mark.django_db
def test_api_vat_comisiones_curso_lista_por_provincia_y_municipio(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API VAT Geografico Comision",
        modalidad=modalidad,
        estado="activo",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="API-GEO-01",
        nombre="Comision API VAT Geografica",
        cupo_total=15,
        fecha_inicio=date(2026, 4, 2),
        fecha_fin=date(2026, 4, 29),
        estado="activa",
    )
    otra_provincia = Provincia.objects.create(nombre="Provincia Geo Comision")
    otro_municipio = Municipio.objects.create(
        nombre="Municipio Geo Comision",
        provincia=otra_provincia,
    )
    otra_localidad = Localidad.objects.create(
        nombre="Localidad Geo Comision",
        municipio=otro_municipio,
    )
    otro_centro = Centro.objects.create(
        nombre="Centro API VAT Comision Alternativo",
        codigo="CFP-ALT-VAT-2",
        provincia=otra_provincia,
        municipio=otro_municipio,
        localidad=otra_localidad,
        calle="10",
        numero=20,
        domicilio_actividad="Calle 10 N° 20",
        telefono="221-8888881",
        celular="221-8888882",
        correo="alternativo-comision@vat.test",
        nombre_referente="Laura",
        apellido_referente="Suarez",
        telefono_referente="221-8888883",
        correo_referente="laura@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    otra_ubicacion = InstitucionUbicacion.objects.create(
        centro=otro_centro,
        localidad=otra_localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 10 N° 20",
        es_principal=True,
    )
    otro_curso = Curso.objects.create(
        centro=otro_centro,
        nombre="Curso API VAT Comision Excluida",
        modalidad=modalidad,
        estado="activo",
    )
    ComisionCurso.objects.create(
        curso=otro_curso,
        ubicacion=otra_ubicacion,
        codigo_comision="API-GEO-02",
        nombre="Comision API VAT Excluida",
        cupo_total=10,
        fecha_inicio=date(2026, 4, 3),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )

    response = vat_api_client.get(
        f"/api/vat/comisiones-curso/?provincia_id={centro.provincia_id}&municipio_id={centro.municipio_id}"
    )

    assert response.status_code == 200
    payload = response.json()
    result_ids = {item["id"] for item in payload["results"]}
    assert comision.id in result_ids
    assert all(item["curso_centro_id"] == centro.id for item in payload["results"])


@pytest.mark.django_db
def test_api_vat_inscripciones_curso_crea_inscripcion_en_comision_curso(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa API Inscripción Curso")
    usuario = User.objects.create_user(
        username="api-inscripcion-curso",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="No Binario")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API Inscripción",
        modalidad=modalidad,
        estado="activo",
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher API Inscripción Curso",
        programa=programa,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="API-INS-CURSO-01",
        nombre="Comisión API Inscripción",
        cupo_total=20,
        fecha_inicio=date(2026, 5, 1),
        fecha_fin=date(2026, 6, 1),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="API",
        nombre="Curso",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=40111222,
        sexo=sexo,
    )

    response = vat_api_client.post(
        "/api/vat/inscripciones-curso/",
        {
            "ciudadano": ciudadano.id,
            "comision_curso": comision.id,
            "estado": "inscripta",
            "origen_canal": "api",
        },
        format="json",
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["comision_curso"] == comision.id
    assert payload["entidad_comision_tipo"] == "comision_curso"
    assert Inscripcion.objects.filter(
        ciudadano=ciudadano,
        comision_curso=comision,
    ).exists()


@pytest.mark.django_db
def test_api_vat_inscripciones_curso_rechaza_comision_legacy(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa API Legacy")
    sexo = Sexo.objects.create(sexo="Femenino")
    sector = Sector.objects.create(nombre="Sector API Legacy")
    plan = PlanVersionCurricular.objects.create(
        provincia=centro.provincia,
        nombre="Plan API Legacy",
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    oferta = OfertaInstitucional.objects.create(
        centro=centro,
        plan_curricular=plan,
        programa=programa,
        nombre_local="Oferta API Legacy",
        ciclo_lectivo=2026,
        estado="publicada",
    )
    comision_legacy = Comision.objects.create(
        oferta=oferta,
        ubicacion=ubicacion,
        codigo_comision="API-INS-LEGACY-01",
        nombre="Comisión Legacy API",
        cupo=20,
        fecha_inicio=date(2026, 5, 1),
        fecha_fin=date(2026, 6, 1),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="API",
        nombre="Legacy",
        fecha_nacimiento=date(1998, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=40111223,
        sexo=sexo,
    )

    response = vat_api_client.post(
        "/api/vat/inscripciones-curso/",
        {
            "ciudadano": ciudadano.id,
            "comision": comision_legacy.id,
            "estado": "inscripta",
            "origen_canal": "api",
        },
        format="json",
    )

    assert response.status_code == 400
    assert response.json() == {
        "comision_curso": [
            "Este endpoint solo admite inscripciones sobre comisiones de curso."
        ]
    }
    assert not Inscripcion.objects.filter(
        ciudadano=ciudadano,
        comision=comision_legacy,
    ).exists()


@pytest.mark.django_db
def test_api_vat_web_cursos_lista_comisiones_curso(vat_api_client, vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Web Cursos")
    usuario = User.objects.create_user(username="api-web-cursos", password="test1234")
    sector = Sector.objects.create(nombre="Servicios API Web")
    plan = PlanVersionCurricular.objects.create(
        provincia=centro.provincia,
        nombre="Plan API Web",
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    TituloReferencia.objects.create(
        plan_estudio=plan,
        nombre="Título API Web",
        activo=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        plan_estudio=plan,
        nombre="Curso Web",
        modalidad=modalidad,
        estado="activo",
        usa_voucher=True,
        costo_creditos=2,
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Web Curso",
        programa=programa,
        cantidad_inicial=8,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="WEB-CURSO-01",
        nombre="Comisión Web Curso",
        cupo_total=30,
        fecha_inicio=date(2026, 7, 1),
        fecha_fin=date(2026, 8, 1),
        estado="activa",
    )

    response = vat_api_client.get(f"/api/vat/web/cursos/?centro_id={centro.id}")

    assert response.status_code == 200
    payload = response.json()
    assert payload["count"] >= 1
    resultado = payload["results"][0]
    assert resultado["id"] == comision.id
    assert resultado["centro_id"] == centro.id
    assert resultado["programa_id"] == programa.id
    assert resultado["usa_voucher"] is True


@pytest.mark.django_db
def test_api_vat_web_prevalidar_inscripcion_informa_falta_voucher(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Web Prevalidación")
    usuario = User.objects.create_user(
        username="api-web-prevalidacion",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="Femenino")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Web Prevalidación",
        modalidad=modalidad,
        estado="activo",
        usa_voucher=True,
        costo_creditos=2,
    )
    voucher_parametria = VoucherParametria.objects.create(
        nombre="Voucher Web Prevalidación",
        programa=programa,
        cantidad_inicial=4,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher_parametria)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="WEB-PRE-01",
        nombre="Comisión Web Prevalidación",
        cupo_total=10,
        fecha_inicio=date(2026, 5, 10),
        fecha_fin=date(2026, 6, 10),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="López",
        nombre="Lucía",
        fecha_nacimiento=date(1995, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111222,
        sexo=sexo,
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/prevalidar/",
        {
            "documento": str(ciudadano.documento),
            "cuil": "27-30111222-8",
            "comision_curso_id": comision.id,
        },
        format="json",
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["puede_inscribirse"] is False
    assert any("voucher activo" in motivo for motivo in payload["motivos"])
    assert payload["comision"]["id"] == comision.id
    assert payload["voucher"]["requerido"] is True
    assert payload["voucher"]["credito_requerido"] == 2


@pytest.mark.django_db
def test_api_vat_web_mi_argentina_flujo_completo_prevalidar_e_inscribir(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Mi Argentina")
    usuario = User.objects.create_user(
        username="api-web-mi-argentina",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="No Binario")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Mi Argentina",
        modalidad=modalidad,
        estado="activo",
        usa_voucher=True,
        costo_creditos=2,
    )
    voucher_parametria = VoucherParametria.objects.create(
        nombre="Voucher Mi Argentina",
        programa=programa,
        cantidad_inicial=6,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher_parametria)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="MIARG-01",
        nombre="Comisión Mi Argentina",
        cupo_total=12,
        fecha_inicio=date(2026, 6, 1),
        fecha_fin=date(2026, 7, 1),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="García",
        nombre="Andrea",
        fecha_nacimiento=date(1997, 8, 15),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=32123456,
        sexo=sexo,
    )
    voucher = Voucher.objects.create(
        parametria=voucher_parametria,
        ciudadano=ciudadano,
        programa=programa,
        cantidad_inicial=6,
        cantidad_usada=0,
        cantidad_disponible=6,
        fecha_vencimiento=date(2026, 12, 31),
        estado="activo",
        asignado_por=usuario,
    )

    response_centros = vat_api_client.get("/api/vat/web/centros/?activo=true")

    assert response_centros.status_code == 200
    centros_payload = response_centros.json()
    assert centros_payload["count"] >= 1
    assert any(item["id"] == centro.id for item in centros_payload["results"])

    response_cursos = vat_api_client.get(f"/api/vat/web/cursos/?centro_id={centro.id}")

    assert response_cursos.status_code == 200
    cursos_payload = response_cursos.json()
    curso_resultado = next(
        item for item in cursos_payload["results"] if item["id"] == comision.id
    )
    assert curso_resultado["centro_id"] == centro.id
    assert curso_resultado["programa_id"] == programa.id
    assert curso_resultado["usa_voucher"] is True

    response_prevalidar = vat_api_client.post(
        "/api/vat/web/inscripciones/prevalidar/",
        {
            "documento": str(ciudadano.documento),
            "cuil": "27-32123456-4",
            "comision_curso_id": comision.id,
        },
        format="json",
    )

    assert response_prevalidar.status_code == 200
    prevalidacion_payload = response_prevalidar.json()
    assert prevalidacion_payload["puede_inscribirse"] is True
    assert prevalidacion_payload["motivos"] == []
    assert prevalidacion_payload["ciudadano"]["id"] == ciudadano.id
    assert prevalidacion_payload["comision"]["id"] == comision.id
    assert prevalidacion_payload["voucher"]["voucher_id"] == voucher.id
    assert prevalidacion_payload["voucher"]["saldo_actual"] == 6
    assert prevalidacion_payload["voucher"]["credito_requerido"] == 2
    assert prevalidacion_payload["voucher"]["saldo_post_inscripcion"] == 4

    response_inscribir = vat_api_client.post(
        "/api/vat/web/inscripciones/",
        {
            "documento": str(ciudadano.documento),
            "comision_curso_id": comision.id,
            "estado": "inscripta",
            "observaciones": "Alta desde Mi Argentina",
        },
        format="json",
    )

    assert response_inscribir.status_code == 201
    inscripcion_payload = response_inscribir.json()
    assert inscripcion_payload["comision_curso"] == comision.id
    assert inscripcion_payload["curso"]["id"] == comision.id
    assert inscripcion_payload["ciudadano"] == ciudadano.id

    response_inscripciones = vat_api_client.get(
        f"/api/vat/web/inscripciones/?documento={ciudadano.documento}"
    )

    assert response_inscripciones.status_code == 200
    inscripciones_payload = response_inscripciones.json()
    assert inscripciones_payload["count"] == 1
    assert inscripciones_payload["results"][0]["comision_curso"] == comision.id

    voucher.refresh_from_db()
    assert voucher.cantidad_disponible == 4


@pytest.mark.django_db
def test_api_vat_web_inscripcion_libre_crea_inscripcion_operativa_sin_ciudadano(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Inscripción Libre",
        modalidad=modalidad,
        estado="activo",
        inscripcion_libre=True,
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="LIBRE-01",
        nombre="Comisión Libre",
        cupo_total=20,
        fecha_inicio=date(2026, 6, 1),
        fecha_fin=date(2026, 7, 1),
        estado="activa",
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/",
        {
            "comision_curso_id": comision.id,
            "estado": "pre_inscripta",
            "datos_postulante": {
                "nombre": "Fabrizzio Nicolas",
                "apellido": "CONIGLIO",
                "tipo_documento": "DNI",
                "documento": "42439852",
                "email": "fconiglio100@gmail.com",
                "telefono": "+54-351-3989965",
            },
        },
        format="json",
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["comision_curso"] == comision.id
    assert payload["estado"] == "pre_inscripta"
    assert payload["ciudadano"] is not None

    ciudadano = Ciudadano.objects.get(pk=payload["ciudadano"])
    assert ciudadano.documento == 42439852
    assert ciudadano.apellido == "CONIGLIO"
    assert ciudadano.fecha_nacimiento == date(1900, 1, 1)

    inscripcion = Inscripcion.objects.get(pk=payload["id"])
    assert inscripcion.comision_curso == comision
    assert inscripcion.ciudadano == ciudadano
    assert inscripcion.programa is None

    solicitud = SolicitudInscripcionPublica.objects.get(inscripcion=inscripcion)
    assert solicitud.comision_curso == comision
    assert solicitud.ciudadano == ciudadano
    assert solicitud.estado == "convertida"
    assert solicitud.datos_postulante["apellido"] == "CONIGLIO"

    response_inscripciones = vat_api_client.get(
        f"/api/vat/web/inscripciones/?documento={ciudadano.documento}"
    )

    assert response_inscripciones.status_code == 200
    inscripciones_payload = response_inscripciones.json()
    assert inscripciones_payload["count"] == 1
    assert inscripciones_payload["results"][0]["id"] == inscripcion.id

    response_cursos = vat_api_client.get(f"/api/vat/web/cursos/?centro_id={centro.id}")

    assert response_cursos.status_code == 200
    cursos_payload = response_cursos.json()
    curso_resultado = next(
        item for item in cursos_payload["results"] if item["id"] == comision.id
    )
    assert curso_resultado["inscripcion_libre"] is True


@pytest.mark.django_db
def test_api_vat_web_inscripcion_libre_usa_cuil_como_documento_si_no_viene_documento(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Inscripción Libre con CUIL",
        modalidad=modalidad,
        estado="activo",
        inscripcion_libre=True,
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="LIBRE-CUIL-01",
        nombre="Comisión Libre con CUIL",
        cupo_total=20,
        fecha_inicio=date(2026, 6, 1),
        fecha_fin=date(2026, 7, 1),
        estado="activa",
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/",
        {
            "comision_curso_id": comision.id,
            "estado": "pre_inscripta",
            "datos_postulante": {
                "nombre": "Leylen Nahir",
                "apellido": "PINTO",
                "fecha_nacimiento": "1993-04-26",
                "genero": "F",
                "email": "leylenpinto718@gmail.com",
                "telefono": "+54-3705-201750",
                "nivel_estudio": "secundario_incompleto",
                "tipo_documento": "DNI",
                "cuil": "27375343520",
                "actual_calle": "Barrio salvador rugeri",
                "actual_numero": "Mz 21 cs 20",
                "actual_entre_calles": "Mz 21 cs 20",
                "actual_municipio_nombre": "Formosa",
                "actual_provincia_nombre": "Formosa",
            },
        },
        format="json",
    )

    assert response.status_code == 201
    payload = response.json()
    ciudadano = Ciudadano.objects.get(pk=payload["ciudadano"])
    assert ciudadano.documento == 27375343520
    assert ciudadano.fecha_nacimiento == date(1993, 4, 26)
    assert "CUIL informado como documento" in (ciudadano.observaciones or "")

    inscripcion = Inscripcion.objects.get(pk=payload["id"])
    solicitud = SolicitudInscripcionPublica.objects.get(inscripcion=inscripcion)
    assert solicitud.datos_postulante["cuil"] == "27375343520"


@pytest.mark.django_db
def test_api_vat_web_inscripciones_informa_lista_espera_si_no_hay_cupo(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Web Estado Espera")
    usuario = User.objects.create_user(
        username="api-web-estado-espera",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="Femenino")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Web Estado Espera",
        modalidad=modalidad,
        estado="activo",
        usa_voucher=True,
        costo_creditos=1,
    )
    voucher_parametria = VoucherParametria.objects.create(
        nombre="Voucher Web Estado Espera",
        programa=programa,
        cantidad_inicial=4,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher_parametria)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="WEB-EST-ESP-01",
        nombre="Comisión Web Estado Espera",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 6, 1),
        fecha_fin=date(2026, 7, 1),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Web Espera",
        fecha_nacimiento=date(1990, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=40111226,
        sexo=sexo,
    )
    aspirante = Ciudadano.objects.create(
        apellido="Aspirante",
        nombre="Web Espera",
        fecha_nacimiento=date(1995, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=40111227,
        sexo=sexo,
    )
    Voucher.objects.create(
        ciudadano=aspirante,
        parametria=voucher_parametria,
        programa=programa,
        cantidad_inicial=3,
        cantidad_usada=0,
        cantidad_disponible=3,
        fecha_vencimiento=date(2026, 12, 31),
        estado="activo",
    )
    Inscripcion.objects.create(
        ciudadano=titular,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/",
        {
            "documento": str(aspirante.documento),
            "comision_curso_id": comision.id,
            "estado": "inscripta",
        },
        format="json",
    )

    payload = response.json()

    assert response.status_code == 201
    assert payload["estado"] == "en_espera"
    assert payload["estado_nombre"] == "En Espera"
    assert payload["en_lista_espera"] is True
    assert payload["curso"]["acepta_lista_espera"] is True


@pytest.mark.django_db
def test_api_vat_web_inscripciones_crea_sobre_comision_curso(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Web Insc")
    usuario = User.objects.create_user(username="api-web-insc", password="test1234")
    sexo = Sexo.objects.create(sexo="Masculino")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Web Insc",
        modalidad=modalidad,
        estado="activo",
        usa_voucher=True,
        costo_creditos=1,
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Web Insc",
        programa=programa,
        cantidad_inicial=4,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="WEB-INSC-01",
        nombre="Comisión Web Inscripción",
        cupo_total=10,
        fecha_inicio=date(2026, 5, 10),
        fecha_fin=date(2026, 6, 10),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="Pérez",
        nombre="Fabricio",
        fecha_nacimiento=date(1999, 4, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=2855739,
        sexo=sexo,
    )
    Voucher.objects.create(
        parametria=voucher,
        ciudadano=ciudadano,
        programa=programa,
        cantidad_inicial=4,
        cantidad_usada=0,
        cantidad_disponible=4,
        fecha_vencimiento=date(2026, 12, 31),
        estado="activo",
        asignado_por=usuario,
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/",
        {
            "ciudadano_id": ciudadano.id,
            "comision_curso_id": comision.id,
            "estado": "inscripta",
        },
        format="json",
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["comision"] == comision.id
    assert payload["comision_curso"] == comision.id
    assert payload["curso"]["id"] == comision.id
    assert Inscripcion.objects.filter(
        ciudadano=ciudadano,
        comision_curso=comision,
    ).exists()


@pytest.mark.django_db
def test_comision_curso_permita_cupo_independiente_del_curso(vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Soldadura Inicial",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="SOLD-01",
        nombre="Comisión mañana",
        cupo_total=25,
        fecha_inicio="2026-03-05",
        fecha_fin="2026-03-25",
        estado="planificada",
    )

    comision.full_clean()


@pytest.mark.django_db
def test_comision_curso_permita_fechas_independientes_del_curso(vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Electricidad",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="ELEC-01",
        nombre="Comisión tarde",
        cupo_total=20,
        fecha_inicio=date(2026, 3, 28),
        fecha_fin=date(2026, 4, 20),
        estado="planificada",
    )

    comision.full_clean()


@pytest.mark.django_db
def test_comision_curso_form_no_expone_codigo_ni_nombre_y_los_autogenera(
    vat_curso_base,
):
    centro, ubicacion, modalidad = vat_curso_base
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso comisión automática",
        modalidad=modalidad,
        estado="planificado",
    )

    form = ComisionCursoForm(
        data={
            "curso": str(curso.id),
            "ubicacion": str(ubicacion.id),
            "cupo_total": 25,
            "fecha_inicio": "2026-04-01",
            "fecha_fin": "2026-04-30",
            "estado": "planificada",
            "observaciones": "",
        }
    )

    assert "codigo_comision" not in form.fields
    assert "nombre" not in form.fields
    assert form.is_valid(), form.errors

    comision = form.save()

    assert comision.codigo_comision.startswith(f"COMCUR-{curso.id}-")
    assert comision.nombre == f"Comisión {curso.nombre}"


@pytest.mark.django_db
def test_comision_curso_create_view_renderiza_formulario(client, vat_curso_base):
    centro, _, modalidad = vat_curso_base
    user = User.objects.create_superuser(
        username="admin-comision-curso-create",
        email="admin-comision-curso-create@vat.test",
        password="test1234",
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso para formulario de comisión",
        modalidad=modalidad,
        estado="planificado",
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_comision_curso_create"),
        {"curso": curso.pk},
    )

    assert response.status_code == 200
    assert "ubicacion" in response.content.decode("utf-8")


@pytest.mark.django_db
def test_comision_curso_create_view_rechaza_fecha_fin_anterior_a_inicio(
    client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    user = User.objects.create_superuser(
        username="admin-comision-curso-fechas",
        email="admin-comision-curso-fechas@vat.test",
        password="test1234",
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con validacion de fechas",
        modalidad=modalidad,
        estado="planificado",
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_create"),
        data={
            "curso": str(curso.id),
            "ubicacion": str(ubicacion.id),
            "cupo_total": 20,
            "fecha_inicio": "2026-04-30",
            "fecha_fin": "2026-04-01",
            "estado": "planificada",
            "observaciones": "",
        },
    )

    assert response.status_code == 200
    assert (
        "La fecha de fin debe ser mayor o igual a la fecha de inicio."
        in response.content.decode("utf-8")
    )
    assert not ComisionCurso.objects.filter(curso=curso).exists()


@pytest.mark.django_db
def test_comision_curso_update_view_renderiza_formulario(client, vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    user = User.objects.create_superuser(
        username="admin-comision-curso-update",
        email="admin-comision-curso-update@vat.test",
        password="test1234",
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso para editar comisión",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        cupo_total=20,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="planificada",
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_comision_curso_update", kwargs={"pk": comision.pk})
    )

    assert response.status_code == 200
    assert "ubicacion" in response.content.decode("utf-8")


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="VAT.urls")
def test_curso_update_ajax_invalido_devuelve_json_y_no_persiste(client, vat_curso_base):
    centro, _, modalidad = vat_curso_base
    sector = Sector.objects.create(nombre="Sector AJAX Curso")
    plan = PlanVersionCurricular.objects.create(
        nombre="Plan AJAX Curso",
        provincia=centro.provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    user = User.objects.create_superuser(
        username="admin-curso-ajax-update",
        email="admin-curso-ajax-update@vat.test",
        password="test1234",
    )
    curso = Curso.objects.create(
        centro=centro,
        plan_estudio=plan,
        modalidad=modalidad,
        nombre="Curso AJAX",
        estado="planificado",
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_curso_update", kwargs={"pk": curso.pk}),
        data={
            "nombre": "Curso AJAX editado",
            "estado": "activo",
            "costo_creditos": "0",
            "observaciones": "No deberia persistir",
        },
        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
    )
    curso.refresh_from_db()
    payload = response.json()

    assert response.status_code == 400
    assert payload["ok"] is False
    assert "plan_estudio" in payload["errors"]
    assert curso.nombre == "Curso AJAX"


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="VAT.urls")
def test_comision_curso_update_ajax_invalido_devuelve_json_y_no_persiste(
    client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    user = User.objects.create_superuser(
        username="admin-comision-ajax-update",
        email="admin-comision-ajax-update@vat.test",
        password="test1234",
    )
    curso = Curso.objects.create(
        centro=centro,
        modalidad=modalidad,
        nombre="Curso AJAX Comision",
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        cupo_total=20,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="planificada",
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_update", kwargs={"pk": comision.pk}),
        data={
            "curso": str(curso.pk),
            "ubicacion": str(ubicacion.pk),
            "cupo_total": 20,
            "fecha_inicio": "2026-04-30",
            "fecha_fin": "2026-04-01",
            "estado": "planificada",
            "observaciones": "No deberia persistir",
        },
        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
    )
    comision.refresh_from_db()
    payload = response.json()

    assert response.status_code == 400
    assert payload["ok"] is False
    assert payload["errors"]["fecha_fin"][0] == (
        "La fecha de fin debe ser mayor o igual a la fecha de inicio."
    )
    assert comision.fecha_inicio == date(2026, 4, 1)


@pytest.mark.django_db
def test_comision_curso_no_permite_ubicacion_de_otro_centro(vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Semipresencial", activo=True)
    centro_a = Centro.objects.create(
        nombre="Centro A",
        codigo="A-001",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="10",
        numero=100,
        domicilio_actividad="Calle 10",
        telefono="221-1000000",
        celular="221-1000001",
        correo="a@vat.test",
        nombre_referente="Ref",
        apellido_referente="A",
        telefono_referente="221-1000002",
        correo_referente="refa@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    centro_b = Centro.objects.create(
        nombre="Centro B",
        codigo="B-001",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="11",
        numero=101,
        domicilio_actividad="Calle 11",
        telefono="221-1100000",
        celular="221-1100001",
        correo="b@vat.test",
        nombre_referente="Ref",
        apellido_referente="B",
        telefono_referente="221-1100002",
        correo_referente="refb@vat.test",
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion_b = InstitucionUbicacion.objects.create(
        centro=centro_b,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 11",
        es_principal=True,
    )

    curso = Curso(
        centro=centro_a,
        nombre="Curso Inválido",
        modalidad=modalidad,
        estado="planificado",
    )
    curso.full_clean()
    curso.save()

    comision = ComisionCurso(
        curso=curso,
        ubicacion=ubicacion_b,
        codigo_comision="CUR-A-01",
        nombre="Comisión inválida",
        cupo_total=10,
        fecha_inicio=date(2026, 1, 1),
        fecha_fin=date(2026, 1, 10),
        estado="planificada",
    )

    with pytest.raises(ValidationError):
        comision.full_clean()


@pytest.mark.django_db
def test_curso_form_rechaza_vouchers_fuera_del_programa(vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    usuario = User.objects.create_user(username="voucher-curso", password="test1234")
    programa_otro = Programa.objects.create(nombre="Programa Otro")
    programa_extra = Programa.objects.create(nombre="Programa Extra")
    sector = Sector.objects.create(nombre="Servicios")
    plan_estudio = PlanVersionCurricular.objects.create(
        provincia=centro.provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    voucher_otro_programa = VoucherParametria.objects.create(
        nombre="Voucher Programa Otro",
        programa=programa_otro,
        cantidad_inicial=3,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    voucher_programa_extra = VoucherParametria.objects.create(
        nombre="Voucher Programa Extra",
        programa=programa_extra,
        cantidad_inicial=3,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )

    form = CursoForm(
        data={
            "plan_estudio": str(plan_estudio.id),
            "nombre": "Curso Test Voucher",
            "estado": "planificado",
            "usa_voucher": "on",
            "voucher_parametrias": [
                str(voucher_otro_programa.id),
                str(voucher_programa_extra.id),
            ],
            "costo_creditos": 1,
            "observaciones": "",
        },
        initial={"centro": centro},
    )

    assert not form.is_valid()
    assert "voucher_parametrias" in form.errors
    assert "programa" not in form.fields


@pytest.mark.django_db
def test_curso_form_plan_estudio_es_primer_campo():
    form = CursoForm()

    assert list(form.fields.keys())[0] == "plan_estudio"
    assert "ubicacion" not in form.fields


@pytest.mark.django_db
def test_curso_form_requiere_costo_creditos_si_usa_voucher(vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Test Costo")
    usuario = User.objects.create_user(username="voucher-costo-1", password="test1234")
    sector = Sector.objects.create(nombre="Administración")
    plan_estudio = PlanVersionCurricular.objects.create(
        provincia=centro.provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Costo",
        programa=programa,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )

    form = CursoForm(
        data={
            "plan_estudio": str(plan_estudio.id),
            "nombre": "Curso sin costo",
            "estado": "planificado",
            "usa_voucher": "on",
            "voucher_parametrias": [str(voucher.id)],
            "costo_creditos": "",
            "observaciones": "",
        },
        initial={"centro": centro},
    )

    assert not form.is_valid()
    assert "costo_creditos" in form.errors


@pytest.mark.django_db
def test_curso_form_default_costo_creditos_si_no_usa_voucher(vat_curso_base):
    centro, ubicacion, modalidad = vat_curso_base
    sector = Sector.objects.create(nombre="Turismo")
    plan_estudio = PlanVersionCurricular.objects.create(
        provincia=centro.provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )

    form = CursoForm(
        data={
            "plan_estudio": str(plan_estudio.id),
            "nombre": "Curso sin voucher",
            "estado": "planificado",
            "costo_creditos": "",
            "observaciones": "",
        },
        initial={"centro": centro},
    )

    assert form.is_valid(), form.errors
    assert form.cleaned_data["costo_creditos"] == 0


@pytest.mark.django_db
def test_curso_form_no_permite_voucher_e_inscripcion_libre_al_mismo_tiempo(
    vat_curso_base,
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Exclusión Curso")
    usuario = User.objects.create_user(
        username="curso-mixto-invalido",
        password="test1234",
    )
    sector = Sector.objects.create(nombre="Servicios")
    plan_estudio = PlanVersionCurricular.objects.create(
        provincia=centro.provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Exclusión Curso",
        programa=programa,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )

    form = CursoForm(
        data={
            "plan_estudio": str(plan_estudio.id),
            "nombre": "Curso mixto inválido",
            "estado": "planificado",
            "usa_voucher": "on",
            "inscripcion_libre": "on",
            "voucher_parametrias": [str(voucher.id)],
            "costo_creditos": 1,
            "observaciones": "",
        },
        initial={"centro": centro},
    )

    assert not form.is_valid()
    assert "inscripcion_libre" in form.errors


@pytest.mark.django_db
def test_curso_form_guarda_plan_estudio(vat_curso_base, vat_plan_estudio_base):
    centro, ubicacion, modalidad = vat_curso_base
    _, _, _, _, titulo, _ = vat_plan_estudio_base
    titulo.plan_estudio.provincia = centro.provincia
    titulo.plan_estudio.save(update_fields=["provincia"])

    form = CursoForm(
        data={
            "plan_estudio": str(titulo.plan_estudio_id),
            "nombre": "Curso con plan",
            "estado": "planificado",
            "costo_creditos": 1,
            "observaciones": "",
        },
        initial={"centro": centro},
    )

    assert form.is_valid(), form.errors

    curso = form.save(commit=False)
    curso.centro = centro
    curso.save()

    assert curso.plan_estudio_id == titulo.plan_estudio_id
    assert curso.modalidad_id == titulo.plan_estudio.modalidad_cursada_id
    assert curso.programa_id is None


@pytest.mark.django_db
def test_curso_programa_se_deriva_desde_vouchers(vat_curso_base):
    centro, _, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Derivado")
    usuario = User.objects.create_user(username="curso-derivado", password="test1234")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con programa derivado",
        modalidad=modalidad,
        estado="planificado",
        usa_voucher=True,
        costo_creditos=1,
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Derivado",
        programa=programa,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )

    curso.voucher_parametrias.add(voucher)

    assert curso.programa_id == programa.id
    assert curso.programa == programa


@pytest.mark.django_db
def test_curso_viewset_filtra_programa_solo_si_el_derivado_es_consistente(
    vat_curso_base,
):
    centro, _, modalidad = vat_curso_base
    user = User.objects.create_user(username="api-vat-programa", password="test1234")
    programa_ok = Programa.objects.create(nombre="Programa OK")
    programa_otro = Programa.objects.create(nombre="Programa Otro API")
    voucher_ok = VoucherParametria.objects.create(
        nombre="Voucher API OK",
        programa=programa_ok,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=user,
        activa=True,
    )
    voucher_otro = VoucherParametria.objects.create(
        nombre="Voucher API Otro",
        programa=programa_otro,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=user,
        activa=True,
    )
    curso_programa_unico = Curso.objects.create(
        centro=centro,
        nombre="Curso programa único",
        modalidad=modalidad,
        estado="planificado",
        usa_voucher=True,
        costo_creditos=1,
    )
    curso_programa_unico.voucher_parametrias.add(voucher_ok)
    curso_programa_mixto = Curso.objects.create(
        centro=centro,
        nombre="Curso programa mixto",
        modalidad=modalidad,
        estado="planificado",
        usa_voucher=True,
        costo_creditos=1,
    )
    curso_programa_mixto.voucher_parametrias.add(voucher_ok, voucher_otro)

    request = RequestFactory().get("/vat/api/cursos/", {"programa_id": programa_ok.id})
    view = CursoViewSet()
    view.action_map = {"get": "list"}
    view.request = view.initialize_request(request)

    queryset = view.get_queryset()

    assert list(queryset.values_list("id", flat=True)) == [curso_programa_unico.id]


@pytest.mark.django_db
def test_curso_form_filtra_plan_estudio_por_provincia_del_centro(vat_curso_base):
    centro, _, _ = vat_curso_base
    provincia_ba = centro.provincia
    provincia_sf = Provincia.objects.create(nombre="Santa Fe")

    sector = Sector.objects.create(nombre="Servicios")
    modalidad = ModalidadCursada.objects.create(nombre="Virtual", activo=True)

    plan_ba = PlanVersionCurricular.objects.create(
        provincia=provincia_ba,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    PlanVersionCurricular.objects.create(
        provincia=provincia_sf,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )

    form = CursoForm(initial={"centro": centro})
    plan_ids = set(form.fields["plan_estudio"].queryset.values_list("id", flat=True))

    assert plan_ba.id in plan_ids
    assert all(
        provincia_id == provincia_ba.id
        for provincia_id in form.fields["plan_estudio"].queryset.values_list(
            "provincia_id", flat=True
        )
    )


@pytest.mark.django_db
def test_centro_cursos_panel_context_incluye_plan_actual_inactivo_para_edicion(
    vat_curso_base,
):
    centro, _, modalidad = vat_curso_base
    sector = Sector.objects.create(nombre="Sector Legacy")
    plan_inactivo = PlanVersionCurricular.objects.create(
        nombre="Plan Inactivo Legacy",
        provincia=centro.provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        activo=False,
    )
    curso = Curso.objects.create(
        centro=centro,
        plan_estudio=plan_inactivo,
        modalidad=modalidad,
        nombre="Curso Legacy",
        estado="planificado",
    )

    context = centro_views._build_cursos_panel_context(
        RequestFactory().get("/"), centro
    )
    plan_ids = set(
        context["curso_form"]
        .fields["plan_estudio"]
        .queryset.values_list("id", flat=True)
    )

    assert curso.plan_estudio_id in plan_ids


@pytest.mark.django_db
def test_centro_detail_difiere_panel_cursos_hasta_abrir_solapa(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Virtual", activo=True)
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-vat-centro-detail",
        email="admin-centro-detail@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP 777",
        codigo="CFP-777",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="12",
        numero=345,
        domicilio_actividad="Calle 12 N° 345",
        telefono="221-7000001",
        celular="221-7000002",
        correo="cfp777@vat.test",
        nombre_referente="Laura",
        apellido_referente="Diaz",
        telefono_referente="221-7000003",
        correo_referente="laura@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 12 N° 345",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Filtrable",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="FIL-01",
        nombre="Comisión Filtrable",
        cupo_total=30,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="planificada",
    )

    client.force_login(user)
    detail_url = reverse("vat_centro_detail", kwargs={"pk": centro.pk})
    panel_url = reverse("vat_centro_cursos_panel", kwargs={"pk": centro.pk})
    response = client.get(detail_url)
    content = response.content.decode("utf-8")

    assert response.status_code == 200
    assert f'data-cursos-panel-url="{panel_url}"' in content
    assert 'id="centroCursosPanelContainer"' in content
    assert 'data-panel-loaded="0"' in content
    assert 'id="tablaCursosCentro"' not in content
    assert 'id="tablaComisionesCursoCentro"' not in content
    assert "loadCursosPanel" in content
    assert "cursosPageSizeSelect.value = '25';" in content
    assert "row.addEventListener('click', lockRowFilter);" in content
    assert "row.addEventListener('keydown', function(event)" in content
    assert "row.addEventListener('mouseenter'" not in content
    assert "row.addEventListener('focus'" not in content
    assert "previewCursoId" not in content
    assert "hoverFilterEnabled" not in content


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="config.urls")
def test_centro_cursos_panel_renderiza_marcadores_para_filtrar_comisiones_por_curso(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Virtual", activo=True)
    sector = Sector.objects.create(nombre="Industria")
    plan = PlanVersionCurricular.objects.create(
        provincia=provincia,
        nombre="Plan Industrial Inicial",
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-vat-centro-detail-panel",
        email="admin-centro-detail-panel@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP 777 Panel",
        codigo="CFP-777-PANEL",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="12",
        numero=345,
        domicilio_actividad="Calle 12 N° 345",
        telefono="221-7000001",
        celular="221-7000002",
        correo="cfp777-panel@vat.test",
        nombre_referente="Laura",
        apellido_referente="Diaz",
        telefono_referente="221-7000003",
        correo_referente="laura@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 12 N° 345",
        es_principal=True,
    )
    _curso = Curso.objects.create(
        centro=centro,
        plan_estudio=plan,
        nombre="Curso Filtrable",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=_curso,
        ubicacion=ubicacion,
        codigo_comision="FIL-01",
        nombre="Comisión Filtrable",
        cupo_total=30,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="planificada",
    )

    client.force_login(user)
    response = client.get(reverse("vat_centro_cursos_panel", kwargs={"pk": centro.pk}))
    content = response.content.decode("utf-8")
    soup = BeautifulSoup(content, "html.parser")
    comisiones_filter_curso = soup.select_one("#comisionesFilterCurso")

    assert response.status_code == 200
    assert 'data-panel-rendered="1"' in content
    assert 'id="tablaCursosCentro"' in content
    assert "<th>Plan Curricular</th>" in content
    assert "<th>Curso FP</th>" in content
    assert 'id="cursosFilterSearch"' in content
    assert 'id="cursosFilterEstado"' in content
    assert 'id="cursosFilterPageSize"' in content
    assert '<option value="25" selected>25</option>' in content
    assert 'id="cursosFilterClear"' in content
    assert 'class="curso-row"' in content
    assert f'data-curso-id="{_curso.id}"' in content
    assert 'data-curso-plan="Plan Industrial Inicial"' in content
    assert (
        'aria-label="Seleccionar curso Curso Filtrable para filtrar comisiones"'
        in content
    )
    assert "<td>Plan Industrial Inicial</td>" in content
    assert 'id="tablaComisionesCursoCentro"' in content
    assert "<th>Código</th>" in content
    assert "<th>Ubicación</th>" in content
    assert "<th>Fecha Inicio</th>" in content
    assert "<th>Fecha Fin</th>" in content
    assert "<th>Observaciones</th>" in content
    assert "FIL-01" in content
    assert "Comisión Filtrable" in content
    assert 'id="comisionesFilterSearch"' in content
    assert 'id="comisionesFilterCurso"' in content
    assert 'id="comisionesFilterEstado"' in content
    assert 'id="comisionesFilterPageSize"' in content
    assert '<option value="25" selected>25</option>' in content
    assert 'id="comisionesFilterClear"' in content
    assert 'class="comision-curso-row"' in content
    assert reverse("vat_comision_curso_detail", kwargs={"pk": comision.pk}) in content
    assert 'title="Gestionar Comisión"' in content
    assert comisiones_filter_curso is not None
    assert "select2" in comisiones_filter_curso.get("class", [])
    assert comisiones_filter_curso.get("data-width") == "100%"


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="config.urls")
def test_centro_cursos_panel_renderiza_selector_de_planes_en_modal_nuevo_curso(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Virtual", activo=True)
    sector = Sector.objects.create(nombre="Servicios")
    titulo = TituloReferencia.objects.create(nombre="Plan con acceso", activo=True)
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-vat-centro-plan-curso",
        email="admin-centro-plan-curso@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP 779",
        codigo="CFP-779",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="14",
        numero=100,
        domicilio_actividad="Calle 14 N° 100",
        telefono="221-7100001",
        celular="221-7100002",
        correo="cfp779@vat.test",
        nombre_referente="Marta",
        apellido_referente="Lopez",
        telefono_referente="221-7100003",
        correo_referente="marta@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    plan = PlanVersionCurricular.objects.create(
        provincia=provincia,
        sector=sector,
        modalidad_cursada=modalidad,
        normativa="Resolución 100/2026",
        activo=True,
    )
    plan.titulos.add(titulo)
    plan_inactivo = PlanVersionCurricular.objects.create(
        provincia=provincia,
        nombre="Plan Inactivo",
        sector=sector,
        modalidad_cursada=modalidad,
        activo=False,
    )
    plan_otra_provincia = PlanVersionCurricular.objects.create(
        provincia=Provincia.objects.create(nombre="Otra Provincia"),
        nombre="Plan Otra Provincia",
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )

    client.force_login(user)
    response = client.get(reverse("vat_centro_cursos_panel", kwargs={"pk": centro.pk}))
    content = response.content.decode("utf-8")
    soup = BeautifulSoup(content, "html.parser")
    selector_sector = soup.select_one("#planCurricularSelectorSector")

    assert response.status_code == 200
    assert 'data-panel-rendered="1"' in content
    assert "Planes Curriculares" not in content
    assert 'title="Nuevo Curso"' in content
    assert 'id="planEstudioSeleccionadoInfo"' in content
    assert 'id="modalPlanCurricularSelector"' in content
    assert 'id="openPlanCurricularSelector"' in content
    assert 'id="planCurricularSelectorSearch"' in content
    assert 'id="planCurricularSelectorSector"' in content
    assert 'id="planCurricularSelectorModalidad"' not in content
    assert 'id="tablaPlanCurricularSelector"' in content
    assert "Plan Curricular" in content
    assert (
        "Se listan todos los planes curriculares activos de la provincia del centro."
        in content
    )
    assert "Seleccionar plan curricular" in content
    assert "Buscar por plan, sector o normativa" in content
    assert f'value="{plan.id}"' in content
    assert f'value="{plan_inactivo.id}"' not in content
    assert f'value="{plan_otra_provincia.id}"' not in content
    assert selector_sector is not None
    assert "select2" in selector_sector.get("class", [])
    assert selector_sector.get("data-width") == "100%"
    assert selector_sector.get("data-dropdown-parent") == "#modalPlanCurricularSelector"


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_detail_muestra_gestion_equivalente(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Presencial", activo=True)
    programa = Programa.objects.create(nombre="Programa Curso VAT")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_user(
        username="referente-comision-curso-detail",
        email="referente-comision-curso-detail@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    permisos = Permission.objects.filter(
        content_type__app_label="VAT",
        codename__in=[
            "view_comisioncurso",
            "change_comisioncurso",
            "delete_comisioncurso",
            "add_inscripcion",
            "add_comisionhorario",
            "change_comisionhorario",
            "delete_comisionhorario",
        ],
    )
    _grant_vat_referente_access(user, *permisos)
    centro = Centro.objects.create(
        nombre="CFP 888",
        codigo="CFP-888",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="13",
        numero=456,
        domicilio_actividad="Calle 13 N 456",
        telefono="221-8000001",
        celular="221-8000002",
        correo="cfp888@vat.test",
        nombre_referente="Ana",
        apellido_referente="Suarez",
        telefono_referente="221-8000003",
        correo_referente="ana@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formacion Profesional",
        situacion="Institucion de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 13 N 456",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con detalle",
        modalidad=modalidad,
        estado="planificado",
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Comision Detalle",
        programa=programa,
        cantidad_inicial=3,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=user,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="DET-01",
        nombre="Comision Detalle",
        cupo_total=25,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 5, 1),
        estado="activa",
    )
    dia = Dia.objects.create(nombre="Martes")
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(9, 0),
        hora_hasta=time(11, 0),
        aula_espacio="Aula 3",
        vigente=True,
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_comision_curso_detail", kwargs={"pk": comision.pk})
    )
    content = response.content.decode("utf-8")

    assert response.status_code == 200
    assert "Curso con detalle" in content
    assert reverse("vat_comision_curso_update", kwargs={"pk": comision.pk}) in content
    assert reverse("vat_comision_curso_delete", kwargs={"pk": comision.pk}) in content
    assert 'data-bs-target="#modalEditarComisionCurso"' in content
    assert 'id="modalEditarComisionCurso"' in content
    assert 'id="formEditarComisionCurso"' in content
    assert 'data-bs-target="#modalEliminarComisionCurso"' in content
    assert 'id="modalEliminarComisionCurso"' in content
    assert 'id="formEliminarComisionCurso"' in content
    assert reverse("vat_inscripcion_rapida_comision_curso") in content
    assert reverse("vat_comision_curso_horario_create") in content
    assert 'data-bs-target="#modalComisionHorario"' in content
    assert 'data-horario-mode="edit"' in content
    assert (
        reverse("vat_comision_curso_horario_update", kwargs={"pk": horario.pk})
        in content
    )
    assert 'data-bs-target="#modalEliminarComisionHorario"' in content
    assert (
        reverse("vat_comision_curso_horario_delete", kwargs={"pk": horario.pk})
        in content
    )
    assert "Informaci" in content
    assert "Inscriptos" in content
    assert "Sesiones" in content
    assert "Horarios" in content
    assert (
        reverse(
            "vat_comision_curso_export_preinscriptos",
            kwargs={"pk": comision.pk},
        )
        in content
    )
    assert (
        reverse(
            "vat_comision_curso_export_inscriptos",
            kwargs={"pk": comision.pk},
        )
        in content
    )


def _build_comision_curso_nomina_export_context(vat_geo_data, suffix):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(
        nombre=f"Presencial Export {suffix}", activo=True
    )
    programa = Programa.objects.create(nombre=f"Programa Export {suffix}")
    sexo = Sexo.objects.create(sexo="Femenino")
    user = User.objects.create_superuser(
        username=f"admin-export-{suffix.lower()}",
        email=f"admin-export-{suffix.lower()}@vat.test",
        password="test1234",
    )
    group, _ = Group.objects.get_or_create(name="CFP")
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre=f"CFP Export {suffix}",
        codigo=f"CFP-EXP-{suffix}",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="18",
        numero=900,
        domicilio_actividad=f"Calle 18 N 900 {suffix}",
        telefono="221-9000001",
        celular="221-9000002",
        correo=f"cfp-export-{suffix.lower()}@vat.test",
        nombre_referente="Ana",
        apellido_referente="Export",
        telefono_referente="221-9000003",
        correo_referente=f"ana-export-{suffix.lower()}@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formacion Profesional",
        situacion="Institucion de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio=f"Calle 18 N 900 {suffix}",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre=f"Curso Export {suffix}",
        modalidad=modalidad,
        estado="planificado",
    )
    voucher = VoucherParametria.objects.create(
        nombre=f"Voucher Export {suffix}",
        programa=programa,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=user,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision=f"EXP-{suffix}",
        nombre=f"Comision Export {suffix}",
        cupo_total=10,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    ciudadana_inscripta = Ciudadano.objects.create(
        apellido="Perez",
        nombre="Ana",
        fecha_nacimiento=date(1990, 5, 17),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111222,
        sexo=sexo,
        email="ana.perez@vat.test",
        telefono="221-1234567",
        cuil_cuit="27-30111222-5",
    )
    ciudadano_espera = Ciudadano.objects.create(
        apellido="Lopez",
        nombre="Juan",
        fecha_nacimiento=date(1988, 7, 9),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=28123456,
        sexo=sexo,
        email="juan.lopez@vat.test",
        telefono="221-7654321",
    )
    Inscripcion.objects.create(
        ciudadano=ciudadana_inscripta,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )
    Inscripcion.objects.create(
        ciudadano=ciudadano_espera,
        comision_curso=comision,
        programa=programa,
        estado="en_espera",
        origen_canal="front_publico",
    )
    return user, comision


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_exporta_nomina_preinscriptos_en_excel(client, vat_geo_data):
    user, comision = _build_comision_curso_nomina_export_context(vat_geo_data, "PRE")
    client.force_login(user)

    response = client.get(
        reverse(
            "vat_comision_curso_export_preinscriptos",
            kwargs={"pk": comision.pk},
        )
    )

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    header = [cell.value for cell in next(worksheet.iter_rows(max_row=1))]
    rows = [
        dict(zip(header, row))
        for row in worksheet.iter_rows(min_row=2, values_only=True)
    ]
    rows_by_name = {row["Nombre"]: row for row in rows}

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "nomina_preinscriptos" in response["Content-Disposition"]
    assert header == [
        "Apellido",
        "Nombre",
        "DNI / CUIL",
        "Fecha de Nacimiento",
        "G\u00e9nero",
        "Comisi\u00f3n",
        "Curso",
        "Centro de Formaci\u00f3n",
        "Estado de Inscripci\u00f3n",
        "Fecha de Inscripci\u00f3n",
        "Canal de Inscripci\u00f3n",
        "Email",
        "Tel\u00e9fono",
    ]
    assert len(rows) == 2
    assert rows_by_name["Ana"]["Apellido"] == "Perez"
    assert rows_by_name["Ana"]["DNI / CUIL"] == "27-30111222-5"
    assert rows_by_name["Ana"]["Estado de Inscripci\u00f3n"] == "Inscripta"
    assert rows_by_name["Ana"]["Canal de Inscripci\u00f3n"] == "Backoffice"
    assert (
        rows_by_name["Ana"]["Centro de Formaci\u00f3n"] == comision.curso.centro.nombre
    )
    assert rows_by_name["Juan"]["Apellido"] == "Lopez"
    assert rows_by_name["Juan"]["DNI / CUIL"] == "28123456"
    assert rows_by_name["Juan"]["Estado de Inscripci\u00f3n"] == "En Espera"
    assert rows_by_name["Juan"]["Canal de Inscripci\u00f3n"] == "Front P\u00fablico"
    assert rows_by_name["Ana"]["Fecha de Inscripci\u00f3n"]
    assert rows_by_name["Juan"]["Fecha de Inscripci\u00f3n"]


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_exporta_nomina_inscriptos_solo_estado_inscripta(
    client, vat_geo_data
):
    user, comision = _build_comision_curso_nomina_export_context(vat_geo_data, "INSC")
    client.force_login(user)

    response = client.get(
        reverse(
            "vat_comision_curso_export_inscriptos",
            kwargs={"pk": comision.pk},
        )
    )

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(min_row=2, values_only=True))

    assert response.status_code == 200
    assert "nomina_inscriptos" in response["Content-Disposition"]
    assert len(rows) == 1
    assert rows[0][0] == "Perez"
    assert rows[0][1] == "Ana"
    assert rows[0][8] == "Inscripta"


def _build_comision_curso_horario_context(vat_geo_data, user, suffix):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(
        nombre=f"Presencial Horario {suffix}", activo=True
    )
    centro = Centro.objects.create(
        nombre=f"CFP Horarios {suffix}",
        codigo=f"CFP-HOR-{suffix}",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="14",
        numero=100,
        domicilio_actividad=f"Calle 14 N° 100 {suffix}",
        telefono="221-1111111",
        celular="221-2222222",
        correo=f"cfphor-{suffix}@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333333",
        correo_referente=f"ana-hor-{suffix}@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio=f"Calle 14 N° 100 {suffix}",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre=f"Curso con horarios {suffix}",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision=f"HOR-{suffix}",
        nombre=f"Comisión Horario {suffix}",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 6),
        fecha_fin=date(2026, 4, 20),
        estado="activa",
    )
    dia = Dia.objects.create(nombre=f"Lunes {suffix}")
    return comision, dia


@pytest.mark.django_db
def test_comision_curso_horario_create_genera_sesiones(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(
        nombre="Presencial Horario", activo=True
    )
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-comision-curso-horario",
        email="admin-comision-curso-horario@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Horarios",
        codigo="CFP-HOR",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="14",
        numero=100,
        domicilio_actividad="Calle 14 N° 100",
        telefono="221-1111111",
        celular="221-2222222",
        correo="cfphor@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333333",
        correo_referente="ana-hor@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 14 N° 100",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con horarios",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="HOR-01",
        nombre="Comisión Horario",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 6),
        fecha_fin=date(2026, 4, 20),
        estado="activa",
    )
    dia = Dia.objects.create(nombre="Lunes")

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_horario_create"),
        data={
            "comision_curso": comision.pk,
            "dia_semana": dia.pk,
            "hora_desde": "09:00",
            "hora_hasta": "11:00",
            "aula_espacio": "Aula 1",
            "vigente": "on",
        },
    )

    horario = ComisionHorario.objects.get(comision_curso=comision)

    assert response.status_code == 302
    assert horario.dia_semana == dia
    assert (
        SesionComision.objects.filter(comision_curso=comision, horario=horario).count()
        == 3
    )


@pytest.mark.django_db
def test_comision_curso_horario_create_rechaza_hora_hasta_menor_a_hora_desde(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(
        nombre="Presencial Horario Invalido", activo=True
    )
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-comision-curso-horario-invalido",
        email="admin-comision-curso-horario-invalido@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Horarios Invalidos",
        codigo="CFP-HOR-INV",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="14",
        numero=101,
        domicilio_actividad="Calle 14 N° 101",
        telefono="221-1111113",
        celular="221-2222224",
        correo="cfphorinv@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333335",
        correo_referente="ana-hor-inv@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 14 N° 101",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con horario invalido",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="HOR-INV-01",
        nombre="Comisión Horario Inválido",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 6),
        fecha_fin=date(2026, 4, 20),
        estado="activa",
    )
    dia = Dia.objects.create(nombre="Martes")

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_horario_create"),
        data={
            "comision_curso": comision.pk,
            "dia_semana": dia.pk,
            "hora_desde": "11:00",
            "hora_hasta": "09:00",
            "aula_espacio": "Aula 2",
            "vigente": "on",
        },
    )

    assert response.status_code == 200
    assert (
        "La hora hasta no puede ser menor a la hora desde."
        in response.content.decode("utf-8")
    )
    assert not ComisionHorario.objects.filter(comision_curso=comision).exists()


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_horario_create_ajax_devuelve_json_con_errores(
    client, vat_geo_data
):
    user = User.objects.create_superuser(
        username="admin-comision-curso-horario-ajax-create",
        email="admin-comision-curso-horario-ajax-create@vat.test",
        password="test1234",
    )
    comision, dia = _build_comision_curso_horario_context(
        vat_geo_data, user, "AJAX-CREATE"
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_horario_create"),
        data={
            "comision_curso": comision.pk,
            "dia_semana": dia.pk,
            "hora_desde": "11:00",
            "hora_hasta": "09:00",
            "aula_espacio": "Aula 2",
            "vigente": "on",
        },
        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
    )

    payload = response.json()

    assert response.status_code == 400
    assert payload["ok"] is False
    assert (
        payload["errors"]["hora_hasta"][0]
        == "La hora hasta no puede ser menor a la hora desde."
    )
    assert not ComisionHorario.objects.filter(comision_curso=comision).exists()


@pytest.mark.django_db
def test_api_vat_inscripciones_curso_envia_a_lista_espera_si_no_hay_cupo(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa API Lista Espera")
    usuario = User.objects.create_user(
        username="api-lista-espera",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="No Binario")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso API Lista Espera",
        modalidad=modalidad,
        estado="activo",
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher API Lista Espera",
        programa=programa,
        cantidad_inicial=5,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="API-ESPERA-01",
        nombre="Comisión API Espera",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 5, 1),
        fecha_fin=date(2026, 6, 1),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Curso",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=40111224,
        sexo=sexo,
    )
    aspirante = Ciudadano.objects.create(
        apellido="Espera",
        nombre="Curso",
        fecha_nacimiento=date(2001, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=40111225,
        sexo=sexo,
    )
    Inscripcion.objects.create(
        ciudadano=titular,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    response = vat_api_client.post(
        "/api/vat/inscripciones-curso/",
        {
            "ciudadano": aspirante.id,
            "comision_curso": comision.id,
            "estado": "inscripta",
            "origen_canal": "api",
        },
        format="json",
    )

    payload = response.json()

    assert response.status_code == 201
    assert payload["estado"] == "en_espera"
    assert Inscripcion.objects.filter(
        ciudadano=aspirante,
        comision_curso=comision,
        estado="en_espera",
    ).exists()


@pytest.mark.django_db
def test_api_vat_web_prevalidar_permite_lista_espera_si_no_hay_cupo(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Web Lista Espera")
    usuario = User.objects.create_user(
        username="api-web-lista-espera",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="Femenino")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Web Lista Espera",
        modalidad=modalidad,
        estado="activo",
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Web Lista Espera",
        programa=programa,
        cantidad_inicial=4,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="WEB-ESPERA-01",
        nombre="Comisión Web Espera",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 5, 10),
        fecha_fin=date(2026, 6, 10),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Web",
        fecha_nacimiento=date(1990, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111223,
        sexo=sexo,
    )
    aspirante = Ciudadano.objects.create(
        apellido="Espera",
        nombre="Web",
        fecha_nacimiento=date(1995, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111224,
        sexo=sexo,
    )
    Inscripcion.objects.create(
        ciudadano=titular,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/prevalidar/",
        {
            "documento": str(aspirante.documento),
            "cuil": "27-30111224-4",
            "comision_curso_id": comision.id,
        },
        format="json",
    )

    payload = response.json()

    assert response.status_code == 200
    assert payload["puede_inscribirse"] is True
    assert payload["motivos"] == []
    assert payload["comision"]["acepta_lista_espera"] is True
    assert payload["comision"]["ingresa_a_lista_espera"] is True
    assert payload["comision"]["cupos_disponibles"] == 0


@pytest.mark.django_db
def test_api_vat_web_prevalidar_permite_lista_espera_sin_voucher_si_no_hay_cupo(
    vat_api_client, vat_curso_base
):
    centro, ubicacion, modalidad = vat_curso_base
    programa = Programa.objects.create(nombre="Programa Web Espera Sin Voucher")
    usuario = User.objects.create_user(
        username="api-web-lista-espera-sin-voucher",
        password="test1234",
    )
    sexo = Sexo.objects.create(sexo="Femenino")
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso Web Espera Voucher",
        modalidad=modalidad,
        estado="activo",
        usa_voucher=True,
        costo_creditos=2,
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Web Espera Sin Voucher",
        programa=programa,
        cantidad_inicial=4,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=usuario,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="WEB-ESPERA-VCH-01",
        nombre="Comisión Web Espera Voucher",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 5, 10),
        fecha_fin=date(2026, 6, 10),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Web Voucher",
        fecha_nacimiento=date(1990, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111233,
        sexo=sexo,
    )
    aspirante = Ciudadano.objects.create(
        apellido="Espera",
        nombre="Web Voucher",
        fecha_nacimiento=date(1995, 2, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=30111234,
        sexo=sexo,
    )
    Inscripcion.objects.create(
        ciudadano=titular,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    response = vat_api_client.post(
        "/api/vat/web/inscripciones/prevalidar/",
        {
            "documento": str(aspirante.documento),
            "cuil": "27-30111234-4",
            "comision_curso_id": comision.id,
        },
        format="json",
    )

    payload = response.json()

    assert response.status_code == 200
    assert payload["puede_inscribirse"] is True
    assert payload["motivos"] == []
    assert payload["comision"]["acepta_lista_espera"] is True
    assert payload["comision"]["ingresa_a_lista_espera"] is True
    assert payload["comision"]["cupos_disponibles"] == 0
    assert payload["voucher"]["requerido"] is True
    assert payload["voucher"]["voucher_id"] is None


@pytest.mark.django_db
def test_inscripcion_curso_no_permite_mover_de_inscripta_a_lista_espera(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(
        nombre="Presencial Bloqueo Espera", activo=True
    )
    programa = Programa.objects.create(nombre="Programa Bloqueo Espera")
    sexo = Sexo.objects.create(sexo="No Binario")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-bloqueo-espera",
        email="admin-bloqueo-espera@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Bloqueo Espera",
        codigo="CFP-BLOQ-ESP",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="18",
        numero=500,
        domicilio_actividad="Calle 18 N° 500",
        telefono="221-1111118",
        celular="221-2222229",
        correo="cfpbloqueoespera@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333340",
        correo_referente="ana-bloqueo-espera@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 18 N° 500",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso bloqueo espera",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="BLOQ-ESP-01",
        nombre="Comisión Bloqueo Espera",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Bloqueo",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=32345689,
        sexo=sexo,
    )
    inscripcion = Inscripcion.objects.create(
        ciudadano=ciudadano,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    client.force_login(user)
    response = client.post(
        reverse(
            "vat_inscripcion_curso_cambiar_estado",
            kwargs={"pk": inscripcion.pk},
        ),
        data={"estado": "en_espera"},
        follow=True,
    )

    inscripcion.refresh_from_db()

    assert response.status_code == 200
    assert inscripcion.estado == "inscripta"
    messages = list(response.context["messages"])
    assert any("cupo ocupado a lista de espera" in str(message) for message in messages)


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_horario_update_ajax_devuelve_json_con_errores(
    client, vat_geo_data
):
    user = User.objects.create_superuser(
        username="admin-comision-curso-horario-ajax-update",
        email="admin-comision-curso-horario-ajax-update@vat.test",
        password="test1234",
    )
    comision, dia = _build_comision_curso_horario_context(
        vat_geo_data, user, "AJAX-UPDATE"
    )
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(9, 0),
        hora_hasta=time(11, 0),
        aula_espacio="Aula 5",
        vigente=True,
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_horario_update", kwargs={"pk": horario.pk}),
        data={
            "comision_curso": comision.pk,
            "dia_semana": dia.pk,
            "hora_desde": "13:00",
            "hora_hasta": "12:00",
            "aula_espacio": "Aula 6",
            "vigente": "on",
        },
        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
    )

    payload = response.json()
    horario.refresh_from_db()

    assert response.status_code == 400
    assert payload["ok"] is False
    assert (
        payload["errors"]["hora_hasta"][0]
        == "La hora hasta no puede ser menor a la hora desde."
    )
    assert horario.hora_desde == time(9, 0)
    assert horario.hora_hasta == time(11, 0)


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_horario_delete_ajax_devuelve_json_ok(client, vat_geo_data):
    user = User.objects.create_superuser(
        username="admin-comision-curso-horario-ajax-delete",
        email="admin-comision-curso-horario-ajax-delete@vat.test",
        password="test1234",
    )
    comision, dia = _build_comision_curso_horario_context(
        vat_geo_data, user, "AJAX-DELETE"
    )
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(9, 0),
        hora_hasta=time(11, 0),
        aula_espacio="Aula 7",
        vigente=True,
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_comision_curso_horario_delete", kwargs={"pk": horario.pk}),
        HTTP_X_REQUESTED_WITH="XMLHttpRequest",
    )

    payload = response.json()

    assert response.status_code == 200
    assert payload["ok"] is True
    assert payload["redirect_url"] == reverse(
        "vat_comision_curso_detail", kwargs={"pk": comision.pk}
    )
    assert not ComisionHorario.objects.filter(pk=horario.pk).exists()


@pytest.mark.django_db
@override_settings(ROOT_URLCONF="tests.urls_vat_comision_horarios")
def test_comision_curso_detail_renderiza_modales_horario_sin_permiso_de_alta(
    client, vat_geo_data
):
    user = User.objects.create_user(
        username="referente-horario-sin-alta",
        email="referente-horario-sin-alta@vat.test",
        password="test1234",
    )
    change_permission = Permission.objects.get(
        content_type__app_label="VAT",
        codename="change_comisionhorario",
    )
    delete_permission = Permission.objects.get(
        content_type__app_label="VAT",
        codename="delete_comisionhorario",
    )
    detail_permission = Permission.objects.get(
        content_type__app_label="VAT",
        codename="view_comisioncurso",
    )
    _grant_vat_referente_access(
        user,
        detail_permission,
        change_permission,
        delete_permission,
    )
    comision, dia = _build_comision_curso_horario_context(
        vat_geo_data, user, "SIN-ALTA"
    )
    horario = ComisionHorario.objects.create(
        comision_curso=comision,
        dia_semana=dia,
        hora_desde=time(9, 0),
        hora_hasta=time(11, 0),
        aula_espacio="Aula 8",
        vigente=True,
    )

    client.force_login(user)
    response = client.get(
        reverse("vat_comision_curso_detail", kwargs={"pk": comision.pk})
    )
    content = response.content.decode("utf-8")

    assert response.status_code == 200
    assert reverse("vat_comision_curso_horario_create") not in content
    assert 'id="modalComisionHorario"' in content
    assert 'id="modalEliminarComisionHorario"' in content
    assert 'data-horario-mode="edit"' in content
    assert (
        reverse("vat_comision_curso_horario_update", kwargs={"pk": horario.pk})
        in content
    )
    assert (
        reverse("vat_comision_curso_horario_delete", kwargs={"pk": horario.pk})
        in content
    )


@pytest.mark.django_db
def test_inscripcion_rapida_comision_curso_crea_inscripcion(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Presencial Insc", activo=True)
    programa = Programa.objects.create(nombre="Programa Inscripción Curso")
    sexo = Sexo.objects.create(sexo="Femenino")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-comision-curso-insc",
        email="admin-comision-curso-insc@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Inscriptos",
        codigo="CFP-INSC",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="15",
        numero=100,
        domicilio_actividad="Calle 15 N° 100",
        telefono="221-1111112",
        celular="221-2222223",
        correo="cfpinsc@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333334",
        correo_referente="ana-insc@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 15 N° 100",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con inscripción",
        modalidad=modalidad,
        estado="planificado",
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Comisión Insc",
        programa=programa,
        cantidad_inicial=10,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=user,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="INSC-01",
        nombre="Comisión Inscriptos",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="Lopez",
        nombre="Juana",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=12345678,
        sexo=sexo,
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_inscripcion_rapida_comision_curso"),
        data={
            "comision": comision.pk,
            "ciudadano_id": ciudadano.pk,
            "observaciones": "Alta rápida",
        },
    )

    payload = response.json()
    inscripcion = Inscripcion.objects.get(comision_curso=comision, ciudadano=ciudadano)

    assert response.status_code == 200
    assert payload["ok"] is True
    assert inscripcion.programa == programa
    assert inscripcion.estado == "inscripta"


@pytest.mark.django_db
def test_inscripcion_rapida_comision_curso_envia_a_lista_espera(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Presencial Espera", activo=True)
    programa = Programa.objects.create(nombre="Programa Espera Curso")
    sexo = Sexo.objects.create(sexo="Masculino")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-comision-curso-espera",
        email="admin-comision-curso-espera@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Espera",
        codigo="CFP-ESP",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="15",
        numero=200,
        domicilio_actividad="Calle 15 N° 200",
        telefono="221-1111112",
        celular="221-2222223",
        correo="cfpespera@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333334",
        correo_referente="ana-espera@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 15 N° 200",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso con espera",
        modalidad=modalidad,
        estado="planificado",
    )
    voucher = VoucherParametria.objects.create(
        nombre="Voucher Comisión Espera",
        programa=programa,
        cantidad_inicial=10,
        fecha_vencimiento=date(2026, 12, 31),
        creado_por=user,
        activa=True,
    )
    curso.voucher_parametrias.add(voucher)
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="ESP-01",
        nombre="Comisión Espera",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Curso",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=12345679,
        sexo=sexo,
    )
    aspirante = Ciudadano.objects.create(
        apellido="Espera",
        nombre="Curso",
        fecha_nacimiento=date(2001, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=12345680,
        sexo=sexo,
    )
    Inscripcion.objects.create(
        ciudadano=titular,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_inscripcion_rapida_comision_curso"),
        data={
            "comision": comision.pk,
            "ciudadano_id": aspirante.pk,
            "observaciones": "Alta rápida en espera",
        },
    )

    payload = response.json()
    inscripcion = Inscripcion.objects.get(comision_curso=comision, ciudadano=aspirante)

    assert response.status_code == 200
    assert payload["ok"] is True
    assert payload["estado"] == "en_espera"
    assert "lista de espera" in payload["message"].lower()
    assert inscripcion.estado == "en_espera"


@pytest.mark.django_db
def test_inscripcion_rapida_comision_curso_libre_sin_programa_crea_inscripcion(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(nombre="Presencial Libre", activo=True)
    sexo = Sexo.objects.create(sexo="Femenino")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-comision-curso-libre",
        email="admin-comision-curso-libre@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Libre",
        codigo="CFP-LIBRE",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="15",
        numero=150,
        domicilio_actividad="Calle 15 N° 150",
        telefono="221-1111112",
        celular="221-2222223",
        correo="cfplibre@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333334",
        correo_referente="ana-libre@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 15 N° 150",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso libre sin programa",
        modalidad=modalidad,
        estado="activo",
        inscripcion_libre=True,
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="LIBRE-FAST-01",
        nombre="Comisión Libre Fast",
        cupo_total=20,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    ciudadano = Ciudadano.objects.create(
        apellido="Libre",
        nombre="Juana",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=22345678,
        sexo=sexo,
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_inscripcion_rapida_comision_curso"),
        data={
            "comision": comision.pk,
            "ciudadano_id": ciudadano.pk,
            "observaciones": "Alta rápida curso libre",
        },
    )

    payload = response.json()
    inscripcion = Inscripcion.objects.get(comision_curso=comision, ciudadano=ciudadano)

    assert response.status_code == 200
    assert payload["ok"] is True
    assert inscripcion.programa is None
    assert inscripcion.estado == "inscripta"


@pytest.mark.django_db
def test_inscripcion_rapida_comision_legacy_envia_a_lista_espera(client, vat_geo_data):
    provincia, municipio, localidad = vat_geo_data
    sexo = Sexo.objects.create(sexo="Femenino")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-comision-legacy-espera",
        email="admin-comision-legacy-espera@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Legacy Espera",
        codigo="CFP-LEG-ESP",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="16",
        numero=300,
        domicilio_actividad="Calle 16 N° 300",
        telefono="221-1111114",
        celular="221-2222225",
        correo="cfplegespera@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333336",
        correo_referente="ana-legacy-espera@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 16 N° 300",
        es_principal=True,
    )
    modalidad = ModalidadCursada.objects.create(nombre="Legacy Espera", activo=True)
    sector = Sector.objects.create(nombre="Sector Legacy Espera")
    plan = PlanVersionCurricular.objects.create(
        provincia=provincia,
        nombre="Plan Legacy Espera",
        sector=sector,
        modalidad_cursada=modalidad,
        activo=True,
    )
    programa = Programa.objects.create(nombre="Programa Legacy Espera")
    oferta = OfertaInstitucional.objects.create(
        centro=centro,
        plan_curricular=plan,
        programa=programa,
        nombre_local="Oferta Legacy Espera",
        ciclo_lectivo=2026,
        estado="publicada",
    )
    comision = Comision.objects.create(
        oferta=oferta,
        ubicacion=ubicacion,
        codigo_comision="LEG-ESP-01",
        nombre="Comisión Legacy Espera",
        cupo=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Legacy",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=22345679,
        sexo=sexo,
    )
    aspirante = Ciudadano.objects.create(
        apellido="Espera",
        nombre="Legacy",
        fecha_nacimiento=date(2001, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=22345680,
        sexo=sexo,
    )
    Inscripcion.objects.create(
        ciudadano=titular,
        comision=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )

    client.force_login(user)
    response = client.post(
        reverse("vat_inscripcion_rapida_comision"),
        data={
            "comision": comision.pk,
            "ciudadano_id": aspirante.pk,
            "observaciones": "Alta rápida legacy en espera",
        },
    )

    payload = response.json()
    inscripcion = Inscripcion.objects.get(comision=comision, ciudadano=aspirante)

    assert response.status_code == 200
    assert payload["ok"] is True
    assert payload["estado"] == "en_espera"
    assert "lista de espera" in payload["message"].lower()
    assert inscripcion.estado == "en_espera"


@pytest.mark.django_db
def test_comision_curso_detail_administra_lista_espera_desde_el_detalle(
    client, vat_geo_data
):
    provincia, municipio, localidad = vat_geo_data
    modalidad = ModalidadCursada.objects.create(
        nombre="Presencial Gestión Espera", activo=True
    )
    programa = Programa.objects.create(nombre="Programa Gestión Espera")
    sexo = Sexo.objects.create(sexo="No Binario")
    group, _ = Group.objects.get_or_create(name="CFP")
    user = User.objects.create_superuser(
        username="admin-gestion-espera",
        email="admin-gestion-espera@vat.test",
        password="test1234",
    )
    user.groups.add(group)
    centro = Centro.objects.create(
        nombre="CFP Gestión Espera",
        codigo="CFP-GES-ESP",
        provincia=provincia,
        municipio=municipio,
        localidad=localidad,
        calle="17",
        numero=400,
        domicilio_actividad="Calle 17 N° 400",
        telefono="221-1111116",
        celular="221-2222227",
        correo="cfpgestionespera@vat.test",
        nombre_referente="Ana",
        apellido_referente="Gomez",
        telefono_referente="221-3333338",
        correo_referente="ana-gestion-espera@vat.test",
        referente=user,
        tipo_gestion="Estatal",
        clase_institucion="Formación Profesional",
        situacion="Institución de ETP",
        activo=True,
    )
    ubicacion = InstitucionUbicacion.objects.create(
        centro=centro,
        localidad=localidad,
        rol_ubicacion="sede_principal",
        domicilio="Calle 17 N° 400",
        es_principal=True,
    )
    curso = Curso.objects.create(
        centro=centro,
        nombre="Curso gestión espera",
        modalidad=modalidad,
        estado="planificado",
    )
    comision = ComisionCurso.objects.create(
        curso=curso,
        ubicacion=ubicacion,
        codigo_comision="GES-ESP-01",
        nombre="Comisión Gestión Espera",
        cupo_total=1,
        acepta_lista_espera=True,
        fecha_inicio=date(2026, 4, 1),
        fecha_fin=date(2026, 4, 30),
        estado="activa",
    )
    titular = Ciudadano.objects.create(
        apellido="Titular",
        nombre="Gestión",
        fecha_nacimiento=date(2000, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=32345679,
        sexo=sexo,
    )
    espera = Ciudadano.objects.create(
        apellido="Espera",
        nombre="Gestión",
        fecha_nacimiento=date(2001, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=32345680,
        sexo=sexo,
    )
    ocupante = Inscripcion.objects.create(
        ciudadano=titular,
        comision_curso=comision,
        programa=programa,
        estado="inscripta",
        origen_canal="backoffice",
    )
    inscripcion_espera = Inscripcion.objects.create(
        ciudadano=espera,
        comision_curso=comision,
        programa=programa,
        estado="en_espera",
        origen_canal="backoffice",
    )

    client.force_login(user)
    detail_response = client.get(
        reverse("vat_comision_curso_detail", kwargs={"pk": comision.pk})
    )
    ocupante.estado = "abandonada"
    ocupante.save(update_fields=["estado", "fecha_modificacion"])
    response = client.post(
        reverse(
            "vat_inscripcion_curso_cambiar_estado",
            kwargs={"pk": inscripcion_espera.pk},
        ),
        data={"estado": "inscripta"},
    )

    inscripcion_espera.refresh_from_db()

    assert detail_response.status_code == 200
    assert "Lista de espera" in detail_response.content.decode("utf-8")
    assert response.status_code == 302
    assert inscripcion_espera.estado == "inscripta"


@pytest.mark.django_db
def test_voucher_parametria_detail_filters_and_paginates_vouchers(vat_admin_client):
    admin_user = User.objects.get(username="admin-vat")
    programa = Programa.objects.create(nombre="Programa Voucher Detail")
    sexo = Sexo.objects.create(sexo="Femenino")
    parametria = VoucherParametria.objects.create(
        nombre="Parametría detalle",
        programa=programa,
        cantidad_inicial=10,
        fecha_vencimiento=date(2028, 4, 6),
        creado_por=admin_user,
        activa=True,
    )

    for index in range(21):
        ciudadano = Ciudadano.objects.create(
            apellido=f"Apellido {index}",
            nombre=f"Nombre {index}",
            fecha_nacimiento=date(2000, 1, 1),
            tipo_documento=Ciudadano.DOCUMENTO_DNI,
            documento=30000000 + index,
            sexo=sexo,
        )
        Voucher.objects.create(
            parametria=parametria,
            ciudadano=ciudadano,
            programa=programa,
            cantidad_inicial=10,
            cantidad_usada=0,
            cantidad_disponible=10,
            fecha_vencimiento=date(2028, 4, 6),
            estado="activo",
        )

    ciudadano_filtrado = Ciudadano.objects.create(
        apellido="Pérez",
        nombre="Ana",
        fecha_nacimiento=date(2001, 1, 1),
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        documento=39999123,
        sexo=sexo,
    )
    Voucher.objects.create(
        parametria=parametria,
        ciudadano=ciudadano_filtrado,
        programa=programa,
        cantidad_inicial=10,
        cantidad_usada=1,
        cantidad_disponible=9,
        fecha_vencimiento=date(2028, 4, 6),
        estado="activo",
    )

    detail_url = reverse("vat_voucher_parametria_detail", kwargs={"pk": parametria.pk})

    response = vat_admin_client.get(detail_url)

    assert response.status_code == 200
    assert response.context["total"] == 22
    assert response.context["vouchers_page_obj"].paginator.per_page == 20
    assert len(response.context["vouchers"]) == 20
    assert response.context["vouchers_is_paginated"] is True

    filtered_response = vat_admin_client.get(detail_url, {"busqueda": "39999123"})

    assert filtered_response.status_code == 200
    assert filtered_response.context["vouchers_total_filtrados"] == 1
    assert len(filtered_response.context["vouchers"]) == 1
    assert (
        filtered_response.context["vouchers"][0].ciudadano_id == ciudadano_filtrado.id
    )

    second_page_response = vat_admin_client.get(detail_url, {"vouchers_page": 2})

    assert second_page_response.status_code == 200
    assert len(second_page_response.context["vouchers"]) == 2
