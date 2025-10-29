import csv
import os
import re
import unicodedata
from collections import defaultdict
from typing import Dict, Iterable, List, Optional, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from comedores.models import Comedor, Programas, TipoDeComedor
from core.models import Localidad, Municipio, Provincia

ACTIVE_STATE = "Activo"

# Headers normalizados -> nombre de campo interno o token especial
HEADER_TO_FIELD: Dict[str, str] = {
    "id": "id_sisoc",
    "id_sisoc": "id_sisoc",
    "idsisoc": "id_sisoc",
    "sisoc_id": "id_sisoc",
    "nombre": "nombre",
    "nombre_del_comedor": "nombre",
    "nombre_comedor": "nombre",
    "nombrecomedor": "nombre",
    "comedor": "nombre",
    "programa": "programa",
    "nombre_programa": "programa",
    "programa_nombre": "programa",
    "tipocomedor": "tipocomedor",
    "tipo": "tipocomedor",
    "tipo_de_comedor": "tipocomedor",
    "provincia": "provincia",
    "municipio": "municipio",
    "localidad": "localidad",
    "partido": "partido",
    "barrio": "barrio",
    "calle": "calle",
    "direccion": "calle",
    "numero": "numero",
    "altura": "numero",
    "entre_calle_1": "entre_calle_1",
    "entre_calle1": "entre_calle_1",
    "entrecalle1": "entre_calle_1",
    "entre_calle_2": "entre_calle_2",
    "entre_calle2": "entre_calle_2",
    "entrecalle2": "entre_calle_2",
    "latitud": "latitud",
    "lat": "latitud",
    "longitud": "longitud",
    "lon": "longitud",
    "lng": "longitud",
    "codigo_postal": "codigo_postal",
    "cod_postal": "codigo_postal",
    "cp": "codigo_postal",
    "comienzo": "comienzo",
    "anio": "comienzo",
    "anio_inicio": "comienzo",
    "codigo_de_proyecto": "codigo_de_proyecto",
    "cod_proyecto": "codigo_de_proyecto",
    "codigo_proyecto": "codigo_de_proyecto",
    "id_externo": "id_externo",
    "idexterno": "id_externo",
}

INT_FIELDS = {"numero", "codigo_postal", "comienzo", "id_externo"}
FLOAT_FIELDS = {"latitud", "longitud"}
STRING_FIELDS = {
    "nombre",
    "calle",
    "entre_calle_1",
    "entre_calle_2",
    "barrio",
    "partido",
    "codigo_de_proyecto",
}


def normalize_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower()
    normalized = normalized.replace(" ", "_").replace("-", "_").replace("/", "_")
    normalized = re.sub(r"[^a-z0-9_]+", "", normalized)
    normalized = re.sub(r"__+", "_", normalized)
    return normalized.strip("_")


def clean_cell(value):
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, float):
        if str(value).lower() == "nan":
            return None
        return value
    return value


def parse_int(value, field_name: str) -> Tuple[Optional[int], Optional[str]]:
    if value is None:
        return None, None
    if isinstance(value, int):
        return value, None
    text = str(value).strip()
    if not text:
        return None, None
    text = text.replace(",", ".")
    try:
        as_float = float(text)
    except ValueError:
        return None, f"No se pudo convertir '{value}' a entero para {field_name}"
    return int(as_float), None


def parse_float(value, field_name: str) -> Tuple[Optional[float], Optional[str]]:
    if value is None:
        return None, None
    if isinstance(value, float):
        return value, None
    text = str(value).strip()
    if not text:
        return None, None
    text = text.replace(",", ".")
    try:
        return float(text), None
    except ValueError:
        return None, f"No se pudo convertir '{value}' a flotante para {field_name}"


def as_string(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def resolve_programa(name: Optional[str]) -> Tuple[Optional[Programas], Optional[str]]:
    if not name:
        return None, None
    programa = Programas.objects.filter(nombre__iexact=name.strip()).first()
    if not programa:
        return None, f"Programa '{name}' no encontrado"
    return programa, None


def resolve_tipocomedor(
    name: Optional[str],
) -> Tuple[Optional[TipoDeComedor], Optional[str]]:
    if not name:
        return None, None
    tipo = TipoDeComedor.objects.filter(nombre__iexact=name.strip()).first()
    if not tipo:
        return None, f"Tipo de comedor '{name}' no encontrado"
    return tipo, None


def resolve_provincia(name: Optional[str]) -> Tuple[Optional[Provincia], Optional[str]]:
    if not name:
        return None, None
    provincia = Provincia.objects.filter(nombre__iexact=name.strip()).first()
    if not provincia:
        return None, f"Provincia '{name}' no encontrada"
    return provincia, None


def resolve_municipio(
    name: Optional[str], provincia: Optional[Provincia]
) -> Tuple[Optional[Municipio], Optional[str]]:
    if not name:
        return None, None
    qs = Municipio.objects.filter(nombre__iexact=name.strip())
    if provincia:
        qs = qs.filter(provincia=provincia)
    municipio = qs.first()
    if municipio:
        return municipio, None
    return None, f"Municipio '{name}' no encontrado"


def resolve_localidad(
    name: Optional[str], municipio: Optional[Municipio]
) -> Tuple[Optional[Localidad], Optional[str]]:
    if not name:
        return None, None
    qs = Localidad.objects.filter(nombre__iexact=name.strip())
    if municipio:
        qs = qs.filter(municipio=municipio)
    localidad = qs.first()
    if localidad:
        return localidad, None
    return None, f"Localidad '{name}' no encontrada"


def build_header_map(headers: Iterable[Optional[str]]) -> Dict[int, str]:
    header_map: Dict[int, str] = {}
    for idx, header_value in enumerate(headers):
        if not header_value:
            continue
        normalized = normalize_header(str(header_value))
        field_name = HEADER_TO_FIELD.get(normalized)
        if field_name:
            header_map[idx] = field_name
    return header_map


def row_has_values(row_data: Dict[str, Optional[object]]) -> bool:
    return any(value is not None for value in row_data.values())


def register_warnings(
    warnings_map: Dict[str, List[str]],
    key: str,
    messages: Iterable[str],
) -> None:
    for message in messages:
        warnings_map[key].append(message)


class Command(BaseCommand):
    help = (
        "Importa comedores desde un archivo Excel o CSV. "
        "Cuando la fila incluye id_sisoc actualiza el registro existente, "
        "y cuando no lo incluye crea un nuevo Comedor."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            help=(
                "Ruta al archivo de origen (Excel .xlsx o CSV) "
                "por ejemplo C:\\Users\\Juanito\\Downloads\\archivo.xlsx."
            ),
        )
        parser.add_argument(
            "--sheet-name",
            dest="sheet_name",
            default=None,
            help="Nombre de la hoja a procesar. Por defecto se usa la primera hoja del libro.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Procesa el archivo sin guardar cambios en la base de datos.",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        sheet_name = options["sheet_name"]
        dry_run = options["dry_run"]

        if not os.path.exists(file_path):
            raise CommandError(f"No se encontró el archivo: {file_path}")

        rows = self._load_rows(file_path, sheet_name)
        if not rows:
            self.stdout.write(self.style.WARNING("El archivo no contiene filas."))
            return

        header_map = build_header_map(rows[0])
        if not header_map:
            raise CommandError(
                "No se reconocieron columnas válidas en el encabezado del archivo."
            )

        updated_ids: List[int] = []
        created_ids: List[int] = []
        failures: List[Dict[str, object]] = []
        warnings: Dict[str, List[str]] = defaultdict(list)

        self.stdout.write(
            self.style.NOTICE(
                f"Columnas reconocidas: {', '.join(sorted(set(header_map.values())))}"
            )
        )
        if dry_run:
            self.stdout.write(
                self.style.WARNING("DRY RUN activo: no se guardarán cambios.")
            )

        for row_index, row_values in enumerate(rows[1:], start=2):
            row_data: Dict[str, Optional[object]] = {}
            for col_idx, cell_value in enumerate(row_values):
                field_name = header_map.get(col_idx)
                if not field_name:
                    continue
                row_data[field_name] = clean_cell(cell_value)

            if not row_has_values(row_data):
                continue

            row_identifier = f"fila {row_index}"
            row_warnings: List[str] = []
            id_sisoc_raw = row_data.get("id_sisoc")

            if id_sisoc_raw is not None:
                comedor_id, error = parse_int(id_sisoc_raw, "id_sisoc")
                if error:
                    failures.append(
                        {
                            "row": row_index,
                            "id_sisoc": id_sisoc_raw,
                            "reason": error,
                        }
                    )
                    continue
                try:
                    comedor = Comedor.objects.get(pk=comedor_id)
                except Comedor.DoesNotExist:
                    failures.append(
                        {
                            "row": row_index,
                            "id_sisoc": comedor_id,
                            "reason": f"No existe Comedor con id {comedor_id}",
                        }
                    )
                    continue

                update_fields = []

                if "programa" in row_data:
                    programa_name = as_string(row_data.get("programa"))
                    programa_obj, warn = resolve_programa(programa_name)
                    if warn:
                        row_warnings.append(warn)
                    comedor.programa = programa_obj
                    update_fields.append("programa")

                comedor.estado_general = ACTIVE_STATE
                update_fields.append("estado_general")

                if not dry_run:
                    comedor.save(update_fields=update_fields)
                updated_ids.append(comedor.id)

                if row_warnings:
                    register_warnings(warnings, f"id {comedor.id}", row_warnings)
                continue

            nombre = as_string(row_data.get("nombre"))
            if not nombre:
                failures.append(
                    {
                        "row": row_index,
                        "reason": "Columna 'nombre' vacía; no se puede crear el Comedor.",
                    }
                )
                continue

            comedor_kwargs: Dict[str, object] = {
                "nombre": nombre,
                "estado_general": ACTIVE_STATE,
            }

            if "programa" in row_data:
                programa_name = as_string(row_data.get("programa"))
                programa_obj, warn = resolve_programa(programa_name)
                if warn:
                    row_warnings.append(warn)
                comedor_kwargs["programa"] = programa_obj

            if "tipocomedor" in row_data:
                tipo_name = as_string(row_data.get("tipocomedor"))
                tipo_obj, warn = resolve_tipocomedor(tipo_name)
                if warn:
                    row_warnings.append(warn)
                comedor_kwargs["tipocomedor"] = tipo_obj

            provincia_obj = None
            if "provincia" in row_data:
                provincia_name = as_string(row_data.get("provincia"))
                provincia_obj, warn = resolve_provincia(provincia_name)
                if warn:
                    row_warnings.append(warn)
                comedor_kwargs["provincia"] = provincia_obj

            municipio_obj = None
            if "municipio" in row_data:
                municipio_name = as_string(row_data.get("municipio"))
                municipio_obj, warn = resolve_municipio(municipio_name, provincia_obj)
                if warn:
                    row_warnings.append(warn)
                comedor_kwargs["municipio"] = municipio_obj

            if "localidad" in row_data:
                localidad_name = as_string(row_data.get("localidad"))
                localidad_obj, warn = resolve_localidad(localidad_name, municipio_obj)
                if warn:
                    row_warnings.append(warn)
                comedor_kwargs["localidad"] = localidad_obj

            for field in STRING_FIELDS:
                if field not in row_data:
                    continue
                value = as_string(row_data.get(field))
                if value is not None:
                    comedor_kwargs[field] = value

            for field in INT_FIELDS:
                if field not in row_data:
                    continue
                value, warn = parse_int(row_data.get(field), field)
                if warn:
                    row_warnings.append(warn)
                if value is not None:
                    comedor_kwargs[field] = value

            for field in FLOAT_FIELDS:
                if field not in row_data:
                    continue
                value, warn = parse_float(row_data.get(field), field)
                if warn:
                    row_warnings.append(warn)
                if value is not None:
                    comedor_kwargs[field] = value

            if dry_run:
                temp_comedor = Comedor(**comedor_kwargs)
                created_ids.append(f"(simulado) {temp_comedor.nombre}")
                if row_warnings:
                    register_warnings(warnings, f"{row_identifier}", row_warnings)
            else:
                with transaction.atomic():
                    comedor = Comedor.objects.create(**comedor_kwargs)
                created_ids.append(comedor.id)
                if row_warnings:
                    register_warnings(warnings, f"id {comedor.id}", row_warnings)

        updated_display = ", ".join(map(str, updated_ids)) if updated_ids else "-"
        created_display = ", ".join(map(str, created_ids)) if created_ids else "-"

        self.stdout.write(
            self.style.SUCCESS(
                f"Comedores actualizados ({len(updated_ids)}): {updated_display}"
            )
        )
        self.stdout.write(
            self.style.SUCCESS(
                f"Comedores creados ({len(created_ids)}): {created_display}"
            )
        )

        if warnings:
            self.stdout.write(self.style.WARNING("Avisos:"))
            for key, messages in warnings.items():
                for message in messages:
                    self.stdout.write(f"  {key}: {message}")

        if failures:
            self.stdout.write(
                self.style.ERROR("No se pudieron procesar las siguientes filas:")
            )
            for failure in failures:
                row = failure.get("row")
                identifier = failure.get("id_sisoc")
                reason = failure.get("reason")
                if identifier is not None:
                    self.stdout.write(f"  Fila {row} (id_sisoc={identifier}): {reason}")
                else:
                    self.stdout.write(f"  Fila {row}: {reason}")
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    "Todas las filas reconocidas se procesaron correctamente."
                )
            )

    def _load_rows(
        self, file_path: str, sheet_name: Optional[str]
    ) -> List[List[Optional[object]]]:
        extension = os.path.splitext(file_path)[1].lower()
        if extension in {".xlsx", ".xlsm"}:
            return self._load_rows_from_excel(file_path, sheet_name)
        if extension == ".csv":
            if sheet_name:
                raise CommandError(
                    "La opción --sheet-name solo aplica para archivos Excel."
                )
            return self._load_rows_from_csv(file_path)
        raise CommandError(
            f"Extensión no soportada '{extension}'. Utilice un archivo .xlsx, .xlsm o .csv."
        )

    def _load_rows_from_excel(
        self, file_path: str, sheet_name: Optional[str]
    ) -> List[List[Optional[object]]]:
        workbook = load_workbook(filename=file_path, data_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise CommandError(
                    f"La hoja '{sheet_name}' no existe en el archivo. Hojas disponibles: {', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def _load_rows_from_csv(self, file_path: str) -> List[List[Optional[object]]]:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as fh:
            sample = fh.read(2048)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(fh, dialect)
            return [list(row) for row in reader]
