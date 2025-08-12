# centrodefamilia/services/informe_cabal_service.py
import logging
from dataclasses import dataclass
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.core.paginator import Paginator
from openpyxl import load_workbook
from centrodefamilia.models import Centro, CabalArchivo, InformeCabalRegistro

logger = logging.getLogger(__name__)

EXPECTED_HEADERS = [
    "NroTarjeta",
    "NroAuto",
    "MTI",
    "NroComercio",
    "RazonSocial",
    "Importe",
    "FechaTRX",
    "MonedaOrigen",
    "ImporteMonOrigen",
    "ImportePesos",
    "CantCuotas",
    "MotivoRechazo",
    "Desc_MotivoRechazo",
    "Disponibles",
]


@dataclass
class PreviewRow:
    fila: int
    data: Dict[str, Any]
    no_coincidente: bool


def _parse_date_ddmmyyyy(value: Any) -> Optional[datetime.date]:
    if value in (None, "", 0):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        value = value.strip()
        try:
            return datetime.strptime(value, "%d/%m/%Y").date()
        except Exception:
            pass
    return None


def _get_cell(row, col_index):
    val = row[col_index].value
    return "" if val is None else val


def read_excel_preview(
    file: UploadedFile, page: int = 1, per_page: int = 25
) -> Tuple[List[PreviewRow], List[int], int]:

    wb = load_workbook(filename=file, data_only=True)
    ws = wb.active

    headers = [
        str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]
    ]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValueError("El encabezado del Excel no coincide con el esperado.")

    rows_raw: List[Tuple[int, Dict[str, Any]]] = []
    codigos_set = set()

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        data = {
            "NroTarjeta": _get_cell(row, 0),
            "NroAuto": _get_cell(row, 1),
            "MTI": _get_cell(row, 2),
            "NroComercio": _get_cell(row, 3),
            "RazonSocial": _get_cell(row, 4),
            "Importe": _get_cell(row, 5),
            "FechaTRX": _get_cell(row, 6),
            "MonedaOrigen": _get_cell(row, 7),
            "ImporteMonOrigen": _get_cell(row, 8),
            "ImportePesos": _get_cell(row, 9),
            "CantCuotas": _get_cell(row, 10),
            "MotivoRechazo": _get_cell(row, 11),
            "Desc_MotivoRechazo": _get_cell(row, 12),
            "Disponibles": _get_cell(row, 13),
        }
        codigo = str(data["NroComercio"]).strip()
        if codigo:
            codigos_set.add(codigo)
        rows_raw.append((i - 1, data))

    total_count = len(rows_raw)

    valid_codes = set(
        Centro.objects.filter(codigo__in=codigos_set).values_list("codigo", flat=True)
    )

    all_rows: List[PreviewRow] = []
    not_matching: List[int] = []
    for fila, data in rows_raw:
        codigo = str(data["NroComercio"]).strip()
        no_coincidente = codigo not in valid_codes
        if no_coincidente:
            not_matching.append(fila)
        all_rows.append(PreviewRow(fila=fila, data=data, no_coincidente=no_coincidente))

    paginator = Paginator(all_rows, per_page)
    page_obj = paginator.get_page(page)
    return list(page_obj.object_list), not_matching, total_count


@transaction.atomic
def persist_file_and_rows(
    file: UploadedFile,
    user,
    force_proceed: bool = False,
) -> Tuple[CabalArchivo, int, int, List[int]]:

    nombre = file.name
    nombre_duplicado = CabalArchivo.objects.filter(nombre_original=nombre).exists()
    advert = nombre_duplicado and not force_proceed
    if advert:
        logger.info("Nombre de archivo duplicado detectado: %s", nombre)
        raise FileExistsError("DUPLICATE_NAME")

    preview_rows, not_matching, total_rows = read_excel_preview(
        file, page=1, per_page=10**9
    )
    codigos_unicos = {
        str(pr.data["NroComercio"]).strip()
        for pr in preview_rows
        if str(pr.data["NroComercio"]).strip()
    }
    centros_map: Dict[str, int] = dict(
        Centro.objects.filter(codigo__in=codigos_unicos).values_list("codigo", "id")
    )

    cabal_archivo = CabalArchivo.objects.create(
        archivo=file,
        nombre_original=nombre,
        usuario=user,
        advertencia_nombre_duplicado=nombre_duplicado,
        total_filas=total_rows,
        total_validas=0,
        total_invalidas=0,
    )

    total_validas = 0
    total_invalidas = 0
    registros_bulk: List[InformeCabalRegistro] = []

    def _to_decimal(v):
        if v in ("", None):
            return None
        try:
            return float(str(v).replace(",", "."))
        except Exception:
            return None

    for pr in preview_rows:
        d = pr.data
        codigo = str(d["NroComercio"]).strip()
        centro_id = centros_map.get(codigo)

        reg = InformeCabalRegistro(
            archivo=cabal_archivo,
            centro_id=centro_id,
            nro_tarjeta=str(d["NroTarjeta"]),
            nro_auto=str(d["NroAuto"]),
            mti=str(d["MTI"]),
            nro_comercio=codigo,
            razon_social=str(d["RazonSocial"]),
            importe=_to_decimal(d["Importe"]),
            fecha_trx=_parse_date_ddmmyyyy(d["FechaTRX"]),
            moneda_origen=str(d["MonedaOrigen"]),
            importe_mon_origen=_to_decimal(d["ImporteMonOrigen"]),
            importe_pesos=_to_decimal(d["ImportePesos"]),
            cant_cuotas=(
                int(d["CantCuotas"]) if str(d["CantCuotas"]).strip().isdigit() else None
            ),
            motivo_rechazo=str(d["MotivoRechazo"]),
            desc_motivo_rechazo=str(d["Desc_MotivoRechazo"]),
            disponibles=str(d["Disponibles"]),
            no_coincidente=pr.no_coincidente,
            fila_numero=pr.fila,
        )
        if pr.no_coincidente:
            total_invalidas += 1
        else:
            total_validas += 1
        registros_bulk.append(reg)

    InformeCabalRegistro.objects.bulk_create(registros_bulk, batch_size=2000)

    cabal_archivo.total_validas = total_validas
    cabal_archivo.total_invalidas = total_invalidas
    cabal_archivo.save(update_fields=["total_validas", "total_invalidas"])

    return cabal_archivo, total_rows, total_validas, not_matching
