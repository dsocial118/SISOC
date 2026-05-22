from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from centrodefamilia.services.consulta_renaper import (  # pylint: disable=no-name-in-module
    consultar_datos_renaper,
)
from ciudadanos.models import Ciudadano, CiudadanosImportJobRow
from comedores.services.comedor_service import ComedorService
from core.models import Sexo

SHEET_NAME = "ciudadanos"
TEMPLATE_HEADERS = ("cuil_o_dni", "sexo")
TEMPLATE_FILENAME = "plantilla_ciudadanos_importacion_masiva.xlsx"
DOCUMENTO_ALIASES = {"cuil_o_dni", "cuil", "dni", "documento"}
SEXO_ALIASES = {"sexo"}
RENAPER_SEXOS = ("M", "F", "X")
SYSTEMIC_RENAPER_ERROR_TYPES = {
    "timeout",
    "auth_error",
    "remote_error",
    "invalid_response",
}
CUIL_WEIGHTS = (5, 4, 3, 2, 7, 6, 5, 4, 3, 2)
SEXO_LABELS = {
    "M": "Masculino",
    "F": "Femenino",
    "X": "X",
}
CIUDADANO_RENAPER_FK_FIELDS = (
    "sexo",
    "provincia",
    "municipio",
    "localidad",
    "nacionalidad",
)


@dataclass(frozen=True)
class ParsedDocumento:
    dni: str
    cuil: str
    input_type: str


@dataclass(frozen=True)
class ParsedCiudadanosImportRow:
    fila: int
    documento_raw: str
    dni: str
    cuil: str
    sexo: str
    input_type: str
    parse_error: str = ""
    error_type: str = ""


def _normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    normalized = normalized.lower().replace(" ", "_").replace("-", "_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _digits_only(value: object) -> str:
    return re.sub(r"\D", "", _clean_cell(value))


def _cuil_verifier(base_digits: str) -> int:
    total = sum(int(digit) * weight for digit, weight in zip(base_digits, CUIL_WEIGHTS))
    verifier = 11 - (total % 11)
    if verifier == 11:
        return 0
    if verifier == 10:
        return 9
    return verifier


def _is_valid_cuil(cuil: str) -> bool:
    if len(cuil) != 11 or not cuil.isdigit():
        return False
    return _cuil_verifier(cuil[:10]) == int(cuil[-1])


def parse_cuil_o_dni(value: object) -> ParsedDocumento:
    digits = _digits_only(value)
    if len(digits) == 11:
        if not _is_valid_cuil(digits):
            raise ValidationError(
                "El CUIL informado tiene digito verificador invalido."
            )
        return ParsedDocumento(
            dni=digits[2:10],
            cuil=digits,
            input_type="cuil",
        )
    if len(digits) == 8 and int(digits or "0") > 0:
        return ParsedDocumento(
            dni=digits,
            cuil="",
            input_type="dni",
        )
    raise ValidationError("El documento debe ser un DNI o CUIL valido.")


def normalize_import_sexo(value: object) -> str:
    text = _clean_cell(value)
    if not text:
        return ""
    normalized = _normalize_header(text)
    aliases = {
        "m": "M",
        "masculino": "M",
        "varon": "M",
        "f": "F",
        "femenino": "F",
        "mujer": "F",
        "x": "X",
        "no_binario": "X",
    }
    sexo = aliases.get(normalized)
    if not sexo:
        raise ValidationError("El sexo informado debe ser M, F o X.")
    return sexo


def _build_header_map(headers: list[str]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if normalized in DOCUMENTO_ALIASES and "documento" not in header_map:
            header_map["documento"] = index
        elif normalized in SEXO_ALIASES and "sexo" not in header_map:
            header_map["sexo"] = index

    if "documento" not in header_map:
        raise ValidationError(
            "El archivo debe incluir la columna cuil_o_dni, cuil, dni o documento."
        )
    return header_map


def _load_rows_from_workbook(uploaded_file) -> tuple[object, list[tuple]]:
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc

    try:
        worksheet = (
            workbook[SHEET_NAME]
            if SHEET_NAME in workbook.sheetnames
            else workbook.active
        )
        rows = list(worksheet.iter_rows(values_only=True))
    except Exception:
        workbook.close()
        raise
    return workbook, rows


def _get_row_value(row: tuple, index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _clean_cell(row[index])


def _parse_import_row(
    row_number: int,
    row: tuple,
    header_map: dict[str, int],
) -> ParsedCiudadanosImportRow | None:
    documento_raw = _get_row_value(row, header_map["documento"])
    sexo_raw = _get_row_value(row, header_map.get("sexo"))
    if not documento_raw and not sexo_raw:
        return None

    dni = ""
    cuil = ""
    input_type = ""
    parse_error = ""
    error_type = ""
    try:
        parsed_documento = parse_cuil_o_dni(documento_raw)
        dni = parsed_documento.dni
        cuil = parsed_documento.cuil
        input_type = parsed_documento.input_type
    except ValidationError as exc:
        parse_error = " ".join(exc.messages)
        error_type = (
            "invalid_cuil" if len(_digits_only(documento_raw)) == 11 else "invalid_dni"
        )

    sexo = ""
    if not parse_error:
        try:
            sexo = normalize_import_sexo(sexo_raw)
        except ValidationError as exc:
            parse_error = " ".join(exc.messages)
            error_type = "invalid_sexo"

    return ParsedCiudadanosImportRow(
        fila=row_number,
        documento_raw=documento_raw,
        dni=dni,
        cuil=cuil,
        sexo=sexo,
        input_type=input_type,
        parse_error=parse_error,
        error_type=error_type,
    )


def load_ciudadanos_import_rows(uploaded_file) -> list[ParsedCiudadanosImportRow]:
    workbook, rows = _load_rows_from_workbook(uploaded_file)
    try:
        if not rows:
            raise ValidationError("El archivo Excel esta vacio.")
        headers = [_clean_cell(value) for value in rows[0]]
        header_map = _build_header_map(headers)

        parsed_rows: list[ParsedCiudadanosImportRow] = []
        for row_number, row in enumerate(rows[1:], start=2):
            parsed_row = _parse_import_row(row_number, row, header_map)
            if parsed_row is not None:
                parsed_rows.append(parsed_row)

        if not parsed_rows:
            raise ValidationError(
                "El archivo Excel no contiene filas con datos para procesar."
            )
        return parsed_rows
    finally:
        workbook.close()


def validate_ciudadanos_import_workbook(uploaded_file) -> None:
    rows = load_ciudadanos_import_rows(uploaded_file)
    if not rows:
        raise ValidationError("El archivo Excel no contiene filas con datos.")
    uploaded_file.seek(0)


def generate_ciudadanos_import_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(list(TEMPLATE_HEADERS))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def get_ciudadanos_import_template_filename() -> str:
    return TEMPLATE_FILENAME


def get_ciudadanos_import_results_filename(job) -> str:
    return f"resultado_importacion_ciudadanos_{job.pk}.xlsx"


def _format_import_result(row: CiudadanosImportJobRow) -> str:
    if row.status in {
        CiudadanosImportJobRow.Status.CREATED,
        CiudadanosImportJobRow.Status.EXISTING,
    }:
        return "OK"
    if row.status == CiudadanosImportJobRow.Status.FAILED:
        return "Fallo"
    return "Pendiente"


def _format_import_row_detail(row: CiudadanosImportJobRow) -> str:
    message = (row.mensaje or "").strip()
    error_type = (row.error_type or "").strip()
    if message and error_type:
        return f"{message} ({error_type})"
    return message or error_type or "-"


def generate_ciudadanos_import_job_results_workbook(job) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "filas"
    worksheet.freeze_panes = "A2"
    worksheet.append(
        [
            "Fila",
            "CUIL/DNI",
            "DNI",
            "Sexo",
            "Resultado",
            "Estado",
            "Intentos sexo",
            "Intentos",
            "Detalle",
            "Ciudadano",
        ]
    )
    rows = job.rows.select_related("ciudadano").order_by("fila", "id")
    for row in rows:
        detalle_url = f"/ciudadanos/ver/{row.ciudadano_id}" if row.ciudadano_id else ""
        worksheet.append(
            [
                row.fila,
                row.documento_raw or "-",
                row.dni or "-",
                row.sexo or "-",
                _format_import_result(row),
                row.get_status_display(),
                row.sexos_intentados or "-",
                row.attempts,
                _format_import_row_detail(row),
                "Ver" if detalle_url else "-",
            ]
        )
        if detalle_url:
            ciudadano_cell = worksheet.cell(row=worksheet.max_row, column=10)
            ciudadano_cell.hyperlink = detalle_url
            ciudadano_cell.style = "Hyperlink"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_ciudadanos_import_error_message(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        return " ".join(exc.messages)
    return "Ocurrio un error inesperado al procesar la fila."


def is_systemic_renaper_error(error_type: str | None) -> bool:
    return str(error_type or "") in SYSTEMIC_RENAPER_ERROR_TYPES


def _get_existing_estandar_by_dni(dni: str) -> Ciudadano | None:
    try:
        documento = int(dni)
    except (TypeError, ValueError):
        return None
    return (
        Ciudadano.objects.filter(
            tipo_registro_identidad=Ciudadano.TIPO_REGISTRO_ESTANDAR,
            tipo_documento=Ciudadano.DOCUMENTO_DNI,
            documento=documento,
        )
        .order_by("id")
        .first()
    )


def _extract_renaper_cuil(result: dict[str, object]) -> str:
    data = result.get("data") or {}
    datos_api = result.get("datos_api") or {}
    for value in (data.get("cuil"), datos_api.get("cuil")):
        digits = _digits_only(value)
        if len(digits) == 11:
            return digits
    return ""


def _apply_sexo_to_renaper_data(
    data: dict[str, object], sexo: str
) -> dict[str, object]:
    data = dict(data)
    if data.get("sexo"):
        return data
    sexo_label = SEXO_LABELS.get(sexo)
    if not sexo_label:
        return data
    sexo_obj = Sexo.objects.filter(sexo=sexo_label).first()
    if sexo_obj:
        data["sexo"] = sexo_obj.pk
    return data


def _build_ciudadano_payload_from_renaper(
    *,
    result: dict[str, object],
    dni: str,
    sexo: str,
) -> tuple[dict[str, object] | None, str | None]:
    data = _apply_sexo_to_renaper_data(result.get("data") or {}, sexo)
    ciudadano_data, error = ComedorService.build_ciudadano_data_from_renaper(data, dni)
    if not ciudadano_data:
        return None, error
    ciudadano_data.update(
        {
            "tipo_registro_identidad": Ciudadano.TIPO_REGISTRO_ESTANDAR,
            "estado_validacion_renaper": Ciudadano.RENAPER_VALIDADO,
            "fecha_validacion_renaper": timezone.now(),
            "datos_renaper": result.get("datos_api") or result.get("data") or {},
            "origen_dato": "renaper",
        }
    )
    renaper_cuil = _extract_renaper_cuil(result)
    if renaper_cuil:
        ciudadano_data["cuil_cuit"] = renaper_cuil
    _normalize_ciudadano_payload_foreign_keys(ciudadano_data)
    return ciudadano_data, None


def _normalize_ciudadano_payload_foreign_keys(
    ciudadano_data: dict[str, object]
) -> None:
    for field_name in CIUDADANO_RENAPER_FK_FIELDS:
        value = ciudadano_data.pop(field_name, None)
        if value in (None, ""):
            continue
        if hasattr(value, "pk"):
            ciudadano_data[field_name] = value
        else:
            ciudadano_data[f"{field_name}_id"] = value


def _lookup_renaper_for_row(
    row: ParsedCiudadanosImportRow,
) -> dict[str, object]:
    sexos = (row.sexo,) if row.sexo else RENAPER_SEXOS
    attempted: list[str] = []
    last_result: dict[str, object] | None = None

    for sexo in sexos:
        attempted.append(sexo)
        result = consultar_datos_renaper(row.dni, sexo)
        result["sexo_consultado"] = sexo
        result["sexos_intentados"] = attempted.copy()
        if result.get("success"):
            return result
        last_result = result
        if is_systemic_renaper_error(result.get("error_type")):
            return result

    if last_result is None:
        return {
            "success": False,
            "error": "No se pudo consultar RENAPER.",
            "error_type": "unexpected_error",
            "sexos_intentados": attempted,
        }
    last_result["sexos_intentados"] = attempted
    return last_result


def _process_successful_renaper_import(
    *,
    row: ParsedCiudadanosImportRow,
    requested_by,
    result: dict[str, object],
    sexos_intentados: str,
) -> dict[str, object]:
    renaper_cuil = _extract_renaper_cuil(result)
    if row.cuil and renaper_cuil and row.cuil != renaper_cuil:
        return {
            "status": "failed",
            "mensaje": "El CUIL informado no coincide con el CUIL devuelto por RENAPER.",
            "error_type": "cuil_mismatch",
            "sexos_intentados": sexos_intentados,
            "ciudadano": None,
            "systemic": False,
            "contacted_renaper": True,
        }

    ciudadano_data, error = _build_ciudadano_payload_from_renaper(
        result=result,
        dni=row.dni,
        sexo=str(result.get("sexo_consultado") or row.sexo or ""),
    )
    if not ciudadano_data:
        return {
            "status": "failed",
            "mensaje": error or "RENAPER no devolvio datos minimos para crear.",
            "error_type": "invalid_response_data",
            "sexos_intentados": sexos_intentados,
            "ciudadano": None,
            "systemic": False,
            "contacted_renaper": True,
        }

    if requested_by is not None:
        ciudadano_data["creado_por"] = requested_by
        ciudadano_data["modificado_por"] = requested_by

    with transaction.atomic():
        existing = _get_existing_estandar_by_dni(row.dni)
        if existing:
            row_result = {
                "status": "existing",
                "mensaje": "Ya existe un ciudadano estandar para el DNI informado.",
                "error_type": "",
                "sexos_intentados": sexos_intentados,
                "ciudadano": existing,
                "systemic": False,
                "contacted_renaper": True,
            }
        else:
            ciudadano = Ciudadano.objects.create(**ciudadano_data)
            row_result = {
                "status": "created",
                "mensaje": "Ciudadano creado desde RENAPER.",
                "error_type": "",
                "sexos_intentados": sexos_intentados,
                "ciudadano": ciudadano,
                "systemic": False,
                "contacted_renaper": True,
            }

    return row_result


def process_ciudadanos_import_row(
    *,
    row: ParsedCiudadanosImportRow,
    requested_by,
) -> dict[str, object]:
    if row.parse_error:
        return {
            "status": "failed",
            "mensaje": row.parse_error,
            "error_type": row.error_type,
            "sexos_intentados": "",
            "ciudadano": None,
            "systemic": False,
            "contacted_renaper": False,
        }

    existing = _get_existing_estandar_by_dni(row.dni)
    if existing:
        return {
            "status": "existing",
            "mensaje": "Ya existe un ciudadano estandar para el DNI informado.",
            "error_type": "",
            "sexos_intentados": row.sexo or "",
            "ciudadano": existing,
            "systemic": False,
            "contacted_renaper": False,
        }

    result = _lookup_renaper_for_row(row)
    sexos_intentados = ",".join(result.get("sexos_intentados") or [])
    if not result.get("success"):
        error_type = str(result.get("error_type") or "renaper_error")
        return {
            "status": "pending" if is_systemic_renaper_error(error_type) else "failed",
            "mensaje": str(
                result.get("error") or "No se encontraron datos en RENAPER."
            ),
            "error_type": error_type,
            "sexos_intentados": sexos_intentados,
            "ciudadano": None,
            "systemic": is_systemic_renaper_error(error_type),
            "contacted_renaper": True,
        }

    return _process_successful_renaper_import(
        row=row,
        requested_by=requested_by,
        result=result,
        sexos_intentados=sexos_intentados,
    )
