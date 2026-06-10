from __future__ import annotations

import logging
import unicodedata
from dataclasses import dataclass
from io import BytesIO

from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email
from openpyxl import Workbook, load_workbook

logger = logging.getLogger("django")

SHEET_NAME = "mailing"


def _normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    normalized = normalized.lower().replace(" ", "_").replace("-", "_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


@dataclass(frozen=True)
class ParsedMailingRow:
    fila: int
    mail: str


def _load_mailing_workbook_rows(uploaded_file) -> list[ParsedMailingRow]:
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc
    try:
        if SHEET_NAME in workbook.sheetnames:
            worksheet = workbook[SHEET_NAME]
        else:
            worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("El archivo Excel esta vacio.")

        headers = [_clean_cell(value) for value in rows[0]]

        # Search for 'mail' column
        mail_index = -1
        for index, header in enumerate(headers):
            if _normalize_header(header) == "mail":
                mail_index = index
                break

        if mail_index == -1:
            raise ValidationError("El archivo debe incluir una columna llamada 'mail'.")

        parsed_rows: list[ParsedMailingRow] = []
        for row_number, row in enumerate(rows[1:], start=2):
            mail = _clean_cell(row[mail_index] if mail_index < len(row) else "")
            if not mail:
                continue

            parsed_rows.append(
                ParsedMailingRow(
                    fila=row_number,
                    mail=mail,
                )
            )

        if not parsed_rows:
            raise ValidationError(
                "El archivo Excel no contiene filas con datos para procesar."
            )

        return parsed_rows
    finally:
        workbook.close()


def validate_mailing_workbook(uploaded_file):
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc

    try:
        worksheet = (
            workbook[SHEET_NAME]
            if SHEET_NAME in workbook.sheetnames
            else workbook.active
        )
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationError("El archivo Excel esta vacio.") from exc

        headers = [_clean_cell(value) for value in header_row]
        mail_index = -1
        for index, header in enumerate(headers):
            if _normalize_header(header) == "mail":
                mail_index = index
                break

        if mail_index == -1:
            raise ValidationError("El archivo debe incluir una columna llamada 'mail'.")

        for row in rows:
            mail = _clean_cell(row[mail_index] if mail_index < len(row) else "")
            if mail:
                uploaded_file.seek(0)
                return

        raise ValidationError(
            "El archivo Excel no contiene filas con datos para procesar."
        )
    finally:
        workbook.close()


def process_mailing_row(
    row: ParsedMailingRow, asunto: str, cuerpo: str
) -> dict[str, object]:
    mail_destino = row.mail
    try:
        validate_email(mail_destino)
    except ValidationError:
        raise ValidationError(f"El mail '{mail_destino}' no es valido.")

    try:
        send_mail(
            subject=asunto,
            message=cuerpo,
            from_email=None,  # Use default
            recipient_list=[mail_destino],
            fail_silently=False,
        )
    except Exception as exc:
        logger.exception("Error enviando mail masivo a %s", mail_destino)
        raise ValidationError(f"Error enviando mail: {str(exc)}")

    return {
        "mail_destino": mail_destino,
        "mensaje": "Enviado correctamente",
    }


def generate_mailing_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(["mail"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_mailing_error_message(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        return " ".join(exc.messages)
    return str(exc)
