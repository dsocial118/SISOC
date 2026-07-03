"""
Servicio para generar padron final del expediente de celiaquia.

La nomina de aprobados se construye desde la base de datos (legajos
aprobados con MATCH de Sintys), no desde el Excel original cargado por la
provincia. El archivo original queda estatico, mientras que los datos de
los ciudadanos pueden corregirse despues de la importacion; la base es la
fuente de verdad.
"""

from datetime import date, datetime
from io import BytesIO
import logging

from openpyxl import Workbook

from celiaquia.models import (
    EstadoCupo,
    ExpedienteCiudadano,
    ResultadoSintys,
    RevisionTecnico,
)
from celiaquia.services.cruce_service import CruceService

logger = logging.getLogger("django")

ESTADO_CUPO_HEADER = "Estado de cupo"

NOMINA_HEADERS = [
    "apellido",
    "nombre",
    "documento",
    "fecha_nacimiento",
    "sexo",
    "nacionalidad",
    "municipio",
    "localidad",
    "calle",
    "altura",
    "codigo_postal",
    "telefono",
    "email",
    "APELLIDO_RESPONSABLE",
    "NOMBRE_RESPONSABLE",
    "CUIT_RESPONSABLE",
    "FECHA_DE_NACIMIENTO_RESPONSABLE",
    "SEXO_RESPONSABLE",
    "DOMICILIO_RESPONSABLE",
    "LOCALIDAD_RESPONSABLE",
    "CELULAR_RESPONSABLE",
    "CORREO_RESPONSABLE",
]

FECHA_COLUMNS = {"fecha_nacimiento", "FECHA_DE_NACIMIENTO_RESPONSABLE"}

FECHA_NUMBER_FORMAT = "DD/MM/YYYY"


class PadronFinalService:
    """Genera nomina final aprobada en Excel para expediente de celiaquia."""

    @staticmethod
    def generar_padron_final_excel(expediente) -> bytes:
        """
        Genera Excel con la nomina de aprobados a partir de la base.

        Incluye los legajos con revision tecnica APROBADO y resultado
        Sintys MATCH, excluyendo a los responsables puros (sus datos se
        vuelcan en las columnas *_RESPONSABLE de cada beneficiario).

        Args:
            expediente: Expediente object

        Returns:
            bytes: Contenido del archivo Excel
        """
        from celiaquia.services.familia_service import FamiliaService

        legajos = list(PadronFinalService._legajos_aprobados(expediente))
        responsables_por_hijo = FamiliaService.obtener_responsables_por_hijo(
            [legajo.ciudadano_id for legajo in legajos]
        )

        rows = []
        sin_documento = []
        for legajo in legajos:
            responsables = responsables_por_hijo.get(legajo.ciudadano_id, [])
            if len(responsables) > 1:
                logger.warning(
                    "padron_final.responsable_ambiguo",
                    extra={
                        "data": {
                            "expediente_id": getattr(expediente, "id", None),
                            "ciudadano_id": legajo.ciudadano_id,
                            "candidatos": len(responsables),
                        }
                    },
                )
            responsable = responsables[0] if responsables else None
            fila = PadronFinalService._fila_beneficiario(legajo, responsable)
            if not fila[NOMINA_HEADERS.index("documento")]:
                sin_documento.append(legajo.ciudadano_id)
            rows.append(fila)

        if sin_documento:
            logger.warning(
                "padron_final.documento_vacio",
                extra={
                    "data": {
                        "expediente_id": getattr(expediente, "id", None),
                        "ciudadanos_sin_documento": sin_documento,
                    }
                },
            )

        return PadronFinalService._build_excel(
            [*NOMINA_HEADERS, ESTADO_CUPO_HEADER], rows
        )

    @staticmethod
    def _legajos_aprobados_qs(expediente):
        """Legajos APROBADO+MATCH del expediente, sin excluir responsables.

        La exclusion de responsables puros se hace en Python con
        ``ExpedienteCiudadano.es_rol_responsable_puro`` (ver
        ``_legajos_aprobados``) para usar la misma regla normalizada que
        cruce y cupo, en vez de un exclude(rol=...) exacto que no tolera
        variantes de mayusculas/espacios.
        """
        return ExpedienteCiudadano.objects.filter(
            expediente=expediente,
            revision_tecnico=RevisionTecnico.APROBADO,
            resultado_sintys=ResultadoSintys.MATCH,
        )

    @staticmethod
    def _legajos_aprobados(expediente):
        qs = (
            PadronFinalService._legajos_aprobados_qs(expediente)
            .select_related(
                "ciudadano",
                "ciudadano__sexo",
                "ciudadano__nacionalidad",
                "ciudadano__municipio",
                "ciudadano__localidad",
            )
            .order_by("ciudadano__apellido", "ciudadano__nombre")
        )
        return [
            legajo
            for legajo in qs
            if not ExpedienteCiudadano.es_rol_responsable_puro(legajo.rol)
        ]

    @staticmethod
    def hay_aprobados(expediente) -> bool:
        """True si el expediente tiene al menos un legajo para el padron."""
        qs = PadronFinalService._legajos_aprobados_qs(expediente)
        return any(
            not ExpedienteCiudadano.es_rol_responsable_puro(rol)
            for rol in qs.values_list("rol", flat=True)
        )

    @staticmethod
    def _texto(value) -> str:
        return "" if value is None else str(value)

    @staticmethod
    def _documento_o_cuit(ciudadano) -> str:
        """Documento a exportar, con fallback a CUIL/CUIT.

        SINTYS identifica primero por CUIL/CUIT y recien despues por
        documento; un ciudadano puede tener ``documento`` vacio y
        ``cuil_cuit`` cargado. Si no hay documento, se deriva el nucleo DNI
        del CUIT (misma logica que el cruce) y, si eso tampoco resuelve
        nada, se exporta el cuil_cuit crudo antes que dejar la celda vacia.
        """
        documento = PadronFinalService._texto(ciudadano.documento)
        if documento:
            return documento
        cuil_cuit = PadronFinalService._texto(ciudadano.cuil_cuit)
        if not cuil_cuit:
            return ""
        return CruceService.extraer_dni_de_cuit(cuil_cuit) or cuil_cuit

    @staticmethod
    def _estado_cupo_label(estado_cupo) -> str:
        """Etiqueta legible del estado de cupo para la columna del padron."""
        if estado_cupo == EstadoCupo.DENTRO:
            return "Con cupo asignado"
        if estado_cupo == EstadoCupo.FUERA:
            return "Lista de espera"
        return "Sin evaluar"

    @staticmethod
    def _fila_beneficiario(legajo, responsable) -> list:
        ciudadano = legajo.ciudadano
        fila = [
            ciudadano.apellido or "",
            ciudadano.nombre or "",
            PadronFinalService._documento_o_cuit(ciudadano),
            ciudadano.fecha_nacimiento,
            CruceService.normalizar_sexo_para_exportacion(ciudadano),
            PadronFinalService._texto(ciudadano.nacionalidad),
            PadronFinalService._texto(ciudadano.municipio),
            PadronFinalService._texto(ciudadano.localidad),
            ciudadano.calle or "",
            ciudadano.altura or "",
            ciudadano.codigo_postal or "",
            ciudadano.telefono or "",
            ciudadano.email or "",
        ]
        if responsable is None:
            fila.extend([""] * 9)
        else:
            domicilio = " ".join(
                PadronFinalService._texto(parte)
                for parte in (responsable.calle, responsable.altura)
                if parte not in (None, "")
            )
            fila.extend(
                [
                    responsable.apellido or "",
                    responsable.nombre or "",
                    PadronFinalService._documento_o_cuit(responsable),
                    responsable.fecha_nacimiento,
                    CruceService.normalizar_sexo_para_exportacion(responsable),
                    domicilio,
                    PadronFinalService._texto(responsable.localidad),
                    responsable.telefono or "",
                    responsable.email or "",
                ]
            )
        fila.append(PadronFinalService._estado_cupo_label(legajo.estado_cupo))
        return fila

    @staticmethod
    def _build_excel(headers, rows) -> bytes:
        fecha_indices = {
            index for index, header in enumerate(headers) if header in FECHA_COLUMNS
        }
        wb = Workbook()
        ws = wb.active
        ws.title = "nomina_aprobados"
        ws.append(headers)
        for row in rows:
            ws.append(row)
            if fecha_indices:
                PadronFinalService._formatear_fechas_fila(ws[ws.max_row], fecha_indices)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def _formatear_fechas_fila(cells, fecha_indices) -> None:
        """Muestra las fechas de nacimiento como fecha sin hora (DD/MM/AAAA)."""
        for index in fecha_indices:
            if index >= len(cells):
                continue
            cell = cells[index]
            value = cell.value
            if isinstance(value, datetime):
                cell.value = value.date()
            if isinstance(cell.value, date):
                cell.number_format = FECHA_NUMBER_FORMAT
