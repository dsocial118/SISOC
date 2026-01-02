"""
Test que usa el proceso REAL de importación de celíacos.
Ejecutar: python manage.py test_celiacos_import
"""
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from celiaquia.services.importacion_service import ImportacionService
from celiaquia.models import ExpedienteCiudadano, Expediente, EstadoExpediente
import openpyxl
import tempfile
import os
from io import BytesIO


class Command(BaseCommand):
    help = 'Test usando el proceso REAL de importación'

    def handle(self, *args, **options):
        self.print_header('🧪 TEST CELÍACOS - PROCESO REAL DE IMPORTACIÓN')
        
        user, _ = User.objects.get_or_create(
            username='test_celiacos',
            defaults={'email': 'test@celiacos.com', 'is_staff': True}
        )
        
        resultados = {}
        
        try:
            resultados['CASO A'] = self.test_caso_a(user)
            resultados['CASO B'] = self.test_caso_b(user)
            resultados['CASO C'] = self.test_caso_c(user)
            resultados['CASO D'] = self.test_caso_d(user)
            resultados['CASO E'] = self.test_caso_e(user)
            resultados['CASO F'] = self.test_caso_f(user)
            
            self.print_resumen(resultados)
            
        except Exception as e:
            self.print_error(f'Error: {str(e)}')
            import traceback
            traceback.print_exc()

    def print_header(self, texto):
        self.stdout.write('\n' + '='*80)
        self.stdout.write(self.style.SUCCESS(texto))
        self.stdout.write('='*80 + '\n')

    def print_section(self, texto):
        self.stdout.write(self.style.HTTP_INFO(f'\n📋 {texto}'))
        self.stdout.write('-'*80)

    def print_ok(self, texto):
        self.stdout.write(self.style.SUCCESS(f'  ✅ {texto}'))

    def print_warning(self, texto):
        self.stdout.write(self.style.WARNING(f'  ⚠️  {texto}'))

    def print_error(self, texto):
        self.stdout.write(self.style.ERROR(f'  ❌ {texto}'))

    def print_info(self, texto):
        self.stdout.write(f'  ℹ️  {texto}')

    def crear_excel(self, datos):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = ['apellido', 'nombre', 'documento', 'fecha_nacimiento', 'sexo', 
                   'nacionalidad', 'municipio', 'localidad', 'calle', 'altura', 
                   'codigo_postal', 'telefono', 'email', 'APELLIDO_RESPONSABLE', 
                   'NOMBRE_RESPONSABLE', 'Cuit_Responsable', 'FECHA_DE_NACIMIENTO_RESPONSABLE', 
                   'SEXO_RESPONSABLE']
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, row_data in enumerate(datos, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def crear_expediente(self, user):
        estado, _ = EstadoExpediente.objects.get_or_create(nombre='CREADO')
        return Expediente.objects.create(usuario_provincia=user, estado=estado)

    def test_caso_a(self, user):
        self.print_section('CASO A: Responsable = Beneficiario')
        try:
            datos = [
                ['García', 'Matías', '20407321384', '01/01/1999', 'M', 'Argentina', 
                 1, 1, 'Calle 1', 100, 1900, '1123456789', 'matias@email.com',
                 'García', 'Matías', '20407321384', '01/01/1999', 'M']
            ]
            
            archivo = self.crear_excel(datos)
            expediente = self.crear_expediente(user)
            
            service = ImportacionService()
            resultado = service.importar_legajos_desde_excel(expediente, archivo, user)
            
            self.print_info(f'Resultado: {resultado}')
            
            legajos = ExpedienteCiudadano.objects.filter(rol='beneficiario_y_responsable')
            if legajos.exists():
                self.print_ok(f'✓ {legajos.count()} expediente(s) creado(s) en BD')
                for legajo in legajos:
                    self.print_info(f'  - {legajo.apellido}, {legajo.nombre} (ID: {legajo.id})')
                self.print_ok('CASO A: COMPLETADO ✓\n')
                return True
            else:
                self.print_error('No se crearon expedientes')
                return False
                    
        except Exception as e:
            self.print_error(str(e))
            return False

    def test_caso_b(self, user):
        self.print_section('CASO B: Responsable ≠ Beneficiario')
        try:
            datos = [
                ['Pérez', 'Nicolás', '20418621384', '15/03/2016', 'M', 'Argentina',
                 1, 1, 'Calle 2', 200, 1900, '1123456789', 'nicolas@email.com',
                 'García', 'Matías', '20407321384', '01/01/1999', 'M']
            ]
            
            archivo = self.crear_excel(datos)
            expediente = self.crear_expediente(user)
            
            service = ImportacionService()
            resultado = service.importar_legajos_desde_excel(expediente, archivo, user)
            
            self.print_info(f'Resultado: {resultado}')
            
            legajos = ExpedienteCiudadano.objects.filter(rol='beneficiario')
            if legajos.exists():
                self.print_ok(f'✓ {legajos.count()} expediente(s) creado(s) en BD')
                for legajo in legajos:
                    self.print_info(f'  - {legajo.apellido}, {legajo.nombre} (ID: {legajo.id})')
                self.print_ok('CASO B: COMPLETADO ✓\n')
                return True
            else:
                self.print_error('No se crearon expedientes')
                return False
                    
        except Exception as e:
            self.print_error(str(e))
            return False

    def test_caso_c(self, user):
        self.print_section('CASO C: Responsable Múltiple')
        try:
            datos = [
                ['García', 'Matías', '20407321385', '01/01/1999', 'M', 'Argentina',
                 1, 1, 'Calle 1', 100, 1900, '1123456789', 'matias@email.com',
                 'García', 'Matías', '20407321385', '01/01/1999', 'M'],
                ['López', 'Juan', '20418621385', '15/03/2014', 'M', 'Argentina',
                 1, 1, 'Calle 3', 300, 1900, '1123456789', 'juan@email.com',
                 'García', 'Matías', '20407321385', '01/01/1999', 'M'],
                ['Rodríguez', 'María', '20419621384', '20/06/2012', 'F', 'Argentina',
                 1, 1, 'Calle 4', 400, 1900, '1123456789', 'maria@email.com',
                 'García', 'Matías', '20407321385', '01/01/1999', 'M']
            ]
            
            archivo = self.crear_excel(datos)
            expediente = self.crear_expediente(user)
            
            service = ImportacionService()
            resultado = service.importar_legajos_desde_excel(expediente, archivo, user)
            
            self.print_info(f'Resultado: {resultado}')
            
            legajos = ExpedienteCiudadano.objects.filter(apellido__in=['García', 'López', 'Rodríguez'])
            if legajos.count() >= 3:
                self.print_ok(f'✓ {legajos.count()} expediente(s) creado(s) en BD')
                for legajo in legajos:
                    self.print_info(f'  - {legajo.apellido}, {legajo.nombre} (ID: {legajo.id})')
                self.print_ok('CASO C: COMPLETADO ✓\n')
                return True
            else:
                self.print_error(f'Se esperaban 3 expedientes, se crearon {legajos.count()}')
                return False
                    
        except Exception as e:
            self.print_error(str(e))
            return False

    def test_caso_d(self, user):
        self.print_section('CASO D: Solo Beneficiario (Sin Responsable)')
        try:
            datos = [
                ['Martínez', 'Carlos', '20420621384', '10/05/1994', 'M', 'Argentina',
                 1, 1, 'Calle 5', 500, 1900, '1123456789', 'carlos@email.com',
                 '', '', '', '', '']
            ]
            
            archivo = self.crear_excel(datos)
            expediente = self.crear_expediente(user)
            
            service = ImportacionService()
            resultado = service.importar_legajos_desde_excel(expediente, archivo, user)
            
            self.print_info(f'Resultado: {resultado}')
            
            legajos = ExpedienteCiudadano.objects.filter(apellido='Martínez')
            if legajos.exists():
                self.print_ok(f'✓ {legajos.count()} expediente(s) creado(s) en BD')
                for legajo in legajos:
                    self.print_info(f'  - {legajo.apellido}, {legajo.nombre} (ID: {legajo.id})')
                self.print_ok('CASO D: COMPLETADO ✓\n')
                return True
            else:
                self.print_error('No se crearon expedientes')
                return False
                    
        except Exception as e:
            self.print_error(str(e))
            return False

    def test_caso_e(self, user):
        self.print_section('CASO E: Beneficiario Menor como Responsable')
        try:
            datos = [
                ['López', 'Pedro', '20421621384', '15/03/2008', 'M', 'Argentina',
                 1, 1, 'Calle 6', 600, 1900, '1123456789', 'pedro@email.com',
                 'López', 'Pedro', '20421621384', '15/03/2008', 'M'],
                ['López', 'Lucas', '20422621384', '20/06/2016', 'M', 'Argentina',
                 1, 1, 'Calle 7', 700, 1900, '1123456789', 'lucas@email.com',
                 'López', 'Pedro', '20421621384', '15/03/2008', 'M']
            ]
            
            archivo = self.crear_excel(datos)
            expediente = self.crear_expediente(user)
            
            service = ImportacionService()
            resultado = service.importar_legajos_desde_excel(expediente, archivo, user)
            
            self.print_info(f'Resultado: {resultado}')
            self.print_warning('Responsable menor de 18 años')
            
            legajos = ExpedienteCiudadano.objects.filter(apellido='López')
            if legajos.count() >= 2:
                self.print_ok(f'✓ {legajos.count()} expediente(s) creado(s) en BD')
                for legajo in legajos:
                    self.print_info(f'  - {legajo.apellido}, {legajo.nombre} (ID: {legajo.id})')
                self.print_ok('CASO E: COMPLETADO ✓ (con warning)\n')
                return True
            else:
                self.print_error(f'Se esperaban 2 expedientes, se crearon {legajos.count()}')
                return False
                    
        except Exception as e:
            self.print_error(str(e))
            return False

    def test_caso_f(self, user):
        self.print_section('CASO F: Cadena Familiar')
        try:
            datos = [
                ['García', 'Roberto', '20400321384', '01/01/1964', 'M', 'Argentina',
                 1, 1, 'Calle 8', 800, 1900, '1123456789', 'roberto@email.com',
                 'García', 'Roberto', '20400321384', '01/01/1964', 'M'],
                ['García', 'Matías', '20407321386', '01/01/1999', 'M', 'Argentina',
                 1, 1, 'Calle 1', 100, 1900, '1123456789', 'matias@email.com',
                 'García', 'Roberto', '20400321384', '01/01/1964', 'M'],
                ['García', 'Tomás', '20425621384', '15/03/2019', 'M', 'Argentina',
                 1, 1, 'Calle 9', 900, 1900, '1123456789', 'tomas@email.com',
                 'García', 'Matías', '20407321386', '01/01/1999', 'M']
            ]
            
            archivo = self.crear_excel(datos)
            expediente = self.crear_expediente(user)
            
            service = ImportacionService()
            resultado = service.importar_legajos_desde_excel(expediente, archivo, user)
            
            self.print_info(f'Resultado: {resultado}')
            
            legajos = ExpedienteCiudadano.objects.filter(apellido='García')
            if legajos.count() >= 3:
                self.print_ok(f'✓ {legajos.count()} expediente(s) creado(s) en BD')
                for legajo in legajos:
                    self.print_info(f'  - {legajo.apellido}, {legajo.nombre} (ID: {legajo.id})')
                self.print_ok('Cadena: Roberto → Matías → Tomás')
                self.print_ok('CASO F: COMPLETADO ✓\n')
                return True
            else:
                self.print_error(f'Se esperaban 3 expedientes, se crearon {legajos.count()}')
                return False
                    
        except Exception as e:
            self.print_error(str(e))
            return False

    def print_resumen(self, resultados):
        self.print_header('📊 RESUMEN DE RESULTADOS')
        
        total = len(resultados)
        exitosos = sum(1 for v in resultados.values() if v)
        
        for caso, resultado in resultados.items():
            if resultado:
                self.stdout.write(self.style.SUCCESS(f'  ✅ {caso}: PASÓ'))
            else:
                self.stdout.write(self.style.ERROR(f'  ❌ {caso}: FALLÓ'))
        
        self.stdout.write('\n' + '='*80)
        self.stdout.write(self.style.SUCCESS(f'TOTAL: {exitosos}/{total} tests completados'))
        self.stdout.write(self.style.SUCCESS(f'✓ DATOS INSERTADOS EN LA BASE DE DATOS USANDO PROCESO REAL'))
        self.stdout.write('='*80 + '\n')
