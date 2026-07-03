"""Tests de exportacion de nomina aprobada Celiaquia.

La nomina de aprobados se genera desde la base (legajos APROBADO + MATCH),
no desde el Excel original cargado por la provincia.
"""

from datetime import date
from io import BytesIO

import pytest
from django.contrib.auth.models import Permission, User
from django.contrib.contenttypes.models import ContentType
from django.urls import reverse
from openpyxl import load_workbook

from celiaquia.models import (
    EstadoExpediente,
    EstadoLegajo,
    Expediente,
    ExpedienteCiudadano,
    ResultadoSintys,
    RevisionTecnico,
)
from celiaquia.services.padron_final_service import (
    NOMINA_HEADERS,
    PadronFinalService,
)
from ciudadanos.models import Ciudadano, GrupoFamiliar
from core.models import Localidad, Municipio, Sexo


def _permission(app_label, codename):
    try:
        return Permission.objects.get(
            content_type__app_label=app_label,
            codename=codename,
        )
    except Permission.DoesNotExist:
        content_type = ContentType.objects.get(app_label="auth", model="user")
        return Permission.objects.create(
            content_type=content_type,
            codename=codename,
            name=codename,
        )


def _grant(user, app_label, codename):
    user.user_permissions.add(_permission(app_label, codename))


def _user(username, *, coord=False):
    user = User.objects.create_user(username=username, password="pass")
    _grant(user, "celiaquia", "view_expediente")
    if coord:
        _grant(user, "auth", "role_coordinadorceliaquia")
    return user


def _workbook_rows(content):
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    return rows[0], rows[1:]


def _expediente(owner, *, estado="CRUCE_FINALIZADO"):
    return Expediente.objects.create(
        usuario_provincia=owner,
        estado=EstadoExpediente.objects.create(nombre=estado),
    )


def _crear_ciudadano(documento, apellido=None, nombre=None, **extra):
    return Ciudadano.objects.create(
        apellido=apellido or f"Apellido {documento}",
        nombre=nombre or f"Nombre {documento}",
        fecha_nacimiento=extra.pop("fecha_nacimiento", "2000-01-01"),
        documento=int(documento),
        **extra,
    )


def _crear_legajo(expediente, documento, *, revision, sintys, rol=None, **extra):
    estado_legajo, _ = EstadoLegajo.objects.get_or_create(nombre="VALIDO")
    ciudadano = _crear_ciudadano(documento, **extra)
    return ExpedienteCiudadano.objects.create(
        expediente=expediente,
        ciudadano=ciudadano,
        estado=estado_legajo,
        revision_tecnico=revision,
        resultado_sintys=sintys,
        rol=rol or ExpedienteCiudadano.ROLE_BENEFICIARIO,
    )


def _col(header, row):
    return row[NOMINA_HEADERS.index(header)]


@pytest.mark.django_db
def test_nomina_aprobados_se_genera_desde_base_y_filtra_aprobados():
    owner = _user("prov-servicio")
    expediente = _expediente(owner)
    municipio = Municipio.objects.create(nombre="San Miguel de Tucumán")
    localidad = Localidad.objects.create(
        nombre="Villa Mariano Moreno", municipio=municipio
    )

    _crear_legajo(
        expediente,
        "20392317989",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
        apellido="ALZUETA",
        nombre="LUCAS",
        calle="Calle",
        altura="1",
        codigo_postal="1900",
        municipio=municipio,
        localidad=localidad,
    )
    _crear_legajo(
        expediente,
        "20392317990",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.NO_MATCH,
    )
    _crear_legajo(
        expediente,
        "20392317991",
        revision=RevisionTecnico.RECHAZADO,
        sintys=ResultadoSintys.MATCH,
    )
    _crear_legajo(
        expediente,
        "20392317992",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
        rol=ExpedienteCiudadano.ROLE_RESPONSABLE,
    )

    header, data_rows = _workbook_rows(
        PadronFinalService.generar_padron_final_excel(expediente)
    )

    assert list(header) == NOMINA_HEADERS
    assert len(data_rows) == 1
    assert _col("documento", data_rows[0]) == "20392317989"
    assert _col("apellido", data_rows[0]) == "ALZUETA"
    assert _col("calle", data_rows[0]) == "Calle"
    assert _col("codigo_postal", data_rows[0]) == "1900"
    # Municipio y localidad se exportan por nombre, no por codigo.
    assert _col("municipio", data_rows[0]) == "San Miguel de Tucumán"
    assert _col("localidad", data_rows[0]) == "Villa Mariano Moreno"


@pytest.mark.django_db
def test_nomina_aprobados_usa_documento_actual_de_la_base():
    """Si el documento se corrige despues de importar, la nomina refleja la base."""
    owner = _user("prov-correccion")
    expediente = _expediente(owner)
    legajo = _crear_legajo(
        expediente,
        "2055724691",  # CUIT mal tipeado en la importacion original
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
    )

    legajo.ciudadano.documento = 20557246918
    legajo.ciudadano.save(update_fields=["documento"])

    _, data_rows = _workbook_rows(
        PadronFinalService.generar_padron_final_excel(expediente)
    )

    assert len(data_rows) == 1
    assert _col("documento", data_rows[0]) == "20557246918"


@pytest.mark.django_db
def test_nomina_aprobados_incluye_datos_del_responsable():
    owner = _user("prov-familia")
    expediente = _expediente(owner)
    sexo_f = Sexo.objects.create(sexo="F")

    legajo_hijo = _crear_legajo(
        expediente,
        "20557246918",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
        apellido="CANIETE",
        nombre="MISAEL",
    )
    legajo_resp = _crear_legajo(
        expediente,
        "27342010844",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.SIN_CRUCE,
        rol=ExpedienteCiudadano.ROLE_RESPONSABLE,
        apellido="SALAS",
        nombre="ANA MARIA",
        fecha_nacimiento="1989-04-29",
        sexo=sexo_f,
        calle="GARMENDIA",
        altura="236",
        telefono="3810000000",
        email="resp@example.com",
    )
    GrupoFamiliar.objects.create(
        ciudadano_1=legajo_resp.ciudadano,
        ciudadano_2=legajo_hijo.ciudadano,
        vinculo=GrupoFamiliar.RELACION_PADRE,
        conviven=True,
        cuidador_principal=True,
    )

    _, data_rows = _workbook_rows(
        PadronFinalService.generar_padron_final_excel(expediente)
    )

    assert len(data_rows) == 1
    fila = data_rows[0]
    assert _col("documento", fila) == "20557246918"
    assert _col("APELLIDO_RESPONSABLE", fila) == "SALAS"
    assert _col("NOMBRE_RESPONSABLE", fila) == "ANA MARIA"
    assert _col("CUIT_RESPONSABLE", fila) == "27342010844"
    assert _col("SEXO_RESPONSABLE", fila) == "Femenino"
    assert _col("DOMICILIO_RESPONSABLE", fila) == "GARMENDIA 236"
    assert _col("CELULAR_RESPONSABLE", fila) == "3810000000"
    assert _col("CORREO_RESPONSABLE", fila) == "resp@example.com"


@pytest.mark.django_db
def test_nomina_aprobados_se_recalcula_con_resultado_sintys_actual():
    owner = _user("prov-reproceso")
    expediente = _expediente(owner)
    _crear_legajo(
        expediente,
        "20392317001",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
        apellido="AAA",
    )
    reprocesado = _crear_legajo(
        expediente,
        "20392317002",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.NO_MATCH,
        apellido="BBB",
    )

    _, data_rows = _workbook_rows(
        PadronFinalService.generar_padron_final_excel(expediente)
    )
    assert [_col("documento", row) for row in data_rows] == ["20392317001"]

    reprocesado.resultado_sintys = ResultadoSintys.MATCH
    reprocesado.save(update_fields=["resultado_sintys"])

    _, data_rows = _workbook_rows(
        PadronFinalService.generar_padron_final_excel(expediente)
    )
    assert [_col("documento", row) for row in data_rows] == [
        "20392317001",
        "20392317002",
    ]


@pytest.mark.django_db
def test_nomina_aprobados_exporta_fecha_nacimiento_sin_hora():
    owner = _user("prov-fechas")
    expediente = _expediente(owner)
    sexo_f = Sexo.objects.create(sexo="F")

    legajo_hijo = _crear_legajo(
        expediente,
        "20392317500",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
        fecha_nacimiento="2010-03-15",
    )
    legajo_resp = _crear_legajo(
        expediente,
        "27342010844",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.SIN_CRUCE,
        rol=ExpedienteCiudadano.ROLE_RESPONSABLE,
        fecha_nacimiento="1980-07-09",
        sexo=sexo_f,
    )
    GrupoFamiliar.objects.create(
        ciudadano_1=legajo_resp.ciudadano,
        ciudadano_2=legajo_hijo.ciudadano,
        vinculo=GrupoFamiliar.RELACION_PADRE,
        conviven=True,
        cuidador_principal=True,
    )

    content = PadronFinalService.generar_padron_final_excel(expediente)
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    fecha_nac_col = NOMINA_HEADERS.index("fecha_nacimiento") + 1
    fecha_resp_col = NOMINA_HEADERS.index("FECHA_DE_NACIMIENTO_RESPONSABLE") + 1
    fecha_nac_cell = ws.cell(row=2, column=fecha_nac_col)
    fecha_resp_cell = ws.cell(row=2, column=fecha_resp_col)

    assert fecha_nac_cell.value.date() == date(2010, 3, 15)
    assert fecha_resp_cell.value.date() == date(1980, 7, 9)
    assert fecha_nac_cell.number_format == "DD/MM/YYYY"
    assert fecha_resp_cell.number_format == "DD/MM/YYYY"


@pytest.mark.django_db
def test_descarga_nomina_aprobados_no_disponible_antes_de_cruce_finalizado(client):
    owner = _user("prov-antes")
    coord = _user("coord-antes", coord=True)
    expediente = _expediente(owner, estado="ASIGNADO")

    client.force_login(coord)
    response = client.get(
        reverse("expediente_padron_final_export", args=[expediente.pk])
    )

    assert response.status_code == 404


@pytest.mark.django_db
def test_descarga_nomina_aprobados_finalizada_devuelve_xlsx(client):
    owner = _user("prov-final")
    coord = _user("coord-final", coord=True)
    expediente = _expediente(owner)
    _crear_legajo(
        expediente,
        "20392317201",
        revision=RevisionTecnico.APROBADO,
        sintys=ResultadoSintys.MATCH,
    )

    client.force_login(coord)
    response = client.get(
        reverse("expediente_padron_final_export", args=[expediente.pk])
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        f"nomina_aprobados_{expediente.pk}.xlsx"
        in response.headers["Content-Disposition"]
    )
    header, data_rows = _workbook_rows(response.content)
    assert list(header) == NOMINA_HEADERS
    assert _col("documento", data_rows[0]) == "20392317201"


@pytest.mark.django_db
def test_detalle_muestra_descarga_solo_con_cruce_finalizado(client):
    owner = _user("prov-detalle")
    coord = _user("coord-detalle", coord=True)
    expediente_asignado = _expediente(owner, estado="ASIGNADO")
    expediente_finalizado = _expediente(owner)

    client.force_login(coord)
    response_asignado = client.get(
        reverse("expediente_detail", args=[expediente_asignado.pk])
    )
    response_finalizado = client.get(
        reverse("expediente_detail", args=[expediente_finalizado.pk])
    )

    assert response_asignado.status_code == 200
    assert response_finalizado.status_code == 200
    assert (
        reverse("expediente_padron_final_export", args=[expediente_asignado.pk])
        not in response_asignado.content.decode()
    )
    assert "Descargar nómina aprobados" in response_finalizado.content.decode()
