"""End-to-end con datos dummy del flujo celiaquía tras los fixes.

Ejercita servicios reales (cruce SINTYS, cupo, padrón), el rechazo RENAPER y el
comando de saneo, verificando el comportamiento por rol y estado:

- beneficiario en/ fuera de SINTYS
- doble rol (beneficiario_y_responsable) cuidador -> ocupa su cupo y va al padrón
- responsable puro -> nunca MATCH, nunca cupo, nunca en el padrón
- hijo validado por el documento de su responsable
- lista de espera (cupo agotado) marcada en el padrón
- rechazo RENAPER -> RECHAZADO + libera cupo + sale del padrón
- saneo -> recupera un doble rol dejado SIN_CRUCE por el bug viejo

Corre con: pytest celiaquia/tests/test_flujo_e2e_celiaquia.py -s
"""

from datetime import date
from io import BytesIO

import pytest
from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from django.test import RequestFactory
from openpyxl import Workbook, load_workbook

from ciudadanos.models import Ciudadano, GrupoFamiliar
from core.models import Provincia
from celiaquia.models import (
    EstadoCupo,
    EstadoExpediente,
    EstadoLegajo,
    Expediente,
    ExpedienteCiudadano,
    ProvinciaCupo,
    ResultadoSintys,
    RevisionTecnico,
)
from celiaquia.services.cruce_service import CruceService
from celiaquia.services.padron_final_service import PadronFinalService
from celiaquia.views.validacion_renaper import ValidacionRenaperView

B = ExpedienteCiudadano.ROLE_BENEFICIARIO
R = ExpedienteCiudadano.ROLE_RESPONSABLE
BR = ExpedienteCiudadano.ROLE_BENEFICIARIO_Y_RESPONSABLE

# documento (numérico) por persona
DOCS = {
    "ANA": 20111111112,  # beneficiaria, en SINTYS
    "BETO": 20222222220,  # beneficiario, NO en SINTYS
    "CARLA": 20333333335,  # doble rol, cuidadora de DANI, en SINTYS
    "DANI": 45444444,  # hijo de CARLA (menor)
    "EMA": 20555555559,  # responsable puro, cuidadora de FEDE, en SINTYS
    "FEDE": 46666666,  # hijo de EMA
    "HUGO": 20777777778,  # doble rol (para el saneo), en SINTYS
    "IVAN": 48888888,  # hijo de HUGO (sin legajo)
}


def _excel(headers, filas):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for fila in filas:
        ws.append(fila)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _ciudadano(nombre, provincia, fnac):
    return Ciudadano.objects.create(
        apellido=nombre,
        nombre=nombre,
        documento=DOCS[nombre],
        fecha_nacimiento=fnac,
        provincia=provincia,
    )


def _legajo(expediente, estado_legajo, ciudadano, rol):
    return ExpedienteCiudadano.objects.create(
        expediente=expediente,
        ciudadano=ciudadano,
        estado=estado_legajo,
        rol=rol,
        revision_tecnico=RevisionTecnico.APROBADO,
    )


def _padron_docs_y_estados(expediente):
    """Devuelve {documento(str): estado_de_cupo_label} leyendo el Excel del padrón."""
    content = PadronFinalService.generar_padron_final_excel(expediente)
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    headers = list(filas[0])
    doc_idx = headers.index("documento")
    return {str(fila[doc_idx]): fila[-1] for fila in filas[1:]}


@pytest.mark.django_db
def test_flujo_completo_celiaquia(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    prov = Provincia.objects.create(nombre="Testlandia")
    ProvinciaCupo.objects.create(provincia=prov, total_asignado=3, usados=0)
    usuario = User.objects.create_superuser("op-e2e", password="x")
    est_exp = EstadoExpediente.objects.create(nombre="ASIGNADO")
    est_leg = EstadoLegajo.objects.create(nombre="VALIDO")

    exp = Expediente.objects.create(usuario_provincia=usuario, estado=est_exp)

    # --- ciudadanos + legajos ---
    ana = _ciudadano("ANA", prov, date(1990, 1, 1))
    beto = _ciudadano("BETO", prov, date(1990, 1, 1))
    carla = _ciudadano("CARLA", prov, date(1985, 1, 1))
    dani = _ciudadano("DANI", prov, date(2015, 1, 1))
    ema = _ciudadano("EMA", prov, date(1980, 1, 1))
    fede = _ciudadano("FEDE", prov, date(2016, 1, 1))

    leg_ana = _legajo(exp, est_leg, ana, B)
    leg_beto = _legajo(exp, est_leg, beto, B)
    leg_carla = _legajo(exp, est_leg, carla, BR)
    leg_dani = _legajo(exp, est_leg, dani, B)
    leg_ema = _legajo(exp, est_leg, ema, R)
    leg_fede = _legajo(exp, est_leg, fede, B)

    # relaciones familiares (cuidador_principal)
    GrupoFamiliar.objects.create(
        ciudadano_1=carla,
        ciudadano_2=dani,
        vinculo=GrupoFamiliar.RELACION_PADRE,
        cuidador_principal=True,
    )
    GrupoFamiliar.objects.create(
        ciudadano_1=ema,
        ciudadano_2=fede,
        vinculo=GrupoFamiliar.RELACION_PADRE,
        cuidador_principal=True,
    )

    # --- Excel masivo original (nómina) ---
    masivo = _excel(
        ["apellido", "nombre", "documento"],
        [
            [n, n, str(DOCS[n])]
            for n in ("ANA", "BETO", "CARLA", "DANI", "EMA", "FEDE", "HUGO")
        ],
    )
    exp.excel_masivo.save("masivo.xlsx", ContentFile(masivo), save=True)

    # --- Archivo SINTYS: ANA, CARLA, EMA, HUGO (BETO no; DANI/FEDE matchean por su responsable) ---
    sintys = _excel(
        ["documento"],
        [[str(DOCS[n])] for n in ("ANA", "CARLA", "EMA", "HUGO")],
    )
    sintys_file = ContentFile(sintys, name="sintys.xlsx")

    # =================== CRUCE ===================
    CruceService.procesar_cruce_por_cuit(exp, sintys_file, usuario)

    for leg in (leg_ana, leg_beto, leg_carla, leg_dani, leg_ema, leg_fede):
        leg.refresh_from_db()

    print("\n=== RESULTADO DEL CRUCE ===")
    for nombre, leg in [
        ("ANA (benef, en SINTYS)", leg_ana),
        ("BETO (benef, NO SINTYS)", leg_beto),
        ("CARLA (doble rol, cuidadora)", leg_carla),
        ("DANI (hijo de CARLA)", leg_dani),
        ("EMA (responsable puro)", leg_ema),
        ("FEDE (hijo de EMA)", leg_fede),
    ]:
        print(
            f"  {nombre:<32} sintys={leg.resultado_sintys:<9} cupo={leg.estado_cupo} activo={leg.es_titular_activo}"
        )

    # Comportamiento esperado del cruce:
    assert leg_ana.resultado_sintys == ResultadoSintys.MATCH
    assert leg_beto.resultado_sintys == ResultadoSintys.NO_MATCH
    assert leg_carla.resultado_sintys == ResultadoSintys.MATCH  # doble rol fix
    assert leg_dani.resultado_sintys == ResultadoSintys.MATCH  # vía responsable
    assert leg_ema.resultado_sintys == ResultadoSintys.SIN_CRUCE  # responsable salteado
    assert leg_fede.resultado_sintys == ResultadoSintys.MATCH  # vía responsable

    # Cupo: 4 matcheados beneficiarios (ANA, CARLA, DANI, FEDE) vs cupo=3 -> 3 DENTRO, 1 FUERA
    matcheados = [leg_ana, leg_carla, leg_dani, leg_fede]
    dentro = [l for l in matcheados if l.estado_cupo == EstadoCupo.DENTRO]
    fuera = [l for l in matcheados if l.estado_cupo == EstadoCupo.FUERA]
    assert len(dentro) == 3 and len(fuera) == 1
    assert ProvinciaCupo.objects.get(provincia=prov).usados == 3
    # EMA (responsable) nunca ocupa cupo
    assert leg_ema.estado_cupo != EstadoCupo.DENTRO

    # =================== PADRÓN ===================
    padron = _padron_docs_y_estados(exp)
    print("\n=== PADRÓN (documento -> estado de cupo) ===")
    for doc, estado in padron.items():
        print(f"  {doc:<14} {estado}")

    # Padrón = APROBADO + MATCH + no responsable = ANA, CARLA, DANI, FEDE
    assert set(padron) == {str(DOCS[n]) for n in ("ANA", "CARLA", "DANI", "FEDE")}
    assert str(DOCS["EMA"]) not in padron  # responsable puro
    assert str(DOCS["BETO"]) not in padron  # no match
    # Columna de estado de cupo: 3 con cupo, 1 en lista de espera
    assert list(padron.values()).count("Con cupo asignado") == 3
    assert list(padron.values()).count("Lista de espera") == 1

    # =================== RENAPER (rechazo) ===================
    objetivo = dentro[0]  # un titular con cupo
    usados_antes = ProvinciaCupo.objects.get(provincia=prov).usados
    request = RequestFactory().post("/x", {})
    request.user = usuario
    ValidacionRenaperView()._guardar_validacion_estado(
        request, exp.pk, objetivo.pk, "2"
    )
    objetivo.refresh_from_db()
    print(f"\n=== RENAPER rechazo sobre doc={objetivo.ciudadano.documento} ===")
    print(f"  revision_tecnico={objetivo.revision_tecnico} cupo={objetivo.estado_cupo}")

    assert objetivo.revision_tecnico == RevisionTecnico.RECHAZADO
    assert objetivo.estado_cupo == EstadoCupo.NO_EVAL
    assert ProvinciaCupo.objects.get(provincia=prov).usados == usados_antes - 1
    padron_post = _padron_docs_y_estados(exp)
    assert str(objetivo.ciudadano.documento) not in padron_post  # salió del padrón

    # =================== SANEO (doble rol dejado SIN_CRUCE por el bug viejo) ===================
    hugo = _ciudadano("HUGO", prov, date(1988, 1, 1))
    ivan = Ciudadano.objects.create(
        apellido="IVAN", nombre="IVAN", documento=DOCS["IVAN"], provincia=prov
    )
    GrupoFamiliar.objects.create(
        ciudadano_1=hugo,
        ciudadano_2=ivan,
        vinculo=GrupoFamiliar.RELACION_PADRE,
        cuidador_principal=True,
    )
    # Simula el estado que dejaba el cruce viejo: APROBADO pero SIN_CRUCE.
    leg_hugo = ExpedienteCiudadano.objects.create(
        expediente=exp,
        ciudadano=hugo,
        estado=est_leg,
        rol=BR,
        revision_tecnico=RevisionTecnico.APROBADO,
        resultado_sintys=ResultadoSintys.SIN_CRUCE,
    )

    from celiaquia.management.commands.sanear_celiaquia import Command

    cmd = Command()
    base = ExpedienteCiudadano.objects.filter(expediente=exp)
    print("\n=== SANEO doble rol (re-evaluación) ===")
    resultado = cmd._reevaluar_doble_rol(base, apply=True, usuario=usuario)
    print(f"  {resultado}")

    leg_hugo.refresh_from_db()
    assert leg_hugo.resultado_sintys == ResultadoSintys.MATCH  # recuperado
    assert str(DOCS["HUGO"]) in _padron_docs_y_estados(exp)  # ahora en el padrón
    print("\n=== OK: el flujo se comporta como se espera ===")
