"""Tests for test nomina sintys export."""

from io import BytesIO

import pytest
from django.contrib.auth.models import Group, User
from openpyxl import load_workbook

from celiaquia.models import (
    Expediente,
    EstadoExpediente,
    EstadoLegajo,
    ExpedienteCiudadano,
)
from celiaquia.services.cruce_service import CruceService
from core.models import Sexo
from ciudadanos.models import Ciudadano


@pytest.mark.django_db
def test_exportar_nomina_sintys(client):
    grupo = Group.objects.create(name="TecnicoCeliaquia")
    user = User.objects.create_user(username="tec", password="pass")
    user.groups.add(grupo)

    estado_exp = EstadoExpediente.objects.create(nombre="ASIGNADO")
    estado_leg = EstadoLegajo.objects.create(nombre="VALIDO")
    sexo = Sexo.objects.create(sexo="Masculino")
    creador = User.objects.create_user(username="prov", password="pass")
    expediente = Expediente.objects.create(usuario_provincia=creador, estado=estado_exp)

    ciudadano = Ciudadano.objects.create(
        apellido="Perez",
        nombre="Juan",
        fecha_nacimiento="2000-01-01",
        documento=12345678,
        tipo_documento=Ciudadano.DOCUMENTO_DNI,
        sexo=sexo,
    )
    ExpedienteCiudadano.objects.create(
        expediente=expediente, ciudadano=ciudadano, estado=estado_leg
    )

    client.force_login(user)
    content = CruceService.generar_nomina_sintys_excel(expediente)

    wb = load_workbook(BytesIO(content))
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    assert header == [
        "Numero_documento",
        "TipoDocumento",
        "nombre",
        "apellido",
        "sexo",
    ]
    row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    assert str(row[0]) == "12345678"
    assert row[1] == "DNI"
    assert row[2] == "Juan"
    assert row[3] == "Perez"
    assert row[4] == "Masculino"
