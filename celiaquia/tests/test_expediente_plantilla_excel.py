"""Tests for test expediente plantilla excel."""

from io import BytesIO

import pytest
from django.urls import reverse
from django.contrib.auth.models import User, Group
from openpyxl import load_workbook

from users.models import Profile
from core.models import Provincia


@pytest.mark.django_db
def test_descargar_plantilla_excel(client):
    grupo = Group.objects.create(name="ProvinciaCeliaquia")
    provincia = Provincia.objects.create(nombre="Buenos Aires")
    user = User.objects.create_user(username="prov", password="pass")
    profile, _ = Profile.objects.get_or_create(user=user)
    profile.es_usuario_provincial = True
    profile.provincia = provincia
    profile.save()
    user.groups.add(grupo)

    client.force_login(user)
    response = client.get(reverse("expediente_plantilla_excel"))
    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    assert header == [
        "apellido",
        "nombre",
        "documento",
        "fecha_nacimiento",
        "sexo",
        "nacionalidad",
        "municipio",
        "localidad",
        "calle",
        "altura",
        "codigo_postal",
        "telefono",
        "email",
        "APELLIDO_RESPONSABLE",
        "NOMBRE_REPSONSABLE",
        "Cuit_Responsable",
        "FECHA_DE_NACIMIENTO_RESPONSABLE",
        "SEXO",
        "DOMICILIO_RESPONSABLE",
        "LOCALIDAD_RESPONSABLE",
        "CELULAR_RESPONSABLE",
        "CORREO_RESPONSABLE",
    ]
