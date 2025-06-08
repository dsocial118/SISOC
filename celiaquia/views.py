from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views.generic import CreateView, DetailView, ListView, UpdateView, View
from django.utils import timezone
from django.core.paginator import Paginator
from celiaquia.forms import ExpedienteForm, PersonaFormularioForm
from celiaquia.models import (
    Expediente, PersonaFormulario, AsignacionTecnico, ArchivoCruce,
    ResultadoCruce, InformePago, EstadoExpediente
)
from django.contrib.auth.models import User
from openpyxl import load_workbook
from io import BytesIO
from django.views import View
from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
import pandas as pd
from django.core.files.storage import default_storage
from celiaquia.models import Expediente, ArchivoCruce, EstadoExpediente
from django.conf import settings
import os
from celiaquia.models import ResultadoCruce
from django.views.generic.edit import FormView
from openpyxl import load_workbook
from .models import PersonaFormulario
from .models import CargaMasivaForm
from django.views import View
from django.http import HttpResponse
from openpyxl import Workbook
from .models import ResultadoCruce

class ExportarResultadosExcelView(View):
    def get(self, request, expediente_id):
        resultados = ResultadoCruce.objects.filter(expediente_id=expediente_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados de Cruce"
        ws.append(["DNI", "Estado", "Motivo de Rechazo"])

        for resultado in resultados:
            ws.append([
                resultado.dni,
                resultado.get_estado_display(),
                resultado.motivo_rechazo or ''
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=resultados_cruce_exp_{expediente_id}.xlsx'
        wb.save(response)
        return response

class CargaMasivaPersonasView(LoginRequiredMixin, FormView):
    form_class = CargaMasivaForm
    template_name = 'celiaquia/carga_masiva_personas.html'

    def form_valid(self, form):
        archivo_excel = form.cleaned_data['archivo']
        expediente = get_object_or_404(Expediente, pk=self.kwargs['pk'])

        try:
            wb = load_workbook(archivo_excel)
            ws = wb.active

            filas_cargadas = 0
            for fila in ws.iter_rows(min_row=2, values_only=True):  # Asumimos encabezados en fila 1
                nombre, apellido, dni, sexo, fecha_nacimiento = fila

                # Validar duplicados en el mismo expediente
                if PersonaFormulario.objects.filter(expediente=expediente, dni=str(dni)).exists():
                    continue

                PersonaFormulario.objects.create(
                    expediente=expediente,
                    nombre=nombre,
                    apellido=apellido,
                    dni=str(dni),
                    sexo=sexo,
                    fecha_nacimiento=fecha_nacimiento
                )
                filas_cargadas += 1

            messages.success(self.request, f"Se cargaron {filas_cargadas} personas correctamente.")
        except Exception as e:
            messages.error(self.request, f"Error al procesar el archivo: {e}")

        return redirect('celiaquia_expedientes_detalle', pk=expediente.pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['expediente'] = get_object_or_404(Expediente, pk=self.kwargs['pk'])
        return context



class ExpedienteListView(LoginRequiredMixin, ListView):
    model = Expediente
    template_name = 'celiaquia/expediente_list.html'
    context_object_name = 'expedientes'


class ExpedienteCreateView(LoginRequiredMixin, CreateView):
    model = Expediente
    fields = ['provincia', 'observaciones']
    template_name = 'celiaquia/expediente_form.html'
    success_url = reverse_lazy('celiaquia_expedientes_listar')

    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        estado = EstadoExpediente.objects.get(nombre="Creado")
        form.instance.estado = estado
        return super().form_valid(form)



class ExpedienteDetailView(DetailView):
    model = Expediente
    template_name = 'celiaquia/expediente_detail.html'
    context_object_name = 'expediente'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expediente = self.object

        resultado = ResultadoCruce.objects.filter(expediente=self.object, estado='aceptado').order_by('dni')
        context['resultado_cruce'] = resultado

        # Personas
        personas = expediente.personas.all().order_by('-id')  # related_name correcto
        paginator = Paginator(personas, 4)
        page = self.request.GET.get('page')
        context['page_obj'] = paginator.get_page(page)

        # Técnicos disponibles (solo si está Finalizado por Provincia)
        if expediente.estado.nombre == "Finalizado por Provincia":
            context['tecnicos'] = User.objects.filter(groups__name='Tecnico Subir Cruces')

        # Técnico asignado (si hay)
        try:
            asignacion = AsignacionTecnico.objects.get(expediente=expediente)
            context['tecnico_asignado'] = asignacion.tecnico
        except AsignacionTecnico.DoesNotExist:
            context['tecnico_asignado'] = None

        # Roles
        context['es_tecnico'] = self.request.user.groups.filter(name="Tecnico Subir Cruces").exists()
        context['puede_subir_cruces'] = (
        (self.request.user.is_superuser or context['es_tecnico'])
        and context['tecnico_asignado'] is not None
            )

        # Verificar si existen los 3 archivos aprobados (uno por organismo)
        cruces_aprobados = expediente.archivos_cruce.filter(tipo='aprobado')
        organismos_requeridos = {'syntis', 'salud', 'renaper'}
        organismos_presentes = set(cruces_aprobados.values_list('organismo', flat=True))
        context['puede_confirmar_nomina'] = organismos_requeridos.issubset(organismos_presentes)

        return context



class ExpedienteUpdateView(LoginRequiredMixin, UpdateView):
    model = Expediente
    form_class = ExpedienteForm
    template_name = 'celiaquia/expediente_form.html'
    success_url = reverse_lazy('celiaquia_expedientes_listar')


class ExpedienteConfirmarView(LoginRequiredMixin, View):
    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        estado = EstadoExpediente.objects.get(nombre="Finalizado por Provincia")
        expediente.estado = estado
        expediente.save()
        messages.success(request, "Expediente confirmado con éxito.")
        return redirect('celiaquia_expedientes_detalle', pk=pk)


class ExpedienteAsignarTecnicoView(LoginRequiredMixin, View):
    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)
        tecnico_id = request.POST.get('tecnico')
        tecnico = get_object_or_404(User, pk=tecnico_id)
        AsignacionTecnico.objects.create(expediente=expediente, tecnico=tecnico)
        estado = EstadoExpediente.objects.get(nombre="Asignado a Técnico")
        expediente.estado = estado
        expediente.save()
        return redirect('celiaquia_expedientes_detalle', pk=pk)


from django.urls import reverse

class FormularioCargarView(LoginRequiredMixin, CreateView):
    model = PersonaFormulario
    form_class = PersonaFormularioForm
    template_name = 'celiaquia/persona_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        expediente = get_object_or_404(Expediente, pk=self.kwargs['pk'])
        context['expediente'] = expediente
        return context

    def form_valid(self, form):
        form.instance.expediente_id = self.kwargs['pk']
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('celiaquia_expedientes_detalle', args=[self.kwargs['pk']])




class CruceArchivosUploadView(LoginRequiredMixin, CreateView):
    model = ArchivoCruce
    fields = ['archivo', 'organismo', 'tipo']
    template_name = 'celiaquia/cruce_upload.html'

    def form_valid(self, form):
        expediente = get_object_or_404(Expediente, pk=self.kwargs['pk'])
        form.instance.expediente = expediente
        expediente.estado = EstadoExpediente.objects.get(nombre="En Proceso de Cruce")
        expediente.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('celiaquia_expedientes_detalle', kwargs={'pk': self.kwargs['pk']})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['expediente'] = get_object_or_404(Expediente, pk=self.kwargs['pk'])
        return context



class CruceResultadoConfirmarView(LoginRequiredMixin, View):
    def post(self, request, pk):
        expediente = get_object_or_404(Expediente, pk=pk)

        # Buscar los 3 archivos aprobados por organismo
        archivos = {}
        for org in ['syntis', 'salud', 'renaper']:
            archivo = ArchivoCruce.objects.filter(
                expediente=expediente,
                organismo=org,
                tipo='aprobado'
            ).first()
            if not archivo:
                messages.error(request, f"Falta archivo aprobado de {org.title()}.")
                return redirect('celiaquia_expedientes_detalle', pk=pk)
            archivos[org] = archivo.archivo.path

        # Leer los tres archivos aprobados
        try:
            df_syntis = pd.read_excel(archivos['syntis'], header=None)
            df_salud = pd.read_excel(archivos['salud'], header=None)
            df_renaper = pd.read_excel(archivos['renaper'], header=None)
        except Exception as e:
            messages.error(request, f"Error leyendo archivos: {str(e)}")
            return redirect('celiaquia_expedientes_detalle', pk=pk)

        # Obtener intersección de DNIs
        dni_syntis = set(df_syntis[0].astype(str).str.strip())
        dni_salud = set(df_salud[0].astype(str).str.strip())
        dni_renaper = set(df_renaper[0].astype(str).str.strip())

        dni_comunes_cruce = dni_syntis & dni_salud & dni_renaper

        # Obtener la nómina de la provincia
        dni_nomina = set(expediente.personas.values_list('dni', flat=True))

        # Resultado final
        resultado_final = dni_comunes_cruce & dni_nomina

        # 🔴 NUEVO: Guardar en la base de datos
        ResultadoCruce.objects.filter(expediente=expediente).delete()  # eliminar anteriores

        for dni in dni_nomina:
            ResultadoCruce.objects.create(
                expediente=expediente,
                dni=dni,
                estado='aceptado' if dni in resultado_final else 'rechazado',
                motivo_rechazo='' if dni in resultado_final else 'No coincide en los 3 archivos'
            )

        # 🔴 También dejar en sesión solo los aceptados (opcional)
        request.session['resultado_cruce'] = list(resultado_final)

        expediente.estado = EstadoExpediente.objects.get(nombre="Cruce Finalizado")
        expediente.save()

        messages.success(request, f"Cruce validado. Coincidencias: {len(resultado_final)}.")
        return redirect('celiaquia_expedientes_detalle', pk=pk)


class ExpedienteDetallePagoView(LoginRequiredMixin, DetailView):
    model = Expediente
    template_name = 'celiaquia/detalle_pago.html'
    context_object_name = 'expediente'


class InformePagoCreateView(LoginRequiredMixin, CreateView):
    model = InformePago
    fields = ['fecha_pago', 'monto', 'observaciones']
    template_name = 'celiaquia/informe_pago_form.html'

    def form_valid(self, form):
        expediente = get_object_or_404(Expediente, pk=self.kwargs['pk'])
        form.instance.expediente = expediente
        form.instance.tecnico = self.request.user
        expediente.estado = EstadoExpediente.objects.get(nombre="Pagado")
        expediente.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('celiaquia_expedientes_detalle', kwargs={'pk': self.kwargs['pk']})
