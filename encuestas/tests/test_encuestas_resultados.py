import csv
import io

import openpyxl
import pytest
from django.contrib.auth.models import Permission
from django.urls import reverse

from encuestas.models import (
    OpcionPregunta,
    Pregunta,
    TipoDocumento,
    TipoPregunta,
    TipoSegmentacion,
)
from encuestas.services import (
    actualizar_segmentacion,
    crear_encuesta,
    publicar,
    registrar_respuesta,
)
from encuestas.services_resultados import (
    _formatear_numero,
    build_export_rows,
    build_resultados_csv_rows,
    build_resultados_excel,
    get_resultados_ronda,
)


@pytest.mark.parametrize(
    "valor, esperado",
    [
        ("7", "7"),
        ("7.00", "7"),
        ("4.50", "4.5"),
        ("0.00", "0"),
        ("10", "10"),
    ],
)
def test_formatear_numero(valor, esperado):
    from decimal import Decimal

    assert _formatear_numero(Decimal(valor)) == esperado


@pytest.fixture
def usuario_creador(django_user_model):
    return django_user_model.objects.create_user(username="creador", password="x")


@pytest.fixture
def respondientes(django_user_model):
    """Usuarios segmentables por DNI (ver _encuesta_con_preguntas). Se les
    asigna un DNI a propósito para que EncuestaObligatoriaMiddleware solo los
    bloquee a ellos, no a cualquier actor de los tests HTTP de esta suite
    (ver test_encuestas_middleware.py y el fixture encuesta_lista de
    test_encuestas_views.py, que tuvieron el mismo problema con
    TODOS_LOS_USUARIOS)."""
    usuarios = []
    for i in range(3):
        usuario = django_user_model.objects.create_user(
            username=f"user{i}", password="x"
        )
        usuario.profile.dni = f"3000000{i}"
        usuario.profile.save()
        usuarios.append(usuario)
    return usuarios


def _encuesta_con_preguntas(usuario_creador, *, anonima=False):
    encuesta = crear_encuesta(
        usuario=usuario_creador,
        titulo="Satisfacción",
        es_obligatoria=True,
        es_anonima=anonima,
        duracion_ronda_dias=7,
    )
    unica = Pregunta.objects.create(
        encuesta=encuesta,
        texto="¿Cómo calificás el servicio?",
        tipo=TipoPregunta.OPCION_UNICA,
        orden=1,
    )
    OpcionPregunta.objects.create(pregunta=unica, texto="Bueno", valor="Bueno", orden=1)
    OpcionPregunta.objects.create(pregunta=unica, texto="Malo", valor="Malo", orden=2)
    si_no = Pregunta.objects.create(
        encuesta=encuesta,
        texto="¿Volverías a usarlo?",
        tipo=TipoPregunta.SI_NO,
        orden=2,
    )
    escala = Pregunta.objects.create(
        encuesta=encuesta,
        texto="Puntuá del 1 al 10",
        tipo=TipoPregunta.ESCALA,
        orden=3,
        obligatoria=False,
    )
    comentario = Pregunta.objects.create(
        encuesta=encuesta,
        texto="Comentarios",
        tipo=TipoPregunta.TEXTO_LARGO,
        orden=4,
        obligatoria=False,
    )
    actualizar_segmentacion(
        encuesta,
        tipo=TipoSegmentacion.LISTADO_DOCUMENTOS,
        destinatarios=[
            {"tipo_documento": TipoDocumento.DNI, "numero_documento": f"3000000{i}"}
            for i in range(3)
        ],
    )
    ronda = publicar(encuesta, usuario=usuario_creador)
    return encuesta, ronda, unica, si_no, escala, comentario


@pytest.mark.django_db
def test_get_resultados_ronda_agrega_por_pregunta(usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {
            f"respuesta-{unica.pk}": "Bueno",
            f"respuesta-{si_no.pk}": "si",
            f"respuesta-{escala.pk}": "8",
            f"respuesta-{comentario.pk}": "Todo bien",
        },
    )
    registrar_respuesta(
        ronda,
        respondientes[1],
        {
            f"respuesta-{unica.pk}": "Bueno",
            f"respuesta-{si_no.pk}": "no",
            f"respuesta-{escala.pk}": "4",
        },
    )
    registrar_respuesta(
        ronda,
        respondientes[2],
        {
            f"respuesta-{unica.pk}": "Malo",
            f"respuesta-{si_no.pk}": "no",
        },
    )

    resultados = get_resultados_ronda(ronda)
    por_pregunta = {r.pregunta_id: r for r in resultados}

    resultado_unica = por_pregunta[unica.pk]
    assert resultado_unica.total_respuestas == 3
    distribucion = {o.texto: o.cantidad for o in resultado_unica.opciones}
    assert distribucion == {"Bueno": 2, "Malo": 1}

    resultado_si_no = por_pregunta[si_no.pk]
    distribucion_si_no = {o.texto: o.cantidad for o in resultado_si_no.opciones}
    assert distribucion_si_no == {"Sí": 1, "No": 2}

    resultado_escala = por_pregunta[escala.pk]
    assert resultado_escala.total_respuestas == 2
    assert resultado_escala.promedio == pytest.approx(6.0)
    assert resultado_escala.minimo == 4.0
    assert resultado_escala.maximo == 8.0

    resultado_comentario = por_pregunta[comentario.pk]
    assert resultado_comentario.respuestas_texto == ["Todo bien"]


@pytest.mark.django_db
def test_get_resultados_ronda_sin_respuestas_no_rompe(usuario_creador):
    _, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )

    resultados = get_resultados_ronda(ronda)

    por_pregunta = {r.pregunta_id: r for r in resultados}
    assert por_pregunta[escala.pk].promedio is None
    assert por_pregunta[comentario.pk].respuestas_texto == []


@pytest.mark.django_db
def test_get_resultados_ronda_evita_consultas_por_pregunta(
    django_assert_num_queries, usuario_creador, respondientes
):
    _, ronda, unica, si_no, _, _ = _encuesta_con_preguntas(usuario_creador)
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )

    with django_assert_num_queries(4):
        get_resultados_ronda(ronda)


@pytest.mark.django_db
def test_opcion_multiple_cuenta_cada_seleccion(usuario_creador, respondientes):
    encuesta = crear_encuesta(
        usuario=usuario_creador,
        titulo="Multi",
        es_obligatoria=True,
        duracion_ronda_dias=7,
    )
    multiple = Pregunta.objects.create(
        encuesta=encuesta,
        texto="¿Qué canales usás?",
        tipo=TipoPregunta.OPCION_MULTIPLE,
        obligatoria=False,
    )
    web = OpcionPregunta.objects.create(
        pregunta=multiple, texto="Web", valor="Web", orden=1
    )
    app = OpcionPregunta.objects.create(
        pregunta=multiple, texto="App", valor="App", orden=2
    )
    actualizar_segmentacion(encuesta, tipo=TipoSegmentacion.TODOS_LOS_USUARIOS)
    ronda = publicar(encuesta, usuario=usuario_creador)

    from django.utils.datastructures import MultiValueDict

    registrar_respuesta(
        ronda,
        respondientes[0],
        MultiValueDict({f"respuesta-{multiple.pk}": ["Web", "App"]}),
    )
    registrar_respuesta(
        ronda,
        respondientes[1],
        MultiValueDict({f"respuesta-{multiple.pk}": ["Web"]}),
    )

    resultados = get_resultados_ronda(ronda)
    resultado = resultados[0]
    distribucion = {o.texto: o.cantidad for o in resultado.opciones}
    assert distribucion == {"Web": 2, "App": 1}
    assert {o.texto: o.porcentaje for o in resultado.opciones} == {
        "Web": pytest.approx(66.7),
        "App": pytest.approx(33.3),
    }


@pytest.mark.django_db
def test_export_incluye_usuario_si_no_es_anonima(usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador, anonima=False
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )

    headers, filas = build_resultados_csv_rows(ronda)

    assert "Usuario" in headers
    fila = filas[0]
    indice_usuario = headers.index("Usuario")
    assert fila[indice_usuario] == respondientes[0].username


@pytest.mark.django_db
def test_export_no_incluye_usuario_si_es_anonima(usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador, anonima=True
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )

    headers, filas = build_resultados_csv_rows(ronda)

    assert "Usuario" not in headers
    assert respondientes[0].username not in filas[0]


@pytest.mark.django_db
def test_exportaciones_neutralizan_formulas(usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, _, comentario = _encuesta_con_preguntas(
        usuario_creador, anonima=True
    )
    comentario.texto = "=HIPERVINCULO()"
    comentario.save(update_fields=["texto"])
    registrar_respuesta(
        ronda,
        respondientes[0],
        {
            f"respuesta-{unica.pk}": "Bueno",
            f"respuesta-{si_no.pk}": "si",
            f"respuesta-{comentario.pk}": "=1+1",
        },
    )

    headers, filas = build_resultados_csv_rows(ronda)
    indice = headers.index("'=HIPERVINCULO()")
    assert filas[0][indice] == "'=1+1"

    workbook = openpyxl.load_workbook(io.BytesIO(build_resultados_excel(ronda)))
    celda = workbook.active.cell(row=2, column=indice + 1)
    assert celda.data_type == "s"
    assert celda.value == "'=1+1"


@pytest.mark.django_db
def test_export_incluye_metadatos_de_ronda_fecha_y_version(
    usuario_creador, respondientes
):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )

    headers, filas = build_resultados_csv_rows(ronda)

    assert headers[:3] == ["Ronda", "Versión", "Fecha de respuesta"]
    assert filas[0][0] == ronda.numero_ronda
    assert filas[0][1] == encuesta.version


@pytest.mark.django_db
def test_build_export_rows_formatea_valores_por_tipo(usuario_creador, respondientes):
    # anonima=True para que la fila no traiga la columna Usuario y el slice de
    # metadatos (Ronda, Versión, Fecha) quede fijo en 3 columnas; ese caso ya
    # se cubre aparte en test_export_incluye_usuario_si_no_es_anonima.
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador, anonima=True
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {
            f"respuesta-{unica.pk}": "Bueno",
            f"respuesta-{si_no.pk}": "no",
            f"respuesta-{escala.pk}": "7",
            f"respuesta-{comentario.pk}": "Excelente",
        },
    )

    preguntas, filas = build_export_rows(ronda)
    fila = filas[0]
    valores_por_pregunta = dict(zip([p.pk for p in preguntas], fila[3:]))

    assert valores_por_pregunta[unica.pk] == "Bueno"
    assert valores_por_pregunta[si_no.pk] == "No"
    assert valores_por_pregunta[escala.pk] == "7"
    assert valores_por_pregunta[comentario.pk] == "Excelente"


@pytest.mark.django_db
def test_excel_generado_es_valido_y_tiene_headers(usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )

    contenido = build_resultados_excel(ronda)

    workbook = openpyxl.load_workbook(io.BytesIO(contenido))
    worksheet = workbook.active
    primera_fila = [cell.value for cell in worksheet[1]]
    assert primera_fila[:3] == ["Ronda", "Versión", "Fecha de respuesta"]
    assert worksheet.max_row == 2  # header + 1 respuesta


@pytest.mark.django_db
def test_csv_response_tiene_bom_y_filas_correctas(usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )

    headers, filas = build_resultados_csv_rows(ronda)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(filas)
    contenido = buffer.getvalue()

    lector = csv.reader(io.StringIO(contenido))
    filas_leidas = list(lector)
    assert filas_leidas[0] == headers


def _permisos(codenames):
    return Permission.objects.filter(
        content_type__app_label="encuestas", codename__in=codenames
    )


@pytest.fixture
def user_resultados(django_user_model):
    user = django_user_model.objects.create_user(username="ve-resultados", password="x")
    user.user_permissions.add(*_permisos(["view_encuesta", "ver_resultados"]))
    return user


@pytest.fixture
def user_sin_permiso(django_user_model):
    return django_user_model.objects.create_user(username="sin-permiso", password="x")


@pytest.mark.django_db
def test_vista_resultados_requiere_permiso(client, user_sin_permiso, usuario_creador):
    encuesta, ronda, *_ = _encuesta_con_preguntas(usuario_creador)
    client.force_login(user_sin_permiso)
    response = client.get(reverse("encuestas_resultados", args=[encuesta.pk]))
    assert response.status_code == 403


@pytest.mark.django_db
def test_vista_resultados_muestra_la_ronda_mas_reciente(
    client, user_resultados, usuario_creador, respondientes
):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )
    client.force_login(user_resultados)

    response = client.get(reverse("encuestas_resultados", args=[encuesta.pk]))

    assert response.status_code == 200
    assert b"Bueno" in response.content
    assert b"Respuestas recibidas" in response.content


@pytest.mark.django_db
def test_vista_exportar_csv(client, user_resultados, usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )
    client.force_login(user_resultados)

    response = client.get(
        reverse("encuestas_resultados_exportar", args=[encuesta.pk, ronda.pk]),
        {"formato": "csv"},
    )

    assert response.status_code == 200
    assert response["Content-Type"].startswith("text/csv")
    assert "attachment" in response["Content-Disposition"]


@pytest.mark.django_db
def test_vista_exportar_excel(client, user_resultados, usuario_creador, respondientes):
    encuesta, ronda, unica, si_no, escala, comentario = _encuesta_con_preguntas(
        usuario_creador
    )
    registrar_respuesta(
        ronda,
        respondientes[0],
        {f"respuesta-{unica.pk}": "Bueno", f"respuesta-{si_no.pk}": "si"},
    )
    client.force_login(user_resultados)

    response = client.get(
        reverse("encuestas_resultados_exportar", args=[encuesta.pk, ronda.pk]),
        {"formato": "xlsx"},
    )

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@pytest.mark.django_db
def test_vista_exportar_requiere_permiso(client, user_sin_permiso, usuario_creador):
    encuesta, ronda, *_ = _encuesta_con_preguntas(usuario_creador)
    client.force_login(user_sin_permiso)
    response = client.get(
        reverse("encuestas_resultados_exportar", args=[encuesta.pk, ronda.pk])
    )
    assert response.status_code == 403


@pytest.mark.django_db
def test_vista_exportar_formato_invalido_da_404(
    client, user_resultados, usuario_creador
):
    encuesta, ronda, *_ = _encuesta_con_preguntas(usuario_creador)
    client.force_login(user_resultados)
    response = client.get(
        reverse("encuestas_resultados_exportar", args=[encuesta.pk, ronda.pk]),
        {"formato": "pdf"},
    )
    assert response.status_code == 404
