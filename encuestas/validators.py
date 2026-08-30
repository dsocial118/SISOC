from __future__ import annotations

import csv
import io
import json
import unicodedata
from dataclasses import dataclass

from django.core.exceptions import ValidationError
from django.core.validators import FileExtensionValidator
from openpyxl import load_workbook

LISTADO_ALLOWED_EXTENSIONS = ("xlsx", "csv")
LISTADO_MAX_SIZE_BYTES = 5 * 1024 * 1024
LISTADO_ACCEPT_ATTR = ".xlsx,.csv"

validate_listado_extension = FileExtensionValidator(
    allowed_extensions=LISTADO_ALLOWED_EXTENSIONS,
    message="Solo se permiten archivos Excel (.xlsx) o CSV.",
)


def validate_listado_file_size(value):
    if value and value.size > LISTADO_MAX_SIZE_BYTES:
        raise ValidationError("El archivo supera el tamaño máximo de 5 MB.")


LISTADO_FILE_VALIDATORS = [validate_listado_extension, validate_listado_file_size]

# Duplicados deliberadamente de encuestas.models (en vez de importarlos) para que
# este módulo no dependa de models.py, que a su vez importa validators.py para los
# validadores del FileField — evita un import circular entre ambos. Los tests de
# sincronización (test_encuestas_services.py) alertan si se desalinean.
TIPOS_DOCUMENTO_VALIDOS = ("dni", "cuit", "cuil")
TIPOS_PREGUNTA_VALIDOS = (
    "texto_corto",
    "texto_largo",
    "opcion_unica",
    "opcion_multiple",
    "escala",
    "si_no",
    "numerico",
    "fecha",
)
TIPOS_PREGUNTA_CON_OPCIONES = ("opcion_unica", "opcion_multiple")
OPERADORES_CONDICION_VALIDOS = ("igual", "distinto")


@dataclass(frozen=True)
class ParsedDestinatarioRow:
    fila: int
    tipo_documento: str
    numero_documento: str


def _normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    return normalized.lower().replace(" ", "_")


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _find_column(headers: list[str], *names: str) -> int:
    for index, header in enumerate(headers):
        if _normalize_header(header) in names:
            return index
    return -1


def _parse_rows(headers: list[str], data_rows) -> list[ParsedDestinatarioRow]:
    tipos_validos = set(TIPOS_DOCUMENTO_VALIDOS)
    tipo_index = _find_column(headers, "tipo_documento", "tipo")
    numero_index = _find_column(
        headers, "numero_documento", "documento", "nro_documento"
    )

    if tipo_index == -1 or numero_index == -1:
        raise ValidationError(
            "El archivo debe incluir las columnas 'tipo_documento' y "
            "'numero_documento'."
        )

    parsed: list[ParsedDestinatarioRow] = []
    for row_number, row in enumerate(data_rows, start=2):
        tipo = _clean_cell(row[tipo_index] if tipo_index < len(row) else "").lower()
        numero = _clean_cell(row[numero_index] if numero_index < len(row) else "")
        if not tipo and not numero:
            continue
        if tipo not in tipos_validos:
            raise ValidationError(
                f"Fila {row_number}: tipo de documento '{tipo}' inválido. Debe ser "
                f"uno de: {', '.join(sorted(tipos_validos))}."
            )
        if not numero.isdigit():
            raise ValidationError(
                f"Fila {row_number}: el número de documento '{numero}' debe ser "
                "numérico."
            )
        parsed.append(
            ParsedDestinatarioRow(
                fila=row_number, tipo_documento=tipo, numero_documento=numero
            )
        )

    if not parsed:
        raise ValidationError("El archivo no contiene filas con datos para procesar.")
    return parsed


def _rows_from_xlsx(uploaded_file) -> list[ParsedDestinatarioRow]:
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("El archivo está vacío.")
        headers = [_clean_cell(value) for value in rows[0]]
        return _parse_rows(headers, rows[1:])
    finally:
        workbook.close()


def _rows_from_csv(uploaded_file) -> list[ParsedDestinatarioRow]:
    uploaded_file.seek(0)
    content = uploaded_file.read()
    if isinstance(content, bytes):
        content = content.decode("utf-8-sig")
    rows = list(csv.reader(io.StringIO(content)))
    if not rows:
        raise ValidationError("El archivo está vacío.")
    headers = [_clean_cell(value) for value in rows[0]]
    return _parse_rows(headers, rows[1:])


def parse_listado_destinatarios(uploaded_file) -> list[ParsedDestinatarioRow]:
    """Parsea un archivo Excel o CSV con columnas tipo_documento/numero_documento."""
    nombre = getattr(uploaded_file, "name", "") or ""
    extension = nombre.rsplit(".", 1)[-1].lower() if "." in nombre else ""
    if extension == "csv":
        return _rows_from_csv(uploaded_file)
    return _rows_from_xlsx(uploaded_file)


@dataclass(frozen=True)
class ParsedCondicion:
    orden_referencia: int
    operador: str
    valor: str


@dataclass(frozen=True)
class ParsedPregunta:
    orden: int
    texto: str
    tipo: str
    obligatoria: bool
    opciones: tuple[str, ...]
    condicion: ParsedCondicion | None


def _parse_condicion(item: dict, indice: int, orden: int) -> ParsedCondicion | None:
    condicion_data = item.get("condicion")
    if not condicion_data:
        return None
    if not isinstance(condicion_data, dict):
        raise ValidationError(f"Pregunta {indice}: condición inválida.")

    orden_referencia = condicion_data.get("orden")
    operador = str(condicion_data.get("operador") or "")
    valor = str(condicion_data.get("valor") or "").strip()

    if not isinstance(orden_referencia, int) or orden_referencia >= orden:
        raise ValidationError(
            f"Pregunta {indice}: la condición debe referenciar a una pregunta "
            "anterior."
        )
    if operador not in OPERADORES_CONDICION_VALIDOS:
        raise ValidationError(f"Pregunta {indice}: operador de condición inválido.")
    if not valor:
        raise ValidationError(f"Pregunta {indice}: falta el valor de la condición.")

    return ParsedCondicion(
        orden_referencia=orden_referencia, operador=operador, valor=valor
    )


def _parse_una_pregunta(
    item: object, indice: int, ordenes_vistos: set[int]
) -> ParsedPregunta:
    if not isinstance(item, dict):
        raise ValidationError(f"Pregunta {indice}: formato inválido.")

    orden = item.get("orden")
    if not isinstance(orden, int) or orden <= 0:
        raise ValidationError(f"Pregunta {indice}: falta un orden válido.")
    if orden in ordenes_vistos:
        raise ValidationError(f"Hay más de una pregunta con el orden {orden}.")
    ordenes_vistos.add(orden)

    texto = str(item.get("texto") or "").strip()
    if not texto:
        raise ValidationError(f"Pregunta {indice}: el texto no puede estar vacío.")

    tipo = str(item.get("tipo") or "")
    if tipo not in TIPOS_PREGUNTA_VALIDOS:
        raise ValidationError(f"Pregunta {indice}: tipo '{tipo}' inválido.")

    opciones = tuple(
        str(opcion).strip()
        for opcion in item.get("opciones") or []
        if str(opcion).strip()
    )
    if tipo in TIPOS_PREGUNTA_CON_OPCIONES and len(opciones) < 2:
        raise ValidationError(
            f"Pregunta {indice}: las preguntas de opción requieren al menos dos "
            "opciones."
        )

    return ParsedPregunta(
        orden=orden,
        texto=texto,
        tipo=tipo,
        obligatoria=bool(item.get("obligatoria", True)),
        opciones=opciones,
        condicion=_parse_condicion(item, indice, orden),
    )


def _validar_condiciones_referencian_preguntas_existentes(
    preguntas: list[ParsedPregunta],
) -> None:
    ordenes_declarados = {pregunta.orden for pregunta in preguntas}
    for pregunta in preguntas:
        if (
            pregunta.condicion
            and pregunta.condicion.orden_referencia not in ordenes_declarados
        ):
            raise ValidationError(
                f"La pregunta con orden {pregunta.orden} referencia una pregunta "
                "que no existe."
            )


def parse_preguntas_payload(raw_json: str) -> list[ParsedPregunta]:
    """Valida y parsea el JSON armado por el editor dinámico de preguntas.

    No se usan formsets de Django: el editor arma preguntas y su lógica de
    condición referenciando otras preguntas por posición (``orden``), algo que
    los formsets no modelan bien antes de guardar. El payload viaja como un
    único campo oculto ``preguntas_json`` (ver encuestas/static o el template
    del formulario) y se valida acá a mano.

    Una encuesta puede guardarse como borrador sin preguntas todavía; la
    exigencia de tener al menos una pregunta se aplica recién al publicar
    (ver services.publicar), no en cada guardado.
    """
    try:
        data = json.loads(raw_json or "[]")
    except (TypeError, ValueError) as exc:
        raise ValidationError(
            "El listado de preguntas no tiene un formato válido."
        ) from exc

    if not isinstance(data, list):
        raise ValidationError("El listado de preguntas no tiene un formato válido.")

    ordenes_vistos: set[int] = set()
    parsed = [
        _parse_una_pregunta(item, indice, ordenes_vistos)
        for indice, item in enumerate(data, start=1)
    ]

    _validar_condiciones_referencian_preguntas_existentes(parsed)
    return parsed
