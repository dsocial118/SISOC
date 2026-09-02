import io
from datetime import date
from unittest.mock import patch

import pytest
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Municipio, Provincia
from pas.models import (
    PasCircuitoMensual,
    PasControlRenaper,
    PasEstado,
    PasIncompatibilidad,
    PasPersona,
)


@pytest.fixture
def usuario_pas(db):
    return get_user_model().objects.create_superuser(
        username="pas-cruces-test",
        email="pas-cruces@example.com",
        password="test",
    )


@pytest.fixture
def persona_pas(db):
    provincia = Provincia.objects.create(nombre="Provincia PAS cruces")
    municipio = Municipio.objects.create(
        nombre="Municipio PAS cruces",
        provincia=provincia,
    )
    estado, _ = PasEstado.objects.get_or_create(nombre="Activo")
    return PasPersona.objects.create(
        id_persona=880001,
        apellidos="García",
        nombres="María",
        dni=30111222,
        cuit="27301112224",
        provincia=provincia,
        municipio=municipio,
        estado=estado,
    )


@pytest.mark.django_db
def test_cruces_muestra_circuito_pendiente(client, usuario_pas):
    client.force_login(usuario_pas)

    response = client.get(reverse("pas_cruces"))

    assert response.status_code == 200
    assert "Circuito administrativo mensual" in response.content.decode()
    assert "Exportación a SINTyS (VIAS)" in response.content.decode()
    assert "Integración no disponible" in response.content.decode()
    assert response.context["pas_areas"][2]["active"] is True
    assert len(response.context["pas_areas"]) == 5
    assert "custom/css/pas.css" in response.content.decode()
    assert "pas-primary-action" in response.content.decode()
    assert "pas-secondary-action" in response.content.decode()
    assert "pas-renaper-update-button" in response.content.decode()
    assert "custom/css/pas_cruces.css" not in response.content.decode()
    assert not PasCircuitoMensual.objects.exists()


@pytest.mark.django_db
def test_cruces_muestra_estado_renaper_e_incompatibilidad(
    client,
    usuario_pas,
    persona_pas,
):
    PasControlRenaper.objects.create(
        persona=persona_pas,
        fecha_consulta=date(2026, 7, 29),
        resultado=PasControlRenaper.Resultado.FALLECIDA,
        sexo_consulta="F",
        error_tipo="fallecido",
    )
    PasIncompatibilidad.objects.create(
        persona=persona_pas,
        categoria=PasIncompatibilidad.Categoria.SUPERVIVENCIA,
        periodo_impacto=date(2026, 8, 1),
        detalle="RENAPER informó fallecimiento.",
    )
    client.force_login(usuario_pas)

    response = client.get(reverse("pas_cruces"))
    contenido = response.content.decode()

    assert response.status_code == 200
    assert "Actualizado 29/07/2026" in contenido
    assert "RENAPER informó fallecimiento." in contenido
    assert "08/2026" in contenido


@pytest.mark.django_db
def test_actualizacion_manual_renaper_fuerza_control_y_redirige(
    client,
    usuario_pas,
):
    client.force_login(usuario_pas)
    with patch(
        "pas.views.sincronizar_supervivencia_pas",
        return_value={
            "total": 3,
            "vigentes": 2,
            "fallecidas": 1,
            "no_encontradas": 0,
            "errores": 0,
            "omitidas": 0,
        },
    ) as sincronizar:
        response = client.post(reverse("pas_cruces_actualizar_renaper"))

    assert response.status_code == 302
    assert response.url == reverse("pas_cruces")
    sincronizar.assert_called_once_with(forzar=True)


@pytest.mark.django_db
def test_exportar_sintys_descarga_xlsx_y_completa_etapa(
    client,
    usuario_pas,
    persona_pas,
    tmp_path,
):
    client.force_login(usuario_pas)

    with override_settings(MEDIA_ROOT=tmp_path):
        response = client.post(reverse("pas_cruces_exportar_sintys"))

    assert response.status_code == 200
    assert response["Content-Type"].endswith("spreadsheetml.sheet")
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    filas = list(workbook["nomina"].iter_rows(values_only=True))
    assert filas[0] == ("numero_cuil", "nombre", "apellido")
    assert filas[1] == ("27301112224", "María", "García")

    circuito = PasCircuitoMensual.objects.get()
    assert circuito.fecha_exportacion_sintys is not None
    assert circuito.exportado_por == usuario_pas
    assert circuito.archivo_exportacion_sintys.name.endswith(".xlsx")


@pytest.mark.django_db
def test_importar_retorno_valida_y_completa_etapa(
    client,
    usuario_pas,
    tmp_path,
):
    client.force_login(usuario_pas)
    archivo = SimpleUploadedFile(
        "retorno-sintys.csv",
        b"numero_cuil,resultado\n27301112224,OK\n",
        content_type="text/csv",
    )

    with override_settings(MEDIA_ROOT=tmp_path):
        response = client.post(
            reverse("pas_cruces_importar_sintys"),
            {"archivo": archivo},
        )

    assert response.status_code == 302
    circuito = PasCircuitoMensual.objects.get()
    assert circuito.fecha_importacion_sintys is not None
    assert circuito.importado_por == usuario_pas
    assert circuito.archivo_retorno_sintys.name.endswith(".csv")


@pytest.mark.django_db
def test_importar_retorno_rechaza_extension_no_permitida(
    client,
    usuario_pas,
):
    client.force_login(usuario_pas)
    archivo = SimpleUploadedFile("retorno.exe", b"contenido")

    response = client.post(
        reverse("pas_cruces_importar_sintys"),
        {"archivo": archivo},
    )

    assert response.status_code == 302
    assert not PasCircuitoMensual.objects.exists()
