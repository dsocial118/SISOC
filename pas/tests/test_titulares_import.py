from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from threading import Barrier

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection, close_old_connections
from django.test import override_settings
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Municipio, Provincia
from pas.models import (
    PasAviso,
    PasEstado,
    PasExportacionTokens,
    PasHistorialEstado,
    PasPersona,
)
from pas.services.ddjj_service import (
    asegurar_invitacion_vigente,
    regenerar_invitacion,
)
from pas.services.titulares_import_service import importar_titulares_csv


@pytest.fixture
def catalogo_importacion():
    provincia = Provincia.objects.create(nombre="Buenos Aires")
    municipio = Municipio.objects.create(nombre="La Plata", provincia=provincia)
    estado = PasEstado.objects.create(nombre="Activo")
    return provincia, municipio, estado


def archivo_csv(contenido):
    return SimpleUploadedFile(
        "titulares.csv", contenido.encode("utf-8"), content_type="text/csv"
    )


@pytest.mark.django_db
def test_importacion_agrega_nuevos_omite_duplicados_y_crea_token(
    catalogo_importacion,
):
    provincia, municipio, estado = catalogo_importacion
    PasPersona.objects.create(
        id_persona=10,
        apellidos="Existente",
        nombres="Ana",
        dni=30111222,
        cuit="27301112220",
        provincia=provincia,
        municipio=municipio,
        estado=estado,
    )
    archivo = archivo_csv(
        "Apellidos;Nombres;DNI;CUIL;Provincia;Municipio\n"
        "Existente;Ana;30111222;27301112220;Buenos Aires;La Plata\n"
        "Nueva;Beatriz;32123456;27321234560;Buenos Aires;La Plata\n"
        "Repetida;Beatriz;32123456;27321234560;Buenos Aires;La Plata\n"
    )

    resultado = importar_titulares_csv(archivo)

    assert resultado == {"creados": 1, "duplicados": 2, "errores": [], "total": 3}
    nueva = PasPersona.objects.get(dni=32123456)
    assert nueva.id_persona == 11
    assert nueva.invitacion_ddjj_vigente is not None


@pytest.mark.django_db
def test_importacion_bloquea_catalogo_pas_antes_de_reservar_identificadores(
    catalogo_importacion, mocker
):
    provincia, municipio, estado = catalogo_importacion
    bloqueo = mocker.patch(
        "pas.services.titulares_import_service.bloquear_catalogo_pas_para_importacion",
        return_value={"activo": estado},
    )
    archivo = archivo_csv(
        "Apellidos;Nombres;DNI;CUIT;Provincia;Municipio\n"
        "Nueva;Beatriz;32123456;27321234560;Buenos Aires;La Plata\n"
    )

    resultado = importar_titulares_csv(archivo)

    assert resultado["creados"] == 1
    bloqueo.assert_called_once_with()


@pytest.mark.django_db(transaction=True)
@pytest.mark.mysql_compat
def test_importaciones_simultaneas_no_reutilizan_id_persona(catalogo_importacion):
    if connection.vendor != "mysql":
        pytest.skip("La contención real se verifica contra MySQL.")

    provincia, municipio, _estado = catalogo_importacion
    inicio = Barrier(2)

    def contenido_csv(desde):
        filas = "".join(
            f"Persona{numero};Prueba;{desde + numero};20{desde + numero}0;"
            f"{provincia.nombre};{municipio.nombre}\n"
            for numero in range(100)
        )
        return "Apellidos;Nombres;DNI;CUIT;Provincia;Municipio\n" + filas

    def importar_concurrente(contenido):
        close_old_connections()
        try:
            inicio.wait(timeout=10)
            return importar_titulares_csv(archivo_csv(contenido))
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=2) as executor:
        resultados = list(
            executor.map(
                importar_concurrente,
                (contenido_csv(33_000_000), contenido_csv(34_000_000)),
                timeout=30,
            )
        )

    assert [resultado["creados"] for resultado in resultados] == [100, 100]
    assert not any(resultado["errores"] for resultado in resultados)
    assert PasPersona.objects.count() == 200
    assert PasPersona.objects.values("id_persona").distinct().count() == 200


@pytest.mark.django_db
def test_importacion_acepta_aliases_y_guarda_datos_estado_y_aviso(
    catalogo_importacion,
):
    _provincia, _municipio, estado_activo = catalogo_importacion
    aviso = PasAviso.objects.create(codigo=1, descripcion="COBRANDO 100%")
    aviso.estados.add(estado_activo)
    archivo = archivo_csv(
        "Apellido;Nombre;DNI;CUIL;Provincia;Municipio;Calle;Altura;Email;"
        "UltimoEstadoPas;AvisoLiquidacion\n"
        "Perez;Lucia;33444555;27334445550;Buenos Aires;La Plata;"
        "Calle 7;123;lucia@example.test;ACTIVO;COBRANDO 100% EL 05/08/2026\n"
    )

    resultado = importar_titulares_csv(archivo)

    assert resultado == {"creados": 1, "duplicados": 0, "errores": [], "total": 1}
    persona = PasPersona.objects.get(dni=33444555)
    assert persona.apellidos == "Perez"
    assert persona.nombres == "Lucia"
    assert persona.domicilio == "Calle 7 123"
    assert persona.correo_electronico == "lucia@example.test"
    assert persona.estado == estado_activo
    assert list(persona.avisos.all()) == [aviso]
    historial = PasHistorialEstado.objects.get(persona=persona)
    assert historial.estado_nuevo == estado_activo
    assert list(historial.avisos_nuevos.all()) == [aviso]


@pytest.mark.django_db
def test_importacion_deja_opcionales_vacios_si_no_estan_en_csv(
    catalogo_importacion,
):
    archivo = archivo_csv(
        "Apellidos;Nombres;DNI;CUIT;Provincia;Municipio\n"
        "SinDato;Mario;35555666;20355556660;Buenos Aires;La Plata\n"
    )

    resultado = importar_titulares_csv(archivo)

    assert resultado["creados"] == 1
    persona = PasPersona.objects.get(dni=35555666)
    assert persona.domicilio == ""
    assert persona.correo_electronico == ""
    assert not persona.avisos.exists()


@pytest.mark.django_db
def test_importacion_aplica_estado_no_activo_y_su_aviso(catalogo_importacion):
    estado_suspendido = PasEstado.objects.create(nombre="Suspendido")
    aviso = PasAviso.objects.create(
        codigo=44,
        descripcion="INCOMPATIBLE ANSES (JUBILACION / PENSION)",
    )
    aviso.estados.add(estado_suspendido)
    archivo = archivo_csv(
        "Apellidos;Nombres;DNI;CUIT;Provincia;Municipio;"
        "UltimoEstadoPas;AvisoLiquidacion\n"
        "Gomez;Elena;37777888;27377778880;Buenos Aires;La Plata;Suspendido;"
        "INCOMPATIBLE ANSES (JUBILACION / PENSION)\n"
    )

    resultado = importar_titulares_csv(archivo)

    assert resultado["creados"] == 1
    persona = PasPersona.objects.get(dni=37777888)
    assert persona.estado == estado_suspendido
    assert list(persona.avisos.all()) == [aviso]
    assert (
        PasHistorialEstado.objects.get(persona=persona).estado_nuevo
        == estado_suspendido
    )


@pytest.mark.django_db
def test_importacion_reporta_aviso_no_asociable_sin_crear_persona(
    catalogo_importacion,
):
    archivo = archivo_csv(
        "Apellidos;Nombres;DNI;CUIT;Provincia;Municipio;AvisoLiquidacion\n"
        "Perez;Juan;36666777;20366667770;Buenos Aires;La Plata;AVISO DESCONOCIDO\n"
    )

    resultado = importar_titulares_csv(archivo)

    assert resultado["creados"] == 0
    assert len(resultado["errores"]) == 1
    assert "no coincide de forma unívoca" in resultado["errores"][0]
    assert not PasPersona.objects.filter(dni=36666777).exists()


@pytest.mark.django_db
def test_regenerar_token_revoca_el_anterior(catalogo_importacion):
    provincia, municipio, estado = catalogo_importacion
    persona = PasPersona.objects.create(
        id_persona=1,
        apellidos="Paz",
        nombres="Luis",
        dni=20111222,
        cuit="20201112220",
        provincia=provincia,
        municipio=municipio,
        estado=estado,
    )
    anterior = asegurar_invitacion_vigente(persona)

    nueva = regenerar_invitacion(persona)

    anterior.refresh_from_db()
    assert anterior.revocada is not None
    assert not anterior.disponible
    assert nueva.disponible
    assert persona.invitacion_ddjj_vigente == nueva


@pytest.mark.django_db
@override_settings(DOMINIO="http://localhost:8002")
def test_exportacion_excel_contiene_cuil_y_token_vigente(client, catalogo_importacion):
    provincia, municipio, estado = catalogo_importacion
    persona = PasPersona.objects.create(
        id_persona=1,
        apellidos="Paz",
        nombres="Luis",
        dni=20111222,
        cuit="20201112220",
        provincia=provincia,
        municipio=municipio,
        estado=estado,
    )
    invitacion = asegurar_invitacion_vigente(persona)
    usuario = get_user_model().objects.create_superuser(
        username="exportador-tokens", email="tokens@example.test", password="test"
    )
    client.force_login(usuario)

    respuesta = client.get(reverse("pas_tokens_exportar"))

    assert respuesta.status_code == 200
    workbook = load_workbook(BytesIO(respuesta.content), read_only=True)
    filas = list(workbook.active.iter_rows(values_only=True))
    assert filas == [
        ("CUIL", "TOKEN"),
        (
            persona.cuit,
            f"http://localhost:8002/pas/ddjj/formulario/{invitacion.token}",
        ),
    ]
    registro = PasExportacionTokens.objects.get()
    assert registro.usuario == usuario
    assert registro.cantidad == 1
    assert registro.fecha is not None


@pytest.mark.django_db
def test_buscador_reemplaza_nuevo_titular_por_importar_y_exportar(
    client, catalogo_importacion
):
    usuario = get_user_model().objects.create_superuser(
        username="admin-padron-pas", email="padron@example.test", password="test"
    )
    client.force_login(usuario)

    respuesta = client.get(reverse("pas_persona_listar"))

    assert respuesta.status_code == 200
    assert b"Importar titulares" in respuesta.content
    assert b"Exportar tokens" in respuesta.content
    assert b"Nuevo titular" not in respuesta.content


@pytest.mark.django_db
def test_importacion_y_exportacion_requieren_permisos(client):
    usuario = get_user_model().objects.create_user(
        username="sin-permisos-padron", password="test"
    )
    client.force_login(usuario)

    assert client.get(reverse("pas_titulares_importar")).status_code == 403
    assert client.get(reverse("pas_tokens_exportar")).status_code == 403


@pytest.mark.django_db
def test_exportacion_rechaza_permiso_generico_de_consulta_pas(client):
    usuario = get_user_model().objects.create_user(
        username="consulta-tokens-padron", password="test"
    )
    usuario.user_permissions.add(Permission.objects.get(codename="view_paspersona"))
    client.force_login(usuario)

    respuesta = client.get(reverse("pas_tokens_exportar"))

    assert respuesta.status_code == 403
    assert PasExportacionTokens.objects.count() == 0

    listado = client.get(reverse("pas_persona_listar"))
    assert listado.status_code == 200
    assert b"Exportar tokens" not in listado.content


@pytest.mark.django_db
def test_exportacion_acepta_permiso_especifico(client):
    usuario = get_user_model().objects.create_user(
        username="exporta-tokens-padron", password="test"
    )
    usuario.user_permissions.add(Permission.objects.get(codename="export_ddjj_tokens"))
    client.force_login(usuario)

    respuesta = client.get(reverse("pas_tokens_exportar"))

    assert respuesta.status_code == 200
    assert respuesta["Content-Type"].endswith("spreadsheetml.sheet")
    assert PasExportacionTokens.objects.get().usuario == usuario


@pytest.mark.django_db
def test_admin_permite_regenerar_token_de_persona(client, catalogo_importacion):
    provincia, municipio, estado = catalogo_importacion
    persona = PasPersona.objects.create(
        id_persona=1,
        apellidos="Admin",
        nombres="Prueba",
        dni=22111222,
        cuit="20221112220",
        provincia=provincia,
        municipio=municipio,
        estado=estado,
    )
    anterior = asegurar_invitacion_vigente(persona)
    usuario = get_user_model().objects.create_superuser(
        username="admin-regenera-token", email="admin@example.test", password="test"
    )
    client.force_login(usuario)

    respuesta = client.post(
        reverse("admin:pas_paspersona_changelist"),
        {
            "action": "regenerar_token_ddjj",
            "_selected_action": [str(persona.pk)],
        },
        follow=True,
    )

    anterior.refresh_from_db()
    assert respuesta.status_code == 200
    assert anterior.revocada is not None
    assert persona.invitacion_ddjj_vigente.token != anterior.token
