from __future__ import annotations

import logging
import unicodedata
from dataclasses import dataclass
from io import BytesIO

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.db import transaction
from django.template.loader import render_to_string
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from core.models import Provincia
from users.models import ProfileTerritorialScope, UserImportJob, UserImportJobRow
from users.services_auth import generate_temporary_password_for_user

User = get_user_model()
logger = logging.getLogger("django")

USER_IMPORT_CREDENTIALS_EMAIL_TEMPLATE = "user/bulk_credentials_email.txt"
USER_IMPORT_CREDENTIALS_EMAIL_SUBJECT = "SISOC - Credenciales de acceso"
USER_IMPORT_TEMPLATE_FILENAME = "plantilla_importacion_usuarios.xlsx"
USER_IMPORT_SHEET_NAME = "usuarios"
USER_IMPORT_REQUIRED_COLUMNS = (
    "nombre",
    "apellido",
    "permisos",
    "provincias",
    "rol",
)
USER_IMPORT_OPTIONAL_COLUMNS = ("correo", "username", "accion_grupos")
USER_IMPORT_KNOWN_COLUMNS = USER_IMPORT_REQUIRED_COLUMNS + USER_IMPORT_OPTIONAL_COLUMNS
USER_IMPORT_TEMPLATE_HEADERS = (
    "Username",
    "Nombre",
    "Apellido",
    "Correo",
    "Permisos",
    "Accion grupos",
    "Provincias",
    "Rol",
)
USERNAME_MAX_LENGTH = 150

GROUP_ACTION_AGREGAR = "agregar"
GROUP_ACTION_QUITAR = "quitar"
GROUP_ACTION_REEMPLAZAR = "reemplazar"
GROUP_ACTIONS = (GROUP_ACTION_AGREGAR, GROUP_ACTION_QUITAR, GROUP_ACTION_REEMPLAZAR)


def _build_login_url() -> str:
    try:
        path = reverse("login")
    except Exception:
        path = "/"
    domain = (
        str(settings.DOMINIO).replace("http://", "").replace("https://", "").rstrip("/")
    )
    scheme = "https" if settings.ENVIRONMENT == "prd" else "http"
    return f"{scheme}://{domain}{path}"


def _normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower().replace(" ", "_").replace("-", "_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _slug_base_desde_email(email: str) -> str:
    local_part = email.split("@", 1)[0]
    normalized = unicodedata.normalize("NFKD", local_part)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    cleaned = "".join(ch if ch.isalnum() else "." for ch in normalized.lower())
    cleaned = ".".join(part for part in cleaned.split(".") if part)
    return cleaned[:USERNAME_MAX_LENGTH] or "usuario"


def _slug_base_desde_nombre(*, nombre: str, apellido: str) -> str:
    raw = f"{apellido} {nombre}".strip()
    normalized = unicodedata.normalize("NFKD", raw)
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    cleaned = "".join(ch if ch.isalnum() else "." for ch in normalized.lower())
    cleaned = ".".join(part for part in cleaned.split(".") if part)
    return cleaned[:USERNAME_MAX_LENGTH] or "usuario"


def _generar_username_unico(base: str) -> str:
    if not User.objects.filter(username__iexact=base).exists():
        return base
    counter = 2
    while True:
        suffix = f"-{counter}"
        candidate = f"{base[:USERNAME_MAX_LENGTH - len(suffix)]}{suffix}"
        if not User.objects.filter(username__iexact=candidate).exists():
            return candidate
        counter += 1


def _parse_semicolon_field(value: str) -> list[str]:
    return [item.strip() for item in value.split(";") if item.strip()]


def _get_active_worksheet(workbook):
    if USER_IMPORT_SHEET_NAME in workbook.sheetnames:
        return workbook[USER_IMPORT_SHEET_NAME]
    return workbook.active


def validate_user_import_workbook(uploaded_file) -> None:
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc

    try:
        worksheet = _get_active_worksheet(workbook)
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationError("El archivo Excel esta vacio.") from exc

        headers = [_normalize_header(v) for v in header_row]
        missing = [col for col in USER_IMPORT_REQUIRED_COLUMNS if col not in headers]
        if missing:
            raise ValidationError(
                f"El archivo debe incluir las columnas obligatorias: {', '.join(missing)}."
            )

        for row in rows:
            if any(_clean_cell(v) for v in row):
                uploaded_file.seek(0)
                return

        raise ValidationError(
            "El archivo Excel no contiene filas con datos para procesar."
        )
    finally:
        workbook.close()


def load_user_import_rows(uploaded_file) -> list[dict]:
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    try:
        worksheet = _get_active_worksheet(workbook)
        raw_rows = list(worksheet.iter_rows(values_only=True))
        if not raw_rows:
            return []

        headers = [_normalize_header(v) for v in raw_rows[0]]
        col_map = {
            col: idx
            for idx, col in enumerate(headers)
            if col in USER_IMPORT_KNOWN_COLUMNS
        }

        parsed = []
        for row_number, row in enumerate(raw_rows[1:], start=2):
            values = {
                col: _clean_cell(row[idx] if idx < len(row) else "")
                for col, idx in col_map.items()
            }
            if not any(values.values()):
                continue
            values["fila"] = row_number
            parsed.append(values)

        return parsed
    finally:
        workbook.close()


def create_user_import_job(
    *,
    uploaded_file,
    requested_by,
    send_credentials: bool,
    is_pwa_import: bool = False,
) -> UserImportJob:
    validate_user_import_workbook(uploaded_file)
    rows = load_user_import_rows(uploaded_file)

    job = UserImportJob(
        requested_by=requested_by,
        original_filename=getattr(uploaded_file, "name", "usuarios.xlsx"),
        send_credentials=send_credentials,
        is_pwa_import=is_pwa_import,
        total_rows=len(rows),
    )
    uploaded_file.seek(0)
    job.archivo.save(job.original_filename, uploaded_file, save=False)
    job.save()

    UserImportJobRow.objects.bulk_create(
        [
            UserImportJobRow(
                job=job,
                fila=row["fila"],
                nombre=row.get("nombre", ""),
                apellido=row.get("apellido", ""),
                email=row.get("correo", ""),
                rol=row.get("rol", ""),
                status=UserImportJobRow.Status.PENDING,
            )
            for row in rows
        ]
    )

    return job


def _enviar_credenciales_import(*, user, password: str) -> bool:
    if not user.email:
        return False
    try:
        from users.services_bulk_credentials import (  # noqa: PLC0415
            BulkCredentialEntry,
        )

        entry = BulkCredentialEntry(
            username=user.username,
            plain_password=password,
            first_name=user.first_name or "",
            last_name=user.last_name or "",
        )
        context = {
            "entries": [entry],
            "is_grouped": False,
            "user_username": entry.username,
            "user_full_name": entry.full_name,
            "plain_password": password,
            "nombre_del_centro": "",
            "login_url": _build_login_url(),
        }
        message = render_to_string(USER_IMPORT_CREDENTIALS_EMAIL_TEMPLATE, context)
        send_mail(
            subject=USER_IMPORT_CREDENTIALS_EMAIL_SUBJECT,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=False,
        )
        return True
    except Exception:
        logger.exception(
            "Fallo enviando credenciales de importacion user_id=%s", user.id
        )
        return False


def _resolver_grupos(permisos_raw: str) -> list:
    grupos = []
    for nombre_grupo in _parse_semicolon_field(permisos_raw):
        grupo = Group.objects.filter(name=nombre_grupo).first()
        if grupo is None:
            raise ValidationError(f"El grupo '{nombre_grupo}' no existe en el sistema.")
        grupos.append(grupo)
    return grupos


def _resolver_provincias(provincias_raw: str) -> list:
    provincias = []
    for nombre_prov in _parse_semicolon_field(provincias_raw):
        prov = Provincia.objects.filter(nombre__iexact=nombre_prov).first()
        if prov is None:
            raise ValidationError(
                f"La provincia '{nombre_prov}' no existe en el sistema."
            )
        provincias.append(prov)
    return provincias


def _is_unrestricted_actor(actor) -> bool:
    if actor is None:
        return True
    if getattr(actor, "is_superuser", False):
        return True
    profile = getattr(actor, "profile", None)
    if not profile:
        return True
    return not profile.grupos_asignables.exists()


def _get_allowed_group_ids(actor) -> set | None:
    if _is_unrestricted_actor(actor):
        return None
    return set(actor.profile.grupos_asignables.values_list("id", flat=True))


def _resolver_usuario_objetivo(row_data: dict):
    username_raw = row_data.get("username", "").strip()
    email_raw = row_data.get("correo", "").strip()
    if username_raw:
        return User.objects.filter(username__iexact=username_raw).first()
    if email_raw:
        return User.objects.filter(email__iexact=email_raw).first()
    return None


def _aplicar_accion_grupos(
    *,
    user,
    grupos: list,
    accion: str,
    allowed_group_ids: set | None,
) -> bool:
    current_group_ids = set(user.groups.values_list("id", flat=True))
    requested_group_ids = {g.pk for g in grupos}

    if accion == GROUP_ACTION_AGREGAR:
        to_add = requested_group_ids - current_group_ids
        if to_add:
            user.groups.add(*to_add)
        return bool(to_add)

    if accion == GROUP_ACTION_QUITAR:
        to_remove = requested_group_ids & current_group_ids
        if to_remove:
            user.groups.remove(*to_remove)
        return bool(to_remove)

    final_group_ids = (
        (current_group_ids - allowed_group_ids | requested_group_ids)
        if allowed_group_ids
        else requested_group_ids
    )
    if final_group_ids != current_group_ids:
        user.groups.set(list(final_group_ids))
        return True
    return False


@dataclass
class _ActualizarUsuarioParams:
    user: object
    email: str
    username_raw: str
    grupos: list
    accion_grupos: str
    allowed_group_ids: set | None


def _procesar_usuario_existente(params: _ActualizarUsuarioParams) -> dict:
    changed = False

    with transaction.atomic():
        if (
            params.username_raw
            and params.email
            and params.user.email.lower() != params.email.lower()
        ):
            params.user.email = params.email
            params.user.save(update_fields=["email"])
            changed = True

        grupos_changed = _aplicar_accion_grupos(
            user=params.user,
            grupos=params.grupos,
            accion=params.accion_grupos,
            allowed_group_ids=params.allowed_group_ids,
        )
        changed = changed or grupos_changed

    status = UserImportJobRow.Status.SKIPPED if not changed else UserImportJobRow.Status.CREATED
    mensaje = (
        f"Usuario {params.user.username}: sin cambios."
        if not changed
        else f"Usuario {params.user.username} actualizado ({params.accion_grupos} grupos)."
    )
    return {
        "status": status,
        "mensaje": mensaje,
        "email": params.user.email,
    }


@dataclass
class _CrearUsuarioParams:
    nombre: str
    apellido: str
    email: str
    username_raw: str
    rol: str
    grupos: list
    provincias_objs: list
    job: object


@dataclass
class _DatosFilaValidados:
    nombre: str
    apellido: str
    email: str
    username_raw: str
    rol: str
    accion_grupos: str
    grupos: list
    provincias_objs: list
    allowed_group_ids: set | None


def _crear_usuario_nuevo(params: _CrearUsuarioParams) -> tuple[User, str]:
    user = User(
        username=params.username_raw,
        email=params.email,
        first_name=params.nombre,
        last_name=params.apellido,
        is_staff=not params.job.is_pwa_import,
        is_active=True,
    )
    user.set_unusable_password()
    user.save()

    if params.grupos:
        user.groups.set(params.grupos)

    profile = user.profile
    profile.rol = params.rol
    if params.provincias_objs:
        profile.es_usuario_provincial = True
    profile.save(update_fields=["rol", "es_usuario_provincial"])

    for prov in params.provincias_objs:
        scope_key = ProfileTerritorialScope.build_scope_key(prov.pk)
        ProfileTerritorialScope.objects.get_or_create(
            profile=profile,
            scope_key=scope_key,
            defaults={
                "provincia_id": prov.pk,
                "municipio": None,
                "localidad": None,
            },
        )

    password = generate_temporary_password_for_user(user=user)
    return user, password


def _validar_y_preparar_fila(
    row_data: dict, job: UserImportJob
) -> _DatosFilaValidados:
    nombre = row_data.get("nombre", "").strip()
    apellido = row_data.get("apellido", "").strip()
    email_raw = row_data.get("correo", "").strip()
    username_raw = row_data.get("username", "").strip()
    rol = row_data.get("rol", "").strip()
    accion_grupos = row_data.get("accion_grupos", "").strip().lower() or GROUP_ACTION_AGREGAR

    if accion_grupos not in GROUP_ACTIONS:
        raise ValidationError(
            f"Accion de grupos invalida: '{accion_grupos}'. "
            f"Los valores validos son: {', '.join(GROUP_ACTIONS)}."
        )

    if not username_raw and not email_raw:
        raise ValidationError(
            "Debe indicar Username o Correo para identificar al usuario de la fila."
        )

    email = ""
    if email_raw:
        try:
            validate_email(email_raw)
        except ValidationError as exc:
            raise ValidationError(
                f"El correo '{email_raw}' no tiene formato valido."
            ) from exc
        email = email_raw.lower()

    grupos = _resolver_grupos(row_data.get("permisos", "").strip())
    allowed_group_ids = _get_allowed_group_ids(job.requested_by)

    if allowed_group_ids is not None and grupos:
        out_of_scope = [g for g in grupos if g.pk not in allowed_group_ids]
        if out_of_scope:
            names = ", ".join(g.name for g in out_of_scope)
            raise ValidationError(f"No tiene permiso para operar los grupos: {names}.")

    provincias_objs = _resolver_provincias(row_data.get("provincias", "").strip())

    return _DatosFilaValidados(
        nombre=nombre,
        apellido=apellido,
        email=email,
        username_raw=username_raw,
        rol=rol,
        accion_grupos=accion_grupos,
        grupos=grupos,
        provincias_objs=provincias_objs,
        allowed_group_ids=allowed_group_ids,
    )


def process_single_user_import_row(
    *, row_data: dict, job: UserImportJob
) -> dict:
    datos = _validar_y_preparar_fila(row_data, job)
    existing_user = _resolver_usuario_objetivo(row_data)

    if existing_user is not None:
        params = _ActualizarUsuarioParams(
            user=existing_user,
            email=datos.email,
            username_raw=datos.username_raw,
            grupos=datos.grupos,
            accion_grupos=datos.accion_grupos,
            allowed_group_ids=datos.allowed_group_ids,
        )
        return _procesar_usuario_existente(params)

    if not datos.nombre or not datos.apellido:
        raise ValidationError("Los campos Nombre y Apellido son obligatorios.")

    username_to_create = datos.username_raw
    if not username_to_create:
        username_base = _slug_base_desde_nombre(
            nombre=datos.nombre, apellido=datos.apellido
        )
        username_to_create = _generar_username_unico(username_base)

    if (
        datos.username_raw
        and User.objects.filter(username__iexact=username_to_create).exists()
    ):
        raise ValidationError(
            f"Ya existe un usuario con el username '{username_to_create}'."
        )

    with transaction.atomic():
        params = _CrearUsuarioParams(
            nombre=datos.nombre,
            apellido=datos.apellido,
            email=datos.email,
            username_raw=username_to_create,
            rol=datos.rol,
            grupos=datos.grupos,
            provincias_objs=datos.provincias_objs,
            job=job,
        )
        user, password = _crear_usuario_nuevo(params)

    if job.send_credentials:
        _enviar_credenciales_import(user=user, password=password)

    return {
        "status": UserImportJobRow.Status.CREATED,
        "mensaje": f"Usuario {user.username} creado correctamente.",
        "email": datos.email,
    }


def build_user_import_error_message(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        return " ".join(exc.messages)
    return "Ocurrio un error inesperado al procesar la fila."


def generate_user_import_template() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = USER_IMPORT_SHEET_NAME
    worksheet.append(list(USER_IMPORT_TEMPLATE_HEADERS))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
