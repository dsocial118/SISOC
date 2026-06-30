from __future__ import annotations

import logging
import os
import signal
import smtplib
import threading
import time
import unicodedata
from contextlib import contextmanager
from dataclasses import dataclass
from io import BytesIO

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email
from django.db import transaction
from django.template.loader import render_to_string
from django.urls import reverse
from openpyxl import Workbook, load_workbook

User = get_user_model()
logger = logging.getLogger("django")

DEFAULT_BULK_CREDENTIALS_SEND_TYPE = "standard"
SHEET_NAME = "credenciales"
BULK_CREDENTIALS_EMAIL_MAX_ATTEMPTS = 2
BULK_CREDENTIALS_EMAIL_RETRY_BACKOFF_SECONDS = 1
BULK_CREDENTIALS_BATCH_TIMEOUT_BUFFER_SECONDS = 5
BULK_CREDENTIALS_MIN_SECONDS_TO_START_NEXT_ROW = 5


class BulkCredentialsEmailTimeoutError(TimeoutError):
    """Timeout controlado para cortar un intento de envio antes del timeout web."""


def _safe_positive_int(value, default: int) -> int:
    try:
        parsed = int(str(value).strip())
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _get_bulk_credentials_email_attempt_timeout_seconds() -> int:
    configured_timeout = int(getattr(settings, "EMAIL_TIMEOUT", 10) or 10)
    return max(1, configured_timeout)


def _get_bulk_credentials_batch_timeout_seconds() -> int:
    worker_timeout = _safe_positive_int(os.getenv("GUNICORN_TIMEOUT", ""), 30)
    return max(
        BULK_CREDENTIALS_MIN_SECONDS_TO_START_NEXT_ROW,
        worker_timeout - BULK_CREDENTIALS_BATCH_TIMEOUT_BUFFER_SECONDS,
    )


def _get_bulk_credentials_batch_deadline() -> float:
    return time.monotonic() + float(_get_bulk_credentials_batch_timeout_seconds())


def _get_remaining_processing_seconds(deadline: float | None) -> float | None:
    if deadline is None:
        return None
    return max(0.0, deadline - time.monotonic())


def _has_enough_batch_time(deadline: float | None) -> bool:
    remaining_seconds = _get_remaining_processing_seconds(deadline)
    if remaining_seconds is None:
        return True
    return remaining_seconds >= BULK_CREDENTIALS_MIN_SECONDS_TO_START_NEXT_ROW


def _get_batch_timeout_message() -> str:
    return (
        "Se alcanzo el tiempo maximo del lote. " "Reintente con una planilla mas chica."
    )


def _get_signal_timeout_guard_components():
    if threading.current_thread() is not threading.main_thread():
        return None

    signal_alarm = getattr(signal, "SIGALRM", None)
    timer_real = getattr(signal, "ITIMER_REAL", None)
    getitimer = getattr(signal, "getitimer", None)
    setitimer = getattr(signal, "setitimer", None)
    if (
        signal_alarm is None
        or timer_real is None
        or not callable(getitimer)
        or not callable(setitimer)
    ):
        return None

    return signal_alarm, timer_real, getitimer, setitimer


@contextmanager
def _mail_send_timeout_guard(timeout_seconds: int):
    timeout_components = _get_signal_timeout_guard_components()
    if timeout_seconds <= 0 or timeout_components is None:
        yield
        return

    signal_alarm, timer_real, getitimer, setitimer = timeout_components
    previous_handler = signal.getsignal(signal_alarm)
    previous_timer = getitimer(timer_real)

    def _raise_mail_timeout(signum, frame):  # pragma: no cover - depende de signal
        raise BulkCredentialsEmailTimeoutError(
            "Timeout enviando credenciales al servidor de correo."
        )

    signal.signal(signal_alarm, _raise_mail_timeout)
    setitimer(timer_real, float(timeout_seconds))
    try:
        yield
    finally:
        setitimer(timer_real, 0)
        signal.signal(signal_alarm, previous_handler)
        if previous_timer != (0.0, 0.0):
            setitimer(timer_real, *previous_timer)


@dataclass(frozen=True)
class BulkCredentialsSendTypeConfig:
    key: str
    label: str
    required_columns: tuple[str, ...]
    template_headers: tuple[str, ...]
    template_filename: str
    email_subject: str
    email_template_name: str
    description: str


@dataclass(frozen=True)
class ParsedCredentialRow:
    fila: int
    data: dict[str, str]

    @property
    def usuario(self) -> str:
        return self.data.get("usuario", "")

    @property
    def mail(self) -> str:
        return self.data.get("mail", "")

    @property
    def password(self) -> str:
        return self.data.get("password", "")

    @property
    def nombre_del_centro(self) -> str:
        return self.data.get("nombre_del_centro", "")


@dataclass(frozen=True)
class BulkCredentialEntry:
    """Una credencial individual dentro de un envío (posiblemente agrupado)."""

    username: str
    plain_password: str
    first_name: str
    last_name: str
    nombre_del_centro: str = ""

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


BULK_CREDENTIALS_SEND_TYPES = {
    "standard": BulkCredentialsSendTypeConfig(
        key="standard",
        label="Estandar",
        required_columns=("usuario", "mail"),
        template_headers=("usuario", "mail"),
        template_filename="plantilla_credenciales_usuarios.xlsx",
        email_subject="SISOC - Credenciales de acceso",
        email_template_name="user/bulk_credentials_email.txt",
        description=(
            "Carga un archivo .xlsx con encabezados usuario y mail. Se valida "
            "el usuario existente y se envia la credencial temporal vigente al "
            "mail informado en la planilla, o al mail del usuario si la celda "
            "mail esta vacia."
        ),
    ),
    "inet": BulkCredentialsSendTypeConfig(
        key="inet",
        label="INET",
        required_columns=("usuario", "mail", "nombre_del_centro"),
        template_headers=("usuario", "mail", "Nombre del Centro"),
        template_filename="plantilla_credenciales_usuarios_inet.xlsx",
        email_subject="Acceso a la plataforma y capacitación virtual – INET",
        email_template_name="user/bulk_credentials_email_inet.txt",
        description=(
            "Carga un archivo .xlsx con encabezados usuario, mail y Nombre "
            "del Centro. Ademas del acceso, el correo incluye la capacitacion "
            "virtual y el video de referencia para INET. Si mail esta vacio, "
            "se usa el mail del usuario."
        ),
    ),
}


def _normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    normalized = normalized.lower().replace(" ", "_").replace("-", "_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_bulk_credentials_send_type_config(
    send_type: str | None = None,
) -> BulkCredentialsSendTypeConfig:
    normalized_send_type = (
        str(send_type or DEFAULT_BULK_CREDENTIALS_SEND_TYPE).strip().lower()
    )
    config = BULK_CREDENTIALS_SEND_TYPES.get(normalized_send_type)
    if not config:
        raise ValidationError("El tipo de envio seleccionado no es valido.")
    return config


def get_bulk_credentials_send_type_choices() -> list[tuple[str, str]]:
    return [
        (config.key, config.label) for config in BULK_CREDENTIALS_SEND_TYPES.values()
    ]


def get_bulk_credentials_send_type_contexts() -> list[dict[str, object]]:
    contexts = []
    for config in BULK_CREDENTIALS_SEND_TYPES.values():
        contexts.append(
            {
                "key": config.key,
                "label": config.label,
                "required_columns": config.template_headers,
                "description": config.description,
            }
        )
    return contexts


def _build_login_url(*, request=None) -> str:
    try:
        path = reverse("login")
    except Exception:
        login_setting = str(getattr(settings, "LOGIN_URL", "") or "").strip()
        path = login_setting if login_setting.startswith("/") else "/"
    if request is not None:
        scheme = "https" if request.is_secure() else "http"
        domain = request.get_host()
    else:
        domain = (
            str(settings.DOMINIO)
            .replace("http://", "")
            .replace("https://", "")
            .rstrip("/")
        )
        scheme = "https" if settings.ENVIRONMENT == "prd" else "http"
    return f"{scheme}://{domain}{path}"


def _build_header_map(
    headers: list[str],
    *,
    send_type_config: BulkCredentialsSendTypeConfig,
) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if (
            normalized in send_type_config.required_columns
            and normalized not in header_map
        ):
            header_map[normalized] = index

    missing = [
        column
        for column in send_type_config.required_columns
        if column not in header_map
    ]
    if missing:
        missing_text = ", ".join(missing)
        raise ValidationError(
            "El archivo debe incluir las columnas obligatorias: " f"{missing_text}."
        )

    return header_map


def _load_workbook_rows(
    uploaded_file,
    *,
    send_type_config: BulkCredentialsSendTypeConfig,
) -> list[ParsedCredentialRow]:
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc
    try:
        if SHEET_NAME in workbook.sheetnames:
            worksheet = workbook[SHEET_NAME]
        else:
            worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("El archivo Excel esta vacio.")

        headers = [_clean_cell(value) for value in rows[0]]
        header_map = _build_header_map(headers, send_type_config=send_type_config)

        parsed_rows: list[ParsedCredentialRow] = []
        for row_number, row in enumerate(rows[1:], start=2):
            values = {
                column: _clean_cell(row[index] if index < len(row) else "")
                for column, index in header_map.items()
            }
            if not any(values.values()):
                continue

            parsed_rows.append(
                ParsedCredentialRow(
                    fila=row_number,
                    data=values,
                )
            )

        if not parsed_rows:
            raise ValidationError(
                "El archivo Excel no contiene filas con datos para procesar."
            )

        return parsed_rows
    finally:
        workbook.close()


def validate_bulk_credentials_workbook(
    uploaded_file,
    *,
    send_type: str | None = None,
) -> BulkCredentialsSendTypeConfig:
    send_type_config = get_bulk_credentials_send_type_config(send_type)
    try:
        uploaded_file.seek(0)
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValidationError("No se pudo leer el archivo Excel cargado.") from exc

    try:
        worksheet = (
            workbook[SHEET_NAME]
            if SHEET_NAME in workbook.sheetnames
            else workbook.active
        )
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise ValidationError("El archivo Excel esta vacio.") from exc

        headers = [_clean_cell(value) for value in header_row]
        _build_header_map(headers, send_type_config=send_type_config)

        for row in rows:
            values = [_clean_cell(value) for value in row]
            if any(values):
                uploaded_file.seek(0)
                return send_type_config

        raise ValidationError(
            "El archivo Excel no contiene filas con datos para procesar."
        )
    finally:
        workbook.close()


def generate_bulk_credentials_template(send_type: str | None = None) -> bytes:
    send_type_config = get_bulk_credentials_send_type_config(send_type)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.append(list(send_type_config.template_headers))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def get_bulk_credentials_template_filename(send_type: str | None = None) -> str:
    send_type_config = get_bulk_credentials_send_type_config(send_type)
    return send_type_config.template_filename


def _send_bulk_credentials_email_once(
    *,
    recipient_email: str,
    entries: list[BulkCredentialEntry],
    login_url: str,
    send_type: str | None = None,
) -> None:
    send_type_config = get_bulk_credentials_send_type_config(send_type)
    # nombre_del_centro y first_name/last_name del primer entry se exponen al
    # template para mantener compatibilidad con los placeholders existentes
    # cuando la lista tiene un solo elemento.
    first = entries[0]
    context = {
        "entries": entries,
        "login_url": login_url,
        "is_grouped": len(entries) > 1,
        # Compatibilidad: campos del primer (o único) usuario.
        "user_username": first.username,
        "user_full_name": first.full_name,
        "plain_password": first.plain_password,
        "nombre_del_centro": first.nombre_del_centro,
    }
    message = render_to_string(send_type_config.email_template_name, context)
    send_mail(
        subject=send_type_config.email_subject,
        message=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[recipient_email],
        fail_silently=False,
    )


EMAIL_ERROR_MESSAGES = (
    (
        (BulkCredentialsEmailTimeoutError, TimeoutError),
        "Timeout al comunicarse con el servidor de correo.",
    ),
    (
        smtplib.SMTPAuthenticationError,
        "El servidor de correo rechazo la autenticacion.",
    ),
    (smtplib.SMTPConnectError, "No se pudo conectar al servidor de correo."),
    (smtplib.SMTPServerDisconnected, "El servidor de correo cerro la conexion."),
    (
        smtplib.SMTPRecipientsRefused,
        "El servidor de correo rechazo el destinatario indicado.",
    ),
    (smtplib.SMTPDataError, "El servidor de correo rechazo el mensaje enviado."),
    (smtplib.SMTPException, "Ocurrio un error del servidor de correo."),
    (OSError, "Ocurrio un error de red al enviar el correo."),
)


def build_bulk_credentials_error_message(exc: Exception) -> str:
    if isinstance(exc, ValidationError):
        return " ".join(exc.messages)
    for error_types, message in EMAIL_ERROR_MESSAGES:
        if isinstance(exc, error_types):
            return message
    return "Ocurrio un error inesperado al procesar la fila."


def _build_email_retry_failure_message(last_error: Exception | None) -> str:
    base_message = build_bulk_credentials_error_message(
        last_error or ValidationError("No se pudo enviar el correo.")
    )
    return f"{base_message} El envio se reintento sin exito."


def send_bulk_credentials_email(
    *,
    recipient_email: str,
    entries: list[BulkCredentialEntry],
    login_url: str,
    send_type: str | None = None,
    max_total_seconds: float | None = None,
) -> None:
    if not entries:
        raise ValidationError("No hay credenciales para enviar.")
    timeout_seconds = _get_bulk_credentials_email_attempt_timeout_seconds()
    total_deadline = (
        time.monotonic() + max_total_seconds
        if max_total_seconds is not None and max_total_seconds > 0
        else None
    )
    last_error = None
    usernames_log = ",".join(entry.username for entry in entries)

    for attempt in range(1, BULK_CREDENTIALS_EMAIL_MAX_ATTEMPTS + 1):
        remaining_total_seconds = _get_remaining_processing_seconds(total_deadline)
        if remaining_total_seconds is not None and remaining_total_seconds < 1:
            raise ValidationError(_get_batch_timeout_message())

        attempt_timeout_seconds = timeout_seconds
        if remaining_total_seconds is not None:
            attempt_timeout_seconds = min(
                timeout_seconds,
                max(1, int(remaining_total_seconds)),
            )

        try:
            with _mail_send_timeout_guard(attempt_timeout_seconds):
                _send_bulk_credentials_email_once(
                    recipient_email=recipient_email,
                    entries=entries,
                    login_url=login_url,
                    send_type=send_type,
                )
            return
        except (
            BulkCredentialsEmailTimeoutError,
            TimeoutError,
            smtplib.SMTPException,
            OSError,
        ) as exc:
            last_error = exc
            logger.warning(
                (
                    "Fallo enviando credenciales por correo. "
                    "usuarios=%s intento=%s/%s tipo=%s"
                ),
                usernames_log,
                attempt,
                BULK_CREDENTIALS_EMAIL_MAX_ATTEMPTS,
                send_type or DEFAULT_BULK_CREDENTIALS_SEND_TYPE,
                exc_info=True,
            )
            if attempt < BULK_CREDENTIALS_EMAIL_MAX_ATTEMPTS:
                remaining_total_seconds = _get_remaining_processing_seconds(
                    total_deadline
                )
                if (
                    remaining_total_seconds is not None
                    and remaining_total_seconds
                    <= BULK_CREDENTIALS_EMAIL_RETRY_BACKOFF_SECONDS + 1
                ):
                    break
                time.sleep(BULK_CREDENTIALS_EMAIL_RETRY_BACKOFF_SECONDS)

    raise ValidationError(
        _build_email_retry_failure_message(last_error)
    ) from last_error


def _validate_row_data(
    row: ParsedCredentialRow,
    *,
    send_type_config: BulkCredentialsSendTypeConfig,
) -> None:
    for column in send_type_config.required_columns:
        if column == "mail":
            continue
        if row.data.get(column):
            continue
        raise ValidationError(f"La columna {column} es obligatoria.")


def _get_recipient_email(*, row: ParsedCredentialRow, user) -> str:
    recipient_email = row.mail.strip()
    if not recipient_email:
        recipient_email = (user.email or "").strip()
        if not recipient_email:
            raise ValidationError(
                "La fila no informa mail y el usuario no tiene un mail cargado."
            )

    try:
        validate_email(recipient_email)
    except ValidationError as exc:
        if row.mail.strip():
            raise ValidationError(
                "El formato del mail informado en la planilla es invalido."
            ) from exc
        raise ValidationError(
            "La fila no informa mail y el usuario tiene un mail invalido cargado."
        ) from exc
    return recipient_email


def _get_user_plain_password(user) -> str:
    profile = getattr(user, "profile", None)
    plain_password = (
        getattr(profile, "temporary_password_plaintext", "") or ""
    ).strip()
    if not plain_password:
        raise ValidationError(
            (
                "El usuario no tiene una contraseña temporal visible para enviar. "
                "Genere o cargue una nueva contraseña temporal antes de reintentar."
            )
        )
    return plain_password


def _resolve_row_for_send(
    row: ParsedCredentialRow,
    *,
    send_type_config: BulkCredentialsSendTypeConfig,
):
    """Valida un row y resuelve user, destinatario y entry para el envío.

    Levanta ValidationError si falta data, no existe el usuario, no hay mail
    válido o el usuario no tiene contraseña temporal visible.
    """
    _validate_row_data(row, send_type_config=send_type_config)
    user = (
        User.objects.select_related("profile")
        .filter(username__iexact=row.usuario)
        .first()
    )
    if not user:
        raise ValidationError("No existe un usuario con ese nombre.")
    recipient_email = _get_recipient_email(row=row, user=user)
    plain_password = _get_user_plain_password(user)
    entry = BulkCredentialEntry(
        username=user.username,
        plain_password=plain_password,
        first_name=user.first_name or "",
        last_name=user.last_name or "",
        nombre_del_centro=row.nombre_del_centro,
    )
    return user, recipient_email, entry


def _row_success_result(
    *, row: ParsedCredentialRow, recipient_email: str, grouped: bool
):
    base_message = "Credenciales enviadas correctamente."
    if grouped:
        base_message = (
            "Credenciales enviadas correctamente "
            "(agrupadas con otras del mismo destinatario)."
        )
    return {
        "fila": row.fila,
        "usuario": row.usuario,
        "mail_destino": recipient_email,
        "estado": "enviada",
        "mensaje": base_message,
        "password_actualizada": False,
    }


def process_bulk_credentials_group(
    *,
    rows: list[ParsedCredentialRow],
    send_type_config: BulkCredentialsSendTypeConfig,
    login_url: str,
    max_total_seconds: float | None = None,
) -> list[dict[str, object]]:
    """Envía un único correo con las credenciales de todas las filas del grupo.

    Todas las filas deben compartir destinatario (ya resuelto). Si alguna falla
    validación se levanta ValidationError sin enviar el correo: el caller debe
    decidir cómo registrar el fallo por fila.
    """
    if not rows:
        raise ValidationError("No hay filas para procesar en el grupo.")

    resolved: list[tuple[ParsedCredentialRow, BulkCredentialEntry]] = []
    recipient_email: str | None = None
    with transaction.atomic():
        for row in rows:
            _user, row_recipient, entry = _resolve_row_for_send(
                row, send_type_config=send_type_config
            )
            if recipient_email is None:
                recipient_email = row_recipient
            elif row_recipient.lower() != recipient_email.lower():
                raise ValidationError(
                    "Las filas del grupo no comparten el mismo destinatario."
                )
            resolved.append((row, entry))

        send_bulk_credentials_email(
            recipient_email=recipient_email,
            entries=[entry for _row, entry in resolved],
            login_url=login_url,
            send_type=send_type_config.key,
            max_total_seconds=max_total_seconds,
        )

    grouped = len(resolved) > 1
    return [
        _row_success_result(row=row, recipient_email=recipient_email, grouped=grouped)
        for row, _entry in resolved
    ]


def process_bulk_credentials_row(
    *,
    row: ParsedCredentialRow,
    send_type_config: BulkCredentialsSendTypeConfig,
    login_url: str,
    max_total_seconds: float | None = None,
) -> dict[str, object]:
    """Procesa una sola fila sin agrupamiento. Mantenido para compatibilidad."""
    results = process_bulk_credentials_group(
        rows=[row],
        send_type_config=send_type_config,
        login_url=login_url,
        max_total_seconds=max_total_seconds,
    )
    return results[0]


def _build_recipient_cache(
    rows: list[ParsedCredentialRow],
) -> dict[str, str]:
    """Pre-carga `username -> email` en una sola query para evitar N lookups.

    Solo se incluyen los usuarios efectivamente referenciados por filas que no
    informan mail explícito. Devuelve un dict indexado por username en minúsculas.
    """
    usernames_needed = {
        row.usuario.strip().lower()
        for row in rows
        if row.usuario and not row.mail.strip()
    }
    if not usernames_needed:
        return {}
    user_qs = User.objects.filter(username__in=list(usernames_needed)).only(
        "username", "email"
    )
    cache = {
        (u.username or "").strip().lower(): (u.email or "").strip().lower()
        for u in user_qs
    }
    return cache


def _row_grouping_key(
    row: ParsedCredentialRow,
    *,
    send_type_config: BulkCredentialsSendTypeConfig,
    recipient_cache: dict[str, str] | None = None,
) -> tuple[str, str] | None:
    """Devuelve la clave de agrupamiento (destinatario, centro) para una fila.

    Si el send_type requiere `nombre_del_centro`, el centro forma parte de la
    clave: dos filas con mismo destinatario pero distinto centro NO se agrupan
    (evita mensajes mezclando datos de centros). Si la fila no puede resolver
    destinatario, retorna None y el row se procesa solo.
    """
    if not row.data.get("usuario"):
        return None

    explicit_mail = row.mail.strip()
    if explicit_mail:
        recipient = explicit_mail.lower()
    else:
        username_key = row.usuario.strip().lower()
        cached = (recipient_cache or {}).get(username_key)
        if cached is None and recipient_cache is None:
            user = (
                User.objects.filter(username__iexact=row.usuario).only("email").first()
            )
            cached = (user.email or "").strip().lower() if user else ""
        recipient = cached or ""

    if not recipient:
        return None

    requires_centro = "nombre_del_centro" in send_type_config.required_columns
    centro_key = ""
    if requires_centro:
        # Normalizamos para que diferencias de espacios/caso no fragmenten el grupo.
        centro_key = " ".join(row.nombre_del_centro.lower().split())

    return recipient, centro_key


def _peek_recipient_email(
    row: ParsedCredentialRow,
    *,
    send_type_config: BulkCredentialsSendTypeConfig,
    recipient_cache: dict[str, str] | None = None,
) -> str | None:
    """Compatibilidad: devuelve solo el destinatario (sin el componente centro)."""
    key = _row_grouping_key(
        row,
        send_type_config=send_type_config,
        recipient_cache=recipient_cache,
    )
    return key[0] if key else None


def _collect_group_indices(
    *,
    rows: list[ParsedCredentialRow],
    start_index: int,
    send_type_config: BulkCredentialsSendTypeConfig,
    skip_indices: set[int],
    recipient_cache: dict[str, str] | None = None,
) -> tuple[list[int], str | None]:
    """Identifica los índices de filas que comparten clave de agrupamiento con
    rows[start_index]. Para INET, la clave incluye `nombre_del_centro`.

    Retorna la lista de índices (incluido start_index) y el destinatario común.
    Si el row inicial no puede resolver clave, retorna ([start_index], None).
    """
    primary_key = _row_grouping_key(
        rows[start_index],
        send_type_config=send_type_config,
        recipient_cache=recipient_cache,
    )
    if primary_key is None:
        return [start_index], None

    primary_recipient = primary_key[0]
    group = [start_index]
    for index in range(start_index + 1, len(rows)):
        if index in skip_indices:
            continue
        candidate_key = _row_grouping_key(
            rows[index],
            send_type_config=send_type_config,
            recipient_cache=recipient_cache,
        )
        if candidate_key == primary_key:
            group.append(index)
    return group, primary_recipient


def process_bulk_credentials_file(
    *,
    uploaded_file,
    send_type: str | None = None,
    request=None,
) -> dict:
    send_type_config = get_bulk_credentials_send_type_config(send_type)
    rows = _load_workbook_rows(
        uploaded_file,
        send_type_config=send_type_config,
    )
    results: list[dict[str, object]] = []
    summary = {
        "procesadas": 0,
        "enviadas": 0,
        "actualizadas": 0,
        "sin_cambios": 0,
        "rechazadas": 0,
    }
    login_url = _build_login_url(request=request)
    processing_deadline = _get_bulk_credentials_batch_deadline()
    results_by_index: dict[int, dict[str, object]] = {}
    handled_indices: set[int] = set()
    recipient_cache = _build_recipient_cache(rows)

    for row_index, row in enumerate(rows):
        if row_index in handled_indices:
            continue
        if not _has_enough_batch_time(processing_deadline):
            timeout_message = _get_batch_timeout_message()
            for pending_index in range(row_index, len(rows)):
                if pending_index in handled_indices:
                    continue
                pending_row = rows[pending_index]
                summary["procesadas"] += 1
                summary["rechazadas"] += 1
                results_by_index[pending_index] = {
                    "fila": pending_row.fila,
                    "usuario": pending_row.usuario,
                    "mail_destino": pending_row.mail,
                    "estado": "rechazada",
                    "mensaje": timeout_message,
                    "password_actualizada": False,
                }
                handled_indices.add(pending_index)
            break

        group_indices, _recipient = _collect_group_indices(
            rows=rows,
            start_index=row_index,
            send_type_config=send_type_config,
            skip_indices=handled_indices,
            recipient_cache=recipient_cache,
        )
        group_rows = [rows[i] for i in group_indices]

        try:
            group_results = process_bulk_credentials_group(
                rows=group_rows,
                send_type_config=send_type_config,
                login_url=login_url,
                max_total_seconds=_get_remaining_processing_seconds(
                    processing_deadline
                ),
            )
            for idx, group_result in zip(group_indices, group_results):
                summary["procesadas"] += 1
                if group_result.get("password_actualizada"):
                    summary["actualizadas"] += 1
                else:
                    summary["sin_cambios"] += 1
                summary["enviadas"] += 1
                results_by_index[idx] = group_result
                handled_indices.add(idx)
        except ValidationError as exc:
            message = build_bulk_credentials_error_message(exc)
            for idx in group_indices:
                fail_row = rows[idx]
                summary["procesadas"] += 1
                summary["rechazadas"] += 1
                results_by_index[idx] = {
                    "fila": fail_row.fila,
                    "usuario": fail_row.usuario,
                    "mail_destino": fail_row.mail,
                    "estado": "rechazada",
                    "mensaje": message,
                    "password_actualizada": False,
                }
                handled_indices.add(idx)
        except Exception as exc:
            message = build_bulk_credentials_error_message(exc)
            logger.exception(
                ("Fallo procesando envio masivo de credenciales. " "tipo=%s filas=%s"),
                send_type_config.key,
                [rows[i].fila for i in group_indices],
            )
            for idx in group_indices:
                fail_row = rows[idx]
                summary["procesadas"] += 1
                summary["rechazadas"] += 1
                results_by_index[idx] = {
                    "fila": fail_row.fila,
                    "usuario": fail_row.usuario,
                    "mail_destino": fail_row.mail,
                    "estado": "rechazada",
                    "mensaje": message,
                    "password_actualizada": False,
                }
                handled_indices.add(idx)

    for idx in range(len(rows)):
        if idx in results_by_index:
            results.append(results_by_index[idx])

    return {
        "summary": summary,
        "rows": results,
        "filename": getattr(uploaded_file, "name", ""),
        "send_type": send_type_config.key,
        "send_type_label": send_type_config.label,
    }
