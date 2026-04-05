from __future__ import annotations

import csv
import unicodedata
from dataclasses import dataclass
from datetime import timedelta
from pathlib import Path

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify
from django.utils import timezone
from openpyxl import load_workbook

from users.models import Profile
from users.services_group_permissions import sync_permissions_for_group

GROUP_NAME = "CFP"
DEFAULT_LAST_NAME = "CFP"
FANTASY_EMAIL_DOMAIN = "vat.local"
HEADER_ALIASES = {
    "username": {"usuario", "username", "user", "legajo"},
    "email": {"email", "correo", "mail", "e_mail"},
    "display_name": {
        "nombre",
        "nombre_centro",
        "nombre_cfp",
        "institucion",
        "establecimiento",
        "centro",
    },
    "first_name": {
        "nombre_completo",
        "nombre_usuario",
        "nombres",
        "first_name",
        "name",
    },
    "last_name": {
        "apellido",
        "apellidos",
        "last_name",
        "lastname",
        "surname",
    },
    "password": {"contrasena", "contraseña", "password", "clave"},
    "rol": {"rol", "role", "perfil"},
}
REQUIRED_FIELDS = ()
STOPWORDS = {
    "de",
    "del",
    "la",
    "las",
    "el",
    "los",
    "y",
    "para",
}
TOKEN_REPLACEMENTS = {
    "escuela": "esc",
    "municipal": "mun",
    "formacion": "form",
    "profesional": "prof",
    "centro": "ctr",
    "educacion": "edu",
    "tecnica": "tec",
    "tecnico": "tec",
    "capacitacion": "cap",
    "trabajo": "trab",
    "adultos": "adult",
    "adulto": "adult",
    "jovenes": "jov",
    "joven": "jov",
    "laboral": "lab",
    "general": "gral",
    "ingeniero": "ing",
}
SPECIAL_PATTERN_REPLACEMENTS = {
    "c.f.p.": "cfp",
    "c.f.p": "cfp",
    "cfp": "cfp",
    "c.e.j.a.": "ceja",
    "c.e.j.a": "ceja",
    "ceja": "ceja",
    "c.c.t.": "cct",
    "c.c.t": "cct",
    "cct": "cct",
    "e.e.t.": "eet",
    "e.e.t": "eet",
    "eet": "eet",
    "nº": "n",
    "n°": "n",
    "no": "n",
}
MAX_USERNAME_LENGTH = 24


@dataclass(frozen=True)
class ParsedRow:
    line_number: int
    username: str
    email: str
    display_name: str
    first_name: str
    last_name: str
    password: str
    rol: str


def normalize_header(value: object) -> str:
    text = str(value or "").strip()
    normalized = unicodedata.normalize("NFKD", text)
    normalized = "".join(
        character
        for character in normalized
        if not unicodedata.combining(character)
    )
    normalized = normalized.lower().replace(" ", "_").replace("-", "_")
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def resolve_header_mapping(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        normalized = normalize_header(header)
        for field_name, aliases in HEADER_ALIASES.items():
            if normalized in aliases and field_name not in mapping:
                mapping[field_name] = index
                break

    missing_fields = [field for field in REQUIRED_FIELDS if field not in mapping]
    if missing_fields:
        missing = ", ".join(missing_fields)
        raise CommandError(
            "El archivo no contiene las columnas requeridas: "
            f"{missing}. Encabezados detectados: {', '.join(headers)}"
        )

    return mapping


def split_full_name(full_name: str) -> tuple[str, str]:
    parts = [part for part in full_name.split() if part]
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return " ".join(parts[:-1]), parts[-1]


def normalize_name_for_username(name: str) -> str:
    normalized = str(name or "").lower()
    for original, replacement in SPECIAL_PATTERN_REPLACEMENTS.items():
        normalized = normalized.replace(original, f" {replacement} ")
    normalized = normalized.replace("n ", " n")
    normalized = slugify(normalized).replace("-", " ")
    return normalized


def shorten_token(token: str) -> str:
    if token == "n":
        return ""
    if token.isdigit():
        return token
    if token.startswith("n") and token[1:].isdigit():
        return token[1:]
    replacement = TOKEN_REPLACEMENTS.get(token)
    if replacement:
        return replacement
    if len(token) <= 4:
        return token
    return token[:4]


def build_username_base(display_name: str) -> str:
    normalized_name = normalize_name_for_username(display_name)
    tokens = []
    for raw_token in normalized_name.split():
        if raw_token in STOPWORDS:
            continue
        short_token = shorten_token(raw_token)
        if short_token:
            tokens.append(short_token)

    if not tokens:
        return "cfpuser"

    username = ""
    for token in tokens[:5]:
        next_value = f"{username}{token}"
        if len(next_value) > MAX_USERNAME_LENGTH:
            break
        username = next_value

    return username or "cfpuser"


def build_batch_safe_username(
    *,
    base_username: str,
    batch_usernames: set[str],
    user_model,
) -> str:
    candidate = base_username[:MAX_USERNAME_LENGTH] or "cfpuser"
    counter = 1

    while candidate in batch_usernames:
        suffix = str(counter)
        trimmed_base = base_username[: MAX_USERNAME_LENGTH - len(suffix)] or "cfpuser"
        candidate = f"{trimmed_base}{suffix}"
        counter += 1

    batch_usernames.add(candidate)
    return candidate


def build_fantasy_email(username: str) -> str:
    return f"{username}@{FANTASY_EMAIL_DOMAIN}"


def resolve_generated_username(
    *,
    parsed_row: ParsedRow,
    batch_usernames: set[str],
    user_model,
) -> str:
    base_username = build_username_base(parsed_row.display_name)
    desired_email = parsed_row.email
    candidate = base_username[:MAX_USERNAME_LENGTH] or "cfpuser"
    counter = 1

    while True:
        if candidate in batch_usernames:
            suffix = str(counter)
            trimmed_base = base_username[: MAX_USERNAME_LENGTH - len(suffix)] or "cfpuser"
            candidate = f"{trimmed_base}{suffix}"
            counter += 1
            continue

        existing_user = user_model.objects.filter(username=candidate).first()
        candidate_email = desired_email or build_fantasy_email(candidate)
        if existing_user is None or existing_user.email == candidate_email:
            batch_usernames.add(candidate)
            return candidate

        suffix = str(counter)
        trimmed_base = base_username[: MAX_USERNAME_LENGTH - len(suffix)] or "cfpuser"
        candidate = f"{trimmed_base}{suffix}"
        counter += 1


def iter_csv_rows(file_path: Path) -> tuple[list[str], list[tuple[int, list[str]]]]:
    with file_path.open(newline="", encoding="utf-8-sig") as csv_file:
        reader = csv.reader(csv_file)
        rows = list(reader)

    if not rows:
        raise CommandError("El archivo CSV está vacío.")

    headers = [clean_cell(value) for value in rows[0]]
    data_rows = [
        (index, [clean_cell(value) for value in row])
        for index, row in enumerate(rows[1:], start=2)
    ]
    return headers, data_rows


def iter_excel_rows(
    file_path: Path, sheet_name: str | None
) -> tuple[list[str], list[tuple[int, list[str]]]]:
    workbook = load_workbook(filename=file_path, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise CommandError(
                f"La hoja '{sheet_name}' no existe. Hojas disponibles: "
                f"{', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise CommandError("El archivo Excel está vacío.")

    headers = [clean_cell(value) for value in rows[0]]
    data_rows = [
        (index, [clean_cell(value) for value in row])
        for index, row in enumerate(rows[1:], start=2)
    ]
    return headers, data_rows


def load_rows(
    file_path: Path, sheet_name: str | None
) -> tuple[list[str], list[tuple[int, list[str]]]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return iter_csv_rows(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return iter_excel_rows(file_path, sheet_name)
    raise CommandError(
        "Formato no soportado. Use un archivo .csv, .xlsx o .xlsm."
    )


def build_parsed_row(
    *,
    line_number: int,
    raw_values: list[str],
    header_mapping: dict[str, int],
    default_password: str,
) -> ParsedRow:
    def get_value(field_name: str) -> str:
        index = header_mapping.get(field_name)
        if index is None or index >= len(raw_values):
            return ""
        return raw_values[index]

    username = get_value("username")
    email = get_value("email")
    display_name = get_value("display_name")
    first_name = get_value("first_name")
    last_name = get_value("last_name")
    password = get_value("password") or default_password
    rol = get_value("rol") or GROUP_NAME

    if not display_name:
        full_name = " ".join(part for part in (first_name, last_name) if part).strip()
        display_name = full_name

    if not first_name and display_name:
        first_name = display_name

    if first_name and not last_name and " " in first_name:
        first_name, _ = split_full_name(first_name)

    last_name = DEFAULT_LAST_NAME

    if not password:
        raise ValueError(
            "debe existir una columna de contraseña o indicar --default-password"
        )
    if not username and not display_name:
        raise ValueError(
            "debe existir la columna Usuario o una columna nombre para generar el username"
        )

    return ParsedRow(
        line_number=line_number,
        username=username,
        email=email,
        display_name=display_name,
        first_name=first_name,
        last_name=last_name,
        password=password,
        rol=rol,
    )


class Command(BaseCommand):
    """Importa usuarios VAT con rol CFP desde un archivo tabular."""

    help = (
        "Crea o actualiza usuarios VAT con grupo CFP a partir de un archivo "
        "CSV o Excel."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "file_path",
            type=str,
            help="Ruta al archivo .csv, .xlsx o .xlsm con los usuarios.",
        )
        parser.add_argument(
            "--sheet-name",
            type=str,
            default=None,
            help="Nombre de la hoja a procesar cuando el archivo es Excel.",
        )
        parser.add_argument(
            "--default-password",
            type=str,
            default="",
            help=(
                "Contraseña temporal a usar cuando la fila no trae una columna "
                "de contraseña."
            ),
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Valida el archivo e informa el resultado sin guardar cambios.",
        )

    def handle(self, *args, **options):
        file_path = Path(options["file_path"])
        sheet_name = options["sheet_name"]
        default_password = (options.get("default_password") or "").strip()
        dry_run = options["dry_run"]

        if not file_path.exists():
            raise CommandError(f"El archivo '{file_path}' no existe.")

        headers, raw_rows = load_rows(file_path, sheet_name)
        header_mapping = resolve_header_mapping(headers)
        user_model = get_user_model()
        batch_usernames: set[str] = set()

        created_count = 0
        updated_count = 0
        skipped_count = 0
        processed_count = 0

        with transaction.atomic():
            group, _ = Group.objects.get_or_create(name=GROUP_NAME)
            sync_permissions_for_group(group)

            for line_number, raw_values in raw_rows:
                if not any(raw_values):
                    continue

                try:
                    parsed_row = build_parsed_row(
                        line_number=line_number,
                        raw_values=raw_values,
                        header_mapping=header_mapping,
                        default_password=default_password,
                    )
                except ValueError as exc:
                    skipped_count += 1
                    self.stdout.write(
                        self.style.WARNING(
                            f"Fila {line_number}: {exc}. Se omite."
                        )
                    )
                    continue

                username = parsed_row.username
                if not username:
                    username = resolve_generated_username(
                        parsed_row=parsed_row,
                        batch_usernames=batch_usernames,
                        user_model=user_model,
                    )
                else:
                    batch_usernames.add(username)

                email = parsed_row.email or build_fantasy_email(username)

                user, created = user_model.objects.get_or_create(
                    username=username,
                    defaults={"email": email},
                )

                user.email = email
                user.first_name = parsed_row.first_name
                user.last_name = parsed_row.last_name
                user.is_active = True
                user.set_password(parsed_row.password)

                if not dry_run:
                    user.save()

                profile, _ = Profile.objects.get_or_create(user=user)
                profile.rol = parsed_row.rol
                profile.must_change_password = True
                profile.password_changed_at = None
                profile.initial_password_expires_at = timezone.now() + timedelta(
                    hours=settings.INITIAL_PASSWORD_MAX_AGE_HOURS
                )
                profile.password_reset_requested_at = None
                profile.temporary_password_plaintext = parsed_row.password

                if not dry_run:
                    profile.save(
                        update_fields=[
                            "rol",
                            "must_change_password",
                            "password_changed_at",
                            "initial_password_expires_at",
                            "password_reset_requested_at",
                            "temporary_password_plaintext",
                        ]
                    )
                    user.groups.add(group)

                processed_count += 1
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            if dry_run:
                transaction.set_rollback(True)

        mode = "Simulación finalizada" if dry_run else "Proceso finalizado"
        self.stdout.write(
            self.style.SUCCESS(
                f"{mode}. Procesados: {processed_count}, creados: {created_count}, "
                f"actualizados: {updated_count}, omitidos: {skipped_count}."
            )
        )